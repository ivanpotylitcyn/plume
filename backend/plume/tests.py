"""Юнит-тесты движка волны 1 — гарантия корректности формул (вместо прод-обкатки).

Каждый тест строит минимальный сценарий и проверяет одну формулу.
"""
import json
import os
import shutil
import tempfile
from decimal import Decimal
from urllib.parse import quote

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, TestCase, override_settings

from plume import admin
from plume import models
from plume import engine
from plume import views


def D(x):
    return Decimal(str(x))


def _cat(code='test', description='Тест'):
    """Категория-заглушка для тестов (волна 15: `Item.category` — обязательный FK).
    Класс изделия в движке логику не ветвит, поэтому одной общей категории хватает."""
    c, _ = models.Category.objects.get_or_create(code=code, defaults={'description': description})
    return c


# Изолированный MEDIA_ROOT для тестов вложений (волна 11): загрузки не пачкают
# рабочий backend/media; чистим на выходе модуля.
_TEST_MEDIA = tempfile.mkdtemp(prefix='plume-test-media-')


def tearDownModule():
    shutil.rmtree(_TEST_MEDIA, ignore_errors=True)


class EngineTestBase(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create(username='t')
        self.main = models.Location.objects.create(code='MAIN', description='Основной склад')
        self.prj = models.Project.objects.create(
            code='P1', description='Проект 1', kind=models.Project.Kind.EXTERNAL)
        self.supplier = models.Counterparty.objects.create(description='Поставщик')

    def make_item(self, code, manufactured=False, kind=None):
        # `kind` — исторический хинт (движок по классу не ветвит); категория —
        # общая заглушка `_cat()`. `manufactured` → ось `native` (волна 15).
        return models.Item.objects.create(
            code=code, description=code, category=_cat(),
            native=manufactured)

    def make_purchase(self, project=None, contractor=None, **kw):
        """Заказ-черновик, готовый к фиксации: **с контрагентом**.

        Ф17: контрагент обязателен к фиксации заказа («у кого купили»), поэтому фикстура
        «настоящий заказ» его несёт. Тесты самого гейта заводят заказ голым
        `engine.create_purchase` и контрагента не ставят.
        """
        p = engine.create_purchase(project or self.prj, self.user, **kw)
        return engine.update_purchase(p, contractor=contractor or self.supplier)

    def receipt_lot(self, item, project, qty, purchase=None, locked=True):
        """Партия, приехавшая по УПД. Волна 19, Ф15: поставка **зафиксирована** —
        иначе её партия на складе не лежит (замок гейтит склад), а фикстура значит
        именно «товар на складе». `locked=False` — для тестов самого гейта."""
        r = models.Receipt.objects.create(
            number=f'UPD-{item.code}-{qty}', date='2026-05-01', contractor=self.supplier,
            project=project, user=self.user, purchase=purchase, locked=locked)
        lot = models.Lot.objects.create(item=item, project=project, origin=r, qty=D(qty))
        engine.rebuild_movements(lot)
        return lot


class RebuildAndStockTests(EngineTestBase):
    def test_receipt_lot_live_qty(self):
        lot = self.receipt_lot(self.make_item('A'), self.prj, 10)
        self.assertEqual(engine.lot_live_qty(lot), D(10))
        self.assertEqual(lot.movements.count(), 1)

    def test_kitting_issue_reduces_qty(self):
        comp = self.make_item('R')
        lot = self.receipt_lot(comp, self.prj, 100)
        dev = self.make_item('DEV', manufactured=True)
        k = models.Kitting.objects.create(project=self.prj, target_item=dev,
                                          user=self.user, qty=D(1),
                                          locked=True)     # Ф15: спаяно = зафиксировано
        models.StockLine.objects.create(document=k, lot=lot,
                                        location=self.main, qty=D(-30))
        engine.rebuild_movements(lot)
        self.assertEqual(engine.lot_live_qty(lot), D(70))

    def test_available_can_be_negative(self):
        comp = self.make_item('R')
        lot = self.receipt_lot(comp, self.prj, 5)
        dev = self.make_item('DEV', manufactured=True)
        k = models.Kitting.objects.create(project=self.prj, target_item=dev,
                                          user=self.user, qty=D(1),
                                          locked=True)
        models.StockLine.objects.create(document=k, lot=lot,
                                        location=self.main, qty=D(-8))
        engine.rebuild_movements(lot)
        self.assertEqual(engine.item_available(comp, self.prj), D(-3))
        self.assertTrue(engine.item_has_negative_lot(comp, self.prj))

    def test_reopen_then_delete_leaves_no_phantom(self):
        """Волна 13 Ф1: отмена = удаление (`cancelled` снят), но проведённую
        комплектацию с born-лотом сперва расфиксируют. Расфиксация чисто сносит
        лот-прибор (не фантом), удаление черновика освобождает компоненты.

        (Историческая грабля Ф1: прямое `posted.delete()` на MySQL упиралось в CHECK
        `exactly_one_origin`. В Ф2b дуга схлопнута в один CASCADE-FK `Lot.origin` — CHECK
        умер; замок «сперва расфиксировать» держит прикладной guard `delete_stock_document`.
        Тут проверяем корректный путь reopen→delete. См. JOURNAL 2026-07-09 Ф1/Ф2b.)"""
        comp = self.make_item('R')
        lot = self.receipt_lot(comp, self.prj, 10)
        dev = self.make_item('DEV', manufactured=True)
        k = models.Kitting.objects.create(project=self.prj, target_item=dev,
                                          user=self.user, qty=D(1))
        engine.add_kitting_line(k, comp, lot, D(4))
        born = engine.lock_kitting(k)              # posted + рождается лот-прибор
        self.assertTrue(models.Lot.objects.filter(pk=born.pk).exists())
        engine.unlock_kitting(k)                    # расфиксировать: born-лот снят
        self.assertFalse(models.Lot.objects.filter(pk=born.pk).exists())  # не фантом
        self.assertFalse(k.locked)
        k.delete()                                  # черновик удаляется свободно
        engine.rebuild_movements(lot)               # компонент освобождён (нет −ISSUE)
        self.assertEqual(engine.lot_live_qty(lot), D(10))

    def test_stockline_rebuild_invariant_across_docs(self):
        """Волна 13 Ф0: единая `StockLine` покрывает 4 бывших таблицы строк-расхода.

        Один лот, тронутый разными документами-владельцами (комплектация/списание/
        передача) через знаковые `StockLine`, даёт те же остаток и движения, что и
        прежние раздельные строки — инвариант остатка при консолидации.
        """
        comp = self.make_item('R')
        lot = self.receipt_lot(comp, self.prj, 100)
        dev = self.make_item('DEV', manufactured=True)
        # Ф15: расход виден складу только у зафиксированных документов — фикстура
        # моделирует свершившийся факт, поэтому все три заведены под замком.
        k = models.Kitting.objects.create(project=self.prj, target_item=dev,
                                          user=self.user, qty=D(1),
                                          locked=True)
        w = models.Writeoff.objects.create(project=self.prj, user=self.user,
                                           number='W-1', date='2026-06-01',
                                           locked=True)
        cust = models.Project.objects.create(
            code='P2', description='Проект 2', kind=models.Project.Kind.EXTERNAL)
        t = models.Transfer.objects.create(project=self.prj, user=self.user,
                                            number='T-1', date='2026-06-01',
                                            locked=True)
        # знаковые строки (− расход) трёх разных документов на один лот
        models.StockLine.objects.create(document=k, lot=lot, location=self.main, qty=D(-30))
        models.StockLine.objects.create(document=w, lot=lot, location=self.main, qty=D(-10))
        models.StockLine.objects.create(document=t, lot=lot, location=self.main, qty=D(-5))
        engine.rebuild_movements(lot)
        # 100 − 30 − 10 − 5 = 55; born-приход + три расхода = 4 движения
        self.assertEqual(engine.lot_live_qty(lot), D(55))
        self.assertEqual(lot.movements.count(), 4)
        srcs = set(lot.movements.values_list('source_type', flat=True))
        self.assertEqual(srcs, {'receipt', 'kitting', 'writeoff', 'transfer'})
        # exclusive-arc: строка ссылается ровно на один документ
        sl = models.StockLine.objects.filter(document=k).get()
        self.assertEqual(sl.doc_kind, 'kitting')
        self.assertLess(sl.qty, D(0))            # хранится со знаком (− расход)


class CoverageTests(EngineTestBase):
    def test_triple_split_segments(self):
        cov = engine._coverage(need=D(10), available=D(4), on_order=D(3))
        self.assertEqual(cov['have'], D(4))
        self.assertEqual(cov['on_order'], D(3))
        self.assertEqual(cov['to_order'], D(3))
        self.assertEqual(cov['status'], 'to_order')

    def test_fully_covered_is_available(self):
        cov = engine._coverage(need=D(10), available=D(12), on_order=D(0))
        self.assertEqual(cov['have'], D(10))
        self.assertEqual(cov['to_order'], D(0))
        self.assertEqual(cov['status'], 'available')

    def test_only_ordered_is_on_order(self):
        cov = engine._coverage(need=D(10), available=D(0), on_order=D(10))
        self.assertEqual(cov['status'], 'on_order')

    def test_negative_available_does_not_credit(self):
        cov = engine._coverage(need=D(10), available=D(-5), on_order=D(0))
        self.assertEqual(cov['have'], D(0))
        self.assertEqual(cov['to_order'], D(10))

    def test_worst_and_best_of(self):
        self.assertEqual(engine._worst_of(['available', 'to_order', 'on_order']),
                         'to_order')
        self.assertEqual(engine._best_of(['to_order', 'on_order']), 'on_order')


class BalanceTests(EngineTestBase):
    """Баланс потребности (2026-08-05): четыре сырых члена, дефицит — невязка."""

    def test_four_members_sum_to_balance(self):
        b = engine._balance(need=D(10), kitted=D(2), in_stock=D(3), on_order=D(1))
        self.assertEqual(b['balance'], D(-4))
        self.assertEqual(b['status'], 'to_order')      # красный: не хватает

    def test_surplus_shows_as_plus(self):
        """То, ради чего затевалось: перебор виден знаком, а не прячется за клампом."""
        b = engine._balance(need=D(6), kitted=D(0), in_stock=D(0), on_order=D(10))
        self.assertEqual(b['balance'], D(4))
        self.assertEqual(b['status'], 'available')     # зелёный: запас 4

    def test_exact_zero_is_orange(self):
        b = engine._balance(need=D(6), kitted=D(6), in_stock=D(0), on_order=D(0))
        self.assertEqual(b['balance'], D(0))
        self.assertEqual(b['status'], 'on_order')      # оранжевый: сошлось впритык

    def test_kitted_counts_as_covered(self):
        """Впаянное закрывает потребность: со склада ушло, но в изделии стоит."""
        b = engine._balance(need=D(6), kitted=D(6), in_stock=D(0), on_order=D(0))
        self.assertEqual(b['balance'], D(0))
        self.assertEqual(b['kitted'], D(6))

    def test_supply_status_separates_from_balance_tone(self):
        """Ноль-баланс оранжевый в колонке, но здоровым зелёный в оси снабжения."""
        zero = engine._balance(need=D(6), kitted=D(6), in_stock=D(0), on_order=D(0))
        self.assertEqual(zero['status'], 'on_order')
        self.assertEqual(engine._supply_status(zero), 'available')
        waiting = engine._balance(need=D(6), kitted=D(0), in_stock=D(0), on_order=D(6))
        self.assertEqual(engine._supply_status(waiting), 'on_order')   # часть ещё едет
        short = engine._balance(need=D(6), kitted=D(0), in_stock=D(0), on_order=D(0))
        self.assertEqual(engine._supply_status(short), 'to_order')

    def test_negative_stock_drags_balance_down(self):
        """Недостача не клампится — она и должна утаскивать баланс в минус."""
        b = engine._balance(need=D(10), kitted=D(0), in_stock=D(-5), on_order=D(0))
        self.assertEqual(b['balance'], D(-15))


class OnOrderTests(EngineTestBase):
    def test_purchased_open_order_minus_received(self):
        item = self.make_item('SCR', kind='material')
        # Ф17: заказ самодостаточен (закупка-план опциональна), контрагент обязателен
        # к фиксации — CHECK `purchase_locked_has_contractor`.
        purchase = models.Purchase.objects.create(
            project=self.prj, user=self.user, contractor=self.supplier,
            locked=True)
        models.PurchaseLine.objects.create(purchase=purchase, item=item, qty=D(40))
        # поступило 15 по этому заказу
        self.receipt_lot(item, self.prj, 15, purchase=purchase)
        self.assertEqual(engine.item_on_order(item, self.prj), D(25))

    def test_draft_purchase_not_counted(self):
        item = self.make_item('SCR', kind='material')
        purchase = models.Purchase.objects.create(
            project=self.prj, user=self.user, locked=False)   # черновик — без контрагента
        models.PurchaseLine.objects.create(purchase=purchase, item=item, qty=D(40))
        self.assertEqual(engine.item_on_order(item, self.prj), D(0))

    def test_manufactured_wip_is_on_order(self):
        board = self.make_item('BRD', manufactured=True)
        models.Kitting.objects.create(project=self.prj, target_item=board,
                                      user=self.user, qty=D(4),
                                      locked=False)
        self.assertEqual(engine.item_on_order(board, self.prj), D(4))


class DeficitTests(EngineTestBase):
    def test_full_deficit_scenario(self):
        device = self.make_item('DEV', manufactured=True, kind='device')
        case = self.make_item('CASE')
        screw = self.make_item('SCR', kind='material')
        models.BomLine.objects.create(parent=device, component=case, qty=D(1))
        models.BomLine.objects.create(parent=device, component=screw, qty=D(4))
        models.ProjectDemand.objects.create(project=self.prj, target_item=device, qty=D(10))

        # CASE: на складе 12 → ✓
        self.receipt_lot(case, self.prj, 12)
        # SCR: заказано 25 (зафиксировано), склада нет → ●25 ▲15
        purchase = models.Purchase.objects.create(
            project=self.prj, user=self.user, contractor=self.supplier,
            locked=True)
        models.PurchaseLine.objects.create(purchase=purchase, item=screw, qty=D(25))

        d = engine.project_deficit(self.prj)
        dm = d['demands'][0]
        self.assertEqual(dm['status'], 'to_order')   # worst-of (SCR не хватает)
        # аккордеон-дерево: CASE/SCR — прямые покупные листья прибора (depth 0)
        lines = {ln['component_code']: ln for ln in dm['tree']}
        self.assertEqual(lines['CASE']['in_stock'], D(12))   # сырой остаток, не клампованный
        self.assertEqual(lines['CASE']['balance'], D(2))     # нужно 10, лежит 12 → запас 2
        self.assertEqual(lines['CASE']['status'], 'available')
        self.assertTrue(lines['CASE']['is_leaf'])
        self.assertEqual(lines['SCR']['need'], D(40))
        self.assertEqual(lines['SCR']['on_order'], D(25))
        self.assertEqual(lines['SCR']['balance'], D(-15))
        self.assertEqual(lines['SCR']['status'], 'to_order')

    def test_surplus_order_reaches_the_view(self):
        """Заказали больше нужды: свод отдаёт сырые числа, перебор виден плюсом."""
        device = self.make_item('DEV', manufactured=True, kind='device')
        screw = self.make_item('SCR', kind='material')
        models.BomLine.objects.create(parent=device, component=screw, qty=D(1))
        models.ProjectDemand.objects.create(project=self.prj, target_item=device, qty=D(6))
        purchase = models.Purchase.objects.create(
            project=self.prj, user=self.user, contractor=self.supplier, locked=True)
        models.PurchaseLine.objects.create(purchase=purchase, item=screw, qty=D(10))

        d = engine.project_deficit(self.prj)
        row = {c['component_code']: c for c in d['components']}['SCR']
        self.assertEqual(row['need'], D(6))
        self.assertEqual(row['on_order'], D(10))         # сырое: столько заказано
        self.assertEqual(row['in_stock'], D(0))
        self.assertEqual(row['kitted'], D(0))
        self.assertEqual(row['balance'], D(4))           # запас 4 виден знаком
        self.assertEqual(row['status'], 'available')
        # то же число и в дереве-аккордеоне (лист несёт баланс)
        leaf = d['demands'][0]['tree'][0]
        self.assertEqual(leaf['balance'], D(4))

    def test_kitted_closes_need_without_stock(self):
        """Собрали приборы: компонент ушёл со склада, но потребность им закрыта."""
        device = self.make_item('DEV', manufactured=True, kind='device')
        screw = self.make_item('SCR', kind='material')
        models.BomLine.objects.create(parent=device, component=screw, qty=D(1))
        models.ProjectDemand.objects.create(project=self.prj, target_item=device, qty=D(6))
        lot = self.receipt_lot(screw, self.prj, 6)        # купили 6
        kit = engine.create_kitting(self.prj, self.user, target_item=device, qty=D(6))
        engine.add_kitting_line(kit, screw, lot, D(6))   # впаяли все 6
        engine.lock_kitting(kit)

        row = {c['component_code']: c
               for c in engine.project_deficit(self.prj)['components']}['SCR']
        self.assertEqual(row['kitted'], D(6))            # стоит в изделии
        self.assertEqual(row['in_stock'], D(0))          # на складе не осталось
        self.assertEqual(row['balance'], D(0))           # и это НЕ дефицит
        self.assertEqual(row['status'], 'on_order')      # оранжевый: сошлось впритык


class PurchaseCoverageTests(EngineTestBase):
    """Ф1b: цвет заказа в списке — покрытие строк лотами приходов."""

    def _purchase_with_line(self, item, qty):
        p = engine.create_purchase(self.prj, self.user)
        models.PurchaseLine.objects.create(purchase=p, item=item, qty=D(qty))
        return p

    def test_empty_purchase_to_order(self):
        p = engine.create_purchase(self.prj, self.user)
        self.assertEqual(engine.purchase_coverage(p), 'to_order')

    def test_nothing_received_to_order(self):
        p = self._purchase_with_line(self.make_item('R1'), 10)
        self.assertEqual(engine.purchase_coverage(p), 'to_order')

    def test_partial_on_order(self):
        item = self.make_item('R1')
        p = self._purchase_with_line(item, 10)
        self.receipt_lot(item, self.prj, 4, purchase=p)
        self.assertEqual(engine.purchase_coverage(p), 'on_order')

    def test_fully_received_available(self):
        item = self.make_item('R1')
        p = self._purchase_with_line(item, 10)
        self.receipt_lot(item, self.prj, 10, purchase=p)
        self.assertEqual(engine.purchase_coverage(p), 'available')

    def test_one_line_open_keeps_on_order(self):
        a, b = self.make_item('A'), self.make_item('B')
        p = engine.create_purchase(self.prj, self.user)
        models.PurchaseLine.objects.create(purchase=p, item=a, qty=D(5))
        models.PurchaseLine.objects.create(purchase=p, item=b, qty=D(5))
        self.receipt_lot(a, self.prj, 5, purchase=p)      # A закрыт, B пуст
        self.assertEqual(engine.purchase_coverage(p), 'on_order')


class ProjectHealthTests(EngineTestBase):
    """Ф1b: цвет проекта в списке — worst-of здоровья (вычисляемая проекция)."""

    def _device_demand(self, qty):
        dev = self.make_item('DEV', manufactured=True, kind='device')
        case = self.make_item('CASE')
        models.BomLine.objects.create(parent=dev, component=case, qty=D(1))
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(qty))
        return dev, case

    def test_internal_project_none(self):
        white = models.Project.objects.create(
            code='W', description='Собственный склад', kind=models.Project.Kind.INTERNAL_STOCK)
        self.assertIsNone(engine.project_health(white))

    def test_empty_project_none(self):
        self.assertIsNone(engine.project_health(self.prj))

    def test_component_to_order_red(self):
        self._device_demand(5)                    # склад пуст → CASE к заказу
        self.assertEqual(engine.project_health(self.prj), 'to_order')

    def test_available_but_unassembled_orange(self):
        _dev, case = self._device_demand(2)
        self.receipt_lot(case, self.prj, 10)      # компонент есть, приборы не собраны
        self.assertEqual(engine.project_health(self.prj), 'on_order')

    def test_all_assembled_green(self):
        dev, case = self._device_demand(1)
        lot = self.receipt_lot(case, self.prj, 5)
        k = models.Kitting.objects.create(project=self.prj, target_item=dev,
                                          user=self.user, qty=D(1), locked=False)
        engine.add_kitting_line(k, case, lot, D(1))
        engine.lock_kitting(k)                     # рождает прибор → собрано
        self.assertEqual(engine.project_health(self.prj), 'available')


class StockMapTests(EngineTestBase):
    def test_map_across_projects_sorted(self):
        white = models.Project.objects.create(
            code='WHITE', description='Собственный склад',
            kind=models.Project.Kind.INTERNAL_STOCK)
        item = self.make_item('CASE')
        self.receipt_lot(item, self.prj, 12)
        inv = models.Inventory.objects.create(project=white, user=self.user,
                                              number='INV-1', date='2026-06-01',
                                              locked=True)   # Ф15: иначе не на складе
        lot = models.Lot.objects.create(item=item, project=white, origin=inv, qty=D(5))
        engine.rebuild_movements(lot)

        m = engine.stock_map(item)
        self.assertEqual(len(m['rows']), 2)
        # белый склад идёт первым (мягкая сортировка)
        self.assertEqual(m['rows'][0]['project_code'], 'WHITE')
        self.assertEqual(m['rows'][0]['available'], D(5))
        self.assertEqual(m['rows'][1]['available'], D(12))

    def test_zero_available_excluded(self):
        item = self.make_item('CASE')
        lot = self.receipt_lot(item, self.prj, 0)
        m = engine.stock_map(item)
        self.assertEqual(m['rows'], [])


class KittingFormTests(EngineTestBase):
    """Волна 2: форма комплектации — призрачные строки, пайка, закрытие."""

    def setUp(self):
        super().setUp()
        self.device = self.make_item('DEV', manufactured=True,
                                     kind='device')
        self.case = self.make_item('CASE')
        self.res = self.make_item('RES')
        # прибор из 1 корпуса и 2 резисторов
        models.BomLine.objects.create(parent=self.device, component=self.case, qty=D(1))
        models.BomLine.objects.create(parent=self.device, component=self.res, qty=D(2))

    def make_kitting(self, qty=2):
        return models.Kitting.objects.create(
            project=self.prj, target_item=self.device, user=self.user,
            qty=D(qty), locked=False)

    def test_target_pick_on_click_born_draft_sets_one_sample(self):
        """Клик-проход Ивана, 500 на выборе НАШЕГО изделия: комплектация рождается по
        клику без кол-ва (Ф12e), и проекция падала на `bl.qty * None`. У покупного
        состава нет — цикл не заходил, и баг выглядел вывернутым наизнанку («покупное
        выбирается, своё роняет»). Выбор цели теперь проставляет 1 образец, как
        рождение сразу с целью."""
        k = engine.create_kitting(self.prj, self.user)     # ни цели, ни кол-ва
        self.assertIsNone(k.qty)
        self.assertEqual(engine.kitting_form(k)['rows'], [])   # без цели — пусто
        engine.update_kitting(k, target_item=self.device)
        k.refresh_from_db()
        self.assertEqual(k.qty, D(1))
        rows = {r['component_code']: r for r in engine.kitting_form(k)['rows']}
        self.assertEqual(rows['RES']['need'], D(2))        # 2×1
        engine.update_kitting(k, qty=D(3))                 # своё кол-во не перетёрто
        self.assertEqual(engine.kitting_form(k)['rows'][0]['need'], D(3))

    def test_kitting_target_must_be_native(self):
        """«Комплектуем только своё» (решение Ивана 2026-07-31): покупное изделие в цель
        не пускаем ни на одном пути — ни при рождении, ни при правке."""
        with self.assertRaises(ValidationError):
            engine.create_kitting(self.prj, self.user, target_item=self.res)
        k = engine.create_kitting(self.prj, self.user)
        with self.assertRaises(ValidationError):
            engine.update_kitting(k, target_item=self.res)
        k.refresh_from_db()
        self.assertIsNone(k.target_item)
        engine.update_kitting(k, target_item=self.device)      # наше — проходит
        k.refresh_from_db()
        self.assertEqual(k.target_item, self.device)

    def test_purchased_target_blocked_at_lock_even_past_the_engine(self):
        """Второй слой правила. CHECK его выразить не может (смотрит в чужую таблицу,
        а MySQL в CHECK подзапросы запрещает), поэтому страхует `clean()`: путь мимо
        движка (прямой ORM, админка) упирается в отказ на фиксации."""
        k = models.Kitting.objects.create(project=self.prj, target_item=self.res,
                                          user=self.user, qty=D(1), locked=False)
        self.receipt_lot(self.res, self.prj, 5)
        lot = models.Lot.objects.filter(item=self.res).first()
        engine.add_kitting_line(k, self.res, lot, D(1))
        with self.assertRaises(ValidationError):
            engine.lock_kitting(k)
        k.refresh_from_db()
        self.assertFalse(k.locked)

    def test_form_survives_target_without_qty(self):
        """Страховка того же места: пустое кол-во у цели с составом — не 500, а «надо 0»
        без призраков. Проекция обязана пережить любой неполный черновик, каким бы
        путём он ни возник (админка, прямой ORM)."""
        k = models.Kitting.objects.create(project=self.prj, target_item=self.device,
                                          user=self.user, qty=None, locked=False)
        rows = engine.kitting_form(k)['rows']
        self.assertEqual(rows[0]['need'], D(0))
        self.assertIsNone(rows[0]['ghost'])

    def test_ghost_rows_before_piercing(self):
        # склад пуст → обе призрачные строки красные (▲ to_order)
        k = self.make_kitting(qty=2)
        c = engine.kitting_form(k)
        rows = {r['component_code']: r for r in c['rows']}
        self.assertEqual(rows['CASE']['need'], D(2))     # 1×2
        self.assertEqual(rows['RES']['need'], D(4))      # 2×2
        self.assertEqual(rows['CASE']['pierced'], D(0))
        self.assertEqual(rows['CASE']['ghost']['status'], 'to_order')
        self.assertEqual(c['worst_status'], 'to_order')
        # Оси компонента — под глиф строки (§7a), как в заказе и закупке-плане.
        self.assertEqual(rows['CASE']['component_native'], False)
        self.assertIn('component_locked', rows['CASE'])

    def test_ghost_available_when_stock_exists(self):
        # есть лот корпуса → призрачная строка зелёная + лот-кандидат
        self.receipt_lot(self.case, self.prj, 10)
        k = self.make_kitting(qty=2)
        c = engine.kitting_form(k)
        row = {r['component_code']: r for r in c['rows']}['CASE']
        self.assertEqual(row['ghost']['status'], 'available')
        self.assertEqual(len(row['ghost']['candidate_lots']), 1)
        self.assertEqual(row['ghost']['candidate_lots'][0]['live_qty'], D(10))

    def test_pierce_creates_line_and_issue(self):
        lot = self.receipt_lot(self.case, self.prj, 10)
        k = self.make_kitting(qty=2)
        engine.add_kitting_line(k, self.case, lot, D(2))
        # Ф15: пайка черновика склад не двигает — компонент ещё целиком свой
        self.assertEqual(engine.lot_live_qty(lot), D(10))
        engine.lock_kitting(k)                   # фиксация материализует −ISSUE
        self.assertEqual(engine.lot_live_qty(lot), D(8))
        c = engine.kitting_form(k)
        row = {r['component_code']: r for r in c['rows']}['CASE']
        self.assertEqual(row['pierced'], D(2))
        self.assertEqual(row['remaining'], D(0))
        self.assertIsNone(row['ghost'])          # покрыто — призрака нет
        self.assertEqual(len(row['real_lines']), 1)

    def test_pierce_rejects_foreign_project_lot(self):
        other = models.Project.objects.create(code='P2', description='Другой')
        lot = self.receipt_lot(self.case, other, 10)
        k = self.make_kitting(qty=1)
        with self.assertRaises(ValidationError):
            engine.add_kitting_line(k, self.case, lot, D(1))

    def test_pierce_rejects_wrong_component_lot(self):
        lot = self.receipt_lot(self.res, self.prj, 10)   # лот резистора
        k = self.make_kitting(qty=1)
        with self.assertRaises(ValidationError):
            engine.add_kitting_line(k, self.case, lot, D(1))   # ждём корпус

    def test_update_line_qty_rebuilds(self):
        lot = self.receipt_lot(self.case, self.prj, 10)
        k = self.make_kitting(qty=2)
        line = engine.add_kitting_line(k, self.case, lot, D(2))
        engine.update_kitting_line(line, D(5))
        engine.lock_kitting(k)                   # Ф15: расход виден с фиксации
        self.assertEqual(engine.lot_live_qty(lot), D(5))

    def test_remove_line_restores_qty(self):
        """Ф15: правка черновика склад не трогает вовсе — ни добавление строки, ни
        её снятие; проверяем через фиксацию (расход) и расфиксацию (возврат)."""
        lot = self.receipt_lot(self.case, self.prj, 10)
        k = self.make_kitting(qty=2)
        line = engine.add_kitting_line(k, self.case, lot, D(3))
        self.assertEqual(engine.lot_live_qty(lot), D(10))   # черновик не двигает
        engine.lock_kitting(k)
        self.assertEqual(engine.lot_live_qty(lot), D(7))
        engine.unlock_kitting(k)                            # расход отпущен
        self.assertEqual(engine.lot_live_qty(lot), D(10))
        engine.remove_kitting_line(line)
        engine.lock_kitting(k)
        self.assertEqual(engine.lot_live_qty(lot), D(10))

    def test_close_births_device_lot_with_cost_snapshot(self):
        case_lot = self.receipt_lot(self.case, self.prj, 10)
        case_lot.unit_cost = D(800); case_lot.save()
        res_lot = self.receipt_lot(self.res, self.prj, 100)
        res_lot.unit_cost = D(1); res_lot.save()
        k = self.make_kitting(qty=2)
        engine.add_kitting_line(k, self.case, case_lot, D(2))   # 2×800
        engine.add_kitting_line(k, self.res, res_lot, D(4))     # 4×1
        lot = engine.lock_kitting(k)
        k.refresh_from_db()
        self.assertTrue(k.locked)
        self.assertEqual(lot.qty, D(2))
        # (2×800 + 4×1) / 2 = 802
        self.assertEqual(lot.unit_cost, D('802.00'))
        self.assertEqual(engine.lot_live_qty(lot), D(2))

    def test_close_only_wip(self):
        k = self.make_kitting(qty=1)
        engine.lock_kitting(k)
        with self.assertRaises(ValidationError):
            engine.lock_kitting(k)

    def test_pierce_blocked_after_close(self):
        lot = self.receipt_lot(self.case, self.prj, 10)
        k = self.make_kitting(qty=1)
        engine.lock_kitting(k)
        with self.assertRaises(ValidationError):
            engine.add_kitting_line(k, self.case, lot, D(1))

    def test_reopen_restores_wip_and_removes_lot(self):
        k = self.make_kitting(qty=1)
        lot = engine.lock_kitting(k)
        engine.unlock_kitting(k)
        k.refresh_from_db()
        self.assertFalse(k.locked)
        self.assertFalse(models.Lot.objects.filter(pk=lot.pk).exists())

    def test_reopen_blocked_when_device_consumed(self):
        k = self.make_kitting(qty=1)
        device_lot = engine.lock_kitting(k)
        # прибор передан заказчику → потомок вниз, переоткрытие запрещено
        transfer = models.Transfer.objects.create(
            project=self.prj, user=self.user, date='2026-06-01', number='TN-1')
        models.StockLine.objects.create(document=transfer, lot=device_lot,
                                        location=self.main, qty=D(-1))
        engine.rebuild_movements(device_lot)
        with self.assertRaises(ValidationError):
            engine.unlock_kitting(k)


class ReceiptFormTests(EngineTestBase):
    """Волна 3: форма прихода — строки-лоты УПД, рождение +RECEIPT, замок."""

    def make_receipt(self, approved=False):
        return models.Receipt.objects.create(
            number='УПД-Т', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user,
            locked=approved)

    def test_add_lot_births_receipt_movement(self):
        r = self.make_receipt()
        case = self.make_item('CASE')
        lot = engine.add_receipt_lot(r, case, D(12), unit_cost=D(800),
                                     lot_name='Корпус Al')
        self.assertEqual(lot.project_id, r.project_id)   # проект наследован
        # Ф15: партия черновой поставки на складе не лежит — движение рождает замок
        self.assertEqual(engine.lot_live_qty(lot), D(0))
        self.assertEqual(lot.movements.count(), 0)
        engine.lock_receipt(r)
        self.assertEqual(engine.lot_live_qty(lot), D(12))
        mv = lot.movements.get()
        self.assertEqual(mv.type, models.StockMovement.Type.RECEIPT)
        self.assertEqual(mv.qty, D(12))

    def test_form_shows_lines_and_total(self):
        r = self.make_receipt()
        engine.add_receipt_lot(r, self.make_item('A'), D(2), unit_cost=D(100))
        engine.add_receipt_lot(r, self.make_item('B'), D(3), unit_cost=D(10))
        c = engine.receipt_form(r)
        self.assertEqual(len(c['lots']), 2)
        self.assertEqual(c['total_cost'], D(230))        # 2×100 + 3×10
        self.assertFalse(c['locked'])

    def test_add_lot_rejects_nonpositive_qty(self):
        r = self.make_receipt()
        with self.assertRaises(ValidationError):
            engine.add_receipt_lot(r, self.make_item('A'), D(0))

    def test_update_lot_qty_rebuilds(self):
        r = self.make_receipt()
        lot = engine.add_receipt_lot(r, self.make_item('A'), D(10))
        engine.update_receipt_lot(lot, qty=D(7))
        engine.lock_receipt(r)                   # Ф15: склад видит с фиксации
        self.assertEqual(engine.lot_live_qty(lot), D(7))

    def test_update_lot_cost_and_name(self):
        r = self.make_receipt()
        lot = engine.add_receipt_lot(r, self.make_item('A'), D(5))
        engine.update_receipt_lot(lot, unit_cost=D(42), lot_name='Ы',
                                  part_number='PN-1')
        lot.refresh_from_db()
        self.assertEqual(lot.unit_cost, D(42))
        self.assertEqual(lot.lot_name, 'Ы')
        self.assertEqual(lot.part_number, 'PN-1')

    def test_remove_lot(self):
        r = self.make_receipt()
        lot = engine.add_receipt_lot(r, self.make_item('A'), D(5))
        engine.remove_receipt_lot(lot)
        self.assertFalse(models.Lot.objects.filter(pk=lot.pk).exists())

    def test_remove_blocked_when_consumed(self):
        r = self.make_receipt()
        comp = self.make_item('R')
        lot = engine.add_receipt_lot(r, comp, D(100))
        dev = self.make_item('DEV', manufactured=True)
        k = models.Kitting.objects.create(project=self.prj, target_item=dev,
                                          user=self.user, qty=D(1),
                                          locked=False)
        engine.add_kitting_line(k, comp, lot, D(30))   # спаяли — потреблён ниже
        with self.assertRaises(ValidationError):
            engine.remove_receipt_lot(lot)

    def test_approve_locks_edits(self):
        r = self.make_receipt()
        lot = engine.add_receipt_lot(r, self.make_item('A'), D(5))
        engine.lock_receipt(r)
        r.refresh_from_db()
        self.assertTrue(r.locked)
        with self.assertRaises(ValidationError):
            engine.update_receipt_lot(lot, qty=D(9))
        with self.assertRaises(ValidationError):
            engine.add_receipt_lot(r, self.make_item('B'), D(1))

    def test_approve_rejects_empty(self):
        r = self.make_receipt()
        with self.assertRaises(ValidationError):
            engine.lock_receipt(r)

    def test_unapprove_reenables_edits(self):
        r = self.make_receipt()
        lot = engine.add_receipt_lot(r, self.make_item('A'), D(5))
        engine.lock_receipt(r)
        self.assertEqual(engine.lot_live_qty(lot), D(5))
        engine.unlock_receipt(r)
        r.refresh_from_db()
        self.assertFalse(r.locked)
        self.assertEqual(engine.lot_live_qty(lot), D(0))   # Ф15: ушла со склада
        engine.update_receipt_lot(lot, qty=D(9))           # снова можно править
        engine.lock_receipt(r)
        self.assertEqual(engine.lot_live_qty(lot), D(9))

    def test_received_lot_feeds_kitting_form(self):
        # сверенный приход РЕЗ → лот виден форме комплектации как кандидат
        # (Ф15: именно сверенный — черновая поставка складу не видна)
        r = self.make_receipt()
        comp = self.make_item('R')
        engine.add_receipt_lot(r, comp, D(50))
        engine.lock_receipt(r)
        dev = self.make_item('DEV', manufactured=True)
        models.BomLine.objects.create(parent=dev, component=comp, qty=D(2))
        k = models.Kitting.objects.create(project=self.prj, target_item=dev,
                                          user=self.user, qty=D(1),
                                          locked=False)
        c = engine.kitting_form(k)
        row = {r['component_code']: r for r in c['rows']}['R']
        self.assertEqual(row['ghost']['status'], 'available')
        self.assertEqual(len(row['ghost']['candidate_lots']), 1)


class PurchaseFormTests(EngineTestBase):
    """Волна 4: форма заказа — строки-обязательства, замок отправки, гашение
    приходом, мост «дефицит → заказ»."""

    def test_create_purchase_leaves_plan_and_contractor_empty(self):
        """Ф17: заказ рождается БЕЗ закупки-плана (и без контрагента).

        До Ф17 `procurement` был NOT NULL, и рождение тихо плодило закупку-пустышку;
        теперь заказ — самостоятельная сущность, план и контрагент выбирают в форме.
        """
        p = engine.create_purchase(self.prj, self.user)
        self.assertFalse(p.locked)
        self.assertIsNone(p.procurement_id)
        self.assertIsNone(p.contractor_id)
        self.assertEqual(p.project_id, self.prj.id)
        self.assertFalse(models.Procurement.objects.exists())   # пустышек не бывает

    def test_add_line_and_form_totals(self):
        p = engine.create_purchase(self.prj, self.user)
        engine.add_purchase_line(p, self.make_item('A'), D(10))
        engine.add_purchase_line(p, self.make_item('B'), D(5))
        c = engine.purchase_form(p)
        self.assertEqual(len(c['rows']), 2)
        self.assertEqual(c['total_ordered'], D(15))
        self.assertEqual(c['total_received'], D(0))
        self.assertTrue(c['editable'])
        self.assertEqual(c['rows'][0]['status'], 'to_order')   # ждём поставки

    def test_add_line_rejects_duplicate_item(self):
        p = engine.create_purchase(self.prj, self.user)
        item = self.make_item('A')
        engine.add_purchase_line(p, item, D(10))
        with self.assertRaises(ValidationError):
            engine.add_purchase_line(p, item, D(3))

    def test_add_line_rejects_nonpositive(self):
        p = engine.create_purchase(self.prj, self.user)
        with self.assertRaises(ValidationError):
            engine.add_purchase_line(p, self.make_item('A'), D(0))

    def test_update_and_remove_line(self):
        p = engine.create_purchase(self.prj, self.user)
        line = engine.add_purchase_line(p, self.make_item('A'), D(10))
        engine.update_purchase_line(line, D(7))
        line.refresh_from_db()
        self.assertEqual(line.qty, D(7))
        engine.remove_purchase_line(line)
        self.assertFalse(models.PurchaseLine.objects.filter(pk=line.pk).exists())

    def test_post_counts_in_on_order(self):
        item = self.make_item('SCR', kind='material')
        p = self.make_purchase()
        engine.add_purchase_line(p, item, D(40))
        self.assertEqual(engine.item_on_order(item, self.prj), D(0))  # draft не в счёте
        engine.lock_purchase(p)
        self.assertEqual(engine.item_on_order(item, self.prj), D(40))

    def test_post_rejects_empty(self):
        p = self.make_purchase()
        with self.assertRaises(ValidationError):
            engine.lock_purchase(p)

    def test_lines_locked_after_post(self):
        p = self.make_purchase()
        line = engine.add_purchase_line(p, self.make_item('A'), D(10))
        engine.lock_purchase(p)
        with self.assertRaises(ValidationError):
            engine.update_purchase_line(line, D(5))
        with self.assertRaises(ValidationError):
            engine.add_purchase_line(p, self.make_item('B'), D(1))

    def test_unpost_reenables_and_drops_from_on_order(self):
        item = self.make_item('SCR', kind='material')
        p = self.make_purchase()
        line = engine.add_purchase_line(p, item, D(40))
        engine.lock_purchase(p)
        engine.unlock_purchase(p)
        self.assertEqual(engine.item_on_order(item, self.prj), D(0))
        engine.update_purchase_line(line, D(30))       # снова можно
        self.assertEqual(line.qty, D(30))

    def test_delete_drops_from_on_order(self):
        """Отмена = удаление (волна 19, Р1): статуса `cancelled` больше нет, снять
        обязательство можно только удалив заказ — и сперва сняв замок."""
        item = self.make_item('SCR', kind='material')
        p = self.make_purchase()
        engine.add_purchase_line(p, item, D(40))
        engine.lock_purchase(p)
        self.assertEqual(engine.item_on_order(item, self.prj), D(40))
        with self.assertRaises(ValidationError):        # утверждённый не удалить
            engine.delete_purchase(p)
        engine.unlock_purchase(p)
        engine.delete_purchase(p)
        self.assertEqual(engine.item_on_order(item, self.prj), D(0))
        self.assertFalse(models.Purchase.objects.filter(pk=p.pk).exists())

    def test_linked_receipt_reduces_on_order_and_closes_line(self):
        item = self.make_item('SCR', kind='material')
        p = self.make_purchase()
        line = engine.add_purchase_line(p, item, D(40))
        engine.lock_purchase(p)
        # приход 15, связанный с заказом → поступило 15, «заказано» 25
        self.receipt_lot(item, self.prj, 15, purchase=p)
        self.assertEqual(engine.item_on_order(item, self.prj), D(25))
        c = engine.purchase_form(p)
        row = c['rows'][0]
        self.assertEqual(row['received'], D(15))
        self.assertEqual(row['remaining'], D(25))
        self.assertEqual(row['status'], 'on_order')     # ● частично
        self.assertEqual(len(c['receipts']), 1)
        # добираем остаток → строка закрыта (✓), «заказано» 0
        self.receipt_lot(item, self.prj, 25, purchase=p)
        self.assertEqual(engine.item_on_order(item, self.prj), D(0))
        row = engine.purchase_form(p)['rows'][0]
        self.assertEqual(row['status'], 'available')

    def test_set_receipt_purchase_rejects_foreign_project(self):
        other = models.Project.objects.create(
            code='P2', description='Проект 2', kind=models.Project.Kind.EXTERNAL)
        p = engine.create_purchase(other, self.user)
        r = models.Receipt.objects.create(
            number='УПД-Х', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        with self.assertRaises(ValidationError):
            engine.set_receipt_purchase(r, p)           # заказ чужого проекта


class ReceiptFromPurchaseTests(EngineTestBase):
    """Волна 19, Ф6: поток «Заказ → УПД» — накладная преднабивается остатком заказа."""

    def setUp(self):
        super().setUp()
        self.item = self.make_item('SCR', kind='material')
        self.other = self.make_item('CAP', kind='material')
        self.p = self.make_purchase()
        engine.add_purchase_line(self.p, self.item, D(40))
        engine.add_purchase_line(self.p, self.other, D(10))
        # Ф17: поставщик поставки приезжает из ЗАКАЗА (Р3 отменена) — он у заказа свой.
        engine.lock_purchase(self.p)

    def test_prefills_lines_from_order(self):
        r = engine.create_receipt_from_purchase(self.p, self.user)
        self.assertEqual(r.purchase_id, self.p.id)
        self.assertEqual(r.project_id, self.prj.id)
        self.assertEqual(r.contractor_id, self.supplier.id)   # Ф17: из заказа
        self.assertFalse(r.locked)                            # черновик
        lots = {lot.item_id: lot for lot in r.lots.all()}
        self.assertEqual(lots[self.item.id].qty, D(40))
        self.assertEqual(lots[self.other.id].qty, D(10))
        # Цена — 0, НЕ estimated_cost: оценка, забытая при фиксации, стала бы фактом
        # трат проекта неотличимо от реальной цены УПД.
        self.assertEqual(lots[self.item.id].unit_cost, D(0))
        self.assertEqual(lots[self.item.id].lot_name, self.item.description)

    def test_draft_purchase_rejected(self):
        engine.unlock_purchase(self.p)
        with self.assertRaises(ValidationError):
            engine.create_receipt_from_purchase(self.p, self.user)

    def test_closed_purchase_rejected(self):
        self.receipt_lot(self.item, self.prj, 40, purchase=self.p)
        self.receipt_lot(self.other, self.prj, 10, purchase=self.p)
        with self.assertRaises(ValidationError):
            engine.create_receipt_from_purchase(self.p, self.user)

    def test_second_call_covers_remainder(self):
        """Поставка частями: зафиксировали первую накладную → вторая идёт на остаток."""
        # номер обязателен к ФИКСАЦИИ, а не к рождению (Ф12e) — здесь задаём сразу
        r1 = engine.create_receipt_from_purchase(self.p, self.user, number='УПД-1')
        # приехало меньше заказанного — правим по бумажной накладной и фиксируем
        lot = r1.lots.get(item=self.item)
        engine.update_receipt_lot(lot, qty=D(15))
        engine.remove_receipt_lot(r1.lots.get(item=self.other))
        engine.lock_receipt(r1)

        r2 = engine.create_receipt_from_purchase(self.p, self.user)
        lots = {lot.item_id: lot for lot in r2.lots.all()}
        self.assertEqual(lots[self.item.id].qty, D(25))       # 40 − 15
        self.assertEqual(lots[self.other.id].qty, D(10))      # не приезжало вовсе

    def test_open_draft_blocks_second_call(self):
        """Черновик заказ не гасит — значит повторный клик плодил бы близнеца."""
        engine.create_receipt_from_purchase(self.p, self.user)
        with self.assertRaises(ValidationError):
            engine.create_receipt_from_purchase(self.p, self.user)

    def test_draft_receipt_does_not_close_line(self):
        """Ф6: закрытость заказа гейтит тот же замок, что склад и деньги (Ф15).

        До правки черновая накладная гасила строку — заказ показывал ✓ до приёмки,
        а позиция исчезала и из «заказано», и со склада разом.
        """
        engine.create_receipt_from_purchase(self.p, self.user)
        row = next(r for r in engine.purchase_form(self.p)['rows']
                   if r['item_id'] == self.item.id)
        self.assertEqual(row['received'], D(0))
        self.assertEqual(row['remaining'], D(40))
        self.assertEqual(row['status'], 'to_order')
        self.assertEqual(row['receipts'], [])
        # и в дефиците позиция всё ещё «заказана» (оранжевый член цел)
        self.assertEqual(engine.item_on_order(self.item, self.prj), D(40))

    def test_line_carries_closing_receipts(self):
        """Обратная связь: строка знает, какими накладными закрыта."""
        engine.create_receipt_from_purchase(self.p, self.user, number='УПД-7')
        r = models.Receipt.objects.get(purchase=self.p)
        engine.update_receipt_lot(r.lots.get(item=self.item), qty=D(15))
        engine.lock_receipt(r)
        row = next(x for x in engine.purchase_form(self.p)['rows']
                   if x['item_id'] == self.item.id)
        self.assertEqual([c['receipt_id'] for c in row['receipts']], [r.id])
        self.assertEqual(row['receipts'][0]['qty'], D(15))
        self.assertEqual(row['received'], D(15))

    def test_http_creates_receipt_and_returns_its_form(self):
        c = Client()
        c.force_login(self.user)
        resp = c.post(f'/api/purchases/{self.p.id}/receipt/')
        self.assertEqual(resp.status_code, 201)
        body = resp.json()
        self.assertEqual(body['purchase_id'], self.p.id)
        self.assertEqual(len(body['lots']), 2)


class PurchaseContractorTests(EngineTestBase):
    """Волна 19, Ф17: контрагент у всех трёх уровней контура.

    Закупка = ЧТО купить · Заказ = У КОГО купить · Поставка = КТО привёз. Каждый уровень
    знает своего контрагента и не зависит от того, есть ли над ним родитель.
    """

    def setUp(self):
        super().setUp()
        self.item = self.make_item('SCR', kind='material')
        self.other_cp = models.Counterparty.objects.create(description='Другой поставщик')

    # ── обязательность на фиксации, а не при рождении ──────────────────────
    def test_lock_without_contractor_refuses_friendly(self):
        """Внятный отказ движка, а не `IntegrityError 3819` от CHECK (урок Ф12e)."""
        p = engine.create_purchase(self.prj, self.user)
        engine.add_purchase_line(p, self.item, D(5))
        with self.assertRaises(ValidationError) as cm:
            engine.lock_purchase(p)
        self.assertIn('контрагента', cm.exception.messages[0])
        p.refresh_from_db()
        self.assertFalse(p.locked)

    def test_lock_passes_once_contractor_chosen(self):
        p = engine.create_purchase(self.prj, self.user)
        engine.add_purchase_line(p, self.item, D(5))
        engine.update_purchase(p, contractor=self.supplier)
        engine.lock_purchase(p)
        self.assertTrue(p.locked)

    # ── самостоятельность уровней ─────────────────────────────────────────
    def test_purchase_without_plan_goes_the_whole_way_to_receipt(self):
        """Заказ без закупки проходит путь целиком — до Ф17 он не знал поставщика."""
        p = self.make_purchase()
        self.assertIsNone(p.procurement_id)
        engine.add_purchase_line(p, self.item, D(7))
        engine.lock_purchase(p)
        r = engine.create_receipt_from_purchase(p, self.user)
        self.assertEqual(r.contractor_id, self.supplier.id)   # из ЗАКАЗА, не сквозь план
        self.assertEqual(r.lots.get(item=self.item).qty, D(7))

    def test_plan_and_order_contractors_live_apart(self):
        """Контрагенты плана и заказа независимы: одна закупка законно идёт к разным.

        Наследование «копией при рождении» отсюда ушло вместе с автосозданием заказов
        (2026-08-05): движок заказы больше не рождает, их заводят руками и контрагента
        ставят там же. Независимость полей — то, ради чего копия и делалась, — осталась.
        """
        plan = engine.create_procurement(self.user)
        engine.add_procurement_line(plan, self.item, D(10))
        engine.update_procurement(plan, contractor=self.supplier)
        pu = engine.create_purchase(self.prj, self.user)
        engine.update_purchase(pu, procurement=plan, contractor=self.supplier)
        engine.set_allocation(plan, pu, self.item, D(10))
        self.assertEqual(pu.contractor_id, self.supplier.id)
        # правка заказа план не трогает…
        engine.update_purchase(pu, contractor=self.other_cp)
        plan.refresh_from_db()
        self.assertEqual(plan.contractor_id, self.supplier.id)
        # …и наоборот
        engine.update_procurement(plan, contractor=None)
        pu.refresh_from_db()
        self.assertEqual(pu.contractor_id, self.other_cp.id)

    # ── расхождение: флаг, а не гейт ──────────────────────────────────────
    def test_mismatch_flag_is_advisory_and_does_not_block_lock(self):
        plan = engine.create_procurement(self.user)
        engine.update_procurement(plan, contractor=self.supplier)
        p = engine.create_purchase(self.prj, self.user)
        engine.add_purchase_line(p, self.item, D(3))
        engine.update_purchase(p, procurement=plan, contractor=self.other_cp)
        self.assertTrue(engine.purchase_form(p)['contractor_mismatch'])
        engine.lock_purchase(p)                    # предупреждение не останавливает
        self.assertTrue(p.locked)

    def test_no_mismatch_when_either_side_empty_or_equal(self):
        """Пустой сверху/снизу — это «не выбран», а не расхождение."""
        plan = engine.create_procurement(self.user)          # у плана контрагента нет
        p = engine.create_purchase(self.prj, self.user)
        engine.update_purchase(p, procurement=plan, contractor=self.supplier)
        self.assertFalse(engine.purchase_form(p)['contractor_mismatch'])
        engine.update_purchase(p, contractor=None)           # нет и у заказа
        self.assertFalse(engine.purchase_form(p)['contractor_mismatch'])
        engine.update_procurement(plan, contractor=self.supplier)
        engine.update_purchase(p, contractor=self.supplier)  # совпали
        self.assertFalse(engine.purchase_form(p)['contractor_mismatch'])

    def test_receipt_mismatch_compares_with_its_purchase(self):
        """У поставки тот же флаг — «кто привёз» против «у кого купили»."""
        p = self.make_purchase()
        engine.add_purchase_line(p, self.item, D(4))
        engine.lock_purchase(p)
        r = engine.create_receipt_from_purchase(p, self.user)
        self.assertFalse(engine.receipt_form(r)['contractor_mismatch'])
        engine.update_receipt(r, contractor=self.other_cp)
        self.assertTrue(engine.receipt_form(r)['contractor_mismatch'])

    # ── HTTP-срез ─────────────────────────────────────────────────────────
    def test_patch_contractor_and_null_plan_http(self):
        p = engine.create_purchase(self.prj, self.user)
        c = Client()
        c.force_login(self.user)
        resp = c.patch(f'/api/purchases/{p.id}/',
                       {'contractor_id': self.supplier.id},
                       content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()['contractor_id'], self.supplier.id)
        # `procurement_id: null` — законное «закупка не выбрана», а не 400
        resp = c.patch(f'/api/purchases/{p.id}/', {'procurement_id': None},
                       content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        self.assertIsNone(resp.json()['procurement_id'])
        # несуществующий контрагент → дружелюбный 400 (не 500)
        bad = c.patch(f'/api/purchases/{p.id}/', {'contractor_id': 99999},
                      content_type='application/json')
        self.assertEqual(bad.status_code, 400)


class TransferFormTests(EngineTestBase):
    """Волна 5: форма передачи — отгрузка партии заказчику (`−ISSUE`), пикер
    отдаваемых лотов, guard чужого проекта, коррекция строк."""

    def setUp(self):
        super().setUp()
        self.device = self.make_item('DEV', manufactured=True,
                                      kind='device')
        # готовое железо на складе проекта: лот 5 (как из комплектации/прихода)
        self.lot = self.receipt_lot(self.device, self.prj, 5)

    def test_add_line_issues_from_lot(self):
        t = engine.create_transfer(self.prj, self.user, 'Н-1')
        engine.add_transfer_line(t, self.lot, D(2))
        # Ф15: черновая накладная ничего не отгружает — партия ещё вся на складе
        self.assertEqual(engine.lot_live_qty(self.lot), D(5))
        self.assertFalse(self.lot.movements.filter(type='ISSUE').exists())
        engine.lock_transfer(t)                                 # «отгружено»
        self.assertEqual(engine.lot_live_qty(self.lot), D(3))   # 5 − 2
        self.assertTrue(self.lot.movements.filter(type='ISSUE', qty=D(-2)).exists())

    def test_form_totals_and_live(self):
        t = engine.create_transfer(self.prj, self.user, 'Н-1')
        engine.add_transfer_line(t, self.lot, D(2), display_name='Прибор зав.№7')
        engine.lock_transfer(t)
        c = engine.transfer_form(t)
        self.assertEqual(c['number'], 'Н-1')
        self.assertEqual(c['total_qty'], D(2))
        self.assertEqual(len(c['lines']), 1)
        row = c['lines'][0]
        self.assertEqual(row['qty'], D(2))
        self.assertEqual(row['lot_live_qty'], D(3))
        self.assertEqual(row['display_name'], 'Прибор зав.№7')

    def test_default_display_name_from_lot(self):
        self.lot.lot_name = 'ЗН-42'
        self.lot.save(update_fields=['lot_name'])
        t = engine.create_transfer(self.prj, self.user, 'Н-1')
        line = engine.add_transfer_line(t, self.lot, D(1))
        self.assertIn('ЗН-42', line.display_name)               # авто-метка лота

    def test_number_required_at_lock_not_at_birth(self):
        """Ф12e: накладная рождается без номера, но не фиксируется без него."""
        t = engine.create_transfer(self.prj, self.user)
        engine.add_transfer_line(t, self.lot, D(1))
        with self.assertRaises(ValidationError):
            engine.lock_transfer(t)
        engine.update_transfer(t, number='Н-7')
        engine.lock_transfer(t)
        self.assertTrue(models.Transfer.objects.get(pk=t.pk).locked)

    def test_add_line_rejects_foreign_project_lot(self):
        other = models.Project.objects.create(
            code='P2', description='Проект 2', kind=models.Project.Kind.EXTERNAL)
        t = engine.create_transfer(other, self.user, 'Н-2')
        with self.assertRaises(ValidationError):
            engine.add_transfer_line(t, self.lot, D(1))         # лот чужого проекта

    def test_add_line_rejects_nonpositive(self):
        t = engine.create_transfer(self.prj, self.user, 'Н-1')
        with self.assertRaises(ValidationError):
            engine.add_transfer_line(t, self.lot, D(0))

    def test_over_issue_drives_negative_not_clamped(self):
        t = engine.create_transfer(self.prj, self.user, 'Н-1')
        engine.add_transfer_line(t, self.lot, D(8))             # больше остатка 5
        engine.lock_transfer(t)
        self.assertEqual(engine.lot_live_qty(self.lot), D(-3))  # недостача информативна

    def test_update_qty_rebuilds_and_remove_restores(self):
        t = engine.create_transfer(self.prj, self.user, 'Н-1')
        line = engine.add_transfer_line(t, self.lot, D(2))
        engine.update_transfer_line(line, qty=D(4))
        engine.lock_transfer(t)
        self.assertEqual(engine.lot_live_qty(self.lot), D(1))   # 5 − 4
        engine.unlock_transfer(t)                               # Ф15: правка — под замком снятым
        engine.update_transfer_line(line, display_name='новое имя')
        line.refresh_from_db()
        self.assertEqual(line.display_name, 'новое имя')
        engine.remove_transfer_line(line)
        self.assertEqual(engine.lot_live_qty(self.lot), D(5))   # источник восстановлен

    def test_available_lots_picker(self):
        # ещё один лот проекта + чужой проект + нулевой остаток → в пикере только живой свой
        other = models.Project.objects.create(
            code='P2', description='Проект 2', kind=models.Project.Kind.EXTERNAL)
        self.receipt_lot(self.device, other, 3)                 # чужой проект
        t = engine.create_transfer(self.prj, self.user, 'Н-1')
        engine.add_transfer_line(t, self.lot, D(5))             # свой лот в ноль
        engine.lock_transfer(t)                                 # Ф15: отгружено фиксацией
        picker = engine.project_available_lots(self.prj)
        self.assertEqual(picker, [])                            # живых своих лотов нет
        self.assertEqual(len(engine.project_available_lots(other)), 1)

    def test_post_locks_edits(self):
        t = engine.create_transfer(self.prj, self.user, 'Н-1')
        line = engine.add_transfer_line(t, self.lot, D(2))
        engine.lock_transfer(t)
        self.assertTrue(engine.transfer_form(t)['locked'])
        with self.assertRaises(ValidationError):
            engine.update_transfer_line(line, qty=D(1))
        with self.assertRaises(ValidationError):
            engine.add_transfer_line(t, self.lot, D(1))
        with self.assertRaises(ValidationError):
            engine.remove_transfer_line(line)

    def test_post_rejects_empty(self):
        t = engine.create_transfer(self.prj, self.user, 'Н-1')
        with self.assertRaises(ValidationError):
            engine.lock_transfer(t)

    def test_unpost_reenables_edits(self):
        t = engine.create_transfer(self.prj, self.user, 'Н-1')
        line = engine.add_transfer_line(t, self.lot, D(2))
        engine.lock_transfer(t)
        engine.unlock_transfer(t)
        self.assertFalse(engine.transfer_form(t)['locked'])
        self.assertEqual(engine.lot_live_qty(self.lot), D(5))  # Ф15: отгрузка отпущена
        engine.update_transfer_line(line, qty=D(3))            # снова можно править
        engine.lock_transfer(t)
        self.assertEqual(engine.lot_live_qty(self.lot), D(2))

    def test_item_movements_projection(self):
        """Лента движений изделия (волна 19, Ф12a): рождение партии + все ордера,
        которые её двигали. Знак сохраняем — расход виден расходом."""
        t = engine.create_transfer(self.prj, self.user, 'Н-7')
        engine.add_transfer_line(t, self.lot, D(2), display_name='Прибор №7')
        rows = engine.item_movements(self.device)
        self.assertEqual(len(rows), 2)                  # рождение + передача
        born = next(r for r in rows if r['event'] == 'born')
        move = next(r for r in rows if r['event'] == 'move')
        self.assertEqual(born['lot_id'], self.lot.id)
        self.assertEqual(born['qty'], self.lot.qty)
        self.assertEqual(move['kind'], models.StockDocument.Kind.TRANSFER)
        self.assertEqual(move['number'], 'Н-7')
        self.assertEqual(move['qty'], D(-2))            # знаковое: отгрузка = расход
        self.assertFalse(move['locked'])


class WriteoffFormTests(EngineTestBase):
    """Волна 6: списание — чистый `−ISSUE` из проекта (серый путь), guard чужого
    проекта, коррекция строк, пересписание в минус (не клампим)."""

    def setUp(self):
        super().setUp()
        self.item = self.make_item('R100')
        self.lot = self.receipt_lot(self.item, self.prj, 10)

    def test_add_line_issues_from_lot(self):
        w = engine.create_writeoff(self.prj, self.user, 'СП-1', reason='брак')
        engine.add_writeoff_line(w, self.lot, D(4))
        # Ф15: черновой акт ничего не списывает — партия целиком на месте
        self.assertEqual(engine.lot_live_qty(self.lot), D(10))
        engine.lock_writeoff(w)                                  # проведён
        self.assertEqual(engine.lot_live_qty(self.lot), D(6))
        self.assertTrue(self.lot.movements.filter(type='ISSUE', qty=D(-4)).exists())

    def test_form_totals(self):
        w = engine.create_writeoff(self.prj, self.user, 'СП-1', reason='брак')
        engine.add_writeoff_line(w, self.lot, D(4))
        engine.lock_writeoff(w)
        c = engine.writeoff_form(w)
        self.assertEqual(c['number'], 'СП-1')
        self.assertEqual(c['reason'], 'брак')
        self.assertEqual(c['total_qty'], D(4))
        self.assertEqual(c['lines'][0]['lot_live_qty'], D(6))

    def test_number_required_at_lock_not_at_birth(self):
        """Ф12e: акт рождается без номера, но не фиксируется без него."""
        w = engine.create_writeoff(self.prj, self.user)
        engine.add_writeoff_line(w, self.lot, D(1))
        with self.assertRaises(ValidationError):
            engine.lock_writeoff(w)
        engine.update_writeoff(w, number='СП-7')
        engine.lock_writeoff(w)
        self.assertTrue(models.Writeoff.objects.get(pk=w.pk).locked)

    def test_add_line_rejects_foreign_project(self):
        other = models.Project.objects.create(
            code='P2', description='Проект 2', kind=models.Project.Kind.EXTERNAL)
        w = engine.create_writeoff(other, self.user, 'СП-2')
        with self.assertRaises(ValidationError):
            engine.add_writeoff_line(w, self.lot, D(1))

    def test_add_line_rejects_nonpositive(self):
        w = engine.create_writeoff(self.prj, self.user, 'СП-1')
        with self.assertRaises(ValidationError):
            engine.add_writeoff_line(w, self.lot, D(0))

    def test_over_writeoff_negative_not_clamped(self):
        w = engine.create_writeoff(self.prj, self.user, 'СП-1')
        engine.add_writeoff_line(w, self.lot, D(14))
        engine.lock_writeoff(w)
        self.assertEqual(engine.lot_live_qty(self.lot), D(-4))

    def test_update_and_remove_restores(self):
        w = engine.create_writeoff(self.prj, self.user, 'СП-1')
        line = engine.add_writeoff_line(w, self.lot, D(4))
        engine.update_writeoff_line(line, D(7))
        engine.lock_writeoff(w)
        self.assertEqual(engine.lot_live_qty(self.lot), D(3))
        engine.unlock_writeoff(w)                      # Ф15: списание отпущено
        engine.remove_writeoff_line(line)
        self.assertEqual(engine.lot_live_qty(self.lot), D(10))


class RequisitionFormTests(EngineTestBase):
    """Волна 6: требование/отпочкование — `−ISSUE` источника + рождение
    лота-потомка (`+RECEIPT`) у получателя с наследованием цены/провенанса."""

    def setUp(self):
        super().setUp()
        self.item = self.make_item('R100')
        self.src = self.receipt_lot(self.item, self.prj, 10)
        self.src.unit_cost = D('2.50')
        self.src.part_number = 'ЗН-9'
        self.src.save(update_fields=['unit_cost', 'part_number'])
        self.white = models.Project.objects.create(
            code='WHITE', description='Собственный склад',
            kind=models.Project.Kind.INTERNAL_STOCK)

    def test_add_line_issues_source_and_births_child(self):
        req = engine.create_requisition(self.white, self.user, 'ТР-1')
        engine.add_requisition_line(req, self.src, D(4))
        # Ф15: черновое требование не двигает ни источник, ни потомка
        self.assertEqual(engine.lot_live_qty(self.src), D(10))
        engine.lock_requisition(req)
        self.assertEqual(engine.lot_live_qty(self.src), D(6))    # источник просел
        born = engine._requisition_born_lot(req, self.src)
        self.assertIsNotNone(born)
        self.assertEqual(born.project_id, self.white.id)
        self.assertEqual(born.qty, D(4))
        self.assertEqual(born.unit_cost, D('2.50'))              # цена унаследована
        self.assertEqual(born.predecessor_id, self.src.id)      # генеалогия
        self.assertEqual(engine.lot_live_qty(born), D(4))       # +RECEIPT у потомка

    def test_form_shows_source_and_born(self):
        req = engine.create_requisition(self.white, self.user, 'ТР-1')
        engine.add_requisition_line(req, self.src, D(4))
        engine.lock_requisition(req)
        c = engine.requisition_form(req)
        self.assertEqual(c['total_qty'], D(4))
        row = c['lines'][0]
        self.assertEqual(row['source_project_code'], self.prj.code)
        self.assertEqual(row['source_live_qty'], D(6))
        self.assertIsNotNone(row['born_lot_id'])

    def test_same_project_rejected(self):
        req = engine.create_requisition(self.prj, self.user, 'ТР-1')
        with self.assertRaises(ValidationError):
            engine.add_requisition_line(req, self.src, D(1))    # источник = получатель

    def test_duplicate_source_rejected(self):
        req = engine.create_requisition(self.white, self.user, 'ТР-1')
        engine.add_requisition_line(req, self.src, D(2))
        with self.assertRaises(ValidationError):
            engine.add_requisition_line(req, self.src, D(1))

    def test_update_syncs_source_and_child(self):
        req = engine.create_requisition(self.white, self.user, 'ТР-1')
        line = engine.add_requisition_line(req, self.src, D(4))
        engine.update_requisition_line(line, D(7))
        engine.lock_requisition(req)
        born = engine._requisition_born_lot(req, self.src)
        self.assertEqual(engine.lot_live_qty(self.src), D(3))
        self.assertEqual(engine.lot_live_qty(born), D(7))

    def test_remove_restores_source_and_deletes_child(self):
        req = engine.create_requisition(self.white, self.user, 'ТР-1')
        line = engine.add_requisition_line(req, self.src, D(4))
        engine.lock_requisition(req)
        self.assertEqual(engine.lot_live_qty(self.src), D(6))
        engine.unlock_requisition(req)                  # Ф15: отпочкование отпущено
        engine.remove_requisition_line(line)
        self.assertEqual(engine.lot_live_qty(self.src), D(10))
        self.assertIsNone(engine._requisition_born_lot(req, self.src))

    def test_all_available_lots_picker(self):
        rows = engine.all_available_lots()
        self.assertTrue(any(r['lot_id'] == self.src.id and
                            r['project_code'] == self.prj.code for r in rows))


class ProjectClosureTests(EngineTestBase):
    """Волна 6: панель закрытия (остаточные лоты → 0) + мягкий замок статуса +
    мосты «списать»/«на баланс»."""

    def setUp(self):
        super().setUp()
        self.item = self.make_item('R100')
        self.lot = self.receipt_lot(self.item, self.prj, 10)

    def test_closure_lists_residuals_and_blocks(self):
        c = engine.project_closure(self.prj)
        self.assertEqual(len(c['residuals']), 1)
        self.assertEqual(c['residual_positive'], D(10))
        self.assertFalse(c['can_close'])
        with self.assertRaises(ValidationError):
            engine.lock_project(self.prj)

    def test_writeoff_bridge_then_close(self):
        # Ф15: мост кладёт остаток в черновой акт, в 0 он уходит на фиксации акта
        w = engine.writeoff_lot(self.prj, self.lot, D(10), self.user)
        self.assertEqual(engine.project_closure(self.prj)['residual_positive'], D(10))
        engine.lock_writeoff(w)
        c = engine.project_closure(self.prj)
        self.assertEqual(c['residuals'], [])
        self.assertTrue(c['can_close'])
        engine.lock_project(self.prj)
        self.prj.refresh_from_db()
        self.assertTrue(self.prj.locked)
        # Ф1c: дата `closed` информационная — фиксация её НЕ штампует.
        self.assertIsNone(self.prj.closed)

    def test_requisition_bridge_moves_to_white(self):
        req = engine.requisition_lot(self.prj, self.lot, D(10), self.user)
        engine.lock_requisition(req)                            # Ф15: перекладка = фиксация
        self.assertEqual(engine.lot_live_qty(self.lot), D(0))
        white = engine._internal_project(models.Project.Kind.INTERNAL_STOCK)
        moved = engine.item_available(self.item, white)
        self.assertEqual(moved, D(10))                          # оказалось на балансе
        self.assertTrue(engine.project_closure(self.prj)['can_close'])

    def test_negative_residual_is_anomaly_and_blocks(self):
        w = engine.create_writeoff(self.prj, self.user, 'СП-1')
        engine.add_writeoff_line(w, self.lot, D(14))            # пересписали → −4
        engine.lock_writeoff(w)
        c = engine.project_closure(self.prj)
        self.assertEqual(c['anomaly_count'], 1)
        self.assertTrue(c['residuals'][0]['anomaly'])
        self.assertFalse(c['can_close'])

    def test_internal_project_not_closable(self):
        white = models.Project.objects.create(
            code='WHITE', description='Собственный склад',
            kind=models.Project.Kind.INTERNAL_STOCK)
        c = engine.project_closure(white)
        self.assertFalse(c['is_external'])
        self.assertFalse(c['can_close'])
        with self.assertRaises(ValidationError):
            engine.lock_project(white)

    def test_unlock_restores_editable(self):
        engine.lock_writeoff(engine.writeoff_lot(self.prj, self.lot, D(10), self.user))
        engine.lock_project(self.prj)
        engine.unlock_project(self.prj)
        self.prj.refresh_from_db()
        self.assertFalse(self.prj.locked)


class HeaderEditTests(EngineTestBase):
    """Волна 6 (докрутка): правка шапки форм — номер/дата/мягкие поля,
    read-only под замком, nullable-дата очищается."""

    def test_transfer_header_edit_and_lock(self):
        dev = self.make_item('DEV', manufactured=True, kind='device')
        lot = self.receipt_lot(dev, self.prj, 5)
        t = engine.create_transfer(self.prj, self.user, 'Н-1')
        engine.update_transfer(t, number='Н-99', date='2026-06-15')
        t.refresh_from_db()
        self.assertEqual(t.number, 'Н-99')
        self.assertEqual(str(t.date), '2026-06-15')
        engine.add_transfer_line(t, lot, D(1))
        engine.lock_transfer(t)
        with self.assertRaises(ValidationError):
            engine.update_transfer(t, number='Н-100')        # под замком нельзя

    def test_draft_header_clears_but_lock_still_demands_it(self):
        """Аудит-1 Б2б-4: «заполнить можно, передумать нельзя» — не наш принцип.
        Номер и дата обязательны к ФИКСАЦИИ (`REQUIRED_HEADER_BY_KIND`), а не к
        черновику: ошибся документом — стирай и заводи заново."""
        lot = self.receipt_lot(self.make_item('DEV2', manufactured=True), self.prj, 5)
        t = engine.create_transfer(self.prj, self.user, 'Н-1')
        engine.update_transfer(t, number='   ', date='')
        t.refresh_from_db()
        self.assertEqual(t.number, '')
        self.assertIsNone(t.date)
        engine.add_transfer_line(t, lot, D(1))
        with self.assertRaises(ValidationError):
            engine.lock_transfer(t)                          # неполную шапку не выпустим
        engine.update_transfer(t, number='Н-2', date='2026-06-15')
        engine.lock_transfer(t)
        t.refresh_from_db()
        self.assertTrue(t.locked)

    def test_receipt_header_locked(self):
        r = models.Receipt.objects.create(number='U-1', date='2026-05-01',
            contractor=self.supplier, project=self.prj, user=self.user)
        engine.add_receipt_lot(r, self.make_item('A'), D(2))
        engine.update_receipt(r, number='U-2')
        r.refresh_from_db()
        self.assertEqual(r.number, 'U-2')
        engine.lock_receipt(r)
        with self.assertRaises(ValidationError):
            engine.update_receipt(r, number='U-3')

    def test_purchase_code_description_and_clear_date(self):
        p = engine.create_purchase(self.prj, self.user, date='2026-05-01', description='x')
        engine.update_purchase(p, code='Нева-1', description='новое', date='')  # '' → NULL (nullable)
        p.refresh_from_db()
        self.assertEqual(p.code, 'Нева-1')
        self.assertEqual(p.description, 'новое')
        self.assertIsNone(p.date)

    def test_kitting_qty_rescales_needs(self):
        comp = self.make_item('R')
        self.receipt_lot(comp, self.prj, 100)
        dev = self.make_item('DEV', manufactured=True)
        models.BomLine.objects.create(parent=dev, component=comp, qty=D(2))
        k = models.Kitting.objects.create(project=self.prj, target_item=dev,
            user=self.user, qty=D(1), locked=False)
        self.assertEqual(engine.kitting_form(k)['rows'][0]['need'], D(2))
        engine.update_kitting(k, qty=D(3))
        self.assertEqual(engine.kitting_form(k)['rows'][0]['need'], D(6))
        engine.lock_kitting(k)
        with self.assertRaises(ValidationError):
            engine.update_kitting(k, qty=D(4))               # не wip — нельзя

    def test_writeoff_and_requisition_header(self):
        w = engine.create_writeoff(self.prj, self.user, 'СП-1')
        engine.update_writeoff(w, number='СП-2', reason='брак')
        w.refresh_from_db()
        self.assertEqual((w.number, w.reason), ('СП-2', 'брак'))
        white = models.Project.objects.create(code='WHITE', description='Склад',
            kind=models.Project.Kind.INTERNAL_STOCK)
        req = engine.create_requisition(white, self.user, 'ТР-1')
        engine.update_requisition(req, number='ТР-2', date='2026-06-01')
        req.refresh_from_db()
        self.assertEqual(req.number, 'ТР-2')


class ScopeDeficitTests(EngineTestBase):
    """Свод по охвату — Σ проектных дефицитов по Item, без перенеттинга (В7; охват В19)."""

    def _device_with_screw(self, screw, qty_per, suffix=''):
        dev = self.make_item(f'DEV{screw.code}{suffix}', manufactured=True,
                             kind='device')
        models.BomLine.objects.create(parent=dev, component=screw, qty=D(qty_per))
        return dev

    def test_rolls_up_by_item_across_projects(self):
        scr = self.make_item('SCR', kind='material')
        dev = self._device_with_screw(scr, 4)
        prj2 = models.Project.objects.create(code='P2', description='Проект 2',
                                             kind=models.Project.Kind.EXTERNAL)
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(10))
        models.ProjectDemand.objects.create(project=prj2, target_item=dev, qty=D(5))

        rows = {r['item_code']: r
                for r in engine.scope_deficit([self.prj, prj2])['rows']}
        row = rows['SCR']
        self.assertEqual(row['need'], D(60))         # 40 + 20
        self.assertEqual(row['to_order'], D(60))     # склада/заказов нет
        self.assertEqual(row['status'], 'to_order')
        self.assertEqual(len(row['by_project']), 2)

    def test_stock_and_order_no_cross_project_netting(self):
        scr = self.make_item('SCR', kind='material')
        dev = self._device_with_screw(scr, 4)
        prj2 = models.Project.objects.create(code='P2', description='Проект 2',
                                             kind=models.Project.Kind.EXTERNAL)
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(10))
        models.ProjectDemand.objects.create(project=prj2, target_item=dev, qty=D(5))
        # склад лежит только в P1 (10 шт) — НЕ должен гасить нужду P2
        self.receipt_lot(scr, self.prj, 10)

        row = {r['item_code']: r
               for r in engine.scope_deficit([self.prj, prj2])['rows']}['SCR']
        self.assertEqual(row['have'], D(10))          # только P1 покрыт
        self.assertEqual(row['to_order'], D(50))      # 30 (P1) + 20 (P2), не 40
        self.assertEqual(row['need'], D(60))

    def test_closed_and_internal_projects_excluded(self):
        scr = self.make_item('SCR', kind='material')
        dev = self._device_with_screw(scr, 2)
        closed = models.Project.objects.create(code='PC', description='Закрытый',
            kind=models.Project.Kind.EXTERNAL, locked=True)
        white = models.Project.objects.create(code='WHITE', description='Склад',
            kind=models.Project.Kind.INTERNAL_STOCK)
        models.ProjectDemand.objects.create(project=closed, target_item=dev, qty=D(3))
        models.ProjectDemand.objects.create(project=white, target_item=dev, qty=D(3))
        # оба в охвате явно — и всё равно не считаются: закрытый проект не закупают,
        # а внутренний склад не потребитель (инвариант арифметики, не фильтр вызова)
        self.assertEqual(engine.scope_deficit([closed, white])['rows'], [])

    def test_intra_project_need_aggregated_across_demands(self):
        # два прибора в одном проекте делят компонент → потребность суммируется,
        # покрытие считается один раз (одна by_project-строка на проект)
        scr = self.make_item('SCR', kind='material')
        dev_a = self._device_with_screw(scr, 4, suffix='A')
        dev_b = self._device_with_screw(scr, 3, suffix='B')
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev_a, qty=D(2))
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev_b, qty=D(2))

        row = {r['item_code']: r
               for r in engine.scope_deficit([self.prj])['rows']}['SCR']
        self.assertEqual(row['need'], D(14))          # 2×4 + 2×3
        self.assertEqual(len(row['by_project']), 1)   # агрегат по проекту

    def test_sorted_worst_first(self):
        red = self.make_item('RED', kind='material')
        green = self.make_item('GRN', kind='material')
        dev = self.make_item('DEVX', manufactured=True, kind='device')
        models.BomLine.objects.create(parent=dev, component=red, qty=D(1))
        models.BomLine.objects.create(parent=dev, component=green, qty=D(1))
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(5))
        self.receipt_lot(green, self.prj, 100)        # GRN покрыт ✓, RED красный ▲

        codes = [r['item_code'] for r in engine.scope_deficit([self.prj])['rows']]
        self.assertEqual(codes, ['RED', 'GRN'])       # красное наверх


class ProcurementFormTests(EngineTestBase):
    """Волна 7: записываемый план закупки — строки, замок отправки, мост, xlsx."""

    def test_create_and_form_totals(self):
        p = engine.create_procurement(self.user, description='весна')
        engine.add_procurement_line(p, self.make_item('A'), D(10))
        engine.add_procurement_line(p, self.make_item('B'), D(5))
        c = engine.procurement_form(p)
        self.assertEqual(len(c['lines']), 2)
        self.assertEqual(c['total_qty'], D(15))
        self.assertTrue(c['editable'])
        self.assertFalse(c['locked'])
        self.assertEqual(c['description'], 'весна')

    def test_estimate_sums_lines_and_names_gaps(self):
        """«Оценка» шапки: Σ(кол-во × оценка изделия) + коды позиций без оценки.

        Позиция без `estimated_cost` в сумму не идёт и НЕ считается нулём молча — её код
        уезжает в `unestimated`, чтобы форма показала «оценка неполна» (2026-08-06).
        Заказ считает ту же величину по своим строкам.
        """
        priced = self.make_item('A')
        priced.estimated_cost = D('12.50')
        priced.save(update_fields=['estimated_cost'])
        blank = self.make_item('B')                    # без оценки — пробел

        p = engine.create_procurement(self.user)
        engine.add_procurement_line(p, priced, D(10))
        engine.add_procurement_line(p, blank, D(5))
        c = engine.procurement_form(p)
        self.assertEqual(c['estimate'], D('125.00'))
        self.assertEqual(c['unestimated'], ['B'])

        pu = self.make_purchase()
        engine.add_purchase_line(pu, priced, D(4))
        self.assertEqual(engine.purchase_form(pu)['estimate'], D('50.00'))
        self.assertEqual(engine.purchase_form(pu)['unestimated'], [])

    def test_intent_money_demand_and_overpay(self):
        """Панель бюджета намерения (2026-08-07): потребность / сумма / разница.

        Потребность берёт количество НЕ из строки, а из нужды проекта (BOM ×
        `ProjectDemand`, разузлование до листьев); строка задаёт только номенклатуру.
        Цена у обоих чисел одна — `estimated_cost`, поэтому переплата = сумма − нужда.
        """
        screw = self.make_item('SCR')
        screw.estimated_cost = D(10)
        screw.save(update_fields=['estimated_cost'])
        dev = self.make_item('DEV', manufactured=True)
        models.BomLine.objects.create(parent=dev, component=screw, qty=D(4))
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(3))

        pu = self.make_purchase()                      # нужно 12 шт (3×4) = 120 ₽
        engine.add_purchase_line(pu, screw, D(20))     # заказали 20 шт = 200 ₽
        c = engine.purchase_form(pu)
        self.assertEqual(c['demand'], D(120))
        self.assertEqual(c['estimate'], D(200))
        self.assertEqual(c['overpay'], D(80))          # + переплата

    def test_intent_money_overpay_negative_is_undersupply(self):
        """Заказали меньше нужды — разница уходит в минус («Недозаказ» в панели)."""
        screw = self.make_item('SCR')
        screw.estimated_cost = D(10)
        screw.save(update_fields=['estimated_cost'])
        dev = self.make_item('DEV', manufactured=True)
        models.BomLine.objects.create(parent=dev, component=screw, qty=D(4))
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(3))

        pu = self.make_purchase()
        engine.add_purchase_line(pu, screw, D(5))      # 50 ₽ против нужды 120 ₽
        self.assertEqual(engine.purchase_form(pu)['overpay'], D(-70))

    def test_intent_money_unestimated_drops_from_both_sums(self):
        """Позиция без оценки выпадает из ОБЕИХ сумм разом — знак разницы не врёт."""
        blank = self.make_item('B')                    # без `estimated_cost`
        dev = self.make_item('DEV', manufactured=True)
        models.BomLine.objects.create(parent=dev, component=blank, qty=D(2))
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(5))

        pu = self.make_purchase()
        engine.add_purchase_line(pu, blank, D(50))
        c = engine.purchase_form(pu)
        self.assertEqual((c['demand'], c['estimate'], c['overpay']), (D(0), D(0), D(0)))
        self.assertEqual(c['unestimated'], ['B'])

    def test_intent_money_ignores_items_outside_bom(self):
        """Строка, которой в составе проекта нет: нужда 0 → вся сумма в переплату."""
        stray = self.make_item('X')
        stray.estimated_cost = D(7)
        stray.save(update_fields=['estimated_cost'])

        pu = self.make_purchase()
        engine.add_purchase_line(pu, stray, D(3))
        c = engine.purchase_form(pu)
        self.assertEqual(c['demand'], D(0))
        self.assertEqual(c['overpay'], D(21))

    def test_intent_money_of_procurement_counts_scope(self):
        """У закупки знаменатель — ОХВАТ (проекты её заказов), сумма по проектам.

        Пока заказов нет, охват пуст: спросить «сколько надо» не у кого, потребность 0 и
        весь план читается переплатой. Появился заказ — появилась и нужда его проекта.
        """
        screw = self.make_item('SCR')
        screw.estimated_cost = D(10)
        screw.save(update_fields=['estimated_cost'])
        dev = self.make_item('DEV', manufactured=True)
        models.BomLine.objects.create(parent=dev, component=screw, qty=D(2))
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(3))

        p = engine.create_procurement(self.user)
        engine.add_procurement_line(p, screw, D(30))
        c = engine.procurement_form(p)
        self.assertEqual(c['demand'], D(0))            # охват пуст — сравнивать не с чем
        self.assertEqual(c['overpay'], D(300))

        pu = self.make_purchase()                      # заказ проекта под этот план
        engine.update_purchase(pu, procurement=p)
        c = engine.procurement_form(p)
        self.assertEqual(c['demand'], D(60))           # 3×2 шт × 10 ₽
        self.assertEqual(c['overpay'], D(240))

    def test_intent_money_of_procurement_skips_closed_project(self):
        """Отсев охвата — общий с витриной «К закупке»: закрытый проект не закупают."""
        screw = self.make_item('SCR')
        screw.estimated_cost = D(10)
        screw.save(update_fields=['estimated_cost'])
        dev = self.make_item('DEV', manufactured=True)
        models.BomLine.objects.create(parent=dev, component=screw, qty=D(2))
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(3))

        p = engine.create_procurement(self.user)
        engine.add_procurement_line(p, screw, D(30))
        pu = self.make_purchase()
        engine.update_purchase(pu, procurement=p)
        self.prj.locked = True
        self.prj.save(update_fields=['locked'])
        self.assertEqual(engine.procurement_form(p)['demand'], D(0))

    def test_line_cost_and_costly_quartile(self):
        """Стоимость строки + оранжевая верхняя четверть набора (2026-08-07).

        Ранжирование — внутри одного списка (в разных документах дорого разное), доля
        `ceil(N/4)`: из четырёх строк дорогая одна. Позиция без оценки в ранжировании не
        участвует и стоимости не имеет (`None`, а не ноль).
        """
        pu = self.make_purchase()
        dev = self.make_item('DEV', manufactured=True)
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(1))
        for code, price, qty in [('A', D(100), D(10)),    # 1000 — самая дорогая
                                 ('B', D(10), D(10)),     # 100
                                 ('C', D(1), D(10)),      # 10
                                 ('D', None, D(10))]:     # без оценки — стоимости нет
            item = self.make_item(code)
            if price is not None:
                item.estimated_cost = price
                item.save(update_fields=['estimated_cost'])
            # нужда с запасом (100 шт против 10 в заказе) — красный «перебор» не мешает
            models.BomLine.objects.create(parent=dev, component=item, qty=D(100))
            engine.add_purchase_line(pu, item, qty)

        rows = {r['item_code']: r for r in engine.purchase_form(pu)['rows']}
        self.assertEqual(rows['A']['cost'], D(1000))
        self.assertIsNone(rows['D']['cost'])
        # ранжируются трое оценённых → ceil(3/4) = 1 дорогая
        self.assertEqual(rows['A']['cost_status'], 'costly')
        self.assertIsNone(rows['B']['cost_status'])
        self.assertIsNone(rows['C']['cost_status'])
        self.assertIsNone(rows['D']['cost_status'])

    def test_overpaid_beats_costly_and_needs_half_over(self):
        """Красный «перебор» сильнее оранжевого и загорается строго за 1.5× нужды."""
        priced = []
        for code in ('A', 'B'):
            item = self.make_item(code)
            item.estimated_cost = D(10)
            item.save(update_fields=['estimated_cost'])
            priced.append(item)
        dev = self.make_item('DEV', manufactured=True)
        for item in priced:
            models.BomLine.objects.create(parent=dev, component=item, qty=D(2))
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(5))

        pu = self.make_purchase()                      # нужда каждой позиции = 10 шт
        engine.add_purchase_line(pu, priced[0], D(16))  # 1.6× — перебор, и она же дороже
        engine.add_purchase_line(pu, priced[1], D(15))  # ровно 1.5× — ещё не перебор
        rows = {r['item_code']: r for r in engine.purchase_form(pu)['rows']}
        self.assertEqual(rows['A']['cost_status'], 'overpaid')   # красный, не 'costly'
        self.assertIsNone(rows['B']['cost_status'])

    def test_purchase_line_carries_project_balance(self):
        """Строка заказа несёт баланс проекта по своему изделию (2026-08-07).

        То же число и та же четвёрка слагаемых, что в «Потребности» проекта и в ячейке
        «Привязки». Свой черновик баланс не двигает (считается только зафиксированное) —
        значит он не гаснет, пока набиваешь заказ.
        """
        screw = self.make_item('SCR')
        dev = self.make_item('DEV', manufactured=True)
        models.BomLine.objects.create(parent=dev, component=screw, qty=D(2))
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(5))
        self.receipt_lot(screw, self.prj, 4)           # на складе 4 при нужде 10

        pu = self.make_purchase()
        engine.add_purchase_line(pu, screw, D(6))
        row = engine.purchase_form(pu)['rows'][0]
        self.assertEqual((row['need'], row['in_stock']), (D(10), D(4)))
        self.assertEqual(row['balance'], D(-6))        # черновик своего заказа не считается
        self.assertEqual(row['balance_status'], 'to_order')

        engine.lock_purchase(pu)                       # зафиксировали — заказ вошёл в баланс
        row = engine.purchase_form(pu)['rows'][0]
        self.assertEqual((row['on_order'], row['balance']), (D(6), D(0)))
        self.assertEqual(row['balance_status'], 'on_order')   # сошлось впритык

    def test_overpay_threshold_rounds_up_to_round_numbers(self):
        """Порог = полторы нормы, округлённые ВВЕРХ до круглого (правка Ивана 2026-08-07).

        Округляем порог, а не потребность: иначе «нужно 13, беру круглые 20» ловилось
        красным (19.5), хотя некратные числа никто не заказывает. Шкала ступеней —
        5 → 10 → 50 → 100 по величине самого порога.
        """
        cases = {1: 5, 3: 5, 7: 15, 13: 20, 40: 60, 100: 150, 333: 500, 0: 0}
        for need, threshold in cases.items():
            self.assertEqual(engine.overpay_threshold(D(need)), D(threshold), need)
        # ...и сам гейт: ровно порог — ещё не перебор, шаг за него — уже да.
        self.assertFalse(engine.is_overpaid(D(20), D(13)))
        self.assertTrue(engine.is_overpaid(D(21), D(13)))

    def test_overpaid_when_item_not_needed_at_all(self):
        """Нужды нет вовсе (позиции нет в BOM) — тоже перебор (решение Ивана)."""
        stray = self.make_item('X')
        stray.estimated_cost = D(7)
        stray.save(update_fields=['estimated_cost'])
        pu = self.make_purchase()
        engine.add_purchase_line(pu, stray, D(3))
        self.assertEqual(engine.purchase_form(pu)['rows'][0]['cost_status'], 'overpaid')

    def test_allocation_costs_on_both_levels(self):
        """«Привязка»: стоимость у строки плана и у ячейки заказа, каждая по своему кол-ву.

        Оранжевого у ячеек не бывает — верхнюю четверть ищем среди строк плана. Красный
        работает на обоих уровнях: у строки плана знаменатель — нужда охвата, у ячейки —
        нужда её проекта.
        """
        screw = self.make_item('SCR')
        screw.estimated_cost = D(10)
        screw.save(update_fields=['estimated_cost'])
        dev = self.make_item('DEV', manufactured=True)
        models.BomLine.objects.create(parent=dev, component=screw, qty=D(2))
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(5))

        p = engine.create_procurement(self.user)
        engine.add_procurement_line(p, screw, D(20))   # нужда охвата 10 шт → 2× = перебор
        pu = self.make_purchase()
        engine.update_purchase(pu, procurement=p)
        engine.set_allocation(p, pu, screw, D(12))     # в заказ 12 при нужде 10 — не перебор

        row = engine.procurement_allocation(p)['rows'][0]
        self.assertEqual(row['cost'], D(200))
        self.assertEqual(row['cost_status'], 'overpaid')
        cell = row['orders'][0]
        self.assertEqual(cell['cost'], D(120))
        self.assertIsNone(cell['cost_status'])

    def test_add_line_rejects_duplicate_and_nonpositive(self):
        p = engine.create_procurement(self.user)
        item = self.make_item('A')
        engine.add_procurement_line(p, item, D(10))
        with self.assertRaises(ValidationError):
            engine.add_procurement_line(p, item, D(3))
        with self.assertRaises(ValidationError):
            engine.add_procurement_line(p, self.make_item('B'), D(0))

    def test_update_and_remove_line(self):
        p = engine.create_procurement(self.user)
        line = engine.add_procurement_line(p, self.make_item('A'), D(10))
        engine.update_procurement_line(line, D(7))
        line.refresh_from_db()
        self.assertEqual(line.qty, D(7))
        engine.remove_procurement_line(line)
        self.assertFalse(models.ProcurementLine.objects.filter(pk=line.pk).exists())

    def test_post_locks_and_rejects_empty(self):
        p = engine.create_procurement(self.user)
        with self.assertRaises(ValidationError):
            engine.lock_procurement(p)                 # пустую нельзя
        line = engine.add_procurement_line(p, self.make_item('A'), D(10))
        engine.lock_procurement(p)
        self.assertTrue(p.locked)
        with self.assertRaises(ValidationError):
            engine.update_procurement_line(line, D(5))  # строки под замком
        with self.assertRaises(ValidationError):
            engine.add_procurement_line(p, self.make_item('B'), D(1))

    def test_unpost_reopens_and_delete_replaces_cancel(self):
        """Замок снимается и возвращается; отмены нет — только удаление (Р1)."""
        p = engine.create_procurement(self.user)
        line = engine.add_procurement_line(p, self.make_item('A'), D(10))
        engine.lock_procurement(p)
        engine.unlock_procurement(p)
        engine.update_procurement_line(line, D(30))    # снова можно
        self.assertEqual(line.qty, D(30))
        engine.lock_procurement(p)
        with self.assertRaises(ValidationError):        # утверждённую не удалить
            engine.delete_procurement(p)
        engine.unlock_procurement(p)
        self.assertFalse(p.locked)
        engine.delete_procurement(p)
        self.assertFalse(models.Procurement.objects.filter(pk=p.pk).exists())

    def test_bridge_tops_up_line_without_doubling(self):
        # Ф13: мост кладёт в ЭТУ закупку и доводит строку ДО наводки (топ-ап), а не
        # плюсует: повторный клик по той же строке витрины ничего не удваивает.
        item = self.make_item('SCR', kind='material')
        p = engine.create_procurement(self.user)
        engine.add_to_procurement(p, item, D(15))
        self.assertEqual(p.lines.get(item=item).qty, D(15))
        engine.add_to_procurement(p, item, D(15))
        self.assertEqual(p.lines.get(item=item).qty, D(15))
        engine.add_to_procurement(p, item, D(25))        # наводка выросла — догоняем
        self.assertEqual(p.lines.get(item=item).qty, D(25))

    def test_bridge_never_cuts_manual_qty(self):
        # набранное руками сверх наводки топ-ап не срезает (мутабельная ДНК: расхождение
        # информативнее молчаливого выравнивания)
        item = self.make_item('SCR', kind='material')
        p = engine.create_procurement(self.user)
        engine.add_procurement_line(p, item, D(100))
        engine.add_to_procurement(p, item, D(10))
        self.assertEqual(p.lines.get(item=item).qty, D(100))

    def test_bridge_refuses_locked_procurement(self):
        item = self.make_item('SCR', kind='material')
        p = engine.create_procurement(self.user)
        engine.add_procurement_line(p, self.make_item('OTH'), D(1))
        engine.lock_procurement(p)
        with self.assertRaises(ValidationError):
            engine.add_to_procurement(p, item, D(5))

    def test_update_header(self):
        p = engine.create_procurement(self.user)
        engine.update_procurement(p, date='2026-07-10', code='ЗАК-1', description='осень')
        p.refresh_from_db()
        self.assertEqual(str(p.date), '2026-07-10')
        self.assertEqual(p.code, 'ЗАК-1')
        self.assertEqual(p.description, 'осень')
        engine.update_procurement(p, date='')          # пустая строка → NULL
        p.refresh_from_db()
        self.assertIsNone(p.date)

    def test_code_soft_uniqueness(self):
        # Волна 19, Ф10 (правило Ивана): занятый код ловим дружелюбно (ValidationError),
        # не IntegrityError/500. Пустой код у нескольких — легально (NULL).
        engine.update_procurement(engine.create_procurement(self.user), code='ДЗЗ-1')
        with self.assertRaises(ValidationError):
            engine.update_procurement(engine.create_procurement(self.user), code='ДЗЗ-1')
        # два без кода не конфликтуют
        engine.create_procurement(self.user)
        engine.create_procurement(self.user)      # не бросает

    def test_update_contractor(self):
        # Волна 19, Ф4: контрагент-поставщик у закупки-плана. Часовой `_UNSET`:
        # не передан → не трогаем; Counterparty → выставить; None → снять.
        p = engine.create_procurement(self.user)
        self.assertIsNone(p.contractor_id)
        engine.update_procurement(p, contractor=self.supplier)
        p.refresh_from_db()
        self.assertEqual(p.contractor_id, self.supplier.id)
        # проекция формы отдаёт контрагента
        cock = engine.procurement_form(p)
        self.assertEqual(cock['contractor_id'], self.supplier.id)
        self.assertEqual(cock['contractor_name'], self.supplier.description)
        # не передан → не трогаем (правка только даты)
        engine.update_procurement(p, date='2026-07-10')
        p.refresh_from_db()
        self.assertEqual(p.contractor_id, self.supplier.id)
        # None → снять
        engine.update_procurement(p, contractor=None)
        p.refresh_from_db()
        self.assertIsNone(p.contractor_id)
        self.assertEqual(engine.procurement_form(p)['contractor_name'], '')

    def test_contractor_set_null_on_counterparty_delete(self):
        # SET_NULL: удаление контрагента не роняет план-закупку — поле опустевает.
        p = engine.create_procurement(self.user)
        engine.update_procurement(p, contractor=self.supplier)
        self.supplier.delete()
        p.refresh_from_db()
        self.assertIsNone(p.contractor_id)

    def _xlsx_rows(self, procurement):
        """Лист бланка списком строк — бланк уходит наружу, читаем его как получатель."""
        from io import BytesIO

        from openpyxl import load_workbook
        data = engine.procurement_xlsx(procurement)
        self.assertTrue(data)                          # непустой байт-поток
        ws = load_workbook(BytesIO(data)).active
        return [[c.value for c in row] for row in ws.iter_rows()]

    def test_xlsx_lists_positions_numbered_in_screen_order(self):
        """Ф4b: у позиции есть «№», и нумерация идёт в порядке экрана, не по id изделия."""
        p = engine.create_procurement(self.user)
        # заводим изделия в обратном алфавитном порядке: если бы файл сортировал по
        # чему-то своему, номера разъехались бы с табом «Строки» (тот же `order_by('id')`)
        engine.add_procurement_line(p, self.make_item('R900'), D(12))
        engine.add_procurement_line(p, self.make_item('R100'), D(5))
        rows = self._xlsx_rows(p)
        head = rows.index(['№', 'Артикул', 'Наименование', 'Кол-во', 'Ед.'])
        self.assertEqual(rows[head + 1][:4], [1, 'R900', 'R900', 12])
        self.assertEqual(rows[head + 2][:4], [2, 'R100', 'R100', 5])
        screen = [ln['item_code'] for ln in engine.procurement_form(p)['lines']]
        self.assertEqual([r[1] for r in rows[head + 1:]], screen)

    def _xlsx_head(self, procurement):
        """Шапка бланка словарём. Она живёт в колонках B/C (A узкая — под «№»), поэтому
        подпись читаем из второй ячейки, а не первой."""
        return {r[1]: r[2] for r in self._xlsx_rows(procurement)
                if r[0] is None and r[1]}

    def test_xlsx_head_carries_both_parties_date_and_author(self):
        """Ф4b: шапка запроса — заказчик (мы) · контрагент · дата · автор."""
        us = models.Counterparty.objects.create(description='Наша сторона', inn='7700000000')
        p = engine.create_procurement(self.user, date='2026-07-30')
        engine.update_procurement(p, contractor=self.supplier)
        engine.add_procurement_line(p, self.make_item('R100'), D(1))
        with self.settings(ORG_COUNTERPARTY_ID=us.id):
            head = self._xlsx_head(p)
            self.assertEqual(head['Заказчик'], 'Наша сторона · ИНН 7700000000')
            self.assertEqual(head['Контрагент'], self.supplier.description)  # ИНН не задан
            self.assertEqual(head['Дата запроса'], '30.07.2026')
            self.assertEqual(head['Автор'], self.user.get_username())
            # та же дата после перечитывания из БД (в памяти она держится строкой из
            # JSON — бланк обязан выдержать оба состояния, документ уходит наружу)
            p.refresh_from_db()
            self.assertEqual(self._xlsx_head(p)['Дата запроса'], '30.07.2026')

    def test_xlsx_head_labels_never_sit_in_the_narrow_number_column(self):
        """Подписи шапки — в колонке B, значения в C: в узкой A (она под «№») Excel
        обрезал бы подпись по границе, потому что соседняя ячейка занята. Поймано
        глазами на реальном файле — ассерты на значения ячеек такого не видят."""
        us = models.Counterparty.objects.create(description='Наша сторона')
        p = engine.create_procurement(self.user, date='2026-07-30')
        engine.add_procurement_line(p, self.make_item('R100'), D(1))
        with self.settings(ORG_COUNTERPARTY_ID=us.id):
            rows = self._xlsx_rows(p)
        table = next(i for i, r in enumerate(rows) if r[0] == '№')
        for r in rows[:table]:                            # всё, что выше таблицы
            self.assertIsNone(r[0], 'колонка A в шапке обязана быть пустой')
            self.assertTrue(r[3] is None and r[4] is None,  # значению есть куда перетечь
                            'в строке шапки заполнены только B и C')

    def test_xlsx_survives_unset_organization_and_empty_head(self):
        """Незаданная строка шапки не рисуется вовсе: пустая графа во внешнем документе
        читается как брак, а не как «значение не задано» (внутри продукта — прочерк)."""
        p = engine.create_procurement(self.user)          # без контрагента и без даты
        engine.add_procurement_line(p, self.make_item('R100'), D(1))
        with self.settings(ORG_COUNTERPARTY_ID=None):     # мы не настроены
            head = self._xlsx_head(p)
            rows = self._xlsx_rows(p)
        self.assertNotIn('Заказчик', head)
        self.assertNotIn('Контрагент', head)
        self.assertNotIn('Дата запроса', head)
        self.assertIn('Автор', head)                      # автор есть всегда (NOT NULL)
        self.assertTrue(any(r[0] == '№' for r in rows))   # таблица позиций на месте

    def test_our_organization_ignores_id_of_deleted_counterparty(self):
        """Настройка указывает на удалённую запись — выгрузка не падает, мы просто
        «не настроены» (окружение и справочник живут раздельно)."""
        gone = models.Counterparty.objects.create(description='Была')
        gone_id = gone.id
        gone.delete()
        with self.settings(ORG_COUNTERPARTY_ID=gone_id):
            self.assertIsNone(engine.our_organization())


class ItemXlsxTests(EngineTestBase):
    """Выгрузка изделия в xlsx (2026-07-30): листы = вкладки формы, кроме «Файлов».

    Живёт рядом с бланком закупки: обе выгрузки рисует один `_xlsx_sheet` (шапка B/C,
    жирные заголовки, сквозной «№»), и расхождение между ними ловится здесь.
    """

    def setUp(self):
        super().setUp()
        self.dev = self.make_item('DEV', manufactured=True)
        self.scr = self.make_item('SCR')
        models.BomLine.objects.create(parent=self.dev, component=self.scr, qty=D(4))

    def _book(self, item, scope=engine.ITEM_XLSX_BOM):
        """Книга словарём `{имя листа: [строки]}` — читаем файл как получатель."""
        from io import BytesIO

        from openpyxl import load_workbook
        data = engine.item_xlsx(item, scope)
        self.assertTrue(data)
        wb = load_workbook(BytesIO(data))
        return {ws.title: [[c.value for c in row] for row in ws.iter_rows()]
                for ws in wb.worksheets}

    def test_bom_scope_is_one_sheet_with_direct_components(self):
        """«Только состав» = один лист «Состав», один уровень BOM (не разузлование)."""
        leaf = self.make_item('LEAF')
        models.BomLine.objects.create(parent=self.scr, component=leaf, qty=D(2))
        book = self._book(self.dev)
        self.assertEqual(list(book), ['Состав'])
        rows = book['Состав']
        head = rows.index(['№', 'Компонент', 'Описание', 'Кол-во', 'Ед.'])
        self.assertEqual([r[1] for r in rows[head + 1:]], ['SCR'])   # без LEAF
        self.assertEqual(rows[head + 1][:4], [1, 'SCR', 'SCR', 4])

    def test_first_sheet_carries_item_head_and_others_are_bare_tables(self):
        """Шапка изделия — только на первом листе; остальные листы = чистые таблицы
        (идентичность им даёт имя файла = `code`)."""
        engine.update_item(self.dev, {'uom': 'шт', 'estimated_cost': D('12.50')})
        book = self._book(self.dev, engine.ITEM_XLSX_ALL)
        rows = book['Состав']
        head = {r[1]: r[2] for r in rows if r[0] is None and r[1]}
        self.assertEqual(head['Код'], 'DEV')
        self.assertEqual(head['Единицы'], 'шт')
        self.assertEqual(head['Оценка'], 12.5)
        self.assertEqual(head['Категория'], self.dev.category.code)
        for name in ('Применение', 'Склад', 'Движения'):
            self.assertIsNotNone(book[name][0][0])       # первая же строка — заголовки
            self.assertEqual(book[name][0][0], '№')

    def test_all_scope_covers_every_tab_but_files(self):
        """«Все вкладки»: состав · применение · склад · движения. Вложений в книге нет
        (файл — снимок таблиц, а не архив)."""
        lot = self.receipt_lot(self.scr, self.prj, 6)
        t = engine.create_transfer(self.prj, self.user, 'Н-7')
        engine.add_transfer_line(t, lot, D(2))
        book = self._book(self.scr, engine.ITEM_XLSX_ALL)
        self.assertEqual(list(book), ['Состав', 'Применение', 'Склад', 'Движения'])
        self.assertNotIn('Файлы', book)
        # применение: где используется покупной винт
        self.assertEqual([r[1] for r in book['Применение'][1:]], ['DEV'])
        # склад: партия с живым остатком (6 приехало − 2 отдано... передача не заперта)
        self.assertEqual(book['Склад'][1][1:5], [f'#{lot.id}', 'P1', 6, 6])
        # движения: рождение помечено, расход виден расходом (знак сохраняем)
        moves = {r[2]: r for r in book['Движения'][1:]}
        self.assertIn('Поставка · партия рождена', moves)
        self.assertEqual(moves['Передача'][6], -2)

    def test_draft_lot_has_no_live_qty_in_stock_sheet(self):
        """Ф15: партия нефиксированного origin показывает прочерк, а не 0 — «едет,
        ещё не принято» (ноль читался бы как «израсходована»)."""
        lot = self.receipt_lot(self.scr, self.prj, 5, locked=False)
        row = self._book(self.scr, engine.ITEM_XLSX_ALL)['Склад'][1]
        self.assertEqual(row[1], f'#{lot.id}')
        self.assertEqual(row[4], '—')

    def test_component_without_bom_still_gets_an_empty_sheet(self):
        """У покупного вкладки «Состав» нет вовсе, а лист есть: пустая таблица честнее
        отказа скачать — видно, что состава нет, а не что кнопка сломалась."""
        rows = self._book(self.make_item('R100'))['Состав']
        self.assertEqual(rows[-1], ['№', 'Компонент', 'Описание', 'Кол-во', 'Ед.'])


class ItemXlsxHttpTests(TestCase):
    """HTTP-путь выгрузки изделия: scope из query, имя файла = `code` (RFC 5987)."""

    def setUp(self):
        get_user_model().objects.create(username='admin', is_superuser=True)
        self.c = Client()
        self.c.force_login(get_user_model().objects.get(is_superuser=True))
        self.item = models.Item.objects.create(code='Нева БУ 1', description='Блок',
            category=_cat(), native=True)

    def test_download_is_xlsx_attachment_named_by_code(self):
        r = self.c.get(f'/api/items/{self.item.id}/xlsx/')
        self.assertEqual(r.status_code, 200)
        self.assertIn('spreadsheetml', r['Content-Type'])
        self.assertTrue(r['Content-Disposition'].startswith('attachment'))
        self.assertIn(quote('Нева БУ 1.xlsx'), r['Content-Disposition'])
        self.assertEqual(r.content[:2], b'PK')            # zip-сигнатура xlsx

    def test_scope_query_picks_sheet_set_and_unknown_falls_back_to_bom(self):
        from io import BytesIO

        from openpyxl import load_workbook
        names = lambda resp: load_workbook(BytesIO(resp.content)).sheetnames
        self.assertEqual(names(self.c.get(f'/api/items/{self.item.id}/xlsx/?scope=all')),
                         ['Состав', 'Применение', 'Склад', 'Движения'])
        self.assertEqual(names(self.c.get(f'/api/items/{self.item.id}/xlsx/?scope=bom')),
                         ['Состав'])
        # опечатка в query — не повод ругаться: отдаём состав
        self.assertEqual(names(self.c.get(f'/api/items/{self.item.id}/xlsx/?scope=xx')),
                         ['Состав'])

    def test_empty_code_falls_back_to_id(self):
        self.item.code = ''
        self.item.save(update_fields=['code'])
        cd = self.c.get(f'/api/items/{self.item.id}/xlsx/')['Content-Disposition']
        self.assertIn(quote(f'изделие-{self.item.id}.xlsx'), cd)


class AllocationTests(EngineTestBase):
    """Привязка (волна 8, переделана 2026-08-05): раскладка плана по ЗАКАЗАМ закупки."""

    def setUp(self):
        super().setUp()
        self.prj2 = models.Project.objects.create(
            code='P2', description='Проект 2', kind=models.Project.Kind.EXTERNAL)
        self.scr = self.make_item('SCR', kind='material')
        dev = self.make_item('DEV', manufactured=True, kind='device')
        models.BomLine.objects.create(parent=dev, component=self.scr, qty=D(4))
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(10))
        models.ProjectDemand.objects.create(project=self.prj2, target_item=dev, qty=D(5))
        self.plan = engine.create_procurement(self.user, description='свод')   # need 40 + 20
        engine.update_procurement(self.plan, contractor=self.supplier)
        engine.add_procurement_line(self.plan, self.scr, D(60))
        # Заказы заводятся РУКАМИ и привязываются к закупке — движок их не рождает.
        self.ord1 = self._order(self.prj)
        self.ord2 = self._order(self.prj2)

    def _order(self, project):
        """Заказ под проект, привязанный к закупке — как это делает человек руками."""
        pu = engine.create_purchase(project, self.user)
        engine.update_purchase(pu, procurement=self.plan, contractor=self.supplier)
        return pu

    def test_scope_is_derived_from_orders(self):
        """Охват больше не хранится: это проекты заказов закупки."""
        self.assertEqual([p.code for p in engine.procurement_scope(self.plan)],
                         ['P1', 'P2'])
        empty = engine.create_procurement(self.user)
        self.assertEqual(list(engine.procurement_scope(empty)), [])   # нет заказов = пусто

    def test_set_allocation_assigns_not_adds(self):
        engine.set_allocation(self.plan, self.ord1, self.scr, D(40))
        self.assertEqual(self.ord1.lines.get(item=self.scr).qty, D(40))
        # повтор — присвоение, а не инкремент (в этом вся разница с прежним пегом)
        engine.set_allocation(self.plan, self.ord1, self.scr, D(25))
        self.assertEqual(self.ord1.lines.get(item=self.scr).qty, D(25))

    def test_zero_removes_the_line(self):
        engine.set_allocation(self.plan, self.ord1, self.scr, D(40))
        engine.set_allocation(self.plan, self.ord1, self.scr, D(0))
        self.assertFalse(self.ord1.lines.filter(item=self.scr).exists())
        # и появляется снова, если вписать число
        engine.set_allocation(self.plan, self.ord1, self.scr, D(7))
        self.assertEqual(self.ord1.lines.get(item=self.scr).qty, D(7))

    def test_allocation_guards(self):
        with self.assertRaises(ValidationError):            # item не в плане
            engine.set_allocation(self.plan, self.ord1, self.make_item('OTH'), D(1))
        with self.assertRaises(ValidationError):            # отрицательное количество
            engine.set_allocation(self.plan, self.ord1, self.scr, D(-1))
        alien = engine.create_procurement(self.user)
        engine.add_procurement_line(alien, self.scr, D(5))
        with self.assertRaises(ValidationError):            # заказ под другой закупкой
            engine.set_allocation(alien, self.ord1, self.scr, D(1))
        engine.set_allocation(self.plan, self.ord1, self.scr, D(10))
        engine.lock_purchase(self.ord1)
        with self.assertRaises(ValidationError):            # зафиксированный заказ
            engine.set_allocation(self.plan, self.ord1, self.scr, D(1))

    def test_matrix_row_has_a_cell_per_order(self):
        """Матрица прямоугольная: строка плана × все заказы закупки, пустое — тоже факт."""
        engine.set_allocation(self.plan, self.ord1, self.scr, D(40))
        row = engine.procurement_allocation(self.plan)['rows'][0]
        cells = {c['purchase_id']: c for c in row['orders']}
        self.assertEqual(set(cells), {self.ord1.id, self.ord2.id})
        self.assertEqual(cells[self.ord1.id]['qty'], D(40))
        self.assertEqual(cells[self.ord2.id]['qty'], D(0))     # заказ есть, ячейка пуста
        self.assertEqual(cells[self.ord1.id]['project_code'], 'P1')
        self.assertEqual(row['allocated'], D(40))
        self.assertEqual(row['remaining'], D(20))

    def test_row_status_by_allocation(self):
        row = lambda: engine.procurement_allocation(self.plan)['rows'][0]
        self.assertEqual(row()['status'], 'to_order')          # не тронута — красный
        engine.set_allocation(self.plan, self.ord1, self.scr, D(40))
        self.assertEqual(row()['status'], 'on_order')          # тронута, не подбита
        engine.set_allocation(self.plan, self.ord2, self.scr, D(20))
        self.assertEqual(row()['status'], 'available')         # разложена в ноль
        engine.set_allocation(self.plan, self.ord2, self.scr, D(30))
        self.assertEqual(row()['remaining'], D(-10))           # перепег законен
        self.assertEqual(row()['status'], 'to_order')          # но красный: не сходится

    def test_cell_carries_project_balance_untouched_by_drafts(self):
        """Баланс в ячейке — тот же, что в «Потребности»: черновая раскладка его не двигает."""
        cells = lambda: {c['purchase_id']: c
                         for c in engine.procurement_allocation(self.plan)['rows'][0]['orders']}
        self.assertEqual(cells()[self.ord1.id]['need'], D(40))
        self.assertEqual(cells()[self.ord1.id]['balance'], D(-40))
        engine.set_allocation(self.plan, self.ord1, self.scr, D(40))
        self.assertEqual(cells()[self.ord1.id]['balance'], D(-40))   # черновик не считается
        engine.lock_purchase(self.ord1)
        self.assertEqual(cells()[self.ord1.id]['balance'], D(0))     # фиксация — момент истины
        self.assertEqual(cells()[self.ord1.id]['balance_status'], 'on_order')

    def test_two_orders_of_one_project_share_the_balance(self):
        """Баланс — свойство ПРОЕКТА: у двух его заказов число одинаково (принято 08-05)."""
        extra = self._order(self.prj)
        engine.set_allocation(self.plan, self.ord1, self.scr, D(30))
        engine.set_allocation(self.plan, extra, self.scr, D(10))
        cells = {c['purchase_id']: c
                 for c in engine.procurement_allocation(self.plan)['rows'][0]['orders']}
        self.assertEqual(cells[self.ord1.id]['balance'], cells[extra.id]['balance'])
        engine.lock_purchase(self.ord1)
        cells = {c['purchase_id']: c
                 for c in engine.procurement_allocation(self.plan)['rows'][0]['orders']}
        self.assertEqual(cells[extra.id]['balance'], D(-10))   # сходится само

    def test_locked_order_stays_visible_in_the_matrix(self):
        """Зафиксированный заказ из привязки НЕ исчезает: «сюда уже заказано» — контекст."""
        engine.set_allocation(self.plan, self.ord1, self.scr, D(40))
        engine.lock_purchase(self.ord1)
        cells = {c['purchase_id']: c
                 for c in engine.procurement_allocation(self.plan)['rows'][0]['orders']}
        self.assertTrue(cells[self.ord1.id]['locked'])
        self.assertEqual(cells[self.ord1.id]['qty'], D(40))

    def test_fan_lists_orders_of_the_plan(self):
        engine.set_allocation(self.plan, self.ord1, self.scr, D(40))
        fan = {f['purchase_id']: f for f in engine.procurement_allocation(self.plan)['fan']}
        self.assertEqual(fan[self.ord1.id]['lines'], 1)
        self.assertEqual(fan[self.ord1.id]['total'], D(40))
        self.assertEqual(fan[self.ord2.id]['lines'], 0)

    def test_plan_list_shows_every_procurement(self):
        """Ф17: список закупок = ВСЕ закупки — прятать больше нечего."""
        engine.create_purchase(self.prj, self.user)         # заказ без плана
        c = Client()
        c.force_login(self.user)
        ids = {row['id'] for row in c.get('/api/procurements/').json()}
        self.assertEqual(ids, {self.plan.id})               # заказ пустышки не создал


class ItemUsageTests(EngineTestBase):
    """Ф5: обратное разузлование — «зачем этот Item проекту» (третий уровень аккордеона)."""

    def _tree(self):
        # DEV ×1 → SUB ×2 → SCR ×3  (плюс SCR ×1 напрямую в DEV)
        scr = self.make_item('SCR', kind='material')
        sub = self.make_item('SUB', manufactured=True)
        dev = self.make_item('DEV', manufactured=True, kind='device')
        models.BomLine.objects.create(parent=sub, component=scr, qty=D(3))
        models.BomLine.objects.create(parent=dev, component=sub, qty=D(2))
        models.BomLine.objects.create(parent=dev, component=scr, qty=D(1))
        return dev, sub, scr

    def test_usage_counts_through_all_levels(self):
        dev, sub, scr = self._tree()
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(10))
        usage = engine.item_usage_in_project(scr, self.prj)
        self.assertEqual(len(usage), 1)                     # одно применение — прибор DEV
        self.assertEqual(usage[0]['target_code'], 'DEV')
        self.assertEqual(usage[0]['per_unit'], D(7))        # 2×3 через SUB + 1 напрямую
        self.assertEqual(usage[0]['demand_qty'], D(10))
        self.assertEqual(usage[0]['total'], D(70))          # совпадает с нуждой проекта

    def test_usage_splits_by_target_device(self):
        dev, sub, scr = self._tree()
        dev2 = self.make_item('DEV2', manufactured=True, kind='device')
        models.BomLine.objects.create(parent=dev2, component=scr, qty=D(5))
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(1))
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev2, qty=D(2))
        usage = {u['target_code']: u for u in engine.item_usage_in_project(scr, self.prj)}
        self.assertEqual(usage['DEV']['total'], D(7))
        self.assertEqual(usage['DEV2']['total'], D(10))
        # сумма применений = нужда проекта по этому Item (сходимость с прямым сводом)
        need = engine.scope_deficit([self.prj])['rows']
        row = {r['item_code']: r for r in need}['SCR']
        self.assertEqual(sum(u['total'] for u in usage.values()), row['need'])

    def test_usage_empty_for_item_outside_bom(self):
        dev, sub, scr = self._tree()
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(1))
        self.assertEqual(engine.item_usage_in_project(self.make_item('OTH'), self.prj), [])


class ClosureHttpTests(TestCase):
    """Волна 6: HTTP-путь через test Client — провязка urls/views + мапинг ошибок."""

    def setUp(self):
        get_user_model().objects.create(username='admin', is_superuser=True)
        self.main = models.Location.objects.create(code='MAIN', description='Основной склад')
        self.prj = models.Project.objects.create(
            code='P1', description='Проект 1', kind=models.Project.Kind.EXTERNAL)
        self.item = models.Item.objects.create(code='R100', description='R100', category=_cat())
        self.sup = models.Counterparty.objects.create(description='П')
        r = models.Receipt.objects.create(number='U-1', date='2026-05-01',
            contractor=self.sup, project=self.prj, locked=True,   # Ф15: партия на складе
            user=get_user_model().objects.first())
        self.lot = models.Lot.objects.create(item=self.item, project=self.prj,
            origin=r, qty=D(10))
        engine.rebuild_movements(self.lot)
        self.c = Client()
        # Волна 12: весь /api/ за логином — HTTP-путь ходит от суперюзера-админа.
        self.c.force_login(get_user_model().objects.get(is_superuser=True))

    def test_writeoff_flow(self):
        r = self.c.post('/api/writeoffs/', {'project_id': self.prj.id,
            'number': 'СП-1', 'reason': 'брак'}, content_type='application/json')
        self.assertEqual(r.status_code, 201)
        wid = r.json()['id']
        r = self.c.post(f'/api/writeoffs/{wid}/lines/',
            {'lot_id': self.lot.id, 'qty': 4}, content_type='application/json')
        self.assertEqual(r.status_code, 201)
        self.assertEqual(float(r.json()['total_qty']), 4.0)
        # чужой проект → 400
        other = models.Project.objects.create(code='P2', description='П2',
            kind=models.Project.Kind.EXTERNAL)
        r2 = self.c.post('/api/writeoffs/', {'project_id': other.id, 'number': 'СП-2'},
            content_type='application/json')
        w2 = r2.json()['id']
        bad = self.c.post(f'/api/writeoffs/{w2}/lines/',
            {'lot_id': self.lot.id, 'qty': 1}, content_type='application/json')
        self.assertEqual(bad.status_code, 400)

    def test_closure_bridges_and_lock(self):
        panel = self.c.get(f'/api/projects/{self.prj.id}/closure/').json()
        self.assertFalse(panel['can_close'])
        self.assertEqual(len(panel['residuals']), 1)
        # мост «на баланс» → черновое требование; Ф15: остаток уйдёт на его фиксации
        r = self.c.post(f'/api/projects/{self.prj.id}/stock-lot/',
            {'lot_id': self.lot.id, 'qty': 10}, content_type='application/json')
        self.assertEqual(r.status_code, 200)
        self.assertFalse(r.json()['can_close'])
        drafts = r.json()['closing_drafts']
        self.assertEqual(len(drafts), 1)                    # панель показывает черновик
        self.assertEqual(drafts[0]['kind'], 'requisition')
        lock = self.c.post(f'/api/requisitions/{drafts[0]["document_id"]}/lock/')
        self.assertEqual(lock.status_code, 200)
        panel = self.c.get(f'/api/projects/{self.prj.id}/closure/').json()
        self.assertTrue(panel['can_close'])
        # закрытие 200 + повторное закрытие → 400
        c1 = self.c.post(f'/api/projects/{self.prj.id}/lock/')
        self.assertEqual(c1.status_code, 200)
        self.assertTrue(c1.json()['locked'])
        c2 = self.c.post(f'/api/projects/{self.prj.id}/lock/')
        self.assertEqual(c2.status_code, 400)
        # переоткрытие
        ro = self.c.post(f'/api/projects/{self.prj.id}/unlock/')
        self.assertFalse(ro.json()['locked'])

    def test_requisition_flow(self):
        white = models.Project.objects.create(code='WHITE', description='Собственный склад',
            kind=models.Project.Kind.INTERNAL_STOCK)
        r = self.c.post('/api/requisitions/', {'project_id': white.id, 'number': 'ТР-1'},
            content_type='application/json')
        self.assertEqual(r.status_code, 201)
        rid = r.json()['id']
        line = self.c.post(f'/api/requisitions/{rid}/lines/',
            {'source_lot_id': self.lot.id, 'qty': 3}, content_type='application/json')
        self.assertEqual(line.status_code, 201)
        self.assertIsNotNone(line.json()['lines'][0]['born_lot_id'])
        picker = self.c.get('/api/available-lots/')
        self.assertEqual(picker.status_code, 200)


class ProcurementHttpTests(TestCase):
    """Волна 7: HTTP-путь — свод, записываемый Procurement, мост, xlsx-бланк."""

    def setUp(self):
        get_user_model().objects.create(username='admin', is_superuser=True)
        self.prj = models.Project.objects.create(
            code='P1', description='Проект 1', kind=models.Project.Kind.EXTERNAL)
        self.scr = models.Item.objects.create(code='SCR', description='Винт',
            category=_cat())
        self.dev = models.Item.objects.create(code='DEV', description='Прибор',
            category=_cat(), native=True)
        models.BomLine.objects.create(parent=self.dev, component=self.scr, qty=D(4))
        models.ProjectDemand.objects.create(project=self.prj, target_item=self.dev,
            qty=D(10))
        self.c = Client()
        # Волна 12: весь /api/ за логином — HTTP-путь ходит от суперюзера-админа.
        self.c.force_login(get_user_model().objects.get(is_superuser=True))

    def test_scope_deficit_and_bridge(self):
        # Витрина живёт ВНУТРИ закупки и считает по её охвату, а охват — проекты её
        # заказов (2026-08-05): пока заказов нет, закупка честно слепа.
        pid = self.c.post('/api/procurements/', {'description': 'весна'},
            content_type='application/json').json()['id']
        blind = self.c.get(f'/api/procurements/{pid}/deficit/')
        self.assertEqual(blind.status_code, 200)
        self.assertEqual(blind.json()['rows'], [])            # пусто = пусто
        oid = self.c.post('/api/purchases/', {'project_id': self.prj.id},
            content_type='application/json').json()['id']
        scope = self.c.patch(f'/api/purchases/{oid}/', {'procurement_id': pid},
            content_type='application/json')
        self.assertEqual(scope.status_code, 200)
        form = self.c.get(f'/api/procurements/{pid}/').json()
        self.assertEqual([p['code'] for p in form['projects']], ['P1'])
        svod = self.c.get(f'/api/procurements/{pid}/deficit/')
        rows = {r['item_code']: r for r in svod.json()['rows']}
        self.assertEqual(float(rows['SCR']['to_order']), 40.0)
        self.assertEqual(float(rows['SCR']['planned']), 0.0)   # в плане ещё ничего
        # мост кладёт позицию В ЭТУ закупку
        add = self.c.post(f'/api/procurements/{pid}/take/',
            {'item_id': self.scr.id, 'qty': 40}, content_type='application/json')
        self.assertEqual(add.status_code, 200)
        self.assertEqual(float(add.json()['total_qty']), 40.0)
        again = self.c.get(f'/api/procurements/{pid}/deficit/').json()
        self.assertEqual(float(again['rows'][0]['planned']), 40.0)

    def test_internal_project_never_reaches_the_scope_deficit(self):
        """Внутренний склад — источник покрытия, а не потребитель.

        Гейта на охвате больше нет (охват вычисляемый, запрещать нечему): отсев живёт в
        самой арифметике свода, куда бы заказ ни был привязан.
        """
        white = models.Project.objects.create(code='WHITE', description='Свой склад',
            kind=models.Project.Kind.INTERNAL_STOCK)
        pid = self.c.post('/api/procurements/', {}, content_type='application/json').json()['id']
        oid = self.c.post('/api/purchases/', {'project_id': white.id},
            content_type='application/json').json()['id']
        self.c.patch(f'/api/purchases/{oid}/', {'procurement_id': pid},
            content_type='application/json')
        form = self.c.get(f'/api/procurements/{pid}/').json()
        self.assertEqual([p['code'] for p in form['projects']], ['WHITE'])  # заказ есть
        svod = self.c.get(f'/api/procurements/{pid}/deficit/')
        self.assertEqual(svod.json()['rows'], [])            # но в расчёт не идёт

    def test_procurement_crud_lock_and_xlsx(self):
        r = self.c.post('/api/procurements/', {'description': 'весна'},
            content_type='application/json')
        self.assertEqual(r.status_code, 201)
        pid = r.json()['id']
        line = self.c.post(f'/api/procurements/{pid}/lines/',
            {'item_id': self.scr.id, 'qty': 12}, content_type='application/json')
        self.assertEqual(line.status_code, 201)
        # дубль item → 400
        dup = self.c.post(f'/api/procurements/{pid}/lines/',
            {'item_id': self.scr.id, 'qty': 1}, content_type='application/json')
        self.assertEqual(dup.status_code, 400)
        # утверждение → строки под замком
        posted = self.c.post(f'/api/procurements/{pid}/lock/')
        self.assertEqual(posted.status_code, 200)
        self.assertTrue(posted.json()['locked'])
        locked = self.c.post(f'/api/procurements/{pid}/lines/',
            {'item_id': self.dev.id, 'qty': 1}, content_type='application/json')
        self.assertEqual(locked.status_code, 400)
        # выгрузка xlsx — бинарное тело, xlsx content-type
        xlsx = self.c.get(f'/api/procurements/{pid}/xlsx/')
        self.assertEqual(xlsx.status_code, 200)
        self.assertIn('spreadsheetml', xlsx['Content-Type'])
        self.assertTrue(xlsx['Content-Disposition'].startswith('attachment'))
        self.assertTrue(xlsx.content[:2] == b'PK')      # zip-сигнатура xlsx

    def test_xlsx_filename_is_procurement_code(self):
        """Имя файла = `code` закупки (2026-07-26), с фолбэком по id у пустого кода.

        Коды человек вводит кириллицей и с пробелами («Нева ДЗЗ 1») — заголовок обязан
        нести RFC 5987 (`filename*=utf-8''…`), иначе имя приедет мусором.
        """
        r = self.c.post('/api/procurements/', {'code': 'Нева ДЗЗ 1'},
            content_type='application/json')
        pid = r.json()['id']
        cd = self.c.get(f'/api/procurements/{pid}/xlsx/')['Content-Disposition']
        self.assertIn("filename*=utf-8''", cd)
        self.assertIn(quote('Нева ДЗЗ 1.xlsx'), cd)
        # Ф12e: рождённая без кода закупка получает фолбэк «Закупка 42» ещё в движке,
        # поэтому имя файла осмысленно уже здесь.
        r2 = self.c.post('/api/procurements/', {'description': 'без кода'},
            content_type='application/json')
        pid2 = r2.json()['id']
        cd2 = self.c.get(f'/api/procurements/{pid2}/xlsx/')['Content-Disposition']
        self.assertIn(quote(f'Закупка {pid2}.xlsx'), cd2)
        # Код можно очистить руками — тогда работает страховочный фолбэк по id.
        engine.update_procurement(models.Procurement.objects.get(pk=pid2), code='')
        cd3 = self.c.get(f'/api/procurements/{pid2}/xlsx/')['Content-Disposition']
        self.assertIn(quote(f'закупка-{pid2}.xlsx'), cd3)
        self.assertNotIn('order', cd3)


class AllocationHttpTests(TestCase):
    """HTTP-путь «Привязки» (2026-08-05): проекция-матрица, присвоение в ячейку, гварды."""

    def setUp(self):
        get_user_model().objects.create(username='admin', is_superuser=True)
        self.prj = models.Project.objects.create(code='P1', description='Проект 1',
            kind=models.Project.Kind.EXTERNAL)
        self.prj2 = models.Project.objects.create(code='P2', description='Проект 2',
            kind=models.Project.Kind.EXTERNAL)
        self.sup = models.Counterparty.objects.create(code='КЭ', description='Поставщик')
        self.scr = models.Item.objects.create(code='SCR', description='Винт',
            category=_cat())
        dev = models.Item.objects.create(code='DEV', description='Прибор',
            category=_cat(), native=True)
        models.BomLine.objects.create(parent=dev, component=self.scr, qty=D(4))
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(10))
        models.ProjectDemand.objects.create(project=self.prj2, target_item=dev, qty=D(5))
        self.c = Client()
        # Волна 12: весь /api/ за логином — HTTP-путь ходит от суперюзера-админа.
        self.c.force_login(get_user_model().objects.get(is_superuser=True))
        self.pid = self.c.post('/api/procurements/', {'description': 'свод'},
            content_type='application/json').json()['id']
        self.c.post(f'/api/procurements/{self.pid}/lines/',
            {'item_id': self.scr.id, 'qty': 60}, content_type='application/json')
        self.ord1 = self._order(self.prj)
        self.ord2 = self._order(self.prj2)

    def _order(self, project):
        oid = self.c.post('/api/purchases/', {'project_id': project.id},
            content_type='application/json').json()['id']
        self.c.patch(f'/api/purchases/{oid}/',
            {'procurement_id': self.pid, 'contractor_id': self.sup.id},
            content_type='application/json')
        return oid

    def test_matrix_projection(self):
        r = self.c.get(f'/api/procurements/{self.pid}/allocation/')
        self.assertEqual(r.status_code, 200)
        row = r.json()['rows'][0]
        self.assertEqual(row['item_code'], 'SCR')
        self.assertEqual(float(row['allocated']), 0.0)
        cells = {c['purchase_id']: c for c in row['orders']}
        self.assertEqual(set(cells), {self.ord1, self.ord2})     # ячейка на каждый заказ
        self.assertEqual(float(cells[self.ord1]['balance']), -40.0)
        self.assertEqual(len(r.json()['fan']), 2)

    def test_allocate_assigns_and_clears(self):
        r = self.c.post(f'/api/procurements/{self.pid}/allocate/',
            {'purchase_id': self.ord1, 'item_id': self.scr.id, 'qty': 40},
            content_type='application/json')
        self.assertEqual(r.status_code, 200)
        self.assertEqual(float(r.json()['rows'][0]['allocated']), 40.0)
        # присвоение, не добавление
        r = self.c.post(f'/api/procurements/{self.pid}/allocate/',
            {'purchase_id': self.ord1, 'item_id': self.scr.id, 'qty': 15},
            content_type='application/json')
        self.assertEqual(float(r.json()['rows'][0]['allocated']), 15.0)
        # ноль снимает строку
        r = self.c.post(f'/api/procurements/{self.pid}/allocate/',
            {'purchase_id': self.ord1, 'item_id': self.scr.id, 'qty': 0},
            content_type='application/json')
        self.assertEqual(float(r.json()['rows'][0]['allocated']), 0.0)
        self.assertEqual(r.json()['rows'][0]['status'], 'to_order')

    def test_allocate_guards(self):
        x = models.Item.objects.create(code='X', description='X', category=_cat())
        bad = self.c.post(f'/api/procurements/{self.pid}/allocate/',
            {'purchase_id': self.ord1, 'item_id': x.id, 'qty': 1},
            content_type='application/json')
        self.assertEqual(bad.status_code, 400)                   # item не в плане
        gone = self.c.post(f'/api/procurements/{self.pid}/allocate/',
            {'purchase_id': 999999, 'item_id': self.scr.id, 'qty': 1},
            content_type='application/json')
        self.assertEqual(gone.status_code, 400)                  # заказа нет
        self.c.post(f'/api/procurements/{self.pid}/allocate/',
            {'purchase_id': self.ord1, 'item_id': self.scr.id, 'qty': 40},
            content_type='application/json')
        self.c.post(f'/api/purchases/{self.ord1}/lock/')
        locked = self.c.post(f'/api/procurements/{self.pid}/allocate/',
            {'purchase_id': self.ord1, 'item_id': self.scr.id, 'qty': 5},
            content_type='application/json')
        self.assertEqual(locked.status_code, 400)                # заказ зафиксирован


class ReferenceCreateTests(EngineTestBase):
    """Канон «＋ Новая» (2026-07-03): создание изделий и проектов из справочников."""

    def test_create_item_defaults_and_fields(self):
        cat = _cat('mcu', 'Микроконтроллеры')
        i = engine.create_item('R100', 'Резистор', category_id=cat.id,
                               uom='шт', native=False, estimated_cost=D('1.50'),
                               temperature='-40-125°C')
        self.assertEqual(i.code, 'R100')
        self.assertEqual(i.category_id, cat.id)
        self.assertEqual(i.temperature, '-40-125°C')
        self.assertEqual(i.estimated_cost, D('1.50'))
        # обрезка пробелов; дефолты uom=шт, native=False
        j = engine.create_item(' B1 ', ' Плата ', category_id=cat.id)
        self.assertEqual(j.code, 'B1')
        self.assertEqual(j.uom, 'шт')
        self.assertFalse(j.native)

    def test_create_item_rejects_dup_and_bad_category(self):
        cat = _cat()
        engine.create_item('R100', 'Резистор', category_id=cat.id)
        # Ф3b: дубль кода ловит ОБЩИЙ `require_unique_code` (тот же, что у закупок/
        # заказов/документов с Ф10) — своей реализации у Item больше нет.
        with self.assertRaises(ValidationError) as ctx:
            engine.create_item('R100', 'Дубль', category_id=cat.id)   # дубль ключа
        self.assertIn('уже занят', ctx.exception.messages[0])
        with self.assertRaises(ValidationError):
            engine.create_item('X3', 'Плохая', category_id=999999)    # неизвестная категория

    def test_create_project_is_external_unlocked(self):
        p = engine.create_project('НИР-1', 'Тема', budget=D('100000'))
        self.assertEqual(p.kind, models.Project.Kind.EXTERNAL)
        self.assertFalse(p.locked)
        self.assertEqual(p.budget, D('100000'))

    def test_create_project_rejects_dup(self):
        engine.create_project('НИР-1', 'Тема')
        with self.assertRaises(ValidationError):
            engine.create_project('НИР-1', 'Дубль')    # дубль кода


class ReferenceCreateHttpTests(TestCase):
    """Канон «＋ Новая»: HTTP-путь создания изделия/проекта."""

    def setUp(self):
        get_user_model().objects.create(username='admin', is_superuser=True)
        self.c = Client()
        # Волна 12: весь /api/ за логином — HTTP-путь ходит от суперюзера-админа.
        self.c.force_login(get_user_model().objects.get(is_superuser=True))

    def test_create_item_http(self):
        cat = _cat()
        r = self.c.post('/api/items/', {'code': 'R100', 'description': 'Резистор',
            'category_id': cat.id, 'native': False},
            content_type='application/json')
        self.assertEqual(r.status_code, 201)
        self.assertEqual(r.json()['code'], 'R100')
        # появляется в списке
        lst = self.c.get('/api/items/').json()
        self.assertTrue(any(i['code'] == 'R100' for i in lst))
        # дубль → 400
        dup = self.c.post('/api/items/', {'code': 'R100', 'description': 'Дубль',
            'category_id': cat.id}, content_type='application/json')
        self.assertEqual(dup.status_code, 400)

    def test_create_project_http(self):
        r = self.c.post('/api/projects/', {'code': 'НИР-1', 'description': 'Тема',
            'budget': '100000'}, content_type='application/json')
        self.assertEqual(r.status_code, 201)
        body = r.json()
        self.assertEqual(body['kind'], 'external')
        self.assertFalse(body['locked'])


class BornByClickTests(EngineTestBase):
    """Волна 19, Ф12e: «＋ Новый» рождает сущность СРАЗУ, поэтому обязательные поля
    обязаны пережить рождение пустыми.

    Правило не «поля стали необязательными», а «обязательность переехала с рождения
    на ФИКСАЦИЮ» — каждый тест проверяет обе половины: родилось пустым И не
    фиксируется, пока не заполнено. Пустой `code` — исключение: он идентичность,
    его добирает фолбэк «Поставка 12» (решение Ивана 2026-07-28)."""

    def test_item_born_without_code_category_description(self):
        i = engine.create_item()
        self.assertEqual(i.code, f'Изделие {i.id}')
        self.assertEqual(i.description, '')
        self.assertIsNone(i.category_id)

    def test_item_without_category_does_not_lock(self):
        i = engine.create_item()
        with self.assertRaises(ValidationError) as ctx:
            engine.lock_item(i)
        self.assertIn('категорию', ctx.exception.messages[0])
        engine.update_item(i, {'category_id': _cat().id})
        engine.lock_item(i)
        self.assertTrue(models.Item.objects.get(pk=i.pk).locked)

    def test_item_lock_refusal_is_400_not_500(self):
        """Отказ фиксации должен доехать до формы строкой, а не пятисоткой.

        `lock_item` до Ф12e не умел отказывать вовсе, поэтому вьюха его не
        оборачивала — новый гейт вылезал `ValidationError` наружу."""
        i = engine.create_item()
        c = Client()
        c.force_login(self.user)
        r = c.post(f'/api/items/{i.id}/lock/')
        self.assertEqual(r.status_code, 400)
        self.assertIn('категорию', r.json()['detail'])

    def test_project_and_location_born_with_fallback_code(self):
        p = engine.create_project()
        self.assertEqual(p.code, f'Проект {p.id}')
        loc = engine.create_location()
        self.assertEqual(loc.code, f'Место {loc.id}')

    def test_fallback_code_steps_aside_from_taken_one(self):
        """Человек мог руками занять ровно тот код, который сгенерит фолбэк."""
        p = engine.create_project()
        engine.update_project(p, {'code': 'своё имя'})       # освободили фолбэк-код
        engine.create_project(f'Проект {p.id}', 'чужак')     # и заняли его руками
        engine.fallback_code(p, 'Проект')
        self.assertEqual(p.code, f'Проект {p.id}-2')

    def test_orders_born_without_number_and_get_kind_code(self):
        for create, label in (
                (engine.create_receipt, 'Поставка'),
                (engine.create_transfer, 'Передача'),
                (engine.create_writeoff, 'Списание'),
                (engine.create_requisition, 'Требование'),
                (engine.create_inventory, 'Инвентаризация'),
                (engine.create_relocation, 'Перемещение')):
            doc = create(self.prj, self.user)
            self.assertEqual(doc.number, '', label)
            self.assertEqual(doc.code, f'{label} {doc.id}', label)

    def test_order_code_space_is_shared_across_kinds(self):
        """Код ордера уникален на ВСЕ семь видов, а `Receipt.objects` фильтрует по
        `kind` — фолбэк обязан искать по родителю, иначе выдаст занятый код."""
        r = engine.create_receipt(self.prj, self.user)
        engine.update_receipt(r, code='своё имя')            # освободили фолбэк-код
        engine.update_writeoff(engine.create_writeoff(self.prj, self.user),
                               code=f'Поставка {r.id}')      # занял ордер ДРУГОГО вида
        engine.fallback_code(r, 'Поставка')
        self.assertEqual(r.code, f'Поставка {r.id}-2')

    def test_receipt_without_contractor_does_not_lock(self):
        r = engine.create_receipt(self.prj, self.user, 'УПД-1')
        engine.add_receipt_lot(r, self.make_item('R1'), D(1), D(1))
        with self.assertRaises(ValidationError):
            engine.lock_receipt(r)              # контрагент обязателен к фиксации

    def test_kitting_born_without_target(self):
        k = engine.create_kitting(self.prj, self.user)
        self.assertIsNone(k.target_item_id)
        self.assertIsNone(k.qty)
        self.assertEqual(k.code, f'Комплектация {k.id}')

    def test_every_list_endpoint_survives_empty_drafts(self):
        """Родить пустой черновик КАЖДОЙ сущности и прочитать ВСЕ списки.

        Пробел, который это закрывает: проверялись detail-проекции, а падали
        СПИСОЧНЫЕ (`_receipt_row` разыменовывал контрагента) — их читает сайдбар,
        то есть 500 ловил бы любой пользователь сразу после клика."""
        c = Client()
        c.force_login(self.user)
        paths = ['projects', 'items', 'locations', 'purchases', 'procurements',
                 'receipts', 'kittings', 'transfers', 'writeoffs',
                 'requisitions', 'inventories', 'relocations']
        for path in paths:
            self.assertEqual(
                c.post(f'/api/{path}/', {}, content_type='application/json')
                .status_code, 201, path)
        for path in paths:                      # каждый список после каждого рождения
            self.assertEqual(c.get(f'/api/{path}/').status_code, 200, path)

    def test_kitting_form_survives_missing_target(self):
        """Форма пустой комплектации открывается: разузловывать пока нечего."""
        form = engine.kitting_form(engine.create_kitting(self.prj, self.user))
        self.assertIsNone(form['target_id'])
        self.assertEqual(form['rows'], [])

    def test_kitting_without_target_does_not_lock(self):
        k = engine.create_kitting(self.prj, self.user)
        with self.assertRaises(ValidationError) as ctx:
            engine.lock_kitting(k)
        self.assertIn('прибор-цель', ctx.exception.messages[0])

    def test_receipt_contractor_is_editable_in_form(self):
        """Поставщик правится В ФОРМЕ, а не только при рождении (Ф12e).

        До Ф12e он задавался лишь в форме создания и в шапке не жил вовсе — после
        сноса той формы поставку было бы не зафиксировать никогда."""
        r = engine.create_receipt(self.prj, self.user, 'УПД-1')
        cp = models.Counterparty.objects.create(description='ООО Поставщик')
        c = Client()
        c.force_login(self.user)
        body = c.patch(f'/api/receipts/{r.id}/', {'contractor_id': cp.id},
                       content_type='application/json').json()
        self.assertEqual(body['contractor_id'], cp.id)
        engine.add_receipt_lot(r, self.make_item('R1'), D(1), D(1))
        engine.lock_receipt(models.Receipt.objects.get(pk=r.pk))   # теперь пускает

    def test_order_born_anchored_to_own_stock_when_project_unset(self):
        """Проект-якорь NOT NULL: фолбэк — белый «Собственный склад», не чужой НИР."""
        anchor = engine.default_document_project()
        self.assertEqual(anchor.kind, models.Project.Kind.INTERNAL_STOCK)
        r = self.client_post_document()
        self.assertEqual(r['project_id'], anchor.id)

    def client_post_document(self):
        c = Client()
        c.force_login(self.user)
        return c.post('/api/receipts/', {}, content_type='application/json').json()


class InventoryFormTests(EngineTestBase):
    """Волна 9: инвентаризация — рождение «найденных» партий (`+RECEIPT`, 4-й
    origin) + серая ре-материализация списанного лота с наследованием провенанса."""

    def setUp(self):
        super().setUp()
        self.item = self.make_item('R100')
        self.grey = models.Project.objects.create(
            code='GREY', description='Свободные неучтённые',
            kind=models.Project.Kind.INTERNAL_WRITEOFF)

    def test_add_lot_births_receipt_movement(self):
        inv = engine.create_inventory(self.prj, self.user, 'ИНВ-1')
        lot = engine.add_inventory_lot(inv, self.item, D(7), unit_cost=D('1.50'),
                                       lot_name='Резистор')
        self.assertEqual(lot.origin_kind, 'inventory')
        self.assertEqual(lot.project_id, self.prj.id)
        self.assertEqual(engine.lot_live_qty(lot), D(0))       # Ф15: акт — черновик
        engine.lock_inventory(inv)                              # проведён
        self.assertEqual(engine.lot_live_qty(lot), D(7))       # +RECEIPT
        self.assertTrue(lot.movements.filter(type='RECEIPT', qty=D(7)).exists())

    def test_form_totals_and_description(self):
        inv = engine.create_inventory(self.prj, self.user, 'ИНВ-1')
        engine.update_inventory(inv, description='пересчёт')
        engine.add_inventory_lot(inv, self.item, D(4), unit_cost=D('2'))
        engine.lock_inventory(inv)
        c = engine.inventory_form(inv)
        self.assertEqual(c['number'], 'ИНВ-1')
        self.assertEqual(c['description'], 'пересчёт')
        self.assertEqual(c['total_cost'], D(8))                # 4 × 2
        self.assertEqual(c['lots'][0]['live_qty'], D(4))

    def test_number_required_at_lock_not_at_birth(self):
        """Ф12e: акт рождается без номера, но не фиксируется без него."""
        inv = engine.create_inventory(self.prj, self.user)
        engine.add_inventory_lot(inv, self.item, D(1), unit_cost=D('1.00'))
        with self.assertRaises(ValidationError):
            engine.lock_inventory(inv)
        engine.update_inventory(inv, number='ИНВ-7')
        engine.lock_inventory(inv)
        self.assertTrue(models.Inventory.objects.get(pk=inv.pk).locked)

    def test_add_lot_rejects_nonpositive_and_negative_cost(self):
        inv = engine.create_inventory(self.prj, self.user, 'ИНВ-1')
        with self.assertRaises(ValidationError):
            engine.add_inventory_lot(inv, self.item, D(0))
        with self.assertRaises(ValidationError):
            engine.add_inventory_lot(inv, self.item, D(1), unit_cost=D(-1))

    def test_update_and_remove_lot(self):
        inv = engine.create_inventory(self.prj, self.user, 'ИНВ-1')
        lot = engine.add_inventory_lot(inv, self.item, D(4))
        engine.update_inventory_lot(lot, qty=D(9), unit_cost=D('3'))
        engine.lock_inventory(inv)
        self.assertEqual(engine.lot_live_qty(lot), D(9))
        engine.unlock_inventory(inv)                 # Ф15: правка — под снятым замком
        engine.remove_inventory_lot(lot)
        self.assertFalse(models.Lot.objects.filter(pk=lot.id).exists())

    def test_remove_blocked_when_consumed(self):
        inv = engine.create_inventory(self.prj, self.user, 'ИНВ-1')
        lot = engine.add_inventory_lot(inv, self.item, D(10))
        # потребим найденный лот передачей → удаление заблокировано
        tr = engine.create_transfer(self.prj, self.user, 'Н-1')
        engine.add_transfer_line(tr, lot, D(3))
        with self.assertRaises(ValidationError):
            engine.remove_inventory_lot(lot)

    def test_rematerialize_written_off_lot_inherits_provenance(self):
        # списываем партию из проекта (серый путь), затем находим и ре-материализуем в GREY
        src = self.receipt_lot(self.item, self.prj, 10)
        src.unit_cost = D('2.50'); src.lot_name = 'Резистор'; src.part_number = 'ЗН-9'
        src.save(update_fields=['unit_cost', 'lot_name', 'part_number'])
        w = engine.create_writeoff(self.prj, self.user, 'СП-1', reason='на серый')
        engine.add_writeoff_line(w, src, D(6))
        engine.lock_writeoff(w)                              # Ф15: списано фиксацией
        self.assertEqual(engine.lot_live_qty(src), D(4))
        # пикер показывает списанный лот с суммой списания
        picker = {r['lot_id']: r for r in engine.written_off_lots()}
        self.assertIn(src.id, picker)
        self.assertEqual(picker[src.id]['written_qty'], D(6))
        # ре-материализация: born-лот в GREY с predecessor и унаследованными полями
        inv = engine.create_inventory(self.grey, self.user, 'ИНВ-G1')
        born = engine.add_inventory_lot(inv, src.item, D(6), unit_cost=src.unit_cost,
                                        lot_name=src.lot_name,
                                        part_number=src.part_number, predecessor=src)
        engine.lock_inventory(inv)
        self.assertEqual(born.project_id, self.grey.id)
        self.assertEqual(born.predecessor_id, src.id)
        self.assertEqual(born.unit_cost, D('2.50'))
        self.assertEqual(engine.lot_live_qty(born), D(6))
        c = engine.inventory_form(inv)
        self.assertEqual(c['lots'][0]['predecessor_id'], src.id)
        self.assertTrue(c['lots'][0]['predecessor_label'])


class InventoryHttpTests(TestCase):
    """Волна 9: HTTP-путь инвентаризации — create/строка/правка/пикер/ре-материализация."""

    def setUp(self):
        get_user_model().objects.create(username='admin', is_superuser=True)
        self.main = models.Location.objects.create(code='MAIN', description='Основной склад')
        self.prj = models.Project.objects.create(
            code='P1', description='Проект 1', kind=models.Project.Kind.EXTERNAL)
        self.grey = models.Project.objects.create(
            code='GREY', description='Свободные неучтённые',
            kind=models.Project.Kind.INTERNAL_WRITEOFF)
        self.item = models.Item.objects.create(code='R100', description='R100', category=_cat())
        self.sup = models.Counterparty.objects.create(description='П')
        r = models.Receipt.objects.create(number='U-1', date='2026-05-01', locked=True,
            contractor=self.sup, project=self.prj, user=get_user_model().objects.first())
        self.lot = models.Lot.objects.create(item=self.item, project=self.prj,
            origin=r, qty=D(10), unit_cost=D('2.50'), part_number='ЗН-9')
        engine.rebuild_movements(self.lot)
        self.c = Client()
        # Волна 12: весь /api/ за логином — HTTP-путь ходит от суперюзера-админа.
        self.c.force_login(get_user_model().objects.get(is_superuser=True))

    def test_inventory_crud_flow(self):
        r = self.c.post('/api/inventories/', {'project_id': self.prj.id,
            'number': 'ИНВ-1'}, content_type='application/json')
        self.assertEqual(r.status_code, 201)
        iid = r.json()['id']
        # строка = найденная партия (+RECEIPT)
        line = self.c.post(f'/api/inventories/{iid}/lots/',
            {'item_id': self.item.id, 'qty': 7, 'unit_cost': '1.5',
             'lot_name': 'Резистор'}, content_type='application/json')
        self.assertEqual(line.status_code, 201)
        body = line.json()
        self.assertEqual(float(body['total_cost']), 10.5)
        self.assertEqual(float(body['lots'][0]['live_qty']), 0.0)   # Ф15: черновик
        posted = self.c.post(f'/api/inventories/{iid}/lock/')
        self.assertEqual(float(posted.json()['lots'][0]['live_qty']), 7.0)
        self.assertEqual(self.c.post(f'/api/inventories/{iid}/unlock/').status_code, 200)
        # нонпозитив qty → 400
        bad = self.c.post(f'/api/inventories/{iid}/lots/',
            {'item_id': self.item.id, 'qty': 0}, content_type='application/json')
        self.assertEqual(bad.status_code, 400)
        # правка шапки
        patch = self.c.patch(f'/api/inventories/{iid}/', {'description': 'обновлено'},
            content_type='application/json')
        self.assertEqual(patch.json()['description'], 'обновлено')

    def test_rematerialize_via_picker(self):
        # списываем часть лота → появляется в пикере ре-материализации
        w = self.c.post('/api/writeoffs/', {'project_id': self.prj.id, 'number': 'СП-1'},
            content_type='application/json').json()
        self.c.post(f"/api/writeoffs/{w['id']}/lines/",
            {'lot_id': self.lot.id, 'qty': 6}, content_type='application/json')
        # Ф15: пока акт черновик, «серого» ещё нет — возвращать нечего
        self.assertEqual(self.c.get('/api/written-off-lots/').json(), [])
        self.assertEqual(self.c.post(f"/api/writeoffs/{w['id']}/lock/").status_code, 200)
        picker = self.c.get('/api/written-off-lots/')
        self.assertEqual(picker.status_code, 200)
        self.assertTrue(any(x['lot_id'] == self.lot.id and float(x['written_qty']) == 6.0
                            for x in picker.json()))
        # ре-материализация в GREY: predecessor → списанный, поля унаследованы
        inv = self.c.post('/api/inventories/', {'project_id': self.grey.id,
            'number': 'ИНВ-G1'}, content_type='application/json').json()
        line = self.c.post(f"/api/inventories/{inv['id']}/lots/",
            {'predecessor_id': self.lot.id, 'qty': 6}, content_type='application/json')
        self.assertEqual(line.status_code, 201)
        row = line.json()['lots'][0]
        self.assertEqual(row['predecessor_id'], self.lot.id)
        self.assertEqual(float(row['unit_cost']), 2.50)       # цена унаследована
        self.assertEqual(row['part_number'], 'ЗН-9')          # PN унаследован


class OrderDeleteHttpTests(TestCase):
    """Волна 13 Ф1b: HTTP-путь post/unpost + DELETE ордеров (friendly-guard)."""

    def setUp(self):
        self.user = get_user_model().objects.create(username='admin', is_superuser=True)
        self.main = models.Location.objects.create(code='MAIN', description='Основной склад')
        self.prj = models.Project.objects.create(
            code='P1', description='Проект 1', kind=models.Project.Kind.EXTERNAL)
        self.item = models.Item.objects.create(code='R100', description='R100', category=_cat())
        self.sup = models.Counterparty.objects.create(description='П')
        r = models.Receipt.objects.create(number='U-1', date='2026-05-01', locked=True,
            contractor=self.sup, project=self.prj, user=self.user)
        self.lot = models.Lot.objects.create(item=self.item, project=self.prj,
            origin=r, qty=D(10))
        engine.rebuild_movements(self.lot)
        self.c = Client()
        self.c.force_login(self.user)

    def test_writeoff_post_unpost_delete_flow(self):
        w = self.c.post('/api/writeoffs/', {'project_id': self.prj.id, 'number': 'СП-1'},
            content_type='application/json').json()
        wid = w['id']
        self.c.post(f'/api/writeoffs/{wid}/lines/', {'lot_id': self.lot.id, 'qty': 4},
            content_type='application/json')
        # провести
        posted = self.c.post(f'/api/writeoffs/{wid}/lock/')
        self.assertEqual(posted.status_code, 200)
        self.assertTrue(posted.json()['locked'])
        # posted — удаление отклонено (сперва расфиксировать)
        blocked = self.c.delete(f'/api/writeoffs/{wid}/')
        self.assertEqual(blocked.status_code, 400)
        # расфиксировать → удалить
        self.assertEqual(self.c.post(f'/api/writeoffs/{wid}/unlock/').status_code, 200)
        gone = self.c.delete(f'/api/writeoffs/{wid}/')
        self.assertEqual(gone.status_code, 204)
        self.assertFalse(models.Writeoff.objects.filter(pk=wid).exists())
        # источник освобождён (нет −ISSUE)
        self.assertEqual(engine.lot_live_qty(self.lot), D(10))

    def test_receipt_delete_draft(self):
        r = models.Receipt.objects.create(number='U-2', date='2026-05-01',
            contractor=self.sup, project=self.prj, user=self.user)
        lot = engine.add_receipt_lot(r, self.item, D(5))
        resp = self.c.delete(f'/api/receipts/{r.id}/')
        self.assertEqual(resp.status_code, 204)
        self.assertFalse(models.Lot.objects.filter(pk=lot.pk).exists())


class ProjectBudgetTests(EngineTestBase):
    """Волна 10: бюджет проекта — потрачено/план/компас + себестоимость/экономия."""

    def make_demand(self, device, qty):
        models.ProjectDemand.objects.create(project=self.prj, target_item=device, qty=D(qty))

    def test_spent_counts_only_receipt_lots(self):
        # приходной лот считается по (цена×кол-во); заём (requisition) — бесплатен
        case = self.make_item('CASE')
        self.receipt_lot(case, self.prj, 1)  # добавит default unit_cost=0
        r = models.Receipt.objects.create(number='U-2', date='2026-05-02', locked=True,
            contractor=self.supplier, project=self.prj, user=self.user)
        paid = models.Lot.objects.create(item=case, project=self.prj, origin=r,
            qty=D(3), unit_cost=D(800))
        engine.rebuild_movements(paid)
        # заём из белого склада → born-лот в prj (origin requisition), цена наследуется
        white = models.Project.objects.create(code='WHT', description='Склад',
            kind=models.Project.Kind.INTERNAL_STOCK)
        src = models.Lot.objects.create(item=case, project=white,
            origin=models.Inventory.objects.create(project=white, user=self.user,
                number='INV-W', date='2026-05-01', locked=True),
            qty=D(5), unit_cost=D(700))
        engine.rebuild_movements(src)
        req = engine.create_requisition(self.prj, self.user, 'ТРБ-1')
        engine.add_requisition_line(req, src, D(2))
        engine.lock_requisition(req)                 # Ф15: заём материализуется фиксацией

        b = engine.project_budget(self.prj)
        self.assertEqual(b['spent'], D(2400))   # только 3×800; заём не в счёт

    def test_plan_estimate_then_replaced_by_fact(self):
        device = self.make_item('DEV', manufactured=True, kind='device')
        screw = self.make_item('SCR', kind='material')
        screw.estimated_cost = D(50)
        screw.save()
        models.BomLine.objects.create(parent=device, component=screw, qty=D(4))
        self.make_demand(device, 10)  # need SCR 40
        self.prj.budget = D(3000)
        self.prj.save()

        # склада/заказа нет → план = оценка 40×50 = 2000, компас = 3000−2000
        b = engine.project_budget(self.prj)
        self.assertEqual(b['spent'], D(0))
        self.assertEqual(b['plan'], D(2000))
        self.assertEqual(b['compass'], D(1000))
        self.assertEqual(b['unestimated'], [])

        # пришёл УПД на все 40 по реальной цене 45 → оценка сменилась фактом
        r = models.Receipt.objects.create(number='U-3', date='2026-05-03', locked=True,
            contractor=self.supplier, project=self.prj, user=self.user)
        lot = models.Lot.objects.create(item=screw, project=self.prj, origin=r,
            qty=D(40), unit_cost=D(45))
        engine.rebuild_movements(lot)
        b = engine.project_budget(self.prj)
        self.assertEqual(b['spent'], D(1800))
        self.assertEqual(b['plan'], D(1800))      # факт заместил оценку → сошлось
        self.assertEqual(b['compass'], D(1200))

    def test_unestimated_flagged_not_silently_zero(self):
        device = self.make_item('DEV', manufactured=True, kind='device')
        screw = self.make_item('SCR', kind='material')  # без estimated_cost
        models.BomLine.objects.create(parent=device, component=screw, qty=D(4))
        self.make_demand(device, 10)

        b = engine.project_budget(self.prj)
        self.assertEqual(b['unestimated'], ['SCR'])
        self.assertEqual(b['plan'], D(0))   # неполон — но флаг поднят

    def test_economy_equals_borrow_value(self):
        # прибор из купленного CASE (спот) + заёмного RES (бесплатно в бюджете,
        # но по реальной цене в себестоимости) → экономия = стоимость заёма
        device = self.make_item('DEV', manufactured=True, kind='device')
        case = self.make_item('CASE')
        res = self.make_item('RES')
        models.BomLine.objects.create(parent=device, component=case, qty=D(1))
        models.BomLine.objects.create(parent=device, component=res, qty=D(2))
        self.make_demand(device, 1)

        # CASE: куплен ровно 1 @ 800
        r = models.Receipt.objects.create(number='U-4', date='2026-05-04', locked=True,
            contractor=self.supplier, project=self.prj, user=self.user)
        case_lot = models.Lot.objects.create(item=case, project=self.prj, origin=r,
            qty=D(1), unit_cost=D(800))
        engine.rebuild_movements(case_lot)
        # RES: заём 2 @ 10 из белого склада
        white = models.Project.objects.create(code='WHT', description='Склад',
            kind=models.Project.Kind.INTERNAL_STOCK)
        src = models.Lot.objects.create(item=res, project=white,
            origin=models.Inventory.objects.create(project=white, user=self.user,
                number='INV-W2', date='2026-05-01', locked=True),
            qty=D(2), unit_cost=D(10))
        engine.rebuild_movements(src)
        req = engine.create_requisition(self.prj, self.user, 'ТРБ-2')
        engine.add_requisition_line(req, src, D(2))
        engine.lock_requisition(req)                 # Ф15: заём материализуется фиксацией
        res_lot = req.lots.first()

        # собираем прибор
        k = models.Kitting.objects.create(project=self.prj, target_item=device,
            user=self.user, qty=D(1), locked=False)
        engine.add_kitting_line(k, case, case_lot, D(1))
        engine.add_kitting_line(k, res, res_lot, D(2))
        engine.lock_kitting(k)

        b = engine.project_budget(self.prj)
        self.assertEqual(b['spent'], D(800))    # только CASE
        self.assertEqual(b['cost'], D(820))     # снимок: 800 + 2×10 (заём по реальной цене)
        self.assertEqual(b['economy'], D(20))   # польза заёма = 2×10

    def test_compass_none_without_budget(self):
        b = engine.project_budget(self.prj)
        self.assertIsNone(b['budget'])
        self.assertIsNone(b['compass'])


class MultiLevelDemandTests(EngineTestBase):
    """Ф5 (волна 16): потребность и план разузловываются до покупных листьев, а не на
    1 уровень. Прибор из подсборок → деньги/дефицит живут на листьях (подсборку купить
    нельзя). Стоимость опущена до листьев — узел (роллап) в план не задваивается."""

    def _tree(self):
        # 3 уровня: DEV → {SUB×2, A×1}; SUB → {A×2, B×3}. Листья A(@100), B(@10).
        # На 1 DEV: A = 2×2 + 1 = 5, B = 2×3 = 6. Лист A виден из двух путей (через
        # SUB и напрямую) → агрегируется.
        a = self.make_item('A', kind='material')
        b = self.make_item('B', kind='material')
        a.estimated_cost = D(100); a.save()
        b.estimated_cost = D(10); b.save()
        sub = self.make_item('SUB', manufactured=True)
        dev = self.make_item('DEV', manufactured=True, kind='device')
        models.BomLine.objects.create(parent=sub, component=a, qty=D(2))
        models.BomLine.objects.create(parent=sub, component=b, qty=D(3))
        models.BomLine.objects.create(parent=dev, component=sub, qty=D(2))
        models.BomLine.objects.create(parent=dev, component=a, qty=D(1))
        return dev, sub, a, b

    def test_leaf_demand_explodes_and_multiplies(self):
        dev, sub, a, b = self._tree()
        leaves, incomplete = engine.project_leaf_demand(self.prj)  # пусто без потребности
        self.assertEqual(leaves, {})
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(10))
        leaves, incomplete = engine.project_leaf_demand(self.prj)
        # только листья A/B, узел SUB не в потребности; кол-во перемножено сквозь дерево
        self.assertEqual({i.code: q for i, q in leaves.items()},
                         {'A': D(50), 'B': D(60)})
        self.assertEqual(incomplete, [])

    def test_deficit_components_are_leaves(self):
        dev, sub, a, b = self._tree()
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(10))
        d = engine.project_deficit(self.prj)
        comps = {c['component_code']: c for c in d['components']}
        self.assertEqual(set(comps), {'A', 'B'})     # свод: SUB не просочился (купить нельзя)
        self.assertEqual(comps['A']['need'], D(50))
        self.assertEqual(comps['B']['need'], D(60))
        # аккордеон — ДЕРЕВО BOM (Ф5b): виден узел SUB + вложенные листья + прямой A.
        # Ключ (код, глубина): A на depth1 (под SUB) и depth0 (прямой) — разные строки.
        tree = {(n['component_code'], n['depth']): n
                for n in d['demands'][0]['tree']}
        self.assertEqual(set(tree), {('SUB', 0), ('A', 1), ('B', 1), ('A', 0)})
        self.assertFalse(tree[('SUB', 0)]['is_leaf'])   # SUB — узел, не лист
        self.assertEqual(tree[('SUB', 0)]['need'], D(20))   # 10 приборов × 2
        self.assertTrue(tree[('A', 1)]['is_leaf'])
        self.assertEqual(tree[('A', 1)]['need'], D(40))     # 20 SUB × 2
        self.assertEqual(tree[('B', 1)]['need'], D(60))     # 20 SUB × 3
        self.assertEqual(tree[('A', 0)]['need'], D(10))     # прямой в приборе

    def test_plan_sums_leaf_costs_no_double_count(self):
        # Ровно репортнутый баг: план подсборки должен пробрасываться в прогноз.
        dev, sub, a, b = self._tree()
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(10))
        bud = engine.project_budget(self.prj)
        # план = 50×100 + 60×10 = 5600 (роллап SUB=230 НЕ добавлен вторым разом)
        self.assertEqual(bud['plan'], D(5600))
        self.assertEqual(bud['unestimated'], [])

    def test_scope_deficit_explodes_to_leaves(self):
        dev, sub, a, b = self._tree()
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(1))
        rows = {r['item_code']: r for r in engine.scope_deficit([self.prj])['rows']}
        self.assertEqual(set(rows), {'A', 'B'})
        self.assertEqual(rows['A']['need'], D(5))
        self.assertEqual(rows['B']['need'], D(6))

    def test_tree_leaves_carry_balance_node_is_structural(self):
        # Листья дерева несут баланс; узел-подсборка — структурная строка без чисел
        # покрытия и БЕЗ статуса (полосы сняты 2026-08-05, смотреть его было некому).
        dev, sub, a, b = self._tree()
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(1))
        self.receipt_lot(a, self.prj, 100)   # лист A покрыт с запасом, лист B — нет
        tree = {(n['component_code'], n['depth']): n
                for n in engine.project_deficit(self.prj)['demands'][0]['tree']}
        self.assertEqual(tree[('A', 1)]['balance'], D(96))        # +96 запас (в узле нужно 4)
        self.assertEqual(tree[('A', 1)]['status'], 'available')
        self.assertEqual(tree[('B', 1)]['balance'], D(-6))        # −6 не хватает
        self.assertEqual(tree[('B', 1)]['status'], 'to_order')
        self.assertNotIn('status', tree[('SUB', 0)])
        self.assertFalse(tree[('SUB', 0)]['is_leaf'])

    def test_produced_without_bom_is_incomplete_zero(self):
        # производимый узел без состава оценить нечем: 0 в план, помечен неполнотой,
        # без краха. (Витрина неполноты — отдельно; здесь важно, что не падает и не врёт.)
        empty = self.make_item('EMPTY', manufactured=True)
        dev = self.make_item('DEV', manufactured=True, kind='device')
        models.BomLine.objects.create(parent=dev, component=empty, qty=D(1))
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(5))
        leaves, incomplete = engine.project_leaf_demand(self.prj)
        self.assertEqual(leaves, {})
        self.assertEqual(incomplete, ['EMPTY'])
        self.assertEqual(engine.project_budget(self.prj)['plan'], D(0))
        self.assertEqual(engine.project_deficit(self.prj)['components'], [])

    def test_leaf_balance_nets_stock_and_order(self):
        # нетинг — на листе (через `_balance`), не на подсборке. Купили часть листа A.
        dev, sub, a, b = self._tree()
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(10))
        self.receipt_lot(a, self.prj, 20)                # A: склад 20 из нужных 50
        comps = {c['component_code']: c
                 for c in engine.project_deficit(self.prj)['components']}
        self.assertEqual(comps['A']['need'], D(50))
        self.assertEqual(comps['A']['in_stock'], D(20))
        self.assertEqual(comps['A']['balance'], D(-30))
        # план оценивает только докупаемое (30×100) + весь B (60×10) = 3600
        self.assertEqual(engine.project_budget(self.prj)['plan'], D(3600))


class ProjectBudgetHttpTests(TestCase):
    """Волна 10: HTTP-путь бюджета проекта."""

    def setUp(self):
        get_user_model().objects.create(username='admin', is_superuser=True)
        self.main = models.Location.objects.create(code='MAIN', description='Основной склад')
        self.prj = models.Project.objects.create(code='P1', description='Проект 1',
            kind=models.Project.Kind.EXTERNAL, budget=D(5000))
        self.sup = models.Counterparty.objects.create(description='П')
        self.c = Client()
        # Волна 12: весь /api/ за логином — HTTP-путь ходит от суперюзера-админа.
        self.c.force_login(get_user_model().objects.get(is_superuser=True))

    def test_budget_projection(self):
        device = models.Item.objects.create(code='DEV', description='DEV',
            category=_cat(), native=True)
        scr = models.Item.objects.create(code='SCR', description='SCR',
            category=_cat(), estimated_cost=D(50))
        models.BomLine.objects.create(parent=device, component=scr, qty=D(2))
        models.ProjectDemand.objects.create(project=self.prj, target_item=device, qty=D(10))
        r = self.c.get(f'/api/projects/{self.prj.id}/budget/')
        self.assertEqual(r.status_code, 200)
        body = r.json()
        self.assertEqual(float(body['plan']), 1000.0)     # 20×50
        self.assertEqual(float(body['compass']), 4000.0)  # 5000−1000
        self.assertEqual(float(body['spent']), 0.0)


@override_settings(MEDIA_ROOT=_TEST_MEDIA)
class AttachmentTests(EngineTestBase):
    """Волна 11: вложения — файл на диск, exclusive-arc владелец, метаданные с сервера."""

    def setUp(self):
        super().setUp()
        self.receipt = models.Receipt.objects.create(
            number='УПД-1', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)

    def _file(self, name='scan.pdf', body=b'%PDF-1.4 test', ctype='application/pdf'):
        return SimpleUploadedFile(name, body, content_type=ctype)

    def test_add_fills_metadata_and_owner(self):
        att = engine.add_attachment('receipt', self.receipt, self._file(),
                                    self.user, description='  скан УПД ')
        self.assertEqual(att.document_id, self.receipt.id)
        self.assertEqual(att.filename, 'scan.pdf')
        self.assertEqual(att.size, len(b'%PDF-1.4 test'))
        self.assertEqual(att.content_type, 'application/pdf')
        self.assertEqual(att.description, 'скан УПД')            # подрезано
        self.assertEqual(att.user_id, self.user.id)
        self.assertIsNone(att.item_id)                     # ровно один владелец (item↔document)

    def test_attachments_for_lists_newest_first(self):
        a1 = engine.add_attachment('receipt', self.receipt, self._file('a.pdf'), self.user)
        a2 = engine.add_attachment('receipt', self.receipt, self._file('b.pdf'), self.user)
        rows = engine.attachments_for('receipt', self.receipt.id)
        self.assertEqual([r['id'] for r in rows], [a2.id, a1.id])
        self.assertEqual(rows[0]['url'], f'/api/attachments/{a2.id}/download/')

    def test_unknown_owner_type_rejected(self):
        # Волна 19, Ф12b: владельцев стало шесть, но список закрытый — склад и
        # категория в него намеренно не входят.
        with self.assertRaises(ValidationError):
            engine.resolve_attachment_owner('location', 1)
        with self.assertRaises(ValidationError):
            engine.attachments_for('bogus', 1)

    def test_non_order_owners_hold_own_field(self):
        """Волна 19, Ф12b: проект/закупка/заказ/контрагент держат СВОЁ поле (в MTI не
        входят, в `document` не схлопываются); дуга по-прежнему из ровно одного."""
        proc = models.Procurement.objects.create(user=self.user)
        purch = models.Purchase.objects.create(procurement=proc, project=self.prj,
                                               user=self.user)
        owners = {'project': self.prj, 'procurement': proc, 'purchase': purch,
                  'counterparty': self.supplier}
        for owner_type, owner in owners.items():
            with self.subTest(owner_type):
                att = engine.add_attachment(owner_type, owner,
                                            self._file(f'{owner_type}.pdf'), self.user)
                self.assertEqual(getattr(att, f'{owner_type}_id'), owner.pk)
                others = set(models.ATTACHMENT_OWNER_FIELDS) - {owner_type}
                for f in others:
                    self.assertIsNone(getattr(att, f'{f}_id'))
                rows = engine.attachments_for(owner_type, owner.pk)
                self.assertEqual([r['id'] for r in rows], [att.id])

    def test_owner_delete_sweeps_files(self):
        """Каскад БД унёс бы строку, оставив файл на диске сиротой — владельцы
        подметают за собой так же, как ордер и изделие."""
        proc = models.Procurement.objects.create(user=self.user)
        purch = models.Purchase.objects.create(procurement=proc, project=self.prj,
                                               user=self.user)
        prj = models.Project.objects.create(code='P-DEL', description='На снос',
                                            kind=models.Project.Kind.EXTERNAL)
        cases = [('purchase', purch, engine.delete_purchase),
                 ('procurement', proc, engine.delete_procurement),
                 ('project', prj, engine.delete_project)]
        for owner_type, owner, delete in cases:       # заказ раньше закупки (PROTECT)
            with self.subTest(owner_type):
                att = engine.add_attachment(owner_type, owner,
                                            self._file(f'{owner_type}.pdf'), self.user)
                path = att.file.path
                delete(owner)
                self.assertFalse(models.Attachment.objects.filter(pk=att.id).exists())
                self.assertFalse(os.path.exists(path))

    def test_office_mime_fits(self):
        """Грабля прода: MIME от Office длиннее прежних 64 символов (xlsx = 65,
        docx = 71, pptx = 73) — вложение отвергалось не по сути, а по длине
        служебной строки. Метаданные больше не решают, приняли файл или нет."""
        office = {
            'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            'pptx': 'application/vnd.openxmlformats-officedocument.presentationml.presentation',
        }
        for ext, ctype in office.items():
            with self.subTest(ext):
                self.assertGreater(len(ctype), 64)         # ради этого тест и живёт
                att = engine.add_attachment(
                    'receipt', self.receipt,
                    self._file(f'КП от 30.07.2026.{ext}', b'PK\x03\x04', ctype), self.user)
                self.assertEqual(att.content_type, ctype)
                self.assertEqual(att.filename, f'КП от 30.07.2026.{ext}')

    def test_absurd_metadata_trimmed_not_rejected(self):
        """Имя и MIME приходят от браузера: сверхдлинные подрезаются под поле,
        загрузку не роняют."""
        att = engine.add_attachment(
            'receipt', self.receipt,
            self._file('и' * 400 + '.pdf', b'%PDF-1.4', 'application/' + 'x' * 400),
            self.user)
        self.assertEqual(len(att.filename), 255)
        self.assertEqual(len(att.content_type), 255)

    def test_oversize_rejected(self):
        big = SimpleUploadedFile('big.bin', b'x' * 10, content_type='application/octet-stream')
        with override_settings(MAX_ATTACHMENT_SIZE=5):
            with self.assertRaises(ValidationError):
                engine.add_attachment('receipt', self.receipt, big, self.user)

    def test_state_tracks_file_on_disk(self):
        """Волна 19, Ф12a: витрина красит глиф вложения по тому, что реально на диске —
        совпадает (ok), перезаписан мимо Plume (changed), пропал (missing)."""
        att = engine.add_attachment('receipt', self.receipt, self._file(), self.user)
        self.assertEqual(engine.attachment_state(att), 'ok')
        with open(att.file.path, 'wb') as f:               # подменили содержимое мимо нас
            f.write(b'%PDF-1.4 tampered longer body')
        self.assertEqual(engine.attachment_state(att), 'changed')
        os.remove(att.file.path)                           # файл унесли, запись осталась
        self.assertEqual(engine.attachment_state(att), 'missing')

    def test_update_and_delete_removes_file(self):
        att = engine.add_attachment('receipt', self.receipt, self._file(), self.user)
        path = att.file.path
        self.assertTrue(os.path.exists(path))
        engine.update_attachment(att, description='новая подпись')
        att.refresh_from_db()
        self.assertEqual(att.description, 'новая подпись')
        engine.delete_attachment(att)
        self.assertFalse(models.Attachment.objects.filter(pk=att.id).exists())
        self.assertFalse(os.path.exists(path))             # файл удалён с диска


@override_settings(MEDIA_ROOT=_TEST_MEDIA)
class AttachmentHttpTests(TestCase):
    """Волна 11: HTTP-путь вложений — multipart upload → list → patch → download → delete."""

    def setUp(self):
        self.user = get_user_model().objects.create(username='admin', is_superuser=True)
        self.supplier = models.Counterparty.objects.create(description='П')
        self.prj = models.Project.objects.create(
            code='P1', description='Проект', kind=models.Project.Kind.EXTERNAL)
        self.receipt = models.Receipt.objects.create(
            number='УПД-1', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        self.c = Client()
        # Волна 12: весь /api/ за логином — HTTP-путь ходит от суперюзера-админа.
        self.c.force_login(get_user_model().objects.get(is_superuser=True))

    def test_full_cycle(self):
        up = SimpleUploadedFile('scan.pdf', b'%PDF data', content_type='application/pdf')
        r = self.c.post(f'/api/attachments/receipt/{self.receipt.id}/',
                        {'file': up, 'description': 'скан'})
        self.assertEqual(r.status_code, 201)
        aid = r.json()['id']
        lst = self.c.get(f'/api/attachments/receipt/{self.receipt.id}/').json()
        self.assertEqual(len(lst), 1)
        self.assertEqual(lst[0]['filename'], 'scan.pdf')
        self.assertEqual(lst[0]['user'], 'admin')          # автор с документа
        pr = self.c.patch(f'/api/attachments/{aid}/', {'description': 'скан УПД №1'},
                          content_type='application/json')
        self.assertEqual(pr.status_code, 200)
        self.assertEqual(pr.json()['description'], 'скан УПД №1')
        dl = self.c.get(f'/api/attachments/{aid}/download/')
        self.assertEqual(dl.status_code, 200)
        self.assertEqual(b''.join(dl.streaming_content), b'%PDF data')
        dr = self.c.delete(f'/api/attachments/{aid}/')
        self.assertEqual(dr.status_code, 204)
        self.assertEqual(self.c.get(f'/api/attachments/receipt/{self.receipt.id}/').json(), [])

    def test_download_disposition_safe_inline_else_attachment(self):
        # PDF — inline (смотреть во вкладке), html — принудительная загрузка (XSS),
        # оба с nosniff.
        pdf = SimpleUploadedFile('scan.pdf', b'%PDF', content_type='application/pdf')
        html = SimpleUploadedFile('bom.html', b'<script>x()</script>',
                                  content_type='text/html')
        pid = self.c.post(f'/api/attachments/receipt/{self.receipt.id}/',
                          {'file': pdf}).json()['id']
        hid = self.c.post(f'/api/attachments/receipt/{self.receipt.id}/',
                          {'file': html}).json()['id']
        dp = self.c.get(f'/api/attachments/{pid}/download/')
        self.assertIn('inline', dp['Content-Disposition'])
        self.assertEqual(dp['X-Content-Type-Options'], 'nosniff')
        dh = self.c.get(f'/api/attachments/{hid}/download/')
        self.assertIn('attachment', dh['Content-Disposition'])
        self.assertEqual(dh['X-Content-Type-Options'], 'nosniff')

    def test_bad_owner_type(self):
        up = SimpleUploadedFile('x.pdf', b'x', content_type='application/pdf')
        r = self.c.post('/api/attachments/location/1/', {'file': up})
        self.assertEqual(r.status_code, 400)

    def test_owner_not_found(self):
        """Тип известен (волна 19, Ф12b), а записи нет — тоже 400, но по другой причине."""
        up = SimpleUploadedFile('x.pdf', b'x', content_type='application/pdf')
        r = self.c.post('/api/attachments/purchase/999/', {'file': up})
        self.assertEqual(r.status_code, 400)

    def test_missing_file(self):
        r = self.c.post(f'/api/attachments/receipt/{self.receipt.id}/',
                        {'description': 'нет файла'})
        self.assertEqual(r.status_code, 400)


class AuthHttpTests(TestCase):
    """Волна 12: логин-экран — вход/выход сессией, гейтинг всего /api/, авторство."""

    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username='ivan', password='s3cret-pass', first_name='Иван')
        self.prj = models.Project.objects.create(
            code='P1', description='Проект', kind=models.Project.Kind.EXTERNAL)
        self.c = Client()

    def test_anonymous_api_is_gated(self):
        # Без логина любой прикладной эндпоинт закрыт (403 — DRF без challenge).
        self.assertEqual(self.c.get('/api/projects/').status_code, 403)

    def test_ping_open_without_login(self):
        r = self.c.get('/api/ping/')
        self.assertEqual(r.status_code, 200)

    def test_me_anonymous_is_401_and_sets_csrf_cookie(self):
        r = self.c.get('/api/auth/me/')
        self.assertEqual(r.status_code, 401)
        self.assertIn('csrftoken', r.cookies)     # токен для последующего POST

    def test_login_flow_and_authorship(self):
        bad = self.c.post('/api/auth/login/',
                          {'username': 'ivan', 'password': 'wrong'},
                          content_type='application/json')
        self.assertEqual(bad.status_code, 400)
        ok = self.c.post('/api/auth/login/',
                         {'username': 'ivan', 'password': 's3cret-pass'},
                         content_type='application/json')
        self.assertEqual(ok.status_code, 200)
        self.assertEqual(ok.json()['username'], 'ivan')
        self.assertEqual(ok.json()['full_name'], 'Иван')     # get_full_name
        # После логина — доступ открыт и авторство пишется реальным юзером.
        me = self.c.get('/api/auth/me/')
        self.assertEqual(me.status_code, 200)
        cr = self.c.post('/api/kittings/', {'project_id': self.prj.id,
            'target_item_id': models.Item.objects.create(
                code='D1', description='Прибор', category=_cat(),
                native=True).id, 'qty': 1},
            content_type='application/json')
        self.assertEqual(cr.status_code, 201)
        k = models.Kitting.objects.get(pk=cr.json()['id'])
        self.assertEqual(k.user, self.user)

    def test_logout_closes_session(self):
        self.c.force_login(self.user)
        self.assertEqual(self.c.get('/api/projects/').status_code, 200)
        out = self.c.post('/api/auth/logout/')
        self.assertEqual(out.status_code, 204)
        self.assertEqual(self.c.get('/api/projects/').status_code, 403)


class BomEditTests(EngineTestBase):
    """Редактор состава (BOM): добавление/правка/удаление, гварды дублей и циклов."""

    def test_add_update_remove_bom_line(self):
        dev = self.make_item('DEV', manufactured=True)
        comp = self.make_item('R')
        line = engine.add_bom_line(dev, comp, D(3))
        self.assertEqual(line.qty, D(3))
        engine.update_bom_line(line, qty=D(5))
        line.refresh_from_db()
        self.assertEqual(line.qty, D(5))
        engine.remove_bom_line(line)
        self.assertFalse(models.BomLine.objects.filter(pk=line.pk).exists())

    def test_bom_rejects_self_and_duplicate_and_nonpositive(self):
        dev = self.make_item('DEV', manufactured=True)
        comp = self.make_item('R')
        with self.assertRaises(ValidationError):
            engine.add_bom_line(dev, dev, D(1))           # сам на себя
        with self.assertRaises(ValidationError):
            engine.add_bom_line(dev, comp, D(0))          # qty <= 0
        engine.add_bom_line(dev, comp, D(1))
        with self.assertRaises(ValidationError):
            engine.add_bom_line(dev, comp, D(2))          # дубль (parent, component)

    def test_bom_rejects_cycle(self):
        a = self.make_item('A', manufactured=True)
        b = self.make_item('B', manufactured=True)
        c = self.make_item('C', manufactured=True)
        engine.add_bom_line(a, b, D(1))
        engine.add_bom_line(b, c, D(1))
        with self.assertRaises(ValidationError):
            engine.add_bom_line(c, a, D(1))               # C ⊃ A замкнул бы цикл A→B→C→A


class ProjectDemandEditTests(EngineTestBase):
    """Редактор потребности проекта (секция «Приборы»)."""

    def test_add_update_remove_demand(self):
        dev = self.make_item('DEV', manufactured=True, kind='device')
        d = engine.add_project_demand(self.prj, dev, D(4))
        self.assertEqual(d.qty, D(4))
        engine.update_project_demand(d, D(7))
        d.refresh_from_db()
        self.assertEqual(d.qty, D(7))
        engine.remove_project_demand(d)
        self.assertFalse(models.ProjectDemand.objects.filter(pk=d.pk).exists())

    def test_demand_rejects_duplicate_and_nonpositive(self):
        dev = self.make_item('DEV', manufactured=True, kind='device')
        with self.assertRaises(ValidationError):
            engine.add_project_demand(self.prj, dev, D(0))
        engine.add_project_demand(self.prj, dev, D(1))
        with self.assertRaises(ValidationError):
            engine.add_project_demand(self.prj, dev, D(2))

    def test_demand_blocked_on_closed_and_internal(self):
        dev = self.make_item('DEV', manufactured=True, kind='device')
        closed = models.Project.objects.create(
            code='PC', description='Закрытый', kind=models.Project.Kind.EXTERNAL,
            locked=True)
        with self.assertRaises(ValidationError):
            engine.add_project_demand(closed, dev, D(1))
        internal = models.Project.objects.create(
            code='WH', description='Склад', kind=models.Project.Kind.INTERNAL_STOCK)
        with self.assertRaises(ValidationError):
            engine.add_project_demand(internal, dev, D(1))

    def test_deficit_components_aggregate(self):
        # Два прибора делят компонент R → сводная потребность суммируется.
        r = self.make_item('R')
        c = self.make_item('C')
        dev1 = self.make_item('DEV1', manufactured=True, kind='device')
        dev2 = self.make_item('DEV2', manufactured=True, kind='device')
        engine.add_bom_line(dev1, r, D(2))
        engine.add_bom_line(dev1, c, D(1))
        engine.add_bom_line(dev2, r, D(3))
        engine.add_project_demand(self.prj, dev1, D(5))   # R: 10
        engine.add_project_demand(self.prj, dev2, D(4))   # R: 12 → всего 22
        out = engine.project_deficit(self.prj)
        agg = {c['component_code']: c for c in out['components']}
        self.assertEqual(agg['R']['need'], D(22))
        self.assertEqual(agg['C']['need'], D(5))
        # Сортировка «горит вперёд»: одинаковый статус (всё к заказу) → по коду.
        codes = [c['component_code'] for c in out['components']]
        self.assertEqual(codes, ['C', 'R'])


class ItemProjectUpdateTests(EngineTestBase):
    """Правка свойств изделия и реквизитов проекта под замком формы (§6)."""

    def test_update_item_fields(self):
        it = self.make_item('X')
        cat2 = _cat('mcu', 'Микроконтроллеры')
        engine.update_item(it, {'description': 'Новое имя', 'category_id': cat2.id,
                                'uom': 'кг', 'estimated_cost': D('12.50'),
                                'temperature': '-40-85°C', 'native': True})
        it.refresh_from_db()
        self.assertEqual(it.description, 'Новое имя')
        self.assertEqual(it.category_id, cat2.id)
        self.assertEqual(it.uom, 'кг')
        self.assertEqual(it.estimated_cost, D('12.50'))
        self.assertEqual(it.temperature, '-40-85°C')
        self.assertTrue(it.native)

    def test_update_item_estimated_cost_can_clear(self):
        it = self.make_item('X')
        it.estimated_cost = D('5'); it.save()
        engine.update_item(it, {'estimated_cost': None})
        it.refresh_from_db()
        self.assertIsNone(it.estimated_cost)

    def test_update_item_rejects_dup_key_and_bad_category(self):
        self.make_item('A')
        it = self.make_item('B')
        with self.assertRaises(ValidationError):
            engine.update_item(it, {'code': 'A'})  # дубль ключа
        with self.assertRaises(ValidationError):
            engine.update_item(it, {'category_id': 999999})  # неизвестная категория
        # Ф12e: очистка описания/категории легальна — заполнить можно, передумать тоже.
        engine.update_item(it, {'description': '   ', 'category_id': None})
        it.refresh_from_db()
        self.assertEqual(it.description, '')
        self.assertIsNone(it.category_id)

    def test_update_item_partial_leaves_others(self):
        it = self.make_item('X')
        it.uom = 'м'; it.save()
        cat0 = it.category_id
        engine.update_item(it, {'description': 'Y'})       # прислали только описание
        it.refresh_from_db()
        self.assertEqual(it.uom, 'м')                      # uom не тронут
        self.assertEqual(it.category_id, cat0)             # категория не тронута

    def test_update_project_fields_and_clear_budget(self):
        engine.update_project(self.prj, {'description': 'Переименован',
                                         'budget': D('1000'), 'started': '2026-01-15'})
        self.prj.refresh_from_db()
        self.assertEqual(self.prj.description, 'Переименован')
        self.assertEqual(self.prj.budget, D('1000'))
        self.assertEqual(str(self.prj.started), '2026-01-15')
        engine.update_project(self.prj, {'budget': None})
        self.prj.refresh_from_db()
        self.assertIsNone(self.prj.budget)

    def test_update_project_allows_clearing_name(self):
        """Ф12e: описание не идентичность (её держит `code`) — очистка легальна."""
        engine.update_project(self.prj, {'description': '  '})
        self.prj.refresh_from_db()
        self.assertEqual(self.prj.description, '')

    def test_update_project_code_rename_and_guards(self):
        # WAVE14 Ф1: код правится в форме, guard как у изделия (не PK — безопасно).
        engine.update_project(self.prj, {'code': 'НОВ-КОД'})
        self.prj.refresh_from_db()
        self.assertEqual(self.prj.code, 'НОВ-КОД')
        with self.assertRaises(ValidationError):               # пустой код
            engine.update_project(self.prj, {'code': '  '})
        models.Project.objects.create(code='ЗАНЯТО', description='Другой')
        with self.assertRaises(ValidationError):               # коллизия кода
            engine.update_project(self.prj, {'code': 'ЗАНЯТО'})

    def test_project_detail_patch_code(self):
        self.c = Client()
        self.c.force_login(self.user)
        r = self.c.patch(f'/api/projects/{self.prj.id}/',
                         {'code': 'ЧЕРЕЗ-API'}, content_type='application/json')
        self.assertEqual(r.status_code, 200)
        self.prj.refresh_from_db()
        self.assertEqual(self.prj.code, 'ЧЕРЕЗ-API')

    def test_project_detail_patch_endpoint(self):
        self.c = Client()
        self.c.force_login(self.user)
        r = self.c.patch(f'/api/projects/{self.prj.id}/',
                         {'budget': '2500'}, content_type='application/json')
        self.assertEqual(r.status_code, 200)
        self.assertEqual(D(str(r.json()['budget'])), D('2500'))
        self.prj.refresh_from_db()
        self.assertEqual(self.prj.budget, D('2500'))

    def test_item_detail_patch_endpoint(self):
        it = self.make_item('Z')
        self.c = Client()
        self.c.force_login(self.user)
        r = self.c.patch(f'/api/items/{it.id}/',
                         {'description': 'Обновлён', 'estimated_cost': '9.9'},
                         content_type='application/json')
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.json()['description'], 'Обновлён')
        it.refresh_from_db()
        self.assertEqual(it.description, 'Обновлён')
        self.assertEqual(it.estimated_cost, D('9.9'))


class UnifiedLockTests(EngineTestBase):
    """Волна 13 Ф1: единый мягкий замок `locked` на всех складских
    документах (свернул `Receipt.approved`/`Transfer.posted`/`Kitting.wip-closed`)."""

    def test_all_docs_default_to_unlocked(self):
        """Плоское создание любого ордера рождает черновик (единый дефолт)."""
        r = models.Receipt.objects.create(
            number='У-1', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        k = models.Kitting.objects.create(
            project=self.prj, target_item=self.make_item('DEV', manufactured=True),
            user=self.user, qty=D(1))
        inv = models.Inventory.objects.create(
            project=self.prj, user=self.user, number='И-1', date='2026-05-01')
        req = models.Requisition.objects.create(
            project=self.prj, user=self.user, number='Т-1', date='2026-05-01')
        t = models.Transfer.objects.create(
            project=self.prj, user=self.user, number='Н-1', date='2026-05-01')
        w = models.Writeoff.objects.create(
            project=self.prj, user=self.user, number='С-1', date='2026-05-01')
        for doc in (r, k, inv, req, t, w):
            self.assertFalse(doc.locked)
            self.assertFalse(doc.locked)

    def test_single_guard_freezes_edits_across_doc_types(self):
        """Один `_require_draft` гейтит правку прихода / передачи / комплектации."""
        # приход
        r = models.Receipt.objects.create(
            number='У-2', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        engine.add_receipt_lot(r, self.make_item('A'), D(5))
        engine.lock_receipt(r)
        self.assertTrue(r.locked)
        with self.assertRaises(ValidationError):
            engine.add_receipt_lot(r, self.make_item('B'), D(1))
        # передача
        dev = self.make_item('DEV', manufactured=True)
        dlot = self.receipt_lot(dev, self.prj, 5)
        t = engine.create_transfer(self.prj, self.user, 'Н-2')
        engine.add_transfer_line(t, dlot, D(1))
        engine.lock_transfer(t)
        t.refresh_from_db()
        self.assertTrue(t.locked)
        with self.assertRaises(ValidationError):
            engine.add_transfer_line(t, dlot, D(1))

    def test_kitting_lock_projection(self):
        """До фронт-среза (Ф1b) форма отдаёт исторические `wip`/`closed`."""
        k = models.Kitting.objects.create(
            project=self.prj, target_item=self.make_item('DEV', manufactured=True),
            user=self.user, qty=D(1))
        self.assertFalse(engine.kitting_form(k)['locked'])
        engine.lock_kitting(k)
        self.assertTrue(engine.kitting_form(k)['locked'])


class DocumentLockAndDeleteRulesTests(EngineTestBase):
    """Фиксация/расфиксация ордера, заморозка правки под замком и единое правило
    удаления: черновик — свободно, зафиксированный — сперва расфиксируй, `PROTECT`
    бережёт потраченные лоты (волна 13, Ф1b; ось `locked` — волна 19, Ф1c)."""

    def _other_project(self):
        return models.Project.objects.create(
            code='P9', description='Проект 9', kind=models.Project.Kind.EXTERNAL)

    # ── post/unpost round-trip + пустой guard ──
    def test_post_unpost_roundtrip_three_docs(self):
        """draft → posted → draft для списания/требования/инвентаризации."""
        # списание
        lot = self.receipt_lot(self.make_item('A'), self.prj, 10)
        w = engine.create_writeoff(self.prj, self.user, 'С-1')
        engine.add_writeoff_line(w, lot, D(3))
        engine.lock_writeoff(w); w.refresh_from_db()
        self.assertTrue(w.locked)
        self.assertTrue(engine.writeoff_form(w)['locked'])
        engine.unlock_writeoff(w); w.refresh_from_db()
        self.assertFalse(w.locked)
        # инвентаризация
        inv = engine.create_inventory(self.prj, self.user, 'И-1')
        engine.add_inventory_lot(inv, self.make_item('B'), D(4))
        engine.lock_inventory(inv); inv.refresh_from_db()
        self.assertTrue(inv.locked)
        self.assertTrue(engine.inventory_form(inv)['locked'])
        engine.unlock_inventory(inv); inv.refresh_from_db()
        self.assertFalse(inv.locked)
        # требование
        src = self.receipt_lot(self.make_item('C'), self._other_project(), 10)
        req = engine.create_requisition(self.prj, self.user, 'Т-1')
        engine.add_requisition_line(req, src, D(2))
        engine.lock_requisition(req); req.refresh_from_db()
        self.assertTrue(req.locked)
        self.assertTrue(engine.requisition_form(req)['locked'])
        engine.unlock_requisition(req); req.refresh_from_db()
        self.assertFalse(req.locked)

    def test_post_empty_doc_refused(self):
        """Пустой ордер нельзя провести (как приход/передача)."""
        w = engine.create_writeoff(self.prj, self.user, 'С-2')
        with self.assertRaises(ValidationError):
            engine.lock_writeoff(w)
        inv = engine.create_inventory(self.prj, self.user, 'И-2')
        with self.assertRaises(ValidationError):
            engine.lock_inventory(inv)
        req = engine.create_requisition(self.prj, self.user, 'Т-2')
        with self.assertRaises(ValidationError):
            engine.lock_requisition(req)

    # ── edit-freeze: проведённый документ read-only ──
    def test_edit_freeze_blocks_edits_on_posted_three_docs(self):
        """posted-ордер гейтит правку шапки И строк (единый `_require_draft`)."""
        # списание
        lot = self.receipt_lot(self.make_item('A'), self.prj, 10)
        w = engine.create_writeoff(self.prj, self.user, 'С-3')
        line = engine.add_writeoff_line(w, lot, D(3))
        engine.lock_writeoff(w)
        with self.assertRaises(ValidationError):
            engine.update_writeoff(w, number='С-3x')
        with self.assertRaises(ValidationError):
            engine.add_writeoff_line(w, lot, D(1))
        with self.assertRaises(ValidationError):
            engine.update_writeoff_line(line, D(2))
        with self.assertRaises(ValidationError):
            engine.remove_writeoff_line(line)
        # инвентаризация
        inv = engine.create_inventory(self.prj, self.user, 'И-3')
        ilot = engine.add_inventory_lot(inv, self.make_item('B'), D(4))
        engine.lock_inventory(inv)
        with self.assertRaises(ValidationError):
            engine.update_inventory(inv, number='И-3x')
        with self.assertRaises(ValidationError):
            engine.add_inventory_lot(inv, self.make_item('B2'), D(1))
        with self.assertRaises(ValidationError):
            engine.update_inventory_lot(ilot, qty=D(9))
        with self.assertRaises(ValidationError):
            engine.remove_inventory_lot(ilot)
        # требование
        src = self.receipt_lot(self.make_item('C'), self._other_project(), 10)
        req = engine.create_requisition(self.prj, self.user, 'Т-3')
        rline = engine.add_requisition_line(req, src, D(2))
        engine.lock_requisition(req)
        with self.assertRaises(ValidationError):
            engine.update_requisition(req, number='Т-3x')
        with self.assertRaises(ValidationError):
            engine.update_requisition_line(rline, D(1))
        with self.assertRaises(ValidationError):
            engine.remove_requisition_line(rline)

    # ── удаление: правило draft/posted ──
    def test_delete_draft_writeoff_rebuilds_source(self):
        """Удаление черновика списания снимает `−ISSUE` — источник возвращает остаток."""
        lot = self.receipt_lot(self.make_item('A'), self.prj, 10)
        w = engine.create_writeoff(self.prj, self.user, 'С-4')
        engine.add_writeoff_line(w, lot, D(4))
        # Ф15: черновик источник не двигал — снимаем его вместе с намерением
        self.assertEqual(engine.lot_live_qty(lot), D(10))
        engine.delete_stock_document(w)
        self.assertFalse(models.Writeoff.objects.filter(pk=w.pk).exists())
        self.assertEqual(engine.lot_live_qty(lot), D(10))   # источник освобождён

    def test_delete_posted_refused_until_unpost(self):
        """posted — «сперва расфиксировать»: удаление отклонено, после unpost — ок."""
        lot = self.receipt_lot(self.make_item('A'), self.prj, 10)
        w = engine.create_writeoff(self.prj, self.user, 'С-5')
        engine.add_writeoff_line(w, lot, D(4))
        engine.lock_writeoff(w)
        with self.assertRaises(ValidationError):
            engine.delete_stock_document(w)
        engine.unlock_writeoff(w)
        engine.delete_stock_document(w)
        self.assertFalse(models.Writeoff.objects.filter(pk=w.pk).exists())

    def test_delete_draft_requisition_drops_born_and_restores_source(self):
        """Удаление черновика требования: born-потомок снят, источник восстановлен."""
        src = self.receipt_lot(self.make_item('C'), self._other_project(), 10)
        req = engine.create_requisition(self.prj, self.user, 'Т-4')
        engine.add_requisition_line(req, src, D(3))
        born = req.lots.get()
        self.assertEqual(engine.lot_live_qty(src), D(10))   # Ф15: черновик не двигал
        engine.delete_stock_document(req)
        self.assertFalse(models.Lot.objects.filter(pk=born.pk).exists())  # born снят
        self.assertEqual(engine.lot_live_qty(src), D(10))                 # источник цел

    def test_delete_receipt_draft_cascades_born_lot(self):
        """Удаление черновика прихода уносит рождённый им лот (born-direct)."""
        r = models.Receipt.objects.create(
            number='У-9', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        lot = engine.add_receipt_lot(r, self.make_item('A'), D(5))
        engine.delete_stock_document(r)
        self.assertFalse(models.Receipt.objects.filter(pk=r.pk).exists())
        self.assertFalse(models.Lot.objects.filter(pk=lot.pk).exists())

    def test_delete_refused_when_born_lot_consumed_downstream(self):
        """`PROTECT` бережёт потраченные лоты: born-лот акта потреблён ниже → отказ."""
        inv = engine.create_inventory(self.prj, self.user, 'И-9')
        found = engine.add_inventory_lot(inv, self.make_item('A'), D(10))
        # потребляем найденный лот списанием (downstream `−ISSUE`)
        w = engine.create_writeoff(self.prj, self.user, 'С-9')
        engine.add_writeoff_line(w, found, D(2))
        with self.assertRaises(ValidationError):
            engine.delete_stock_document(inv)
        self.assertTrue(models.Inventory.objects.filter(pk=inv.pk).exists())


class DocumentKindStampTests(EngineTestBase):
    """Вид ордера живёт дискриминатором `kind` в единой таблице `StockDocument`:
    `save()` штампует его по классу, а менеджер proxy-вида сужает выборку до своего.

    Класс родился в волне 13 (Ф2a) про MTI-ядро и после Ф14 (снос MTI) переписан:
    три теста из четырёх проверяли равенство PK ребёнка и родителя — у proxy это
    верно **по построению** и сломаться не может. Живое здесь — штамп вида и
    `_KindManager`: на нём держатся 108 обращений `models.<Вид>.objects` в движке."""

    def _one_of_each(self):
        r = models.Receipt.objects.create(
            number='У-1', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        k = models.Kitting.objects.create(
            project=self.prj, target_item=self.make_item('DEV', manufactured=True),
            user=self.user, qty=D(1))
        inv = models.Inventory.objects.create(
            project=self.prj, user=self.user, number='И-1', date='2026-05-01')
        req = models.Requisition.objects.create(
            project=self.prj, user=self.user, number='Т-1', date='2026-05-01')
        t = models.Transfer.objects.create(
            project=self.prj, user=self.user, number='Н-1', date='2026-05-01')
        w = models.Writeoff.objects.create(
            project=self.prj, user=self.user, number='С-1', date='2026-05-01')
        return {'receipt': r, 'kitting': k, 'inventory': inv,
                'requisition': req, 'transfer': t, 'writeoff': w}

    def test_kind_stamped_on_each_doc_type(self):
        """`save()` штампует свой `kind` в родителя на плоской вставке каждого типа."""
        for kind, doc in self._one_of_each().items():
            self.assertEqual(doc.kind, kind)
            self.assertTrue(models.StockDocument.objects.filter(
                pk=doc.pk, kind=kind).exists())

    def test_kind_manager_shows_only_own_kind(self):
        """`_KindManager` держит иллюзию семи таблиц: менеджер вида видит только свои
        документы, а базовая таблица — все. Иллюзия load-bearing: на ней пережил снос
        MTI весь движок (`Receipt.objects.filter(...)` и per-kind функции)."""
        docs = self._one_of_each()
        self.assertEqual(models.StockDocument.objects.count(), len(docs))
        for kind, doc in docs.items():
            proxy = type(doc)
            self.assertEqual([d.pk for d in proxy.objects.all()], [doc.pk],
                             f'{proxy.__name__}.objects видит чужие виды')
        # и наоборот: чужой вид по своему же id не находится
        self.assertFalse(
            models.Receipt.objects.filter(pk=docs['writeoff'].pk).exists())


class DocumentOwnershipFkTests(EngineTestBase):
    """Владение — по одному FK на `StockDocument`, а не дугой из типизированных FK.

    `Lot.origin` и `StockLine.document` — единственный FK каждый (Check «ровно один
    origin» умер вместе с дугой); у `Attachment` дуга жива, но ордера в ней занимают
    один путь из шести. Проекция движений при этом читает вид из `document.kind`
    (волна 13, Ф2b; шесть путей владельца вложения — волна 19, Ф12b)."""

    def _constraint_names(self, model):
        return {c.name for c in model._meta.constraints}

    def test_lot_origin_is_single_fk_no_arc(self):
        """`Lot.origin` — один FK на родителя; старых 4 FK и Check origin нет."""
        r = models.Receipt.objects.create(
            number='У-1', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        lot = engine.add_receipt_lot(r, self.make_item('A'), D(5))
        self.assertEqual(lot.origin_id, r.pk)
        self.assertEqual(lot.origin_kind, models.StockDocument.Kind.RECEIPT)
        field_names = {f.name for f in models.Lot._meta.get_fields()}
        self.assertIn('origin', field_names)
        self.assertFalse({'receipt', 'kitting', 'inventory', 'requisition'}
                         & field_names)
        self.assertNotIn('lot_exactly_one_origin', self._constraint_names(models.Lot))

    def test_stockline_document_is_single_fk_no_arc(self):
        """`StockLine.document` — один FK; старых 4 FK и Check document нет."""
        lot = self.receipt_lot(self.make_item('R'), self.prj, 100)
        w = engine.create_writeoff(self.prj, self.user, 'С-1')
        line = engine.add_writeoff_line(w, lot, D(3))
        self.assertEqual(line.document_id, w.pk)
        self.assertEqual(line.doc_kind, models.StockDocument.Kind.WRITEOFF)
        field_names = {f.name for f in models.StockLine._meta.get_fields()}
        self.assertIn('document', field_names)
        self.assertFalse({'kitting', 'transfer', 'writeoff', 'requisition'}
                         & field_names)
        self.assertNotIn('stockline_exactly_one_document',
                         self._constraint_names(models.StockLine))

    def test_movement_projection_source_preserved_after_collapse(self):
        """Проекция движений неизменна: source_type = document.kind, source_id = id
        родителя (то же, что раньше давали `origin_kind`/`{kind}_id`)."""
        lot = self.receipt_lot(self.make_item('R'), self.prj, 50)
        w = engine.create_writeoff(self.prj, self.user, 'С-2')
        engine.add_writeoff_line(w, lot, D(4))
        engine.lock_writeoff(w)                     # Ф15: движение есть у проведённого
        born = lot.movements.get(type=models.StockMovement.Type.RECEIPT)
        self.assertEqual(born.source_type, models.StockDocument.Kind.RECEIPT)
        self.assertEqual(born.source_id, lot.origin_id)
        issue = lot.movements.get(type=models.StockMovement.Type.ISSUE)
        self.assertEqual(issue.source_type, models.StockDocument.Kind.WRITEOFF)
        self.assertEqual(issue.source_id, w.pk)

    def test_attachment_owner_is_one_path_of_six(self):
        """Шесть видов ордера схлопнуты в один FK `document`, остальные владельцы держат
        своё поле: 'receipt' → `document`, 'item' → `item`; API-строки owner_type те же;
        ровно один задан (Check жив, теперь на шесть путей — волна 19, Ф12b)."""
        r = models.Receipt.objects.create(
            number='У-2', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        item = self.make_item('A')
        f = SimpleUploadedFile('s.pdf', b'%PDF-1.4', content_type='application/pdf')
        att_doc = engine.add_attachment('receipt', r, f, self.user)
        self.assertEqual(att_doc.document_id, r.pk)
        self.assertIsNone(att_doc.item_id)
        f2 = SimpleUploadedFile('d.pdf', b'%PDF-1.4', content_type='application/pdf')
        att_item = engine.add_attachment('item', item, f2, self.user)
        self.assertEqual(att_item.item_id, item.pk)
        self.assertIsNone(att_item.document_id)
        # список по виду ордера строг: чужой вид → пусто (id глобально уникален)
        self.assertEqual(len(engine.attachments_for('receipt', r.pk)), 1)
        self.assertEqual(len(engine.attachments_for('transfer', r.pk)), 0)
        field_names = {f.name for f in models.Attachment._meta.get_fields()}
        self.assertFalse({'transfer', 'kitting', 'inventory', 'writeoff',
                          'requisition'} & field_names)
        # не-ордерные владельцы, наоборот, поля имеют (Ф12b)
        self.assertLessEqual({'project', 'procurement', 'purchase', 'counterparty'},
                             field_names)
        self.assertIn('attachment_exactly_one_owner',
                      self._constraint_names(models.Attachment))

    def test_reverse_accessors_read_from_document(self):
        """Реверсы дуг объявлены на `StockDocument`, но читаются с любого вида:
        `receipt.lots`, `writeoff.lines`, `receipt.attachments`."""
        r = models.Receipt.objects.create(
            number='У-3', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        lot = engine.add_receipt_lot(r, self.make_item('A'), D(7))
        self.assertEqual(list(r.lots.all()), [lot])
        w = engine.create_writeoff(self.prj, self.user, 'С-3')
        line = engine.add_writeoff_line(w, lot, D(2))
        self.assertEqual(list(w.lines.all()), [line])
        f = SimpleUploadedFile('s.pdf', b'%PDF-1.4', content_type='application/pdf')
        att = engine.add_attachment('receipt', r, f, self.user)
        self.assertEqual(list(r.attachments.all()), [att])


class DocumentHeaderFieldsTests(EngineTestBase):
    """Все колонки ордера — общая шапка И специфика вида — живут в ОДНОЙ таблице.

    Волна 13 (Ф2c) подняла общие поля `project`/`user`/`date`/`number` с шести детей
    в родителя, волна 19 добавила туда `code`/`description` (Ф10) и увела туда же
    специфику видов (Ф14, снос MTI). Своих колонок у видов не осталось вовсе —
    именно это здесь и проверяется; реверс общего поля — `project.documents`."""

    def _one_of_each(self):
        r = models.Receipt.objects.create(
            number='У-1', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        k = models.Kitting.objects.create(
            project=self.prj, target_item=self.make_item('DEV', manufactured=True),
            user=self.user, qty=D(1), date='2026-05-02')
        inv = models.Inventory.objects.create(
            project=self.prj, user=self.user, number='И-1', date='2026-05-03',
            description='примечание акта')
        req = models.Requisition.objects.create(
            project=self.prj, user=self.user, number='Т-1', date='2026-05-04')
        t = models.Transfer.objects.create(
            project=self.prj, user=self.user, number='Н-1', date='2026-05-05')
        w = models.Writeoff.objects.create(
            project=self.prj, user=self.user, number='С-1', date='2026-05-06',
            reason='порча')
        return {'receipt': r, 'kitting': k, 'inventory': inv,
                'requisition': req, 'transfer': t, 'writeoff': w}

    def test_all_columns_live_on_one_table(self):
        """Шапка и специфика — колонки `StockDocument`; у видов своих колонок НЕТ.

        До Ф14 второй половиной этого теста было «специфика осталась на детях» — она
        осталась зелёной и после сноса MTI (proxy наследует поля родителя), продолжая
        утверждать снесённое. Проверка перевёрнута: своих полей у вида ноль — регресс,
        вернувший детскую колонку, теперь виден.
        """
        parent = {f.name for f in models.StockDocument._meta.get_fields()}
        for name in ('project', 'user', 'date', 'number', 'code', 'description',
                     'contractor', 'purchase', 'target_item', 'qty', 'reason'):
            self.assertIn(name, parent)
        for child in (models.Receipt, models.Kitting, models.Inventory,
                      models.Requisition, models.Transfer, models.Writeoff,
                      models.Relocation):
            own = {f.name for f in child._meta.get_fields()
                   if getattr(f, 'model', None) is child}
            self.assertFalse(own, f'{child.__name__} завёл свою колонку: {own}')
            self.assertTrue(child._meta.proxy, f'{child.__name__} перестал быть proxy')

    def test_reverse_accessor_is_documents(self):
        """Реверс общего `project` — `project.documents` (единый по всем видам),
        типизированный дочерний фильтр по родительскому полю тоже работает."""
        docs = self._one_of_each()
        self.assertEqual(self.prj.documents.count(), len(docs))
        self.assertEqual(
            models.Writeoff.objects.filter(project=self.prj).count(), 1)
        self.assertEqual(
            self.prj.documents.filter(kind=models.StockDocument.Kind.TRANSFER).count(), 1)

    def test_kind_manager_filters_and_orders_by_header(self):
        """Менеджер вида фильтрует/сортирует по полю шапки прозрачно
        (как в движке `Writeoff.objects.filter(project=…)`)."""
        self._one_of_each()
        r2 = models.Receipt.objects.create(
            number='У-2', date='2026-06-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        latest = models.Receipt.objects.filter(project=self.prj).order_by('-date').first()
        self.assertEqual(latest, r2)


class HeaderRequiredByKindTests(EngineTestBase):
    """Обязательность шапки — по виду: строгим видам нужны дата и номер, комплектации
    не нужно ничего. Единый kind-driven источник (`REQUIRED_HEADER_BY_KIND`/`clean`)
    гейтит и админ-форму (`full_clean → clean`), и фиксацию (`_require_header`);
    правило восстанавливает per-kind NOT NULL, ослабленный подъёмом полей в общую
    таблицу (волна 13, Ф2c/Ф2d)."""

    def test_strict_kinds_require_date_and_number(self):
        """Строгие виды требуют дату+номер; kitting свободен. Ф2e: relocation стал
        строгим (реальный документ с номером), только kitting остаётся свободным."""
        req = models.StockDocument.REQUIRED_HEADER_BY_KIND
        K = models.StockDocument.Kind
        for kind in (K.RECEIPT, K.INVENTORY, K.REQUISITION, K.TRANSFER, K.WRITEOFF,
                     K.RELOCATION):
            self.assertEqual(set(req[kind]), {'date', 'number'})
        self.assertEqual(req[K.KITTING], ())

    def test_clean_rejects_blank_number(self):
        """Админ-путь: Transfer с пустым номером ловится model.clean() (ошибка по полю)."""
        t = models.Transfer(project=self.prj, user=self.user, number='', date='2026-05-01')
        with self.assertRaises(ValidationError) as cm:
            t.clean()
        self.assertIn('number', cm.exception.message_dict)

    def test_clean_rejects_null_date(self):
        """Inventory без даты — ошибка по полю `date`."""
        inv = models.Inventory(project=self.prj, user=self.user, number='И-1', date=None)
        with self.assertRaises(ValidationError) as cm:
            inv.clean()
        self.assertIn('date', cm.exception.message_dict)

    def test_clean_passes_complete_header(self):
        """Полная шапка строгого вида проходит без ошибок."""
        models.Receipt(project=self.prj, user=self.user, contractor=self.supplier,
                       number='У-1', date='2026-05-01').clean()

    def test_kitting_exempt_from_header(self):
        """Kitting освобождён: пустой номер/дата в clean() не ошибка (как до Ф2c)."""
        models.Kitting(project=self.prj, user=self.user, qty=D(1), number='', date=None,
                       target_item=self.make_item('DEV', manufactured=True)).clean()

    def test_post_gates_incomplete_header(self):
        """Проведение не выпускает неполный ордер, минуя create-guard (прямой ORM).
        Гейт после empty-check: строку добавляем, чтобы дойти до валидации шапки."""
        lot = self.receipt_lot(self.make_item('R'), self.prj, 100)
        w = models.Writeoff.objects.create(project=self.prj, user=self.user,
                                            number='', date='2026-05-01', reason='порча')
        engine.add_writeoff_line(w, lot, D(5))
        with self.assertRaises(ValidationError):
            engine.lock_writeoff(w)
        w.number = 'С-1'
        w.save(update_fields=['number'])
        engine.lock_writeoff(w)
        self.assertTrue(w.locked)

    def test_approve_receipt_gated_on_missing_date(self):
        """lock_receipt гейтит отсутствующую дату (прямой ORM-обход дефолта create)."""
        r = models.Receipt.objects.create(project=self.prj, user=self.user,
                                           contractor=self.supplier, number='У-9', date=None)
        models.Lot.objects.create(item=self.make_item('B'), project=self.prj,
                                  origin=r, qty=D(3))
        with self.assertRaises(ValidationError):
            engine.lock_receipt(r)
        r.date = '2026-05-01'
        r.save(update_fields=['date'])
        engine.lock_receipt(r)
        self.assertTrue(r.locked)


class RelocationAndLocationStockTests(EngineTestBase):
    """Остаток считается по паре `(лот, локация)`, а ход перемещения — пара знаковых
    строк (`−q`@источник, `+q`@приёмник), сохраняющая тотал лота (волна 13, Ф2e)."""

    def setUp(self):
        super().setUp()
        # self.main (код MAIN) уже есть; добавим второе место — станок пайки.
        self.sold = models.Location.objects.create(code='105', description='Место пайки')
        self.case = self.make_item('CASE')
        self.lot = self.receipt_lot(self.case, self.prj, 12)   # рождён на self.main

    def _reloc_move(self, qty=4, lock=True):
        """Ход перемещения. Ф15: по умолчанию **проводим** — до фиксации перемещение
        склад не двигает, а тесты ниже проверяют именно расщепление по местам."""
        r = engine.create_relocation(self.prj, self.user, number='ПЕР-1',
                                     date='2026-06-05')
        engine.add_relocation_line(r, self.lot, D(qty),
                                   from_location=self.main, to_location=self.sold)
        if lock:
            engine.lock_relocation(r)
        return r

    def test_conserves_total_splits_locations(self):
        """Тотал лота сохранён (12), распределение расщеплено: 8@103 + 4@105."""
        self._reloc_move(4)
        self.assertEqual(engine.lot_live_qty(self.lot), D(12))          # тотал цел
        self.assertEqual(engine.lot_live_qty(self.lot, self.main), D(8))
        self.assertEqual(engine.lot_live_qty(self.lot, self.sold), D(4))

    def test_lot_locations_breakdown(self):
        self._reloc_move(4)
        by = {r['location_id']: r['qty'] for r in engine.lot_locations(self.lot)}
        self.assertEqual(by, {self.main.id: D(8), self.sold.id: D(4)})

    def test_item_available_by_location(self):
        self._reloc_move(4)
        self.assertEqual(engine.item_available(self.case, self.prj), D(12))
        self.assertEqual(engine.item_available(self.case, self.prj, self.main), D(8))
        self.assertEqual(engine.item_available(self.case, self.prj, self.sold), D(4))

    def test_available_lots_by_location(self):
        self._reloc_move(4)
        at_main = engine.available_lots(self.case, self.prj, self.main)
        at_sold = engine.available_lots(self.case, self.prj, self.sold)
        self.assertEqual(at_main[0]['live_qty'], D(8))
        self.assertEqual(at_sold[0]['live_qty'], D(4))
        # без локации — тотал (байт-в-байт со старым контрактом формы)
        self.assertEqual(engine.available_lots(self.case, self.prj)[0]['live_qty'], D(12))

    def test_stock_map_by_location(self):
        self._reloc_move(4)
        row = next(r for r in engine.stock_map(self.case)['rows']
                   if r['project_id'] == self.prj.id)
        self.assertEqual(row['available'], D(12))
        by = {b['location_id']: b['available'] for b in row['by_location']}
        self.assertEqual(by, {self.main.id: D(8), self.sold.id: D(4)})

    def test_add_line_creates_signed_pair(self):
        r = self._reloc_move(4, lock=False)          # строки есть и в черновике
        lines = sorted(r.lines.all(), key=lambda l: l.qty)
        self.assertEqual(len(lines), 2)
        self.assertEqual(lines[0].qty, D(-4))   # источник
        self.assertEqual(lines[0].location_id, self.main.id)
        self.assertEqual(lines[1].qty, D(4))    # приёмник
        self.assertEqual(lines[1].location_id, self.sold.id)

    def test_update_line_adjusts_both(self):
        r = self._reloc_move(4, lock=False)
        engine.update_relocation_line(r, self.lot, qty=D(7))
        engine.lock_relocation(r)                    # Ф15: провели — места разъехались
        self.assertEqual(engine.lot_live_qty(self.lot, self.main), D(5))
        self.assertEqual(engine.lot_live_qty(self.lot, self.sold), D(7))
        self.assertEqual(engine.lot_live_qty(self.lot), D(12))

    def test_remove_line_restores(self):
        r = self._reloc_move(4, lock=False)
        engine.remove_relocation_line(r, self.lot)
        self.assertEqual(r.lines.count(), 0)
        self.assertEqual(engine.lot_live_qty(self.lot, self.main), D(12))
        self.assertEqual(engine.lot_live_qty(self.lot, self.sold), D(0))

    def test_form_move(self):
        r = self._reloc_move(4)
        cp = engine.relocation_form(r)
        self.assertEqual(cp['total_qty'], D(4))
        self.assertEqual(len(cp['moves']), 1)
        m = cp['moves'][0]
        self.assertEqual(m['qty'], D(4))
        self.assertEqual(m['from_location_id'], self.main.id)
        self.assertEqual(m['to_location_id'], self.sold.id)
        self.assertEqual(m['from_live_qty'], D(8))
        self.assertEqual(m['to_live_qty'], D(4))

    def test_guards(self):
        r = engine.create_relocation(self.prj, self.user, number='ПЕР-2',
                                     date='2026-06-05')
        # одно и то же место
        with self.assertRaises(ValidationError):
            engine.add_relocation_line(r, self.lot, D(1),
                                       from_location=self.main, to_location=self.main)
        # чужой проект
        other = models.Project.objects.create(code='P2', description='P2',
                                               kind=models.Project.Kind.EXTERNAL)
        foreign = self.receipt_lot(self.make_item('X'), other, 5)
        with self.assertRaises(ValidationError):
            engine.add_relocation_line(r, foreign, D(1),
                                       from_location=self.main, to_location=self.sold)
        # неположительное кол-во
        with self.assertRaises(ValidationError):
            engine.add_relocation_line(r, self.lot, D(0),
                                       from_location=self.main, to_location=self.sold)
        # дубль лота
        engine.add_relocation_line(r, self.lot, D(2),
                                   from_location=self.main, to_location=self.sold)
        with self.assertRaises(ValidationError):
            engine.add_relocation_line(r, self.lot, D(1),
                                       from_location=self.main, to_location=self.sold)

    def test_post_gates_empty_and_header(self):
        # пустое перемещение не проводится
        empty = engine.create_relocation(self.prj, self.user, number='ПЕР-3',
                                         date='2026-06-05')
        with self.assertRaises(ValidationError):
            engine.lock_relocation(empty)
        # шапка обязательна на проведении (relocation стал строгим, Ф2e); прямой ORM
        # обходит create-guard пустым номером
        r = models.Relocation.objects.create(project=self.prj, user=self.user,
                                              number='', date='2026-06-05')
        engine.add_relocation_line(r, self.lot, D(2),
                                   from_location=self.main, to_location=self.sold)
        with self.assertRaises(ValidationError):
            engine.lock_relocation(r)
        r.number = 'ПЕР-4'
        r.save(update_fields=['number'])
        engine.lock_relocation(r)
        self.assertTrue(r.locked)
        # под замком правка запрещена
        with self.assertRaises(ValidationError):
            engine.add_relocation_line(r, self.lot, D(1),
                                       from_location=self.main, to_location=self.sold)

    def test_kind_stamp_and_mti(self):
        r = self._reloc_move(4)
        self.assertEqual(r.kind, models.StockDocument.Kind.RELOCATION)
        parent = models.StockDocument.objects.get(id=r.id)
        self.assertEqual(parent.kind, 'relocation')
        self.assertEqual(parent.id, r.id)   # PK == id родителя (унификация Ф2a)

    def test_source_lots_picker(self):
        self._reloc_move(4)
        picker = engine.relocation_source_lots(self.prj)
        row = next(p for p in picker if p['lot_id'] == self.lot.id)
        self.assertEqual(row['live_qty'], D(12))
        by = {b['location_id']: b['qty'] for b in row['by_location']}
        self.assertEqual(by, {self.main.id: D(8), self.sold.id: D(4)})

    def test_delete_restores_and_conserves(self):
        r = self._reloc_move(4, lock=False)          # удаляют только расфиксированное
        engine.delete_stock_document(r)
        self.assertFalse(models.Relocation.objects.filter(id=r.id).exists())
        self.assertEqual(engine.lot_live_qty(self.lot), D(12))
        self.assertEqual(engine.lot_live_qty(self.lot, self.sold), D(0))


class LotIdentifiersTests(EngineTestBase):
    """У партии два независимых идентификатора: `lot_name` (человеческий, из УПД) и
    `part_number` (машинный — MPN/децимальный). Писатели, формы и метка лота разводят
    их порознь и не подменяют один другим (волна 13, Ф2f)."""

    def setUp(self):
        super().setUp()
        self.item = self.make_item('R100')
        self.receipt = self.make_receipt()

    def make_receipt(self, approved=False):
        return models.Receipt.objects.create(
            number='UPD-2f', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user,
            locked=approved)

    def test_born_lot_carries_both_identifiers(self):
        lot = engine.add_receipt_lot(self.receipt, self.item, D(5),
                                     lot_name='Резистор 10к',
                                     part_number='RES-10K-0805')
        self.assertEqual(lot.lot_name, 'Резистор 10к')
        self.assertEqual(lot.part_number, 'RES-10K-0805')
        row = engine.receipt_form(self.receipt)['lots'][0]
        self.assertEqual(row['lot_name'], 'Резистор 10к')
        self.assertEqual(row['part_number'], 'RES-10K-0805')

    def test_update_separates_identifiers(self):
        lot = engine.add_receipt_lot(self.receipt, self.item, D(5))
        engine.update_receipt_lot(lot, part_number='PN-1')   # только PN
        lot.refresh_from_db()
        self.assertEqual(lot.part_number, 'PN-1')
        self.assertEqual(lot.lot_name, '')                   # имя не тронуто
        engine.update_receipt_lot(lot, lot_name='Имя')       # только имя
        lot.refresh_from_db()
        self.assertEqual(lot.lot_name, 'Имя')
        self.assertEqual(lot.part_number, 'PN-1')

    def test_lot_label_prefers_lot_name_then_part_number(self):
        lot = engine.add_receipt_lot(self.receipt, self.item, D(1),
                                     part_number='PN-ONLY')
        # только PN → метка берёт PN (нет человеческого имени)
        self.assertIn('PN-ONLY', engine._lot_label(lot))
        engine.update_receipt_lot(lot, lot_name='Человек')
        lot.refresh_from_db()
        # появилось имя → приоритет у него
        label = engine._lot_label(lot)
        self.assertIn('Человек', label)
        self.assertNotIn('PN-ONLY', label)

    def test_requisition_child_inherits_both(self):
        src = engine.add_receipt_lot(self.receipt, self.item, D(10),
                                     lot_name='Исходник', part_number='PN-SRC')
        white = models.Project.objects.create(
            code='WHITE', description='Собственный склад',
            kind=models.Project.Kind.INTERNAL_STOCK)
        req = engine.create_requisition(white, self.user, 'ТР-1')
        engine.add_requisition_line(req, src, D(4))
        born = models.Lot.objects.get(origin=req)
        self.assertEqual(born.lot_name, 'Исходник')
        self.assertEqual(born.part_number, 'PN-SRC')


class CounterpartyRolesTests(EngineTestBase):
    """Контрагент — одна сущность, а не два справочника: у поставки он поставщик, у
    передачи заказчик, направление читается из ВИДА документа (волна 13, Ф2f+; одна
    колонка — волна 19, Ф14). Ролей-флагов у справочника нет (волна 20, Ф3) —
    сторона выводится из фактов, см. `CounterpartySidesTests`."""

    def test_receipt_form_emits_contractor(self):
        r = models.Receipt.objects.create(
            number='U-g', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        cp = engine.receipt_form(r)
        self.assertEqual(cp['contractor_id'], self.supplier.id)
        self.assertEqual(cp['contractor_name'], self.supplier.description)

    def test_create_transfer_with_customer(self):
        cust = models.Counterparty.objects.create(
            description='Заказчик')
        t = engine.create_transfer(self.prj, self.user, 'Н-1', contractor=cust)
        self.assertEqual(t.contractor_id, cust.id)
        cp = engine.transfer_form(t)
        self.assertEqual(cp['contractor_id'], cust.id)
        self.assertEqual(cp['contractor_name'], 'Заказчик')

    def test_transfer_contractor_optional_and_settable(self):
        t = engine.create_transfer(self.prj, self.user, 'Н-2')   # без получателя
        self.assertIsNone(t.contractor_id)
        self.assertEqual(engine.transfer_form(t)['contractor_name'], '')
        cust = models.Counterparty.objects.create(
            description='Поздний')
        engine.update_transfer(t, contractor=cust)               # проставить позже
        t.refresh_from_db()
        self.assertEqual(t.contractor_id, cust.id)
        engine.update_transfer(t, contractor=None)               # снять (nullable)
        t.refresh_from_db()
        self.assertIsNone(t.contractor_id)

    def test_update_transfer_sentinel_keeps_contractor(self):
        """Часовой `_UNSET`: правка номера/даты не сбрасывает получателя."""
        cust = models.Counterparty.objects.create(
            description='Стойкий')
        t = engine.create_transfer(self.prj, self.user, 'Н-3', contractor=cust)
        engine.update_transfer(t, number='Н-3-ред')              # contractor не передан
        t.refresh_from_db()
        self.assertEqual(t.contractor_id, cust.id)
        self.assertEqual(t.number, 'Н-3-ред')

    def test_counterparties_endpoint_hides_nobody(self):
        """Ф3: список — ВЕСЬ справочник, фильтра по роли больше нет.

        Прежний `?role=` прятал запись, у которой «роль не та», и с места ошибки это
        не читалось. Параметр остался разве что в старых закладках — он обязан быть
        безвредным."""
        models.Counterparty.objects.create(description='Ничей')
        c = Client()
        c.force_login(self.user)
        rows = c.get('/api/counterparties/').json()
        names = {r['description'] for r in rows}
        self.assertIn('Поставщик', names)
        self.assertIn('Ничей', names)                       # без фактов — но в списке
        self.assertNotIn('is_supplier', rows[0])            # флагов в проекции нет
        stale = {r['description'] for r in
                 c.get('/api/counterparties/?role=supplier').json()}
        self.assertEqual(stale, names)                      # забытый параметр не сужает


class CounterpartyFormTests(EngineTestBase):
    """Волна 20 — витрина контрагента: две стороны документооборота, четыре списка.

    Закупочная сторона (закупки → заказы → поставки) и передачная (передачи заказчику)
    считаются раздельно; сторона без движений отдаётся `None` — «её нет», и вью не рисует
    панель нулей. Материальный итог берёт только ЗАФИКСИРОВАННЫЕ документы (гейт Ф15,
    как «потрачено» проекта).
    """

    def setUp(self):
        super().setUp()
        self.comp = self.make_item('C1')

    def _receipt(self, qty, cost, locked=True, contractor=None, uom=None):
        """Поставка с одной партией по цене — единица берётся из изделия."""
        item = self.comp
        if uom:
            item = models.Item.objects.create(code=f'C-{uom}', description=uom,
                                              category=_cat(), uom=uom)
        r = models.Receipt.objects.create(
            number=f'УПД-{qty}-{cost}', date='2026-05-01',
            contractor=contractor or self.supplier,
            project=self.prj, user=self.user, locked=locked)
        lot = models.Lot.objects.create(item=item, project=self.prj, origin=r,
                                        qty=D(qty), unit_cost=D(cost))
        engine.rebuild_movements(lot)
        return r, lot

    def test_both_sides_none_without_documents(self):
        """Свежий контрагент: движений нет ни на одной стороне — обе панели `None`."""
        fresh = engine.create_counterparty(description='Пустой')
        form = engine.counterparty_form(fresh)
        self.assertIsNone(form['supply'])
        self.assertIsNone(form['shipment'])
        self.assertEqual(form['procurements'], [])
        self.assertEqual(form['receipts'], [])

    def test_supply_integral_counts_contour_and_material(self):
        proc = engine.create_procurement(self.user)
        engine.update_procurement(proc, contractor=self.supplier)
        self.make_purchase()                       # заказ на self.supplier (пустой)
        self._receipt(10, 100)                     # зафиксированная поставка
        self._receipt(3, 50, uom='м')              # другая единица — вектор, не скаляр
        supply = engine.counterparty_form(self.supplier)['supply']
        self.assertEqual(supply['procurements'], 1)
        self.assertEqual(supply['purchases'], 1)
        self.assertEqual(supply['receipts'], 2)
        self.assertEqual(supply['lots'], 2)
        self.assertEqual(supply['qty_by_uom'],
                         [{'uom': 'м', 'qty': D(3)}, {'uom': 'шт', 'qty': D(10)}])
        self.assertEqual(supply['total'], D(10) * D(100) + D(3) * D(50))

    def test_supply_material_ignores_draft_receipts(self):
        """Черновой УПД — намерение, не факт: в интеграл он не идёт, в таб идёт."""
        self._receipt(10, 100, locked=False)
        form = engine.counterparty_form(self.supplier)
        self.assertEqual(form['supply']['receipts'], 1)   # в счёте документов виден
        self.assertEqual(form['supply']['lots'], 0)       # в материальном итоге — нет
        self.assertEqual(form['supply']['total'], D(0))
        # «поставок 1, привёз 0» без объяснения читалось бы как баг — панель подписана
        self.assertEqual(form['supply']['draft_receipts'], 1)
        self.assertEqual(len(form['receipts']), 1)        # таб показывает всё

    def test_open_purchases_counts_unclosed_orders(self):
        p = self.make_purchase()
        engine.add_purchase_line(p, self.comp, D(5))
        supply = engine.counterparty_form(self.supplier)['supply']
        self.assertEqual(supply['open_purchases'], 1)     # ни одного лота
        engine.lock_purchase(p)
        _r, lot = self._receipt(5, 10)
        lot.origin.purchase = p
        lot.origin.save(update_fields=['purchase'])
        supply = engine.counterparty_form(self.supplier)['supply']
        self.assertEqual(supply['purchases'], 1)
        self.assertEqual(supply['open_purchases'], 0)     # строка закрыта целиком

    def test_shipment_integral_from_locked_transfers(self):
        cust = engine.create_counterparty(description='Заказчик')
        device = self.make_item('DEV', manufactured=True)
        r, lot = self._receipt(5, 200)
        lot.item = device
        lot.save(update_fields=['item'])
        t = engine.create_transfer(self.prj, self.user, 'Н-1', contractor=cust)
        engine.add_transfer_line(t, lot, D(2))
        form = engine.counterparty_form(cust)
        self.assertIsNone(form['supply'])                  # заказчик ничего не привозил
        self.assertEqual(form['shipment']['transfers'], 1)
        self.assertEqual(form['shipment']['lots'], 0)      # черновик ещё не отгрузил
        self.assertEqual(form['shipment']['draft_transfers'], 1)
        engine.lock_transfer(t)
        shipment = engine.counterparty_form(cust)['shipment']
        self.assertEqual(shipment['lots'], 1)
        self.assertEqual(shipment['qty_by_uom'], [{'uom': 'шт', 'qty': D(2)}])
        self.assertEqual(shipment['total'], D(2) * D(200))  # по цене лота-источника

    def test_both_sides_live_together(self):
        """Одно юрлицо в двух ролях — панелей две сразу (роли не исключают друг друга)."""
        both = engine.create_counterparty(description='И то, и то')
        self._receipt(4, 25, contractor=both)
        _r, lot = self._receipt(4, 25)
        t = engine.create_transfer(self.prj, self.user, 'Н-2', contractor=both)
        engine.add_transfer_line(t, lot, D(1))
        engine.lock_transfer(t)
        form = engine.counterparty_form(both)
        self.assertIsNotNone(form['supply'])
        self.assertIsNotNone(form['shipment'])

    def test_tabs_list_all_four_levels(self):
        proc = engine.create_procurement(self.user, code='ЗАК-1')
        engine.update_procurement(proc, contractor=self.supplier)
        engine.add_procurement_line(proc, self.comp, D(7))
        p = self.make_purchase(code='ЗАКАЗ-1')
        engine.add_purchase_line(p, self.comp, D(7))
        self._receipt(7, 11)
        cust = engine.create_counterparty(description='Заказчик-2')
        _r, lot = self._receipt(2, 5)
        t = engine.create_transfer(self.prj, self.user, 'Н-3', contractor=cust)
        engine.add_transfer_line(t, lot, D(2))
        engine.lock_transfer(t)

        form = engine.counterparty_form(self.supplier)
        self.assertEqual([r['code'] for r in form['procurements']], ['ЗАК-1'])
        self.assertEqual(form['procurements'][0]['qty'], D(7))
        self.assertEqual([r['code'] for r in form['purchases']], ['ЗАКАЗ-1'])
        self.assertEqual(form['purchases'][0]['coverage'], 'to_order')  # ждём поставки
        self.assertEqual(form['purchases'][0]['project_code'], self.prj.code)
        self.assertEqual(len(form['receipts']), 2)
        self.assertEqual(form['transfers'], [])            # передачи ушли заказчику
        row = engine.counterparty_form(cust)['transfers'][0]
        self.assertEqual(row['qty'], D(2))                 # магнитуда знаковой строки
        self.assertEqual(row['total'], D(2) * D(5))

    def test_birth_without_code_gets_fallback(self):
        """Ф12e: рождение по клику — код фолбэком, титул формы не пустует."""
        c = engine.create_counterparty()
        self.assertEqual(c.code, f'Контрагент {c.pk}')
        named = engine.create_counterparty(code='КОМПЭЛ', description='ООО Компэл')
        self.assertEqual(named.code, 'КОМПЭЛ')
        with self.assertRaises(ValidationError):
            engine.create_counterparty(code='КОМПЭЛ')       # мягкая уникальность

    def test_update_dna_under_lock(self):
        engine.update_counterparty(self.supplier, code='  КОД  ',
                                   description=' ООО Ромашка ', inn=' 7712345678 ')
        self.supplier.refresh_from_db()
        self.assertEqual(self.supplier.code, 'КОД')
        self.assertEqual(self.supplier.description, 'ООО Ромашка')
        self.assertEqual(self.supplier.inn, '7712345678')
        engine.update_counterparty(self.supplier, inn='')    # передумать вправе
        self.supplier.refresh_from_db()
        self.assertEqual(self.supplier.inn, '')
        engine.update_counterparty(self.supplier, code='')   # пустой код → NULL
        self.supplier.refresh_from_db()
        self.assertIsNone(self.supplier.code)
        engine.update_counterparty(self.supplier, description='Тронули описание')
        self.supplier.refresh_from_db()
        self.assertIsNone(self.supplier.code)                # часовой не выдумал код

    def test_delete_blocked_by_documents_and_orders(self):
        with_receipt = engine.create_counterparty(description='С поставкой')
        self._receipt(1, 1, contractor=with_receipt)
        with self.assertRaises(ValidationError):
            engine.delete_counterparty(with_receipt)
        with_order = engine.create_counterparty(description='С заказом')
        self.make_purchase(contractor=with_order)
        with self.assertRaises(ValidationError):
            engine.delete_counterparty(with_order)

    def test_delete_empties_plan_contractor(self):
        """`Procurement.contractor` — SET_NULL (Ф17): план не держит справочник."""
        cp = engine.create_counterparty(description='Только план')
        proc = engine.create_procurement(self.user)
        engine.update_procurement(proc, contractor=cp)
        engine.delete_counterparty(cp)
        proc.refresh_from_db()
        self.assertIsNone(proc.contractor_id)

    def test_sides_by_facts_replace_role_flags(self):
        """Ф3: сторона контрагента — ФАКТ документооборота, а не флаг справочника.

        `counterparty_sides` — ось СПИСКА (глиф `fold-*` и порядок в пикере), и она
        обязана давать ровно то же, что интегралы формы: «есть сторона» = «интеграл
        не `None`». Любой из трёх закупочных документов зажигает закупочную сторону —
        в том числе план и заказ, по которым ещё ничего не приехало.
        """
        def sides(cp):
            row = engine.counterparty_sides(
                models.Counterparty.objects.filter(pk=cp.pk)).get()
            return row.has_supply, row.has_shipment

        naked = engine.create_counterparty(description='Ничей')
        self.assertEqual(sides(naked), (False, False))      # пустой — законное состояние

        planned = engine.create_counterparty(description='Только план')
        proc = engine.create_procurement(self.user)
        engine.update_procurement(proc, contractor=planned)
        self.assertEqual(sides(planned), (True, False))     # заказа ещё нет — сторона есть

        ordered = engine.create_counterparty(description='Только заказ')
        self.make_purchase(contractor=ordered)
        self.assertEqual(sides(ordered), (True, False))

        self.assertEqual(sides(self.supplier), (False, False))
        self._receipt(3, 7)                                 # привёз — стал закупочной
        self.assertEqual(sides(self.supplier), (True, False))

        cust = engine.create_counterparty(description='Заказчик')
        _r, lot = self._receipt(2, 5)
        t = engine.create_transfer(self.prj, self.user, 'Н-с', contractor=cust)
        engine.add_transfer_line(t, lot, D(1))
        self.assertEqual(sides(cust), (False, True))        # черновик — тоже факт стороны

        both = engine.create_counterparty(description='И то, и то')
        self._receipt(1, 1, contractor=both)
        t2 = engine.create_transfer(self.prj, self.user, 'Н-с2', contractor=both)
        engine.add_transfer_line(t2, lot, D(1))
        self.assertEqual(sides(both), (True, True))
        # Ось списка и витрина формы — одна правда.
        form = engine.counterparty_form(both)
        self.assertEqual((form['supply'] is not None, form['shipment'] is not None),
                         sides(both))


@override_settings(MEDIA_ROOT=_TEST_MEDIA)
class CounterpartyDeleteSweepsFilesTests(EngineTestBase):
    """Долг из семени волны 20: у контрагента не было пути удаления вообще, а «карточка
    предприятия» — вложение (Ф12b). Каскад БД унёс бы строку и оставил файл сиротой."""

    def test_delete_removes_attachment_files(self):
        cp = engine.create_counterparty(description='С карточкой')
        att = engine.add_attachment(
            'counterparty', cp,
            SimpleUploadedFile('card.pdf', b'%PDF-1.4 card', content_type='application/pdf'),
            self.user)
        path = att.file.path
        self.assertTrue(os.path.exists(path))
        engine.delete_counterparty(cp)
        self.assertFalse(models.Attachment.objects.filter(pk=att.id).exists())
        self.assertFalse(os.path.exists(path))


class CounterpartyHttpTests(TestCase):
    """Волна 20 — HTTP-контур режима «Контрагенты»: рождение → форма → правка → снос."""

    def setUp(self):
        self.user = get_user_model().objects.create(username='t')
        self.client = Client()
        self.client.force_login(self.user)

    def test_born_by_click_then_form(self):
        born = self.client.post('/api/counterparties/', {},
                                content_type='application/json')
        self.assertEqual(born.status_code, 201)
        cid = born.json()['id']
        self.assertEqual(born.json()['code'], f'Контрагент {cid}')
        # Ф3: рождённый по клику пуст с обеих сторон — ни панелей, ни табов контура.
        # Это ожидаемая картина, а не дыра: контрагента заводят под первый заказ.
        self.assertEqual(born.json()['has_supply'], False)
        self.assertEqual(born.json()['has_shipment'], False)
        form = self.client.get(f'/api/counterparties/{cid}/').json()
        self.assertNotIn('role', form)
        self.assertIsNone(form['supply'])
        self.assertIsNone(form['shipment'])
        self.assertEqual(form['transfers'], [])

    def test_quick_create_uses_typed_name_as_code(self):
        """«Завести "X"» из пикера: напечатанное имя — это и ярлык (`code`) тоже."""
        created = self.client.post('/api/counterparties/',
                                   {'description': 'Амети́ст'},
                                   content_type='application/json').json()
        self.assertEqual(created['code'], 'Амети́ст')

    def test_patch_dna(self):
        """ДНК контрагента — ровно три поля (Ф3: роль снесена, сторона выводится)."""
        cid = self.client.post('/api/counterparties/', {},
                               content_type='application/json').json()['id']
        form = self.client.patch(
            f'/api/counterparties/{cid}/',
            {'code': 'КОМПЭЛ', 'description': 'ООО Компэл', 'inn': '7712345678'},
            content_type='application/json').json()
        self.assertEqual(form['code'], 'КОМПЭЛ')
        self.assertEqual(form['description'], 'ООО Компэл')
        self.assertEqual(form['inn'], '7712345678')

    def test_patch_duplicate_code_is_friendly_400(self):
        self.client.post('/api/counterparties/', {'code': 'ЗАНЯТО', 'description': 'Раз'},
                         content_type='application/json')
        cid = self.client.post('/api/counterparties/', {'description': 'Два'},
                               content_type='application/json').json()['id']
        bad = self.client.patch(f'/api/counterparties/{cid}/', {'code': 'ЗАНЯТО'},
                                content_type='application/json')
        self.assertEqual(bad.status_code, 400)
        self.assertIn('занят', bad.json()['detail'])

    def test_delete_empty_then_guarded(self):
        cid = self.client.post('/api/counterparties/', {'description': 'На снос'},
                               content_type='application/json').json()['id']
        self.assertEqual(self.client.delete(f'/api/counterparties/{cid}/').status_code, 204)
        # с поставкой — friendly-guard, а не 500
        cp = engine.create_counterparty(description='Держит')
        prj = models.Project.objects.create(code='P1', description='Проект 1',
                                            kind=models.Project.Kind.EXTERNAL)
        models.Receipt.objects.create(number='УПД-9', date='2026-05-01', contractor=cp,
                                      project=prj, user=self.user)
        blocked = self.client.delete(f'/api/counterparties/{cp.id}/')
        self.assertEqual(blocked.status_code, 400)
        self.assertIn('удаление заблокировано', blocked.json()['detail'])


class OrderAdminTests(EngineTestBase):
    """Админка ордера — ОДИН пункт на семь видов, и она не подменяет движок.

    До Ф16 здесь было восемь регистраций на одну таблицу (read-only обзор + семь
    per-kind админок) — пережиток MTI, который Ф14 сделала бессмысленным. Теперь
    список и форма одни, вид рулит составом полей и инлайнов, а всё, что двигает
    склад (замок, удаление, строки, партии), из админки недоступно: инвариант
    «движения ⟺ документ зафиксирован» держит движок, и молча обойти его нельзя.
    """

    def _admin(self):
        from django.contrib import admin as dj_admin
        return dj_admin.site._registry[models.StockDocument]

    def _request(self):
        from django.test import RequestFactory
        su = get_user_model().objects.create_superuser('su_adm', 'a@a.tld', 'x')
        req = RequestFactory().get('/admin/')
        req.user = su
        return req

    def test_seven_kind_admins_collapsed_to_one(self):
        """В сайдбаре ордер ровно один: proxy-виды не зарегистрированы."""
        from django.contrib import admin as dj_admin
        self.assertIn(models.StockDocument, dj_admin.site._registry)
        for proxy in (models.Receipt, models.Kitting, models.Inventory,
                      models.Requisition, models.Transfer, models.Writeoff,
                      models.Relocation):
            self.assertNotIn(proxy, dj_admin.site._registry,
                             f'{proxy.__name__} снова отдельным пунктом сайдбара')

    def test_form_shape_follows_kind(self):
        """Поля и инлайны формы зависят от вида: специфика чужого вида не показывается
        (её колонка обязана быть пустой — CHECK `doc_*_only_*`)."""
        ma, req = self._admin(), self._request()
        r = models.Receipt.objects.create(
            number='ПР-f', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        k = models.Kitting.objects.create(
            project=self.prj, user=self.user, qty=D(1),
            target_item=self.make_item('DEV-f', manufactured=True))
        w = models.Writeoff.objects.create(
            project=self.prj, user=self.user, number='СП-f', date='2026-05-02')

        receipt_fields = ma.get_fields(req, r)
        self.assertIn('contractor', receipt_fields)
        self.assertIn('purchase', receipt_fields)
        self.assertNotIn('target_item', receipt_fields)
        self.assertNotIn('reason', receipt_fields)

        kitting_fields = ma.get_fields(req, k)
        self.assertIn('target_item', kitting_fields)
        self.assertIn('qty', kitting_fields)
        self.assertNotIn('contractor', kitting_fields)
        self.assertNotIn('number', kitting_fields)      # у комплектации номера нет

        self.assertIn('reason', ma.get_fields(req, w))

        # инлайны: строки — у видов, которые расходуют; born-лоты — у рождающих
        self.assertEqual(ma.get_inlines(req, r), [admin.BornLotsInline])
        self.assertEqual(ma.get_inlines(req, k),
                         [admin.KittingLinesInline, admin.BornLotsInline])
        self.assertEqual(ma.get_inlines(req, w), [admin.StockLinesInline])
        self.assertEqual(ma.get_inlines(req, None), [])   # на форме добавления вида нет

    def test_lock_and_delete_belong_to_engine(self):
        """Замок read-only, удаления нет, вид у существующего ордера не меняется."""
        ma, req = self._admin(), self._request()
        r = models.Receipt.objects.create(
            number='ПР-g', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        self.assertIn('locked', ma.get_readonly_fields(req, r))
        self.assertIn('kind', ma.get_readonly_fields(req, r))
        self.assertIn('locked', ma.get_readonly_fields(req, None))
        self.assertNotIn('kind', ma.get_readonly_fields(req, None))  # при заводе — выбор
        self.assertFalse(ma.has_delete_permission(req, r))
        self.assertNotIn('delete_selected', ma.get_actions(req))

    def test_engine_owned_tables_are_read_only(self):
        """Партии, движения, строки и вложения админка только показывает."""
        from django.contrib import admin as dj_admin
        req = self._request()
        for model in (models.Lot, models.StockMovement, models.StockLine,
                      models.Attachment):
            ma = dj_admin.site._registry[model]
            self.assertFalse(ma.has_add_permission(req), model.__name__)
            self.assertFalse(ma.has_change_permission(req), model.__name__)
            self.assertFalse(ma.has_delete_permission(req), model.__name__)

    def test_changelist_and_form_http(self):
        """HTTP-путь: смешанный список видит все виды, форма вида открывается."""
        r = models.Receipt.objects.create(
            number='ПР-http', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        models.Writeoff.objects.create(
            number='СП-http', date='2026-05-02', reason='порча',
            project=self.prj, user=self.user)
        su = get_user_model().objects.create_superuser(
            username='root', email='r@e.x', password='x')
        c = Client()
        c.force_login(su)
        lst = c.get('/admin/plume/stockdocument/')
        self.assertEqual(lst.status_code, 200)
        self.assertContains(lst, 'ПР-http')              # оба вида вперемешку
        self.assertContains(lst, 'СП-http')
        form = c.get(f'/admin/plume/stockdocument/{r.pk}/change/')
        self.assertEqual(form.status_code, 200)
        self.assertContains(form, 'ПР-http')
        # удалять ордер из админки нельзя — кнопки нет
        self.assertNotContains(form, f'/admin/plume/stockdocument/{r.pk}/delete/')

    def test_change_form_renders_for_every_kind(self):
        """Форма открывается у всех семи видов. Инлайн-витрина ломается конфигом
        (поля/`readonly_fields`/`max_num`), и ломается она в РЕНДЕРЕ — 500 админки
        сюита не видит, пока страницу не запросили."""
        su = get_user_model().objects.create_superuser(
            username='root2', email='r2@e.x', password='x')
        c = Client()
        c.force_login(su)
        Kind = models.StockDocument.Kind
        docs = {
            Kind.RECEIPT: models.Receipt.objects.create(
                number='ПР-r', date='2026-05-01', contractor=self.supplier,
                project=self.prj, user=self.user),
            Kind.KITTING: models.Kitting.objects.create(
                project=self.prj, user=self.user, qty=D(1),
                target_item=self.make_item('DEV-r', manufactured=True)),
            Kind.INVENTORY: models.Inventory.objects.create(
                number='ИН-r', date='2026-05-01', project=self.prj, user=self.user),
            Kind.REQUISITION: models.Requisition.objects.create(
                number='ТР-r', date='2026-05-01', project=self.prj, user=self.user),
            Kind.TRANSFER: models.Transfer.objects.create(
                number='НК-r', date='2026-05-01', project=self.prj, user=self.user),
            Kind.WRITEOFF: models.Writeoff.objects.create(
                number='СП-r', date='2026-05-01', project=self.prj, user=self.user),
            Kind.RELOCATION: models.Relocation.objects.create(
                number='ПМ-r', date='2026-05-01', project=self.prj, user=self.user),
        }
        for kind, doc in docs.items():
            resp = c.get(f'/admin/plume/stockdocument/{doc.pk}/change/')
            self.assertEqual(resp.status_code, 200, f'форма вида {kind} не открылась')
        # и форма заведения нового ордера (вида ещё нет — специфики на ней тоже)
        self.assertEqual(c.get('/admin/plume/stockdocument/add/').status_code, 200)


class DocumentAuthorTests(EngineTestBase):
    """Автор — поле документа, а не системный штамп: переназначается на всех ордерах
    (плюс заказ и закупка), пока документ расфиксирован. Формы несут `user_id`/
    `user_name`, `update_*` принимают часового `_UNSET`, `/api/users/` кормит пикер
    (волна 13, Ф2j)."""

    def setUp(self):
        super().setUp()
        # второй автор с человеческим именем — цель переназначения авторства
        self.author2 = get_user_model().objects.create(
            username='ivan', first_name='Иван', last_name='Пэ')

    def test_form_emits_author(self):
        r = models.Receipt.objects.create(
            number='U-j', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.author2)
        cp = engine.receipt_form(r)
        self.assertEqual(cp['user_id'], self.author2.id)
        self.assertEqual(cp['user_name'], 'Иван Пэ')       # get_full_name()

    def test_purchase_and_procurement_carry_author(self):
        p = engine.create_purchase(self.prj, self.author2)
        self.assertEqual(engine.purchase_form(p)['user_id'], self.author2.id)
        proc = engine.create_procurement(self.author2)
        self.assertEqual(engine.procurement_form(proc)['user_id'], self.author2.id)

    def test_update_changes_author(self):
        r = models.Receipt.objects.create(
            number='U-j2', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        engine.update_receipt(r, user=self.author2)
        r.refresh_from_db()
        self.assertEqual(r.user_id, self.author2.id)

    def test_author_sentinel_keeps_current(self):
        """Часовой `_UNSET`: правка номера/даты не сбрасывает автора."""
        r = models.Receipt.objects.create(
            number='U-j3', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.author2)
        engine.update_receipt(r, number='U-j3-ред')        # user не передан
        r.refresh_from_db()
        self.assertEqual(r.user_id, self.author2.id)
        self.assertEqual(r.number, 'U-j3-ред')

    def test_author_none_rejected(self):
        """Автор обязателен (FK NOT NULL) — явный `None` отклоняется."""
        r = models.Receipt.objects.create(
            number='U-j4', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        with self.assertRaises(ValidationError):
            engine.update_receipt(r, user=None)

    def test_author_edit_gated_by_lock(self):
        """Проведённый ордер (edit-freeze) не отдаёт авторство на правку."""
        lot = self.receipt_lot(self.make_item('Rj'), self.prj, 10)
        w = engine.create_writeoff(self.prj, self.user, 'СП-j', date='2026-05-01')
        engine.add_writeoff_line(w, lot, D(2))
        engine.lock_writeoff(w)
        with self.assertRaises(ValidationError):
            engine.update_writeoff(w, user=self.author2)
        w.refresh_from_db()
        self.assertEqual(w.user_id, self.user.id)          # автор не сдвинулся

    def test_users_endpoint_lists_active(self):
        inactive = get_user_model().objects.create(username='ghost', is_active=False)
        c = Client()
        c.force_login(self.user)
        ids = {u['id'] for u in c.get('/api/users/').json()}
        self.assertIn(self.author2.id, ids)
        self.assertNotIn(inactive.id, ids)                 # неактивные скрыты

    def test_patch_user_id_changes_author_http(self):
        r = models.Receipt.objects.create(
            number='U-j5', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        c = Client()
        c.force_login(self.user)
        resp = c.patch(f'/api/receipts/{r.id}/', {'user_id': self.author2.id},
                       content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()['user_id'], self.author2.id)
        r.refresh_from_db()
        self.assertEqual(r.user_id, self.author2.id)
        # несуществующий пользователь → дружелюбный 400 (не 500)
        bad = c.patch(f'/api/receipts/{r.id}/', {'user_id': 99999},
                      content_type='application/json')
        self.assertEqual(bad.status_code, 400)


class OrderAnchorTests(EngineTestBase):
    """Структурные якоря шапки (`project` у всех ордеров и заказа, `target_item` у
    комплектации, `procurement` у заказа) меняются только у «пустого» документа:
    за якорем следуют лоты и строки, поэтому при непустом — дружелюбный отказ.
    Часовой `_UNSET`, `None`-отказ (FK NOT NULL), гейт замком (волна 13, Ф2k)."""

    def setUp(self):
        super().setUp()
        self.prj2 = models.Project.objects.create(
            code='P9', description='Проект 9', kind=models.Project.Kind.EXTERNAL)

    # ── project-якорь ──────────────────────────────────────────────────────
    def test_project_changes_on_empty_order(self):
        w = engine.create_writeoff(self.prj, self.user, 'СП-k', date='2026-05-01')
        engine.update_writeoff(w, project=self.prj2)
        w.refresh_from_db()
        self.assertEqual(w.project_id, self.prj2.id)

    def test_project_refused_when_lines_exist(self):
        lot = self.receipt_lot(self.make_item('Rk'), self.prj, 10)
        w = engine.create_writeoff(self.prj, self.user, 'СП-k2', date='2026-05-01')
        engine.add_writeoff_line(w, lot, D(2))
        with self.assertRaises(ValidationError):
            engine.update_writeoff(w, project=self.prj2)
        w.refresh_from_db()
        self.assertEqual(w.project_id, self.prj.id)      # якорь не сдвинулся

    def test_project_refused_when_born_lots_exist(self):
        r = models.Receipt.objects.create(
            number='U-k', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        engine.add_receipt_lot(r, self.make_item('Rk3'), D(5))   # рождает born-лот
        with self.assertRaises(ValidationError):
            engine.update_receipt(r, project=self.prj2)

    def test_project_sentinel_keeps_current(self):
        """Часовой `_UNSET`: правка номера не сбрасывает проект-якорь."""
        r = models.Receipt.objects.create(
            number='U-k2', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        engine.update_receipt(r, number='U-k2-ред')       # project не передан
        r.refresh_from_db()
        self.assertEqual(r.project_id, self.prj.id)
        self.assertEqual(r.number, 'U-k2-ред')

    def test_project_none_rejected(self):
        r = models.Receipt.objects.create(
            number='U-k3', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        with self.assertRaises(ValidationError):
            engine.update_receipt(r, project=None)        # FK NOT NULL

    def test_project_edit_gated_by_lock(self):
        """Проведённый ордер (edit-freeze) не отдаёт проект на правку."""
        lot = self.receipt_lot(self.make_item('Rk4'), self.prj, 10)
        w = engine.create_writeoff(self.prj, self.user, 'СП-k4', date='2026-05-01')
        engine.add_writeoff_line(w, lot, D(2))
        engine.lock_writeoff(w)
        with self.assertRaises(ValidationError):
            engine.update_writeoff(w, project=self.prj2)

    # ── target_item-якорь (комплектация) ───────────────────────────────────
    def test_target_item_changes_on_empty_kitting(self):
        dev = self.make_item('DEV-k', manufactured=True)
        dev2 = self.make_item('DEV-k2', manufactured=True)
        k = models.Kitting.objects.create(project=self.prj, target_item=dev,
                                          user=self.user, qty=D(1))
        engine.update_kitting(k, target_item=dev2)
        k.refresh_from_db()
        self.assertEqual(k.target_item_id, dev2.id)

    def test_target_item_refused_when_lines_exist(self):
        comp = self.make_item('Rk5')
        lot = self.receipt_lot(comp, self.prj, 10)
        dev = self.make_item('DEV-k3', manufactured=True)
        dev2 = self.make_item('DEV-k4', manufactured=True)
        k = models.Kitting.objects.create(project=self.prj, target_item=dev,
                                          user=self.user, qty=D(1))
        engine.add_kitting_line(k, comp, lot, D(3))
        with self.assertRaises(ValidationError):
            engine.update_kitting(k, target_item=dev2)
        k.refresh_from_db()
        self.assertEqual(k.target_item_id, dev.id)

    # ── Purchase: project + procurement ────────────────────────────────────
    def test_purchase_project_changes_without_receipts(self):
        p = engine.create_purchase(self.prj, self.user)
        engine.update_purchase(p, project=self.prj2)
        p.refresh_from_db()
        self.assertEqual(p.project_id, self.prj2.id)

    def test_purchase_project_refused_with_receipts(self):
        p = engine.create_purchase(self.prj, self.user)
        models.Receipt.objects.create(                    # приход, привязан к заказу
            number='U-k6', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user, purchase=p)
        with self.assertRaises(ValidationError):
            engine.update_purchase(p, project=self.prj2)
        p.refresh_from_db()
        self.assertEqual(p.project_id, self.prj.id)

    def test_purchase_procurement_changes(self):
        p = engine.create_purchase(self.prj, self.user)
        proc2 = engine.create_procurement(self.user)
        engine.update_purchase(p, procurement=proc2)
        p.refresh_from_db()
        self.assertEqual(p.procurement_id, proc2.id)
        self.assertEqual(engine.purchase_form(p)['procurement_id'], proc2.id)

    def test_purchase_procurement_clears_to_none(self):
        """Ф17: закупка-план опциональна — якорь снимается, заказ живёт дальше."""
        p = engine.create_purchase(self.prj, self.user)
        engine.update_purchase(p, procurement=engine.create_procurement(self.user))
        engine.update_purchase(p, procurement=None)
        p.refresh_from_db()
        self.assertIsNone(p.procurement_id)
        self.assertIsNone(engine.purchase_form(p)['procurement_id'])

    # ── HTTP-срез ──────────────────────────────────────────────────────────
    def test_patch_project_id_http(self):
        w = engine.create_writeoff(self.prj, self.user, 'СП-k7', date='2026-05-01')
        c = Client()
        c.force_login(self.user)
        resp = c.patch(f'/api/writeoffs/{w.id}/', {'project_id': self.prj2.id},
                       content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()['project_id'], self.prj2.id)
        # несуществующий проект → дружелюбный 400 (не 500)
        bad = c.patch(f'/api/writeoffs/{w.id}/', {'project_id': 99999},
                      content_type='application/json')
        self.assertEqual(bad.status_code, 400)

    def test_patch_kitting_target_http(self):
        dev = self.make_item('DEV-k5', manufactured=True)
        dev2 = self.make_item('DEV-k6', manufactured=True)
        k = models.Kitting.objects.create(project=self.prj, target_item=dev,
                                          user=self.user, qty=D(1))
        c = Client()
        c.force_login(self.user)
        resp = c.patch(f'/api/kittings/{k.id}/', {'target_id': dev2.id},
                       content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()['target_id'], dev2.id)


class RelocationHttpTests(EngineTestBase):
    """HTTP-слой перемещения и справочника мест: создание → пикер лотов/мест →
    добавление хода → правка → фиксация/расфиксация → удаление. Инвариант тот же,
    что у движка: тотал лота сохранён — перемещение двигает распределение по
    `(лот, локация)`, а не остаток (волна 13, Ф3)."""

    def setUp(self):
        super().setUp()
        self.user.is_superuser = True
        self.user.save()
        self.sold = models.Location.objects.create(code='105', description='Место пайки')
        self.item = self.make_item('R100')
        self.lot = self.receipt_lot(self.item, self.prj, 12)  # born @ MAIN
        self.c = Client()
        self.c.force_login(self.user)

    def _create(self):
        r = self.c.post('/api/relocations/',
                        {'project_id': self.prj.id, 'number': 'ПЕР-1'},
                        content_type='application/json')
        self.assertEqual(r.status_code, 201)
        return r.json()['id']

    def test_locations_endpoint_lists_places(self):
        rows = self.c.get('/api/locations/').json()
        codes = {row['code'] for row in rows}
        self.assertEqual(codes, {'MAIN', '105'})

    def test_source_lots_picker_shows_live_lot_with_breakdown(self):
        rid = self._create()
        rows = self.c.get(f'/api/relocations/{rid}/source-lots/').json()
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['lot_id'], self.lot.id)
        self.assertEqual(D(str(rows[0]['live_qty'])), D(12))
        # разбивка по местам: весь остаток на MAIN
        self.assertEqual(rows[0]['by_location'][0]['code'], 'MAIN')

    def test_add_move_splits_distribution_total_preserved(self):
        rid = self._create()
        resp = self.c.post(f'/api/relocations/{rid}/lines/',
                           {'lot_id': self.lot.id, 'qty': 5,
                            'from_location_id': self.main.id,
                            'to_location_id': self.sold.id},
                           content_type='application/json')
        self.assertEqual(resp.status_code, 201)
        ck = resp.json()
        self.assertEqual(len(ck['moves']), 1)
        self.assertEqual(self.c.post(f'/api/relocations/{rid}/lock/').status_code, 200)
        # Ф15: провели → 7 @ MAIN, 5 @ 105, тотал 12
        self.assertEqual(engine.lot_live_qty(self.lot, self.main), D(7))
        self.assertEqual(engine.lot_live_qty(self.lot, self.sold), D(5))
        self.assertEqual(engine.lot_live_qty(self.lot), D(12))

    def test_same_source_and_dest_rejected(self):
        rid = self._create()
        resp = self.c.post(f'/api/relocations/{rid}/lines/',
                           {'lot_id': self.lot.id, 'qty': 5,
                            'from_location_id': self.main.id,
                            'to_location_id': self.main.id},
                           content_type='application/json')
        self.assertEqual(resp.status_code, 400)

    def test_foreign_project_lot_rejected(self):
        other = models.Project.objects.create(code='P2', description='Проект 2',
            kind=models.Project.Kind.EXTERNAL)
        foreign = self.receipt_lot(self.item, other, 3)
        rid = self._create()
        resp = self.c.post(f'/api/relocations/{rid}/lines/',
                           {'lot_id': foreign.id, 'qty': 1,
                            'from_location_id': self.main.id,
                            'to_location_id': self.sold.id},
                           content_type='application/json')
        self.assertEqual(resp.status_code, 400)

    def test_update_and_delete_move_keyed_by_lot(self):
        rid = self._create()
        self.c.post(f'/api/relocations/{rid}/lines/',
                    {'lot_id': self.lot.id, 'qty': 5,
                     'from_location_id': self.main.id,
                     'to_location_id': self.sold.id},
                    content_type='application/json')
        # правка кол-ва хода (ключ хода — лот)
        upd = self.c.patch(f'/api/relocations/{rid}/lines/{self.lot.id}/',
                          {'qty': 8}, content_type='application/json')
        self.assertEqual(upd.status_code, 200)
        self.assertEqual(self.c.post(f'/api/relocations/{rid}/lock/').status_code, 200)
        self.assertEqual(engine.lot_live_qty(self.lot, self.sold), D(8))
        # удаление хода (после расфиксации) → распределение вернулось (всё на MAIN)
        self.assertEqual(self.c.post(f'/api/relocations/{rid}/unlock/').status_code, 200)
        rm = self.c.delete(f'/api/relocations/{rid}/lines/{self.lot.id}/')
        self.assertEqual(rm.status_code, 200)
        self.assertEqual(engine.lot_live_qty(self.lot, self.main), D(12))

    def test_post_unpost_delete_flow(self):
        rid = self._create()
        self.c.post(f'/api/relocations/{rid}/lines/',
                    {'lot_id': self.lot.id, 'qty': 5,
                     'from_location_id': self.main.id,
                     'to_location_id': self.sold.id},
                    content_type='application/json')
        posted = self.c.post(f'/api/relocations/{rid}/lock/')
        self.assertEqual(posted.status_code, 200)
        self.assertTrue(posted.json()['locked'])
        # posted — удаление отклонено (сперва расфиксировать)
        self.assertEqual(self.c.delete(f'/api/relocations/{rid}/').status_code, 400)
        # добавление хода под замком отклонено
        blocked = self.c.post(f'/api/relocations/{rid}/lines/',
                              {'lot_id': self.lot.id, 'qty': 1,
                               'from_location_id': self.main.id,
                               'to_location_id': self.sold.id},
                              content_type='application/json')
        self.assertEqual(blocked.status_code, 400)
        # расфиксировать → удалить; тотал лота цел
        self.assertEqual(self.c.post(f'/api/relocations/{rid}/unlock/').status_code, 200)
        self.assertEqual(self.c.delete(f'/api/relocations/{rid}/').status_code, 204)
        self.assertFalse(models.Relocation.objects.filter(pk=rid).exists())
        self.assertEqual(engine.lot_live_qty(self.lot), D(12))
        self.assertEqual(engine.lot_live_qty(self.lot, self.main), D(12))

    def test_empty_relocation_cannot_be_posted(self):
        rid = self._create()
        resp = self.c.post(f'/api/relocations/{rid}/lock/')
        self.assertEqual(resp.status_code, 400)

    def test_patch_header_number_and_project_anchor(self):
        rid = self._create()
        # № правится свободно
        upd = self.c.patch(f'/api/relocations/{rid}/', {'number': 'ПЕР-9'},
                          content_type='application/json')
        self.assertEqual(upd.status_code, 200)
        self.assertEqual(upd.json()['number'], 'ПЕР-9')
        # проект-якорь: у пустого ордера сменить можно
        other = models.Project.objects.create(code='P3', description='Проект 3',
            kind=models.Project.Kind.EXTERNAL)
        moved = self.c.patch(f'/api/relocations/{rid}/', {'project_id': other.id},
                            content_type='application/json')
        self.assertEqual(moved.status_code, 200)
        self.assertEqual(moved.json()['project_id'], other.id)


class LocationEntityTests(EngineTestBase):
    """Место хранения — полноценная сущность «Склады»: что на нём лежит и его ДНК
    (код/описание/вид), включая guard на дубль кода (волна 13, Ф4)."""

    def setUp(self):
        super().setUp()
        self.user.is_superuser = True
        self.user.save()
        self.sold = models.Location.objects.create(code='105', description='Место пайки')
        self.item = self.make_item('R100')
        self.lot = self.receipt_lot(self.item, self.prj, 12)  # born @ MAIN
        self.c = Client()
        self.c.force_login(self.user)

    def test_location_stock_lists_live_lots_with_project(self):
        rows = engine.location_stock(self.main)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['lot_id'], self.lot.id)
        self.assertEqual(rows[0]['qty'], D(12))
        self.assertEqual(rows[0]['project_code'], self.prj.code)
        # второе место пусто
        self.assertEqual(engine.location_stock(self.sold), [])

    def test_location_stock_reflects_relocation_split(self):
        rel = engine.create_relocation(self.prj, self.user, 'ПЕР-1')
        engine.add_relocation_line(rel, self.lot, D(5), self.main, self.sold)
        engine.lock_relocation(rel)                 # Ф15: расщепление — с фиксации
        main_rows = engine.location_stock(self.main)
        sold_rows = engine.location_stock(self.sold)
        self.assertEqual(main_rows[0]['qty'], D(7))
        self.assertEqual(sold_rows[0]['qty'], D(5))

    def test_create_location_and_duplicate_code_rejected(self):
        loc = engine.create_location('201', 'Архив', kind='хранилище')
        self.assertEqual(loc.kind, 'хранилище')
        with self.assertRaises(ValidationError):
            engine.create_location('201', 'Дубль')

    def test_update_location_dna_and_duplicate_guard(self):
        engine.update_location(self.sold, description='Пайка-2', kind='цех')
        self.sold.refresh_from_db()
        self.assertEqual(self.sold.description, 'Пайка-2')
        self.assertEqual(self.sold.kind, 'цех')
        # код на занятый — дружелюбный отказ
        with self.assertRaises(ValidationError):
            engine.update_location(self.sold, code='MAIN')

    def test_http_location_form_and_patch(self):
        resp = self.c.get(f'/api/locations/{self.main.id}/')
        self.assertEqual(resp.status_code, 200)
        body = resp.json()
        self.assertEqual(body['code'], 'MAIN')
        self.assertEqual(len(body['stock']), 1)
        # PATCH вида (свободный текст)
        patched = self.c.patch(f'/api/locations/{self.main.id}/',
                              {'kind': 'основной'}, content_type='application/json')
        self.assertEqual(patched.status_code, 200)
        self.assertEqual(patched.json()['kind'], 'основной')

    def test_http_create_location(self):
        resp = self.c.post('/api/locations/',
                          {'code': '301', 'description': 'Резерв'},
                          content_type='application/json')
        self.assertEqual(resp.status_code, 201)
        self.assertTrue(models.Location.objects.filter(code='301').exists())
        # дубль кода → 400
        dup = self.c.post('/api/locations/',
                        {'code': '301', 'description': 'Дубль'},
                        content_type='application/json')
        self.assertEqual(dup.status_code, 400)


class EntityDeleteTests(EngineTestBase):
    """WAVE14 Ф2: консистентное удаление справочных сущностей из UI (Изделие/Склад/
    Заказ/Закупка/Проект) — единый friendly-guard движка, как у ордеров."""

    # ── Изделие ──
    def test_delete_item_free_when_unlinked(self):
        it = self.make_item('FREE')
        engine.delete_item(it)
        self.assertFalse(models.Item.objects.filter(pk=it.pk).exists())

    def test_delete_item_blocked_by_lot(self):
        it = self.make_item('WITHLOT')
        self.receipt_lot(it, self.prj, 5)
        with self.assertRaises(ValidationError):
            engine.delete_item(it)
        self.assertTrue(models.Item.objects.filter(pk=it.pk).exists())

    def test_delete_item_blocked_when_used_in_bom(self):
        parent = self.make_item('P', manufactured=True)
        comp = self.make_item('C')
        engine.add_bom_line(parent, comp, D(2))
        with self.assertRaises(ValidationError):
            engine.delete_item(comp)                 # входит в чужой BOM
        # родителя (свой BOM — каскад) сносим свободно
        parent_id = parent.pk
        engine.delete_item(parent)
        self.assertFalse(models.Item.objects.filter(pk=parent_id).exists())
        self.assertFalse(models.BomLine.objects.filter(parent_id=parent_id).exists())

    def test_delete_item_blocked_by_demand(self):
        dev = self.make_item('DEV', manufactured=True)
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(1))
        with self.assertRaises(ValidationError):
            engine.delete_item(dev)

    # ── Склад ──
    def test_delete_location_free_when_empty(self):
        loc = models.Location.objects.create(code='EMPTY', description='Пустой склад')
        engine.delete_location(loc)
        self.assertFalse(models.Location.objects.filter(pk=loc.pk).exists())

    def test_delete_location_blocked_by_movements(self):
        self.receipt_lot(self.make_item('X'), self.prj, 3)   # рождает движение на MAIN
        with self.assertRaises(ValidationError):
            engine.delete_location(self.main)
        self.assertTrue(models.Location.objects.filter(pk=self.main.pk).exists())

    # ── Заказ ──
    def test_delete_purchase_draft_cascades_lines(self):
        p = engine.create_purchase(self.prj, self.user)
        engine.add_purchase_line(p, self.make_item('A'), D(4))
        pid = p.pk
        engine.delete_purchase(p)
        self.assertFalse(models.Purchase.objects.filter(pk=pid).exists())
        self.assertFalse(models.PurchaseLine.objects.filter(purchase_id=pid).exists())

    def test_delete_purchase_blocked_when_sent(self):
        p = self.make_purchase()
        engine.add_purchase_line(p, self.make_item('A'), D(4))
        engine.lock_purchase(p)
        with self.assertRaises(ValidationError):
            engine.delete_purchase(p)                # отправлен — сперва в черновик

    def test_delete_purchase_blocked_by_receipt(self):
        p = engine.create_purchase(self.prj, self.user)
        self.receipt_lot(self.make_item('A'), self.prj, 5, purchase=p)
        with self.assertRaises(ValidationError):
            engine.delete_purchase(p)                # привязан приход

    # ── Закупка (план) ──
    def test_delete_procurement_draft_free(self):
        proc = engine.create_procurement(self.user)
        engine.add_procurement_line(proc, self.make_item('A'), D(3))
        pid = proc.pk
        engine.delete_procurement(proc)
        self.assertFalse(models.Procurement.objects.filter(pk=pid).exists())
        self.assertFalse(models.ProcurementLine.objects.filter(procurement_id=pid).exists())

    def test_delete_procurement_blocked_by_purchase(self):
        proc = engine.create_procurement(self.user)
        p = engine.create_purchase(self.prj, self.user)
        engine.update_purchase(p, procurement=proc)      # Ф17: план выбирают, а не плодят
        with self.assertRaises(ValidationError):
            engine.delete_procurement(proc)          # привязан заказ

    # ── Проект ──
    def test_delete_project_free_when_empty(self):
        prj = models.Project.objects.create(
            code='EMPTY-PRJ', description='Пустой', kind=models.Project.Kind.EXTERNAL)
        engine.delete_project(prj)
        self.assertFalse(models.Project.objects.filter(pk=prj.pk).exists())

    def test_delete_project_blocked_by_lot(self):
        self.receipt_lot(self.make_item('X'), self.prj, 3)
        with self.assertRaises(ValidationError):
            engine.delete_project(self.prj)

    def test_delete_project_blocked_by_demand(self):
        dev = self.make_item('DEV', manufactured=True)
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(1))
        with self.assertRaises(ValidationError):
            engine.delete_project(self.prj)

    def test_delete_project_internal_forbidden(self):
        stock = models.Project.objects.create(
            code='WHITE', description='Собственный склад',
            kind=models.Project.Kind.INTERNAL_STOCK)
        with self.assertRaises(ValidationError):
            engine.delete_project(stock)


class EntityDeleteHttpTests(TestCase):
    """WAVE14 Ф2: HTTP-путь DELETE справочных сущностей (204 успех / 400 friendly-guard)."""

    def setUp(self):
        self.user = get_user_model().objects.create(username='admin', is_superuser=True)
        self.main = models.Location.objects.create(code='MAIN', description='Основной склад')
        self.prj = models.Project.objects.create(
            code='P1', description='Проект 1', kind=models.Project.Kind.EXTERNAL)
        self.sup = models.Counterparty.objects.create(description='П')
        self.c = Client()
        self.c.force_login(self.user)

    def test_item_delete_204_and_guard_400(self):
        it = models.Item.objects.create(code='FREE', description='FREE', category=_cat())
        self.assertEqual(self.c.delete(f'/api/items/{it.id}/').status_code, 204)
        self.assertFalse(models.Item.objects.filter(pk=it.pk).exists())
        # с лотом → 400
        it2 = models.Item.objects.create(code='WL', description='WL', category=_cat())
        r = models.Receipt.objects.create(number='U-1', date='2026-05-01',
            contractor=self.sup, project=self.prj, user=self.user)
        lot = models.Lot.objects.create(item=it2, project=self.prj, origin=r, qty=D(1))
        engine.rebuild_movements(lot)
        self.assertEqual(self.c.delete(f'/api/items/{it2.id}/').status_code, 400)

    def test_location_delete_204_and_guard_400(self):
        loc = models.Location.objects.create(code='EMPTY', description='Пустой')
        self.assertEqual(self.c.delete(f'/api/locations/{loc.id}/').status_code, 204)
        r = models.Receipt.objects.create(number='U-2', date='2026-05-01', locked=True,
            contractor=self.sup, project=self.prj, user=self.user)   # Ф15: движение есть
        lot = models.Lot.objects.create(
            item=models.Item.objects.create(code='M', description='M', category=_cat()),
            project=self.prj, origin=r, qty=D(1))
        engine.rebuild_movements(lot)               # движение на MAIN
        self.assertEqual(self.c.delete(f'/api/locations/{self.main.id}/').status_code, 400)

    def test_project_delete_204_and_guard_400(self):
        empty = models.Project.objects.create(
            code='E', description='E', kind=models.Project.Kind.EXTERNAL)
        self.assertEqual(self.c.delete(f'/api/projects/{empty.id}/').status_code, 204)
        r = models.Receipt.objects.create(number='U-3', date='2026-05-01',
            contractor=self.sup, project=self.prj, user=self.user)
        lot = models.Lot.objects.create(
            item=models.Item.objects.create(code='Q', description='Q', category=_cat()),
            project=self.prj, origin=r, qty=D(1))
        engine.rebuild_movements(lot)
        self.assertEqual(self.c.delete(f'/api/projects/{self.prj.id}/').status_code, 400)

    def test_purchase_delete_204(self):
        p = engine.create_purchase(self.prj, self.user)
        self.assertEqual(self.c.delete(f'/api/purchases/{p.id}/').status_code, 204)
        self.assertFalse(models.Purchase.objects.filter(pk=p.pk).exists())

    def test_procurement_delete_204(self):
        proc = engine.create_procurement(self.user)
        self.assertEqual(self.c.delete(f'/api/procurements/{proc.id}/').status_code, 204)
        self.assertFalse(models.Procurement.objects.filter(pk=proc.pk).exists())


# Полная 8-колоночная схема библиотеки (парсер ищет нужные колонки по имени, но
# тесты кормят реальный заголовок — терпимость к лишним колонкам заодно проверена).
_LIB_HEADER = ['Design Item Id', 'Comment', 'Description', 'Footprint Path',
               'Footprint Ref', 'Library Path', 'Library Ref', 'Temperature']


def _lib_csv(items):
    """CP1251-байты CSV библиотеки из списка `(code, description, temperature)`.
    Разделитель `;`, финальный LF (как в реальных выгрузках Altium)."""
    lines = [';'.join(_LIB_HEADER)]
    for did, desc, temp in items:
        lines.append(';'.join([did, '', desc, '', '', '', '', temp]))
    return ('\n'.join(lines) + '\n').encode('cp1251')


class LibraryParseTests(TestCase):
    """Волна 15 Ф1: парсер CSV библиотеки (CP1251, `;`, мульти-файл)."""

    def test_parse_basic(self):
        raw = _lib_csv([('CAP-1', 'Конденсатор 1', '-55-125°C'),
                        ('CAP-2', 'Конденсатор 2', '-40-85°C')])
        cat, rows = engine.parse_library_file('csv/capacitors.csv', raw)
        self.assertEqual(cat, 'capacitors')
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0], {'code': 'CAP-1', 'description': 'Конденсатор 1',
                                   'temperature': '-55-125°C', 'category': 'capacitors'})

    def test_trailing_blank_line_skipped(self):
        raw = _lib_csv([('X', 'Икс', '')]) + b'\n'      # лишний финальный перевод
        _, rows = engine.parse_library_file('sensors.csv', raw)
        self.assertEqual(len(rows), 1)

    def test_missing_column(self):
        bad = 'Design Item Id;Comment\nX;y\n'.encode('cp1251')
        with self.assertRaises(ValidationError):
            engine.parse_library_file('mcu.csv', bad)

    def test_empty_key(self):
        raw = _lib_csv([('', 'без ключа', '')])
        with self.assertRaises(ValidationError):
            engine.parse_library_file('mcu.csv', raw)

    def test_duplicate_key_in_file(self):
        raw = _lib_csv([('D', 'раз', ''), ('D', 'два', '')])
        with self.assertRaises(ValidationError):
            engine.parse_library_file('mcu.csv', raw)

    def test_cp1251_decode(self):
        # Кириллица в CP1251 читается верно (в отличие от наивного UTF-8-декода).
        raw = _lib_csv([('X', 'Микросхема ЦАП', '-40-125°C')])
        _, rows = engine.parse_library_file('mcu.csv', raw)
        self.assertEqual(rows[0]['description'], 'Микросхема ЦАП')
        self.assertEqual(rows[0]['temperature'], '-40-125°C')

    def test_multi_file_aggregate(self):
        files = [('capacitors.csv', _lib_csv([('C1', 'к', '')])),
                 ('mcu.csv', _lib_csv([('M1', 'м', '')]))]
        parsed = engine.parse_library(files)
        self.assertEqual(set(parsed['categories']), {'capacitors', 'mcu'})
        self.assertEqual(len(parsed['rows']), 2)

    def test_duplicate_category(self):
        files = [('capacitors.csv', _lib_csv([('C1', 'к', '')])),
                 ('capacitors.csv', _lib_csv([('C2', 'к', '')]))]
        with self.assertRaises(ValidationError):
            engine.parse_library(files)

    def test_duplicate_key_across_files(self):
        files = [('capacitors.csv', _lib_csv([('DUP', 'к', '')])),
                 ('mcu.csv', _lib_csv([('DUP', 'м', '')]))]
        with self.assertRaises(ValidationError):
            engine.parse_library(files)

    def test_empty_upload(self):
        with self.assertRaises(ValidationError):
            engine.parse_library([])


class LibraryDiffTests(TestCase):
    """Волна 15 Ф2: диф загруженной библиотеки против справочника."""

    def _cat(self, code):
        return engine.ensure_category(code)

    def _item(self, did, cat, desc='desc', temp='', native=False):
        return models.Item.objects.create(
            code=did, description=desc, category=self._cat(cat),
            temperature=temp, native=native)

    def _diff(self, files):
        return engine.library_diff(engine.parse_library(files))

    def _by_key(self, rows):
        return {r['code']: r for r in rows}

    def test_new_changed_same(self):
        self._item('CAP-1', 'capacitors', desc='старое', temp='-40-85°C')
        c2 = self._item('CAP-2', 'capacitors', desc='совпадает', temp='-55-125°C')
        c2.synced = True; c2.save()          # совпадает И уже помечено библиотечным → same
        files = [('capacitors.csv', _lib_csv([
            ('CAP-1', 'новое', '-40-85°C'),     # description изменился
            ('CAP-2', 'совпадает', '-55-125°C'), # без изменений
            ('CAP-3', 'новый', '')])),           # нет в БД
        ]
        rows = self._by_key(self._diff(files))
        self.assertEqual(rows['CAP-1']['status'], 'changed')
        self.assertIn('description', rows['CAP-1']['changes'])
        self.assertEqual(rows['CAP-2']['status'], 'same')
        self.assertEqual(rows['CAP-3']['status'], 'new')

    def test_mark_when_unsynced_matches_library(self):
        # Ф3a: содержимое совпадает, но изделие ещё не помечено библиотечным → mark.
        self._item('CAP-9', 'capacitors', desc='совп', temp='')   # synced=False по умолчанию
        rows = self._by_key(self._diff([('capacitors.csv', _lib_csv([('CAP-9', 'совп', '')]))]))
        self.assertEqual(rows['CAP-9']['status'], 'mark')
        self.assertFalse(rows['CAP-9']['current']['synced'])

    def test_gone_when_unused(self):
        self._item('CAP-OLD', 'capacitors')      # в БД, не в загрузке, не используется
        rows = self._by_key(self._diff([('capacitors.csv', _lib_csv([('CAP-1', 'к', '')]))]))
        self.assertEqual(rows['CAP-OLD']['status'], 'gone')

    def test_orphan_when_used(self):
        used = self._item('CAP-USED', 'capacitors')
        parent = self._item('DEV', 'capacitors', native=True)
        models.BomLine.objects.create(parent=parent, component=used, qty=D(1))
        # DEV тоже в капаситорах и не в загрузке → сам gone/orphan; CAP-USED — orphan.
        rows = self._by_key(self._diff([('capacitors.csv', _lib_csv([('CAP-1', 'к', '')]))]))
        self.assertEqual(rows['CAP-USED']['status'], 'orphan')

    def test_missing_scoped_by_category(self):
        # Изделие класса mcu не считается пропавшим, если грузили только capacitors.
        self._item('MCU-X', 'mcu')
        rows = self._by_key(self._diff([('capacitors.csv', _lib_csv([('CAP-1', 'к', '')]))]))
        self.assertNotIn('MCU-X', rows)

    def test_category_change_detected(self):
        self._item('MULTI', 'capacitors')
        rows = self._by_key(self._diff([('mcu.csv', _lib_csv([('MULTI', 'desc', '')]))]))
        self.assertEqual(rows['MULTI']['status'], 'changed')
        self.assertIn('category', rows['MULTI']['changes'])


class LibraryApplyTests(TestCase):
    """Волна 15 Ф2: применение подтверждённых строк дифа."""

    def _item(self, did, cat, desc='desc', temp='', native=False):
        return models.Item.objects.create(
            code=did, description=desc, category=engine.ensure_category(cat),
            temperature=temp, native=native)

    def test_apply_creates_new(self):
        files = [('sensors.csv', _lib_csv([('S-1', 'Датчик', '-40-85°C')]))]
        parsed = engine.parse_library(files)
        summary = engine.apply_library_diff(parsed, ['S-1'])
        self.assertEqual(summary['created'], 1)
        item = models.Item.objects.get(code='S-1')
        self.assertFalse(item.native)                  # импорт → покупное
        self.assertTrue(item.synced)                     # рождается библиотечным
        self.assertFalse(item.locked)                    # но НЕ запертым — готово под цену
        self.assertIsNone(item.estimated_cost)           # цена — за Plume
        self.assertEqual(item.category.code, 'sensors')  # категория заведена на лету
        self.assertEqual(item.category.description, 'Датчики') # канон LIBRARY_CATEGORIES

    def test_apply_updates_changed(self):
        self._item('S-1', 'sensors', desc='старое', temp='')
        parsed = engine.parse_library([('sensors.csv', _lib_csv([('S-1', 'новое', '-40-85°C')]))])
        engine.apply_library_diff(parsed, ['S-1'])
        item = models.Item.objects.get(code='S-1')
        self.assertEqual(item.description, 'новое')
        self.assertEqual(item.temperature, '-40-85°C')
        self.assertTrue(item.synced)                     # синк подтвердил библиотечность

    def test_apply_deletes_gone(self):
        self._item('S-OLD', 'sensors')
        parsed = engine.parse_library([('sensors.csv', _lib_csv([('S-1', 'к', '')]))])
        summary = engine.apply_library_diff(parsed, ['S-OLD'])
        self.assertEqual(summary['deleted'], 1)
        self.assertFalse(models.Item.objects.filter(code='S-OLD').exists())

    def test_apply_only_confirmed(self):
        self._item('S-OLD', 'sensors')                   # gone, НЕ подтверждаем
        parsed = engine.parse_library([('sensors.csv', _lib_csv([
            ('S-1', 'новый A', ''), ('S-2', 'новый B', '')]))])
        summary = engine.apply_library_diff(parsed, ['S-1'])   # только S-1
        self.assertEqual(summary['created'], 1)
        self.assertTrue(models.Item.objects.filter(code='S-1').exists())
        self.assertFalse(models.Item.objects.filter(code='S-2').exists())
        self.assertTrue(models.Item.objects.filter(code='S-OLD').exists())

    def test_apply_mark_sets_synced(self):
        # Ф3a: совпадающее по содержимому непомеченное изделие → mark → synced=True,
        # locked не трогаем (решение Ивана 2026-07-24).
        it = self._item('S-1', 'sensors', desc='датчик', temp='')   # synced=False, совпадает
        parsed = engine.parse_library([('sensors.csv', _lib_csv([('S-1', 'датчик', '')]))])
        summary = engine.apply_library_diff(parsed, ['S-1'])
        self.assertEqual(summary['marked'], 1)
        self.assertEqual(summary['updated'], 0)   # содержимое не трогали
        it.refresh_from_db()
        self.assertTrue(it.synced)
        self.assertFalse(it.locked)               # замок не тронут

    def test_apply_mark_only_confirmed(self):
        # Не подтверждённый mark — не помечается.
        it = self._item('S-1', 'sensors', desc='датчик', temp='')
        parsed = engine.parse_library([('sensors.csv', _lib_csv([('S-1', 'датчик', '')]))])
        summary = engine.apply_library_diff(parsed, [])   # ничего не подтверждаем
        self.assertEqual(summary['marked'], 0)
        it.refresh_from_db()
        self.assertFalse(it.synced)

    def test_orphan_not_deleted_even_if_confirmed(self):
        used = self._item('S-USED', 'sensors')
        parent = self._item('DEV', 'sensors', native=True)
        models.BomLine.objects.create(parent=parent, component=used, qty=D(1))
        parsed = engine.parse_library([('sensors.csv', _lib_csv([('S-1', 'к', '')]))])
        summary = engine.apply_library_diff(parsed, ['S-USED'])   # подтверждаем сироту
        self.assertEqual(summary['deleted'], 0)          # orphan — не действие
        self.assertTrue(models.Item.objects.filter(code='S-USED').exists())


class ItemStatusTests(EngineTestBase):
    """Волна 17, фаза 1: статус изделия `draft ⇄ posted` (фиксация), гейт мутаций,
    синк библиотеки → posted, проброс статуса в сериализацию."""


    def test_manual_item_defaults_draft(self):
        i = engine.create_item('R100', 'Резистор', category_id=_cat().id)
        self.assertFalse(i.locked)
        self.assertFalse(i.synced)                # ручное — не из библиотеки

    def test_post_unpost_idempotent(self):
        i = self.make_item('R')
        engine.lock_item(i)
        i.refresh_from_db()
        self.assertTrue(i.locked)
        engine.lock_item(i)                       # повторно — без падения
        i.refresh_from_db()
        self.assertTrue(i.locked)
        engine.unlock_item(i)
        i.refresh_from_db()
        self.assertFalse(i.locked)
        engine.unlock_item(i)                     # повторно — без падения
        i.refresh_from_db()
        self.assertFalse(i.locked)

    def test_gate_blocks_property_edit_when_posted(self):
        i = self.make_item('R')
        engine.lock_item(i)
        with self.assertRaises(ValidationError):
            engine.update_item(i, {'description': 'нельзя'})
        # расфиксировали — правка снова проходит
        engine.unlock_item(i)
        engine.update_item(i, {'description': 'можно'})
        i.refresh_from_db()
        self.assertEqual(i.description, 'можно')

    def test_gate_blocks_bom_edits_when_parent_posted(self):
        dev = self.make_item('DEV', manufactured=True)
        comp = self.make_item('R')
        line = engine.add_bom_line(dev, comp, D(2))
        engine.lock_item(dev)
        with self.assertRaises(ValidationError):
            engine.add_bom_line(dev, self.make_item('C'), D(1))
        with self.assertRaises(ValidationError):
            engine.update_bom_line(line, qty=D(5))
        with self.assertRaises(ValidationError):
            engine.remove_bom_line(line)
        line.refresh_from_db()
        self.assertEqual(line.qty, D(2))          # не изменилось

    def test_gate_blocks_delete_when_posted(self):
        i = self.make_item('R')
        engine.lock_item(i)
        with self.assertRaises(ValidationError):
            engine.delete_item(i)
        self.assertTrue(models.Item.objects.filter(pk=i.pk).exists())
        engine.unlock_item(i)
        engine.delete_item(i)
        self.assertFalse(models.Item.objects.filter(pk=i.pk).exists())

    def test_library_new_is_synced_unlocked(self):
        # Ф3a: новое библиотечное — synced=True, но locked=False (готово под цену).
        parsed = engine.parse_library([('sensors.csv', _lib_csv([('S-1', 'Датчик', '')]))])
        engine.apply_library_diff(parsed, ['S-1'])
        item = models.Item.objects.get(code='S-1')
        self.assertTrue(item.synced)
        self.assertFalse(item.locked)

    def test_library_changed_marks_synced_clears_stale_lock(self):
        # заведено руками и залочено (наследие волны 17), затем совпало с библиотекой →
        # синк метит synced И снимает стухший замок (инвариант synced ⟹ not locked)
        models.Item.objects.create(code='S-1', description='старое',
                                   category=engine.ensure_category('sensors'), locked=True)
        parsed = engine.parse_library([('sensors.csv', _lib_csv([('S-1', 'новое', '-40-85°C')]))])
        engine.apply_library_diff(parsed, ['S-1'])
        item = models.Item.objects.get(code='S-1')
        self.assertTrue(item.synced)
        self.assertFalse(item.locked)             # стухший замок снят
        self.assertEqual(item.description, 'новое')

    def test_library_mark_clears_stale_lock(self):
        # содержимое совпадает, но изделие залочено со времён волны 17 → mark снимает
        # замок вместе с пометкой synced (инвариант synced ⟹ not locked)
        models.Item.objects.create(code='S-1', description='датчик',
                                   category=engine.ensure_category('sensors'), locked=True)
        parsed = engine.parse_library([('sensors.csv', _lib_csv([('S-1', 'датчик', '')]))])
        summary = engine.apply_library_diff(parsed, ['S-1'])
        self.assertEqual(summary['marked'], 1)
        item = models.Item.objects.get(code='S-1')
        self.assertTrue(item.synced)
        self.assertFalse(item.locked)

    def test_library_deletes_gone_locked(self):
        # ушедшее из библиотеки изделие — зафиксировано; синк расфиксирует и удалит (гейт не мешает)
        gone = models.Item.objects.create(code='S-OLD', description='ушло',
                                          category=engine.ensure_category('sensors'),
                                          locked=True)
        parsed = engine.parse_library([('sensors.csv', _lib_csv([('S-1', 'к', '')]))])
        summary = engine.apply_library_diff(parsed, ['S-OLD'])
        self.assertEqual(summary['deleted'], 1)
        self.assertFalse(models.Item.objects.filter(pk=gone.pk).exists())

    def test_status_in_item_serialization(self):
        i = self.make_item('R')
        engine.lock_item(i)
        i.refresh_from_db()
        self.assertTrue(views._item_row(i)['locked'])
        self.assertTrue(views._item_detail_payload(i)['locked'])

    def test_component_status_in_bom_and_demand(self):
        dev = self.make_item('DEV', manufactured=True)
        comp = self.make_item('R')                # покупной лист
        engine.lock_item(comp)
        engine.add_bom_line(dev, comp, D(2))
        # BOM в форме изделия несёт статус компонента
        bom = views._item_detail_payload(dev)['bom']
        self.assertTrue(bom[0]['component_locked'])
        # Потребность проекта: свод по листьям + дерево несут статус компонента
        engine.add_project_demand(self.prj, dev, D(1))
        deficit = engine.project_deficit(self.prj)
        leaf = next(c for c in deficit['components'] if c['component_id'] == comp.id)
        self.assertTrue(leaf['component_locked'])
        tree_leaf = next(n for n in deficit['demands'][0]['tree'] if n['component_id'] == comp.id)
        self.assertTrue(tree_leaf['component_locked'])

    def test_http_post_unpost_endpoints(self):
        get_user_model().objects.filter(username='t').update(
            is_staff=True, is_superuser=True)
        c = Client()
        c.force_login(self.user)
        i = self.make_item('R')
        r = c.post(f'/api/items/{i.id}/lock/')
        self.assertEqual(r.status_code, 200)
        self.assertTrue(r.json()['locked'])
        i.refresh_from_db()
        self.assertTrue(i.locked)
        r = c.post(f'/api/items/{i.id}/unlock/')
        self.assertEqual(r.status_code, 200)
        self.assertFalse(r.json()['locked'])


class ItemAxesTests(EngineTestBase):
    """Ф3a (волна 19): три оси native/synced/locked — инвариант + матрица правки."""

    def test_synced_implies_not_native_constraint(self):
        # CheckConstraint: библиотечное (`synced`) не может быть нашим (`native`).
        from django.db import IntegrityError, transaction
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                models.Item.objects.create(
                    code='BAD', description='x', category=_cat(),
                    synced=True, native=True)

    def test_synced_implies_not_locked_constraint(self):
        # CheckConstraint: библиотечное (`synced`) не запирается (`locked`) — две оси
        # защиты взаимоисключающи.
        from django.db import IntegrityError, transaction
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                models.Item.objects.create(
                    code='BAD2', description='x', category=_cat(),
                    synced=True, locked=True)

    def test_synced_item_edits_price_only(self):
        # Библиотечное расфиксированное: правится ТОЛЬКО оценочная стоимость.
        i = self.make_item('R')
        i.synced = True; i.save()
        engine.update_item(i, {'estimated_cost': D('1.50')})   # цена — можно
        i.refresh_from_db()
        self.assertEqual(i.estimated_cost, D('1.50'))
        with self.assertRaises(ValidationError):               # прочее — нельзя
            engine.update_item(i, {'description': 'ручная правка'})

    def test_manual_item_edits_all(self):
        # Ручное (synced=False) расфиксированное: правится всё.
        i = self.make_item('R')
        engine.update_item(i, {'description': 'новое имя', 'estimated_cost': D('2')})
        i.refresh_from_db()
        self.assertEqual(i.description, 'новое имя')


class RollupCostTests(TestCase):
    """Волна 15 Ф4: рекурсивный роллап оценочной стоимости по BOM."""

    def _item(self, did, native=False, cost=None):
        return models.Item.objects.create(
            code=did, description=did, category=_cat(),
            native=native, estimated_cost=(D(cost) if cost is not None else None))

    def _bom(self, parent, component, qty):
        models.BomLine.objects.create(parent=parent, component=component, qty=D(qty))

    def test_recursive_rollup_writes_all_produced(self):
        l1 = self._item('L1', cost=10)
        l2 = self._item('L2', cost=5)
        board = self._item('BOARD', native=True)
        dev = self._item('DEV', native=True)
        self._bom(board, l1, 2)         # 2×10
        self._bom(board, l2, 3)         # 3×5 → плата 35
        self._bom(dev, board, 1)        # 1×35
        self._bom(dev, l1, 1)           # 1×10 → прибор 45
        res = engine.rollup_estimated_cost(dev)
        self.assertEqual(res['estimated_cost'], D(45))
        board.refresh_from_db(); dev.refresh_from_db()
        self.assertEqual(board.estimated_cost, D(35))   # промежуточный узел тоже переоценён
        self.assertEqual(dev.estimated_cost, D(45))
        self.assertEqual(set(res['updated']), {'BOARD', 'DEV'})
        self.assertEqual(res['incomplete'], [])

    def test_incomplete_flags_unknown_leaf(self):
        leaf = self._item('L-NOCOST')                   # покупной без цены
        dev = self._item('DEV', native=True)
        self._bom(dev, leaf, 2)
        res = engine.rollup_estimated_cost(dev)
        self.assertEqual(res['estimated_cost'], D(0))   # неизвестное считаем 0
        self.assertIn('L-NOCOST', res['incomplete'])

    def test_produced_without_bom_is_incomplete(self):
        dev = self._item('DEV', native=True)
        res = engine.rollup_estimated_cost(dev)
        self.assertIn('DEV', res['incomplete'])
        self.assertEqual(res['estimated_cost'], D(0))

    def test_non_produced_rejected(self):
        leaf = self._item('L', cost=10)
        with self.assertRaises(ValidationError):
            engine.rollup_estimated_cost(leaf)

    def test_cycle_guarded(self):
        a = self._item('A', native=True)
        b = self._item('B', native=True)
        self._bom(a, b, 1)
        models.BomLine.objects.create(parent=b, component=a, qty=D(1))  # цикл в обход guard'а
        with self.assertRaises(ValidationError):
            engine.rollup_estimated_cost(a)


class LibrarySyncHttpTests(TestCase):
    """Волна 15 Ф6: HTTP-путь синка — диф (без записи) → применение → пересчёт цены."""

    def setUp(self):
        get_user_model().objects.create(username='admin', is_superuser=True)
        self.c = Client()
        self.c.force_login(get_user_model().objects.get(is_superuser=True))

    def _upload(self, name, items):
        return SimpleUploadedFile(name, _lib_csv(items), content_type='text/csv')

    def test_diff_then_apply(self):
        models.Item.objects.create(code='CAP-OLD', description='старьё',
                                   category=engine.ensure_category('capacitors'))
        up = self._upload('capacitors.csv', [('CAP-1', 'Конденсатор', '-55-125°C')])
        r = self.c.post('/api/library/diff/', {'files': up})
        self.assertEqual(r.status_code, 200)
        body = r.json()
        self.assertEqual(body['categories'], ['capacitors'])
        by = {row['code']: row for row in body['rows']}
        self.assertEqual(by['CAP-1']['status'], 'new')
        self.assertEqual(by['CAP-OLD']['status'], 'gone')
        # применяем: заводим новый, старый удаляем
        up2 = self._upload('capacitors.csv', [('CAP-1', 'Конденсатор', '-55-125°C')])
        a = self.c.post('/api/library/apply/',
                        {'files': up2, 'confirmed': json.dumps(['CAP-1', 'CAP-OLD'])})
        self.assertEqual(a.status_code, 200)
        self.assertEqual(a.json(), {'created': 1, 'updated': 0, 'marked': 0, 'deleted': 1})
        self.assertTrue(models.Item.objects.filter(code='CAP-1').exists())
        self.assertFalse(models.Item.objects.filter(code='CAP-OLD').exists())

    def test_diff_creates_missing_categories(self):
        """Волна 22: сверка заводит недостающие КЛАССЫ (изделия не трогает) — иначе
        таб «Категории» формы синхронизации нечего было бы показывать до применения."""
        self.assertFalse(models.Category.objects.filter(code='inductors').exists())
        up = self._upload('inductors.csv', [('L-4R7', 'Дроссель 4.7 мкГн', '')])
        r = self.c.post('/api/library/diff/', {'files': up})
        self.assertEqual(r.status_code, 200)
        cat = models.Category.objects.get(code='inductors')
        # Класса нет в `LIBRARY_CATEGORIES` — всплывает с сырым описанием, юзер правит.
        self.assertEqual(cat.description, 'inductors')
        # Изделия при этом НЕ заведены: диф остаётся дифом.
        self.assertFalse(models.Item.objects.filter(code='L-4R7').exists())

    def test_diff_keeps_edited_category_description(self):
        """Повторная сверка не затирает описание, которое человек уже правил."""
        cat = engine.ensure_category('inductors')
        cat.description = 'Индуктивности'
        cat.save(update_fields=['description'])
        up = self._upload('inductors.csv', [('L-4R7', 'Дроссель', '')])
        self.c.post('/api/library/diff/', {'files': up})
        cat.refresh_from_db()
        self.assertEqual(cat.description, 'Индуктивности')

    def test_patch_category_description(self):
        """Волна 22: описание класса правится из продукта (таб «Категории»), а не
        только из админки. `code` — ключ синка и в теле не принимается."""
        cat = engine.ensure_category('capacitors')
        r = self.c.patch(f'/api/categories/{cat.id}/',
                         data=json.dumps({'description': '  Конденсаторы SMD  ',
                                          'code': 'hacked'}),
                         content_type='application/json')
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.json()['description'], 'Конденсаторы SMD')
        cat.refresh_from_db()
        self.assertEqual(cat.description, 'Конденсаторы SMD')
        self.assertEqual(cat.code, 'capacitors')

    def test_patch_category_description_too_long(self):
        cat = engine.ensure_category('capacitors')
        r = self.c.patch(f'/api/categories/{cat.id}/',
                         data=json.dumps({'description': 'я' * 129}),
                         content_type='application/json')
        self.assertEqual(r.status_code, 400)
        cat.refresh_from_db()
        self.assertEqual(cat.description, 'Конденсаторы')

    def test_recalc_cost_endpoint(self):
        cat = _cat()
        leaf = models.Item.objects.create(code='L', description='L', category=cat,
                                          estimated_cost=D(7))
        dev = models.Item.objects.create(code='D', description='D', category=cat,
                                         native=True)
        models.BomLine.objects.create(parent=dev, component=leaf, qty=D(3))
        r = self.c.post(f'/api/items/{dev.id}/recalc-cost/')
        self.assertEqual(r.status_code, 200)
        body = r.json()
        self.assertEqual(float(body['estimated_cost']), 21.0)
        self.assertEqual(body['rollup']['updated'], ['D'])
        dev.refresh_from_db()
        self.assertEqual(dev.estimated_cost, D(21))

    def test_recalc_cost_rejects_non_produced(self):
        item = models.Item.objects.create(code='P', description='P',
                                          category=_cat(), estimated_cost=D(5))
        r = self.c.post(f'/api/items/{item.id}/recalc-cost/')
        self.assertEqual(r.status_code, 400)


class LotOriginProjectionTests(EngineTestBase):
    """Волна 19, Ф12c: глиф партии (§7a) = форма по origin + цвет по остатку, поэтому
    `origin` обязан быть В КАЖДОЙ проекции, где строка = партия. Раньше вид рождения
    знали только форма изделия и внутренний склад — списки склада/ордеров молчали."""

    def setUp(self):
        super().setUp()
        self.item = self.make_item('R100')
        self.lot = self.receipt_lot(self.item, self.prj, 10)   # origin = поставка

    def test_location_stock_carries_origin(self):
        rows = engine.location_stock(self.main)
        self.assertEqual(rows[0]['origin'], models.StockDocument.Kind.RECEIPT)

    def test_project_closure_residuals_carry_origin(self):
        row = engine.project_closure(self.prj)['residuals'][0]
        self.assertEqual(row['origin'], models.StockDocument.Kind.RECEIPT)

    def test_order_lines_carry_origin_of_source_lot(self):
        wo = engine.create_writeoff(self.prj, self.user, 'СП-1')
        engine.add_writeoff_line(wo, self.lot, D(2))
        self.assertEqual(engine.writeoff_form(wo)['lines'][0]['origin'],
                         models.StockDocument.Kind.RECEIPT)
        tr = engine.create_transfer(self.prj, self.user, 'НАК-1')
        engine.add_transfer_line(tr, self.lot, D(1))
        self.assertEqual(engine.transfer_form(tr)['lines'][0]['origin'],
                         models.StockDocument.Kind.RECEIPT)

    def test_purchase_rows_carry_item_axes(self):
        """Строка заказа несёт оси изделия — глиф строки один (форма = изделие,
        цвет = закрытость), как в форме закупки-плана."""
        purchase = engine.create_purchase(self.prj, self.user)
        engine.add_purchase_line(purchase, self.item, D(5))
        row = engine.purchase_form(purchase)['rows'][0]
        self.assertEqual(
            (row['item_native'], row['item_synced'], row['item_locked']),
            (self.item.native, self.item.synced, self.item.locked))


class DraftDoesNotMoveStockTests(EngineTestBase):
    """Волна 19, Ф15: **замок гейтит склад** — черновик ничего не двигает.

    Правило одно на семь видов ордера: пока `locked=False`, документ существует и
    правится, но его партии не лежат на складе, его строки не расходуют чужие, а
    бюджет проекта его не видит. Фиксация материализует, расфиксация снимает.
    Единственная точка врезки — `rebuild_movements`; всё остальное (остатки, дефицит,
    карта складов, деньги) читает движения и подчиняется автоматически.
    """

    def setUp(self):
        super().setUp()
        self.item = self.make_item('R100')

    # ── рождение партии: поставка / инвентаризация ──
    def test_draft_receipt_lot_is_not_on_stock(self):
        r = models.Receipt.objects.create(
            number='У-1', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        lot = engine.add_receipt_lot(r, self.item, D(10), unit_cost=D(100))
        self.assertTrue(models.Lot.objects.filter(pk=lot.pk).exists())  # партия есть
        self.assertEqual(lot.movements.count(), 0)                      # склада нет
        self.assertEqual(engine.item_available(self.item, self.prj), D(0))

    def test_lock_materializes_unlock_takes_back(self):
        r = models.Receipt.objects.create(
            number='У-1', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        lot = engine.add_receipt_lot(r, self.item, D(10))
        engine.lock_receipt(r)
        self.assertEqual(engine.item_available(self.item, self.prj), D(10))
        engine.unlock_receipt(r)
        self.assertEqual(engine.item_available(self.item, self.prj), D(0))
        self.assertEqual(lot.movements.count(), 0)

    def test_draft_inventory_lot_is_not_on_stock(self):
        inv = engine.create_inventory(self.prj, self.user, 'ИНВ-1')
        lot = engine.add_inventory_lot(inv, self.item, D(5))
        self.assertEqual(engine.lot_live_qty(lot), D(0))
        engine.lock_inventory(inv)
        self.assertEqual(engine.lot_live_qty(lot), D(5))

    # ── расход существующей партии: все знаковые виды ──
    def test_draft_consumers_do_not_touch_source(self):
        """Пять расходных видов в черновике не двигают источник; фиксация двигает."""
        lot = self.receipt_lot(self.item, self.prj, 100)
        other = models.Project.objects.create(
            code='P2', description='Проект 2', kind=models.Project.Kind.EXTERNAL)
        dev = self.make_item('DEV', manufactured=True)

        k = models.Kitting.objects.create(project=self.prj, target_item=dev,
                                          user=self.user, qty=D(1))
        engine.add_kitting_line(k, self.item, lot, D(2))
        t = engine.create_transfer(self.prj, self.user, 'Н-1')
        engine.add_transfer_line(t, lot, D(3))
        w = engine.create_writeoff(self.prj, self.user, 'СП-1')
        engine.add_writeoff_line(w, lot, D(4))
        req = engine.create_requisition(other, self.user, 'ТР-1')
        engine.add_requisition_line(req, lot, D(5))
        rel = engine.create_relocation(self.prj, self.user, 'ПЕР-1')
        engine.add_relocation_line(rel, lot, D(6), self.main,
                                   models.Location.objects.create(code='105',
                                                                  description='Пайка'))
        # все пять — черновики: остаток нетронут, движение ровно одно (рождение)
        self.assertEqual(engine.lot_live_qty(lot), D(100))
        self.assertEqual(lot.movements.count(), 1)

        engine.lock_kitting(k); engine.lock_transfer(t); engine.lock_writeoff(w)
        engine.lock_requisition(req); engine.lock_relocation(rel)
        # 100 − 2 − 3 − 4 − 5 = 86 (перемещение тотал не меняет: −6/+6)
        self.assertEqual(engine.lot_live_qty(lot), D(86))

    def test_issue_from_draft_receipt_goes_negative(self):
        """Решение Ивана: расход непринятой партии **пускаем в минус** — не клампим,
        недостача информативнее (канон мутабельной ДНК)."""
        r = models.Receipt.objects.create(
            number='У-1', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        lot = engine.add_receipt_lot(r, self.item, D(10))
        t = engine.create_transfer(self.prj, self.user, 'Н-1')
        engine.add_transfer_line(t, lot, D(4))
        engine.lock_transfer(t)                      # накладная проведена, УПД — нет
        self.assertEqual(engine.lot_live_qty(lot), D(-4))

    # ── производные проекции подчиняются автоматически ──
    def test_draft_receipt_does_not_cover_deficit(self):
        dev = self.make_item('DEV', manufactured=True)
        models.BomLine.objects.create(parent=dev, component=self.item, qty=D(1))
        models.ProjectDemand.objects.create(project=self.prj, target_item=dev, qty=D(10))
        r = models.Receipt.objects.create(
            number='У-1', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        engine.add_receipt_lot(r, self.item, D(10))
        row = engine.project_deficit(self.prj)['components'][0]
        self.assertEqual(row['in_stock'], D(0))              # черновик не на складе
        self.assertEqual(row['balance'], D(-10))             # и потребность не покрывает
        engine.lock_receipt(r)
        row = engine.project_deficit(self.prj)['components'][0]
        self.assertEqual(row['in_stock'], D(10))
        self.assertEqual(row['balance'], D(0))               # сверен → покрыл ровно
        self.assertEqual(row['supply'], 'available')         # снабжение здорово

    def test_draft_receipt_does_not_spend_budget(self):
        """Деньги — единственное чтение мимо движений, поэтому гейт продублирован в
        `_project_spent`: черновой УПД не двигает бюджет проекта (Ф15)."""
        r = models.Receipt.objects.create(
            number='У-1', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        engine.add_receipt_lot(r, self.item, D(3), unit_cost=D(800))
        self.assertEqual(engine.project_budget(self.prj)['spent'], D(0))
        engine.lock_receipt(r)
        self.assertEqual(engine.project_budget(self.prj)['spent'], D(2400))

    def test_closure_panel_shows_draft_closing_documents(self):
        """Мост панели кладёт остаток в ЧЕРНОВОЙ документ — панель это показывает и
        называет причиной отказа, иначе кнопка выглядела бы несработавшей."""
        lot = self.receipt_lot(self.item, self.prj, 10)
        w = engine.writeoff_lot(self.prj, lot, D(10), self.user)
        panel = engine.project_closure(self.prj)
        self.assertFalse(panel['can_close'])
        self.assertEqual(len(panel['closing_drafts']), 1)
        self.assertEqual(panel['closing_drafts'][0]['document_id'], w.id)
        self.assertEqual(panel['closing_drafts'][0]['qty'], D(10))
        self.assertIn('черновик', panel['blocker'].lower())
        engine.lock_writeoff(w)
        panel = engine.project_closure(self.prj)
        self.assertEqual(panel['closing_drafts'], [])
        self.assertTrue(panel['can_close'])

    def test_closure_blocker_names_both_halves_of_work(self):
        """Аудит-1 Б2б-5: черновик разобрал часть остатка — панель обязана назвать и
        вторую половину. Одинокое «зафиксируйте черновики» читается как «больше делать
        нечего», хотя неразобранное никуда не делось."""
        a = self.receipt_lot(self.item, self.prj, 10)
        self.receipt_lot(self.make_item('R300'), self.prj, 4)
        engine.writeoff_lot(self.prj, a, D(10), self.user)
        panel = engine.project_closure(self.prj)
        self.assertEqual(panel['residual_in_drafts'], D(10))
        self.assertEqual(panel['residual_unsorted'], D(4))
        self.assertIn('10', panel['blocker'])
        self.assertIn('4', panel['blocker'])

    def test_item_screen_keeps_draft_lot_but_marks_it(self):
        """Таб «Склад» изделия показывает партии НЕЗАВИСИМО от движений — так виден
        входящий поток («10 едет, 0 принято»). Значит проекция обязана отдать замок
        origin-документа: иначе вью не отличит непринятую партию от израсходованной
        (обе с остатком 0) и подпишет её «исчерпана»."""
        r = models.Receipt.objects.create(
            number='У-1', date='2026-05-01', contractor=self.supplier,
            project=self.prj, user=self.user)
        engine.add_receipt_lot(r, self.item, D(10))
        row = views._item_detail_payload(self.item)['lots'][0]
        self.assertEqual(row['qty_born'], D(10))     # рождено — видно
        self.assertEqual(row['live_qty'], D(0))      # на складе — нет
        self.assertFalse(row['origin_locked'])       # и вью знает, почему
        engine.lock_receipt(r)
        row = views._item_detail_payload(self.item)['lots'][0]
        self.assertEqual(row['live_qty'], D(10))
        self.assertTrue(row['origin_locked'])

    def test_bridge_reuses_only_draft_document(self):
        """Мост переиспользует лишь расфиксированный акт — зафиксированный правке не
        подлежит, поэтому для следующего остатка заводится новый."""
        a = self.receipt_lot(self.item, self.prj, 4)
        b = self.receipt_lot(self.make_item('R200'), self.prj, 6)
        w1 = engine.writeoff_lot(self.prj, a, D(4), self.user)
        w2 = engine.writeoff_lot(self.prj, b, D(6), self.user)
        self.assertEqual(w1.id, w2.id)                    # черновик переиспользован
        engine.lock_writeoff(w1)
        c = self.receipt_lot(self.make_item('R300'), self.prj, 2)
        w3 = engine.writeoff_lot(self.prj, c, D(2), self.user)
        self.assertNotEqual(w3.id, w1.id)                 # под замком — новый акт

    def test_bridge_refuses_lot_already_in_closing_draft(self):
        """Повторный клик по остатку не заводит ВТОРОЙ акт на тот же лот: после Ф15
        остаток не гаснет до фиксации, клик выглядит несработавшим — и два акта
        увели бы лот в минус. Вместо документа человек получает адрес первого."""
        lot = self.receipt_lot(self.item, self.prj, 10)
        w = engine.writeoff_lot(self.prj, lot, D(10), self.user)
        with self.assertRaises(ValidationError) as e:
            engine.writeoff_lot(self.prj, lot, D(10), self.user)
        self.assertIn(w.code, e.exception.messages[0])
        # и второй мост тоже: черновик один на лот, чьего бы вида он ни был
        with self.assertRaises(ValidationError):
            engine.requisition_lot(self.prj, lot, D(10), self.user)
        self.assertEqual(models.Writeoff.objects.filter(project=self.prj).count(), 1)
        engine.lock_writeoff(w)                            # зафиксировали — путь открыт
        self.assertEqual(engine.lot_live_qty(lot), D(0))

    def test_draft_writeoff_is_not_offered_for_rematerialization(self):
        """Пикер инвентаризации предлагает вернуть только реально списанное:
        черновой акт — намерение, а не факт (Ф15), и «серого» за ним ещё нет."""
        lot = self.receipt_lot(self.item, self.prj, 10)
        w = engine.create_writeoff(self.prj, self.user, 'СП-1')
        engine.add_writeoff_line(w, lot, D(10))
        self.assertEqual(engine.written_off_lots(), [])
        engine.lock_writeoff(w)
        rows = engine.written_off_lots()
        self.assertEqual([r['lot_id'] for r in rows], [lot.id])
        self.assertEqual(rows[0]['written_qty'], D(10))


class AccountTests(TestCase):
    """Волна 21, Ф1 — аккаунт: профиль-приставка, ДНК, тема, пароль, «свои» документы.

    Форма адресуется БЕЗ `pk` (`/api/account/`), поэтому «а вдруг откроет чужую» здесь
    не сценарий, а структурная невозможность: движок работает с `request.user`.
    """

    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username='ivan', password='k7-Plume-pass', first_name='Иван')
        self.client = Client()
        self.client.force_login(self.user)

    # ── профиль-приставка ──

    def test_profile_is_born_lazily(self):
        """Сигнала на рождение пользователя нет (он ломает `loaddata`) — профиль
        появляется при первом обращении и ровно один."""
        self.assertEqual(models.UserProfile.objects.count(), 0)
        self.assertEqual(engine.profile_of(self.user).theme, models.DEFAULT_THEME)
        engine.profile_of(self.user)
        self.assertEqual(models.UserProfile.objects.count(), 1)

    def test_me_carries_theme(self):
        """Тема приезжает первым же запросом старта — без второго round-trip."""
        me = self.client.get('/api/auth/me/').json()
        self.assertEqual(me['theme'], models.DEFAULT_THEME)
        # …но не в справочнике пикера авторства: чужие настройки вью знать незачем.
        self.assertNotIn('theme', self.client.get('/api/users/').json()[0])

    # ── ДНК Django ──

    def test_form_shows_dna_and_no_username_field_to_patch(self):
        form = self.client.get('/api/account/').json()
        self.assertEqual(form['username'], 'ivan')
        self.assertEqual(form['full_name'], 'Иван')
        self.assertEqual(form['documents'], [])

    def test_patch_name_and_email(self):
        form = self.client.patch(
            '/api/account/',
            {'first_name': 'Иван', 'last_name': 'Потылицын', 'email': 'i@example.com'},
            content_type='application/json').json()
        self.assertEqual(form['last_name'], 'Потылицын')
        self.assertEqual(form['full_name'], 'Иван Потылицын')
        self.assertEqual(form['email'], 'i@example.com')

    def test_patch_bad_email_is_friendly_400(self):
        bad = self.client.patch('/api/account/', {'email': 'не почта'},
                                content_type='application/json')
        self.assertEqual(bad.status_code, 400)
        self.assertIn('Почта', bad.json()['detail'])

    def test_empty_email_is_legal(self):
        """Django свою пустую почту допускает — «заполнить можно, передумать нельзя»
        продукт не заводит нигде."""
        self.client.patch('/api/account/', {'email': 'i@example.com'},
                          content_type='application/json')
        form = self.client.patch('/api/account/', {'email': ''},
                                 content_type='application/json').json()
        self.assertEqual(form['email'], '')

    # ── тема интерфейса ──

    def test_theme_patch_and_unknown_slug_is_400(self):
        form = self.client.patch('/api/account/', {'theme': 'light'},
                                 content_type='application/json').json()
        self.assertEqual(form['theme'], 'light')
        bad = self.client.patch('/api/account/', {'theme': 'ide-dark'},
                                content_type='application/json')
        self.assertEqual(bad.status_code, 400)
        self.assertIn('Неизвестная тема', bad.json()['detail'])
        # отказ ничего не записал
        self.assertEqual(engine.profile_of(self.user).theme, 'light')

    def test_theme_slug_is_guarded_by_engine_not_by_check(self):
        """Схема слаг НЕ стережёт (сознательно: тема = набор файлов вью, и новая тема
        не должна стоить миграции) — стережёт движок, единственным входом."""
        models.UserProfile.objects.create(user=self.user, theme='что-угодно')
        with self.assertRaises(ValidationError):
            engine.set_theme(self.user, 'что-угодно')

    # ── смена пароля ──

    def test_password_change_keeps_session_alive(self):
        ok = self.client.post('/api/account/password/',
                              {'current': 'k7-Plume-pass', 'new': 'q9-Plume-next',
                               'repeat': 'q9-Plume-next'},
                              content_type='application/json')
        self.assertEqual(ok.status_code, 204)
        self.user.refresh_from_db()
        self.assertTrue(self.user.check_password('q9-Plume-next'))
        # сессия подписана хэшем пароля: без `update_session_auth_hash` человек
        # вылетел бы на логин сразу после успешной смены.
        self.assertEqual(self.client.get('/api/account/').status_code, 200)

    def test_password_wrong_current_is_400(self):
        bad = self.client.post('/api/account/password/',
                               {'current': 'мимо', 'new': 'q9-Plume-next',
                                'repeat': 'q9-Plume-next'},
                               content_type='application/json')
        self.assertEqual(bad.status_code, 400)
        self.assertIn('Текущий пароль', bad.json()['detail'])
        self.user.refresh_from_db()
        self.assertTrue(self.user.check_password('k7-Plume-pass'))

    def test_password_repeat_mismatch_is_400(self):
        bad = self.client.post('/api/account/password/',
                               {'current': 'k7-Plume-pass', 'new': 'q9-Plume-next',
                                'repeat': 'q9-Plume-nekst'},
                               content_type='application/json')
        self.assertEqual(bad.status_code, 400)
        self.assertIn('не совпадают', bad.json()['detail'])

    def test_password_weak_is_400_by_django_validators(self):
        """Своих правил стойкости не изобретаем — отказывают штатные
        `AUTH_PASSWORD_VALIDATORS`, те же, что у админки."""
        bad = self.client.post('/api/account/password/',
                               {'current': 'k7-Plume-pass', 'new': '12345678',
                                'repeat': '12345678'},
                               content_type='application/json')
        self.assertEqual(bad.status_code, 400)
        self.user.refresh_from_db()
        self.assertTrue(self.user.check_password('k7-Plume-pass'))

    # ── три ленты «своих» документов ──

    def test_form_carries_only_own_documents(self):
        """«Свои» — это авторство, а не права: реверс `user.documents` и родня. Модели
        «кто чьи документы видит» в продукте нет."""
        other = get_user_model().objects.create_user(username='petr')
        prj = models.Project.objects.create(code='P1', description='Проект 1',
                                            kind=models.Project.Kind.EXTERNAL)
        cp = engine.create_counterparty(code='КОМПЭЛ', description='ООО Компэл')
        mine = models.Receipt.objects.create(code='Поставка моя', number='У-1',
                                             date='2026-05-01', contractor=cp,
                                             project=prj, user=self.user)
        models.Receipt.objects.create(code='Поставка чужая', number='У-2',
                                      date='2026-05-02', contractor=cp,
                                      project=prj, user=other)
        engine.create_procurement(code='Закупка моя', user=self.user)
        engine.create_purchase(prj, self.user, code='Заказ мой')

        form = self.client.get('/api/account/').json()
        self.assertEqual([d['id'] for d in form['documents']], [mine.id])
        self.assertEqual(form['documents'][0]['kind'], 'receipt')
        self.assertEqual(form['documents'][0]['project_code'], 'P1')
        self.assertEqual([p['code'] for p in form['procurements']], ['Закупка моя'])
        self.assertEqual([p['code'] for p in form['purchases']], ['Заказ мой'])
