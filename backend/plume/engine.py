"""Движок-СППР plume.

Чистые функции-проекции над документами (один движок на всю линзу). Ничего не
кэшируем: всё вычислимое держим свежим (данных мало, без Celery).

Волна 1:
- `rebuild_movements(lot)` — пересборка StockMovement партии из её документов.
- `lot_live_qty` / `item_available` — живые остатки.
- `project_deficit(project)` — дефицит проекта (надо − склад − заказано),
  1 уровень BOM, тройной разбор ✓/●/▲, worst-of цвет.
- `stock_map(item)` — карта остатков Item по всем складам-проектам (north-star).

Волна 2 (записываемое ядро + форма Kitting):
- `kitting_form(kitting)` — проекция формы сборки: BOM 1 уровень, реальные
  (пробитые) строки + призрачные строки, покрашенные по доступности + лоты-кандидаты.
- `lock_kitting` / `unlock_kitting` — рождение/снятие лота-прибора (мягкий замок).

Волна 3 (записываемый приход / УПД):
- `receipt_form(receipt)` — проекция формы прихода: строки-лоты УПД (в модели
  отдельной ReceiptLine нет — строки прихода это его лоты) + живой остаток + сумма.
- `add/update/remove_receipt_lot`, `lock/unlock_receipt` — рождение лотов
  (`+RECEIPT`) и мягкий замок «сверено со сканом».

Волна 4 (записываемый заказ / Purchase + связь с приходом):
- `purchase_form(purchase)` — шапка + строки (заказано/поступило/остаток) + приходы;
  `create_purchase`, `add/update/remove_purchase_line`, `send/unsend/cancel/restore_purchase`.
- `set_receipt_purchase(receipt, purchase)` — связь `Receipt↔Purchase` (гашение заказа).
- `lines_estimate(lines)` — прогноз денег документа-намерения: `Σ(qty×estimated_cost)`
  + коды позиций без оценки. Общий для заказа и закупки-плана (поле «Оценка» в шапке).

Волна 5 (записываемая передача / Transfer — отгрузка заказчику):
- `transfer_form(transfer)` — шапка накладной + строки-лоты (отдаём партию заказчику,
  `−ISSUE`) + живой остаток источника + итог; `project_available_lots(project)` — пикер
  отдаваемых лотов (live>0). `create_transfer`, `add/update/remove_transfer_line`.
- Мягкий замок «отгружено» (единый `locked`, волна 13 Ф1):
  `lock_transfer`/`unlock_transfer` — под замком форма read-only; снятие ничего не
  разрушает (guard по потомкам не нужен). Лента `item_movements(item)` — ВСЕ ордера,
  коснувшиеся изделия (рождения партий + движения), для его экрана; заменила узкую
  `item_shipments` в волне 19 (Ф12a).

Волна 6 (закрытие проекта — сведение остатков в 0 + мягкий замок):
- `writeoff_form` / `create_writeoff` / `add|update|remove_writeoff_line` — списание
  (`−ISSUE`, лот покидает учёт; серый путь). `requisition_form` / `create_requisition`
  / `add|update|remove_requisition_line` — требование/отпочкование (`−ISSUE` источника +
  рождение лота-потомка в проекте-получателе, `+RECEIPT`; белый путь / заём).
- `project_closure(project)` — панель сведения остаточных лотов (live≠0) в 0 +
  готовность; `lock_project`/`unlock_project` — мягкий замок-веха «проект отработан».
- Мосты панели: `writeoff_lot` (списать остаток) / `requisition_lot` (на баланс → белый).

Волна 7 (планирование закупок — записываемый Procurement; охват — волна 19, Ф13):
- `scope_deficit(projects)` — свод по оси Item через **заданный набор** проектов
  (`Σ` проектных дефицитов, без перенеттинга между проектами). Витрина.
- `procurement_scope(proc)` — охват закупки (проекты её заказов, вычисляемый);
  `procurement_deficit(proc)` = свод по охвату (таб «К закупке»; заменил экран
  «Командный свод», у которого не было ни кода, ни замка).
- `procurement_form` / `create_procurement` / `add|update|remove_procurement_line`
  / `send|unsend|cancel|restore_procurement` — записываемый план закупки под охват
  проектов (мягкий замок `locked`). Нарезка на `Purchase` — волна 8.
- `add_to_procurement(...)` — мост «витрина → строки плана» (топ-ап до наводки);
  `procurement_xlsx(...)` — xlsx-бланк поставщику.

Волна 8 (нарезка плана на проектные заказы; переделана 2026-08-05):
- `procurement_allocation(proc)` — проекция «Привязка»: строка плана × ЗАКАЗЫ закупки
  (не проекты), в ячейке — разложенное количество и баланс проекта по этому Item
  (`_balance`, то же число, что в «Потребности»); плюс веер заказов.
- `set_allocation(proc, purchase, item, qty)` — присвоение количества в ячейку
  (`0` снимает строку). Заказы заводятся руками в своей форме и привязываются к
  закупке; движок их больше не рождает.

Волна 10 (бюджет/экономия — north-star окупаемости линзы):
- `project_budget(project)` — проекция денег проекта: **потрачено** (факт по
  `Receipt`-лотам, заём/свои бесплатны, только покупные) + **план** (прогноз «факт
  где есть, оценка где нет» через `estimated_cost`) + компас `budget − план` +
  позиции без оценки; **себестоимость** (Σ снимков лотов-приборов верхних целей,
  заём по реальной цене) + **экономия** = себестоимость − потрачено. Чистая витрина.
- `intent_money(lines, need)` (2026-08-07) — те же деньги, но у документа-намерения:
  **потребность** (нужда проекта × цена — сколько надо по-настоящему), **сумма
  документа** (`lines_estimate`) и **переплата** = их разница со знаком. Знаменатель
  даёт `demand_map(projects)` (нужда по листьям через набор проектов): у заказа это его
  проект, у закупки — её охват (`counts_in_scope`, тот же набор, что у свода).
- `line_cost(item, qty)` + `mark_costs(rows)` (2026-08-07) — деньги ОТДЕЛЬНОЙ строки в
  «Строках» заказа и «Привязке» закупки: стоимость (без оценки — `None`, не ноль) и
  `cost_status` = `overpaid` (взято сверх порога `overpay_threshold` = полторы нормы,
  округлённые вверх до круглого: 5/10/50/100) > `costly` (верхняя четверть набора по
  стоимости). Ранжирование — внутри одного списка; порог едет в строку (`overpay_at`),
  чтобы вью объясняла цвет числом, а не словами.

Волна 13 Ф2e (мультисклад + перемещение):
- Остаток по паре `(лот, локация)`: `lot_live_qty(lot, location)` / `item_available(…,
  location)` / `available_lots(…, location)` (опциональный фильтр, по умолчанию — тотал,
  байт-в-байт), `lot_locations(lot)` — разбивка по местам; `stock_map` несёт аддитивный
  `by_location`.
- Перемещение (`Relocation`): `create_relocation` / `relocation_form` / `add|update|
  remove_relocation_line` (пара знаковых `StockLine` `−q`@источник/`+q`@приёмник на ход,
  тотал лота сохранён) / `post|unlock_relocation` / `relocation_source_lots` (пикер с
  разбивкой по местам). HTTP/React — следующим заходом («вьюхи потом»).

Волна 19, Ф15 (замок гейтит склад):
- `rebuild_movements` пропускает в `StockMovement` только зафиксированные документы:
  черновик создаётся/правится/удаляется, **ничего никому не двигая**; фиксация делает
  его видимым всем. Правило не новое — так с волны 2 работала комплектация (лот-прибор
  ждал замка) и фильтр «готово» в дефиците; фаза сняла шесть исключений.
- `rebuild_document_movements(doc)` — материализация/снятие склада на `lock_*`/`unlock_*`
  (партии документа + партии-источники его строк). Деньги: `_project_spent` считает
  только лоты зафиксированных поставок (единственное чтение денег мимо движений).

Волна 21 (аккаунт + тема интерфейса):
- `profile_of(user)` — ленивая приставка настроек (`UserProfile`), `set_theme` —
  единственный вход к теме (движок знает СЛАГ и список допустимых, про цвета не знает).
- `user_form(user)` — проекция формы аккаунта: ДНК Django + тема + три ленты «своих»
  документов; `update_user` / `change_password` — правка ДНК и пароля.

Следующие волны: логин-экран, UI вложений (`Attachment`).
"""
import csv
import io
import os
from datetime import date as dt_date, datetime, timedelta, timezone as dt_timezone
from decimal import ROUND_CEILING, ROUND_HALF_UP, Decimal
from uuid import uuid4

from django.conf import settings
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError
from django.core.validators import EmailValidator
from django.db import transaction
from django.db.models import Exists, OuterRef, Sum
from django.db.models.deletion import ProtectedError
from django.utils import timezone

from . import models

ZERO = Decimal('0')


def _num(value):
    """Кол-во в текст сообщения: без хвостовых нулей и без экспоненты («10», «2.5»).
    `normalize()` даёт `1E+1` на круглых числах — формат `f` возвращает его обратно."""
    return f'{Decimal(value).normalize():f}'


# --------------------------------------------------------------------------- #
#  Категории изделий (волна 15) — канон внешней библиотеки компонентов
# --------------------------------------------------------------------------- #
# Стем имени CSV-файла → рус. описание. Синк и сид зовут `ensure_category`;
# неизвестный стем всплывает с description=code (сырой — юзер правит в аппе/админке).
# Стартовый набор — только эти 5 (прочие классы прибор/крепёж/деталь юзер добавит сам).
# Волна 19, Ф10: иконки убраны (per-категорийный глиф отпал — режимы Изделия/Компоненты).
LIBRARY_CATEGORIES = {
    'capacitors': 'Конденсаторы',
    'mcu':        'Микроконтроллеры',
    'regulators': 'Стабилизаторы',
    'sensors':    'Датчики',
    'interfaces': 'Интерфейсы',
}


def ensure_category(code):
    """get_or_create категории по стему CSV-файла. Канон `description` — из
    `LIBRARY_CATEGORIES`; неизвестный код всплывает с сырым `description=code`.
    Существующую (юзер уже правил описание) НЕ перезаписываем."""
    code = (code or '').strip()
    description = LIBRARY_CATEGORIES.get(code, code)
    cat, _ = models.Category.objects.get_or_create(
        code=code, defaults={'description': description})
    return cat


# --------------------------------------------------------------------------- #
#  Единый мягкий замок складского документа (волна 13, Ф1)
# --------------------------------------------------------------------------- #
def _require_unlocked(doc, msg=None):
    """Единый guard правки: документ должен быть расфиксирован (замок снят).

    Свернул разнородные `_require_wip`/`_require_unapproved`/`_require_unposted` —
    один мягкий замок на всех ордерах, `locked=True` = edit-freeze. С волны 19 (Ф1)
    обслуживает и Закупку с Заказом — та же ось; своих `_require_*_draft` больше
    нет. Ф1c сделала ось булевой. `msg` уточняет формулировку по сущности.
    """
    if doc.locked:
        raise ValidationError(
            msg or 'Документ зафиксирован — расфиксируйте для правки.')


def _require_header(doc):
    """Гейт полноты шапки на фиксации — условная валидация специфики по виду
    (волна 13, Ф2d). Единый kind-driven источник правила живёт на модели
    (`StockDocument.clean` + `REQUIRED_HEADER_BY_KIND`); фиксация не выпускает
    неполный ордер независимо от пути его создания (API/админ/прямой ORM).

    Ф12e добавила сюда специфику вида. Обязательность `contractor`/`target_item`
    Ф14 выразила CHECK'ами по `locked` — но прикладного гейта у них не было, и до
    Ф12e это не проявлялось: до неё оба поля требовались при рождении, и дойти до
    фиксации с пустым было неоткуда. Теперь это штатное состояние черновика, и без
    гейта пользователь получал бы `IntegrityError 3819` вместо внятного отказа.
    """
    doc.clean()
    kind = doc.KIND or doc.kind
    if kind == models.DocumentKind.RECEIPT and doc.contractor_id is None:
        raise ValidationError('Перед фиксацией выберите поставщика.')
    if kind == models.DocumentKind.KITTING and (
            doc.target_item_id is None or doc.qty is None):
        raise ValidationError('Перед фиксацией выберите прибор-цель и кол-во.')


def rebuild_document_movements(doc):
    """Пересобрать движения ВСЕХ партий, которых касается документ (волна 19, Ф15).

    Замок гейтит склад, поэтому смена `locked` — событие склада: фиксация обязана
    материализовать движения документа, расфиксация — снять. Касается двух семейств
    партий: рождённых документом (`doc.lots`, origin-`+RECEIPT`) и упомянутых его
    знаковыми строками (`doc.lines`, расход/приём источника). Зовётся из всех
    `lock_*`/`unlock_*`; сама пересборка чистая, порядок вызовов не важен.
    """
    lots = set(doc.lots.all())
    lots |= {sl.lot for sl in doc.lines.select_related('lot')}
    for lot in lots:
        rebuild_movements(lot)


def lock_document(doc, rows, empty_msg):
    """Поставить единый мягкий замок (edit-freeze формы) + материализовать склад.

    `rows` — менеджер строк документа (`doc.lots` для born-only, `doc.lines` для
    расходных): нельзя зафиксировать пустой документ. Ф15: фиксация — момент, когда
    документ становится виден складу (до неё его строки не двигали ничего).
    """
    if not rows.exists():
        raise ValidationError(empty_msg)
    _require_header(doc)
    doc.locked = True
    doc.save(update_fields=['locked'])
    rebuild_document_movements(doc)
    return doc


def unlock_document(doc):
    """Снять единый мягкий замок: форма снова правима, движения документа сняты.

    Ничего не разрушает (строки и рождённые лоты остаются) — guard по потомкам не
    нужен, в отличие от расфиксации комплектации, где снимается лот-прибор. Ф15:
    расфиксация возвращает документ в черновики и **со склада убирает** — расход
    источников отпускается, рождённые партии перестают лежать на складе.
    """
    doc.locked = False
    doc.save(update_fields=['locked'])
    rebuild_document_movements(doc)
    return doc


def delete_stock_document(doc):
    """Удалить складской ордер (отмена = удаление, канон В13 Ф1 — не копим надгробия).

    Правило удаления (единое на 6 ордеров):
    - **locked** — «сперва расфиксировать»: удаляем только расфиксированный;
    - **не locked** — свободно, но `PROTECT` бережёт потраченные лоты: если рождённый
      документом лот (`doc.lots`, born-direct) уже потреблён/передан/отпочкован ниже
      — дружелюбный отказ вместо сырого `ProtectedError`.

    Механика (обходит грабли CHECK `lot_exactly_one_origin`, см. JOURNAL Ф1a): born-лоты
    и их движения сносим **явно** (как `unlock_kitting`), затем документ (каскад
    расходных `StockLine` + вложений), затем пересобираем движения лотов-источников
    (снять их `−ISSUE`). Файлы вложений чистим отдельно — каскад БД их бы осиротил.
    """
    if doc.locked:
        raise ValidationError(
            'Документ зафиксирован — сперва расфиксируйте его для удаления.')
    born = list(doc.lots.all()) if hasattr(doc, 'lots') else []
    for lot in born:
        if _lot_consumed_downstream(lot):
            raise ValidationError(
                'Рождённый документом лот уже потреблён/передан ниже — '
                'удаление заблокировано (сперва снимите потребление).')
    # лоты-источники расходных строк — их движения пересобрать после каскада строк
    source_lots = ({sl.lot for sl in doc.lines.select_related('lot')}
                   if hasattr(doc, 'lines') else set())
    source_lots -= set(born)                       # born сносим сами, не пересобираем
    for att in doc.attachments.all():              # физические файлы (каскад их сиротит)
        delete_attachment(att)
    for lot in born:                               # явный снос born (обход CHECK-грабли)
        lot.movements.all().delete()
        lot.delete()
    doc.delete()                                   # каскад: StockLine + строки-вложения
    for lot in source_lots:                        # снять −ISSUE удалённых строк
        rebuild_movements(lot)


# Волна 19, Ф1c: слой совместимости `wip`/`closed` (комплектация отдавала эти
# строки с волны 2) снят — наружу идёт та же `locked`, что у всех сущностей.
# Фронт переехал вместе с бэком, третьего словаря статуса в API больше нет.


# --------------------------------------------------------------------------- #
#  Склад: пересборка движений и живые остатки
# --------------------------------------------------------------------------- #
def _main_location():
    """Дом приходного движения. В MVP — один «Основной склад» (код MAIN)."""
    loc = models.Location.objects.filter(code='MAIN').first()
    return loc or models.Location.objects.order_by('id').first()


def rebuild_movements(lot):
    """Пересобрать StockMovement партии из её документов (чистая пересборка).

    origin-`+RECEIPT` берёт рождённое количество из `Lot.qty`; расходные `ISSUE`
    выводятся из строк-потребителей, ссылающихся на партию.

    **Волна 19, Ф15 — замок гейтит склад.** В движения попадает только то, что
    пришло от **зафиксированного** документа: черновик спокойно создаётся,
    заполняется и удаляется, ничего никому не двигая; фиксация делает его видимым
    всем (`lock_*` пересобирает движения, `unlock_*` их снимает). Точка врезки одна
    — эта функция единственный писатель `StockMovement`, поэтому остатки, дефицит,
    карта складов и бюджет подчиняются правилу автоматически.
    """
    lot.movements.all().delete()
    main = _main_location()
    rows = []

    # origin: рождение партии (+). Дуга схлопнута в один FK (Ф2b): вид и id берём
    # из родителя `StockDocument` (`kind` == прежнее имя origin-FK; id == прежнему).
    # Ф15: партия черновика на складе не лежит — материализуется на фиксации.
    if lot.origin_id and lot.qty and lot.origin.locked:
        rows.append(models.StockMovement(
            lot=lot, location=main, type=models.StockMovement.Type.RECEIPT,
            qty=lot.qty, source_type=lot.origin.kind, source_id=lot.origin_id,
        ))

    # движение существующего лота: единые знаковые строки `StockLine` (волна 13, Ф0)
    # свернули 4 таблицы строк-расхода (комплектация/передача/списание/отпочкование).
    # `StockLine.qty` уже со знаком (− расход); source_type/id — из документа-владельца
    # (`document.kind`/id; дуга схлопнута в один FK в Ф2b). Ф15: строка расфиксированного
    # документа тоже не двигает склад; отмена = удаление документа (каскад строк+лотов).
    for sl in lot.stock_lines.select_related('location', 'document'):
        if not sl.document.locked:
            continue
        rows.append(models.StockMovement(
            lot=lot, location=sl.location,
            type=(models.StockMovement.Type.RECEIPT if sl.qty > 0
                  else models.StockMovement.Type.ISSUE),
            qty=sl.qty, source_type=sl.document.kind, source_id=sl.document_id,
        ))

    models.StockMovement.objects.bulk_create(rows)
    return rows


def rebuild_all():
    """Пересобрать движения для всех партий (сид/тесты/детектор дрейфа)."""
    for lot in models.Lot.objects.all():
        rebuild_movements(lot)


def lot_live_qty(lot, location=None):
    """Живой остаток партии = сумма её движений (Lot.qty + Σ расход).

    Волна 13, Ф2e (мультисклад): опциональный `location` сужает до остатка партии
    **в этом месте хранения** (пара `(лот, локация)`). По умолчанию (None) — тотал по
    всем локациям, как раньше (перемещение `−q/+q` его сохраняет — двигает лишь
    распределение). Может быть отрицательным (недостача) — не клампим.
    """
    qs = lot.movements.all()
    if location is not None:
        qs = qs.filter(location=location)
    return qs.aggregate(s=Sum('qty'))['s'] or ZERO


def lot_locations(lot):
    """Разбивка остатка партии по местам хранения (волна 13, Ф2e).

    Возвращает строки `{location_id, code, name, qty}` с ненулевым остатком —
    «где физически лежит этот лот». Тотал строк == `lot_live_qty(lot)`.
    """
    rows = []
    agg = (lot.movements.values('location').annotate(q=Sum('qty'))
           .order_by('location'))
    loc_ids = [r['location'] for r in agg if r['q']]
    locs = {loc.id: loc for loc in models.Location.objects.filter(id__in=loc_ids)}
    for r in agg:
        if not r['q']:
            continue
        loc = locs.get(r['location'])
        rows.append({
            'location_id': r['location'],
            'code': loc.code if loc else '',
            'description': loc.description if loc else '',
            'qty': r['q'],
        })
    return rows


# ── Место хранения как сущность (волна 13 Ф4): что лежит на складе + правка ДНК ──
def location_stock(location):
    """Лоты с живым остатком > 0 на данном месте хранения (В13 Ф4).

    Инверсия `lot_locations` («где лежит лот») → «что лежит на этом складе», с
    проектом-владельцем каждого лота (проект — свойство лота, живёт всю жизнь).
    Агрегат движений `(лот)` на этой локации; отрицательные/нулевые прячем —
    показываем физически присутствующее.
    """
    agg = (models.StockMovement.objects.filter(location=location)
           .values('lot').annotate(q=Sum('qty')).order_by('lot'))
    lot_ids = [r['lot'] for r in agg if r['q'] and r['q'] > 0]
    lots = {lot.id: lot for lot in models.Lot.objects
            .filter(id__in=lot_ids).select_related('item', 'project', 'origin')}
    rows = []
    for r in agg:
        if not r['q'] or r['q'] <= 0:
            continue
        lot = lots.get(r['lot'])
        if lot is None:
            continue
        rows.append({
            'lot_id': lot.id, 'lot_label': _lot_label(lot),
            'part_number': lot.part_number, 'lot_name': lot.lot_name,
            'origin': lot.origin_kind,          # глиф партии (§7a): форма = откуда родилась
            'item_id': lot.item_id, 'item_code': lot.item.code,
            'item_description': lot.item.description, 'uom': lot.item.uom, 'qty': r['q'],
            'project_id': lot.project_id, 'project_code': lot.project.code,
            'project_name': lot.project.description,
        })
    return rows


def location_form(location):
    """Проекция экрана склада: ДНК (код/описание/вид) + что на нём лежит."""
    return {
        'id': location.id, 'code': location.code, 'description': location.description,
        'kind': location.kind, 'stock': location_stock(location),
    }


def create_location(code=None, description='', kind=''):
    """Завести место хранения (В13 Ф4). Код уникален (дружелюбная проверка до
    IntegrityError); пустой — фолбэком «Место 12» (Ф12e)."""
    code = (code or '').strip()
    if code and models.Location.objects.filter(code=code).exists():
        raise ValidationError('Место с таким кодом уже есть.')
    fields = dict(description=(description or '').strip(),
                  kind=(kind or '').strip())
    if code:
        return models.Location.objects.create(code=code, **fields)
    return create_with_fallback_code(models.Location, 'Место', **fields)


def update_location(location, code=None, description=None, kind=None):
    """Правка ДНК места хранения (В13 Ф4) — мутабельная, под интерфейсным замком.
    Часовые `None` (поле не передано); пустой код отклоняем.

    Ф12e: описание разрешено пустым — сущность рождается по клику незаполненной, и
    запрет очистки означал бы, что заполнить поле можно, а передумать нельзя.
    Идентичность держит `code` (он и остался обязательным)."""
    if code is not None:
        code = code.strip()
        if not code:
            raise ValidationError('Код места хранения обязателен.')
        if models.Location.objects.filter(code=code).exclude(pk=location.pk).exists():
            raise ValidationError('Место с таким кодом уже есть.')
        location.code = code
    if description is not None:
        location.description = description.strip()
    if kind is not None:
        location.kind = kind.strip()
    location.save()
    return location


def delete_location(location):
    """Удалить склад (WAVE14 Ф2). Домен: склад с движениями бережём — friendly-guard.
    Ссылки — движения (`StockMovement`) и строки движения (`StockLine`), обе PROTECT;
    пустой справочный склад сносим свободно."""
    if (models.StockMovement.objects.filter(location=location).exists()
            or models.StockLine.objects.filter(location=location).exists()):
        raise ValidationError('На складе есть движения — удаление заблокировано.')
    try:
        location.delete()
    except ProtectedError:
        raise ValidationError('Склад связан с движениями — удаление заблокировано.')


def item_available(item, project, location=None):
    """Доступный остаток Item в проекте — Σ живых остатков своих лотов.

    Волна 13, Ф2e: опциональный `location` сужает до остатка в этом месте хранения.
    Может быть отрицательным (недостача) — не клампим, это информативно.
    """
    qs = models.StockMovement.objects.filter(lot__item=item, lot__project=project)
    if location is not None:
        qs = qs.filter(location=location)
    return qs.aggregate(s=Sum('qty'))['s'] or ZERO


def item_kitted(item, project):
    """Впаяно: Σ количеств Item, ушедших в комплектации проекта (положительное число).

    Четвёртый член баланса потребности (2026-08-05). Берём расход по **зафиксированным**
    комплектациям — тот же замок, что гейтит склад: `StockMovement` собирается только по
    запертым ордерам (`rebuild_movements`), поэтому фильтра по `locked` здесь не нужно,
    он уже в природе таблицы. Черновой акт не двигает ни склад, ни это число — обе
    колонки видят одну и ту же реальность и не могут разъехаться.

    Знак: движения расхода отрицательны, наружу отдаём магнитуду — в балансе это
    слагаемое покрытия («уже стоит в изделии»), а не вычитаемое.
    """
    total = models.StockMovement.objects.filter(
        lot__item=item, lot__project=project,
        source_type=models.DocumentKind.KITTING,
        type=models.StockMovement.Type.ISSUE,
    ).aggregate(s=Sum('qty'))['s'] or ZERO
    return -total


def item_has_negative_lot(item, project):
    """Есть ли лот Item в проекте с отрицательным остатком (аномалия «подбей лоты»)."""
    for lot in models.Lot.objects.filter(item=item, project=project):
        if lot_live_qty(lot) < 0:
            return True
    return False


# --------------------------------------------------------------------------- #
#  «Заказано» (оранжевый член): открытый заказ или wip-комплектация
# --------------------------------------------------------------------------- #
def _line_received(line):
    """Поступило по строке заказа = Σ Lot.qty лотов её item по связанным приходам.

    Документ = УПД правда: поступившее — приход (`+RECEIPT` через `Receipt.purchase`),
    не текущий остаток (получили 100 → спаяли 40 → заказ закрыт на 100).

    Ф14: путь был `origin__receipt__purchase` — JOIN через детскую таблицу MTI. После
    схлопывания `purchase` живёт на самом ордере, а непоставки его не имеют вовсе
    (CHECK `doc_purchase_only_receipt`), поэтому сужать по виду отдельно не нужно.

    Ф6: считаем только **зафиксированные** поставки (`origin__locked`). Фильтра не было,
    и это расходилось с собственным обещанием Ф15 («черновая поставка не в потрачено и
    не на складе — её позиции остаются в в пути / к заказу»): черновой УПД гасил заказ,
    хотя товар не приехал. До Ф6 черновик, привязанный к заказу, был редкостью; кнопка
    «Создать УПД» делает его штатным первым шагом, и заказ показывал бы ✓ до приёмки.
    Один и тот же замок теперь гейтит склад, деньги и закрытость заказа.
    """
    return models.Lot.objects.filter(
        item=line.item, origin__purchase=line.purchase, origin__locked=True,
    ).aggregate(s=Sum('qty'))['s'] or ZERO


def _purchased_on_order(item, project):
    """Σ max(0, PurchaseLine.qty − поступившее) по открытым (sent) заказам проекта."""
    total = ZERO
    lines = models.PurchaseLine.objects.filter(
        item=item, purchase__project=project,
        purchase__locked=True,
    ).select_related('purchase')
    for line in lines:
        total += max(ZERO, line.qty - _line_received(line))
    return total


def _manufactured_in_progress(item, project):
    """Σ кол-во в производимых расфиксированных комплектациях, делающих этот Item в проекте."""
    agg = models.Kitting.objects.filter(
        target_item=item, project=project, locked=False,
    ).aggregate(s=Sum('qty'))
    return agg['s'] or ZERO


def item_on_order(item, project):
    """Оранжевый член, обобщённый по типу Item (покупной/производимый)."""
    if item.native:
        return _manufactured_in_progress(item, project)
    return _purchased_on_order(item, project)


# --------------------------------------------------------------------------- #
#  Тройной разбор строки и цвет
# --------------------------------------------------------------------------- #
def _contractor_view(cp):
    """Контрагент наружу: `code` первичен, `description` — расшифровка (2026-08-05).

    [[code-identity-principle]]: идентичность несёт код («КОМПЭЛ»), а не длинное имя
    из ЕГРЮЛ — в шапках и списках он и короче, и узнаваемее. Код у контрагента
    nullable (авто-фолбэка нет), поэтому отдаём обе строки: вью показывает код, при
    пустом падает на описание и мягко просит код завести.
    """
    if cp is None:
        return {'contractor_id': None, 'contractor_code': '', 'contractor_name': ''}
    return {'contractor_id': cp.id, 'contractor_code': cp.code or '',
            'contractor_name': cp.description}


def _coverage(need, available, on_order):
    """Разложить потребность на ✓ есть · ● заказано · ▲ заказать (сегменты).

    Сегменты **клампованы потребностью**: сумма трёх членов ровно `need`. Цена клампа —
    невидимый перебор (заказали 10 при нужде 6 → в ● всё равно 6), поэтому витрина
    потребности проекта на разборе больше НЕ стоит: она считает баланс (`_balance`), где
    перебор виден знаком. Разбор остался там, где закрытие нужды и есть вопрос: свод «К
    закупке» и призрачная строка комплектации.
    """
    have = min(need, max(ZERO, available))
    ordered = min(need - have, max(ZERO, on_order))
    to_order = need - have - ordered
    if to_order > 0:
        status = 'to_order'      # ▲ красный — нужна работа
    elif ordered > 0:
        status = 'on_order'      # ● оранжевый — запущен процесс, ждём
    else:
        status = 'available'     # ✓ зелёный — покрыто складом
    return {
        'need': need, 'have': have, 'on_order': ordered, 'to_order': to_order,
        'status': status,
    }


def _balance(need, kitted, in_stock, on_order):
    """Баланс потребности: `(впаяно + склад + в заказе) − надо`, со знаком (2026-08-05).

    Четыре члена вместо трёх сегментов, и **ни один не клампится**: числа сырые, а
    дефицит — не отдельная формула, а невязка. Отсюда главное свойство: перебор виден
    так же ясно, как недобор, просто знаком. `−4` — не хватает четырёх, `+4` — запас.

    Тон — ось «как дела», а не «что за член» (`Status` переиспользуем как словарь
    цветов, [[engine-view-seam]]: движок даёт смысл, тема выбирает знак):
    `to_order` красный — не хватает; `on_order` оранжевый — сходится в ноль, запаса нет;
    `available` зелёный — есть запас.

    Что складываем — принципиально: **впаянное** входит в покрытие наравне со складом.
    Компонент, ушедший в изделие, потребность уже закрыл, хотя на складе его нет; без
    этого члена собранный проект показывал бы дефицит на всё, что израсходовал.
    """
    balance = kitted + in_stock + on_order - need
    if balance < 0:
        status = 'to_order'
    elif balance == 0:
        status = 'on_order'
    else:
        status = 'available'
    return {
        'need': need, 'kitted': kitted, 'in_stock': in_stock, 'on_order': on_order,
        'balance': balance, 'status': status,
    }


def _supply_status(row):
    """«Как дела» по строке баланса — ось снабжения, НЕ тон колонки «Баланс».

    Два разных вопроса не смешиваем. Колонка отвечает «сходится ли» (ноль — оранжевый:
    впритык, запаса нет), а цвет прибора и проекта — «нужна ли работа»: ровно сошлось —
    это хорошо, зелёное. Отсюда: не хватает → красный; хватает, но часть ещё едет →
    оранжевый; всё на месте → зелёный.
    """
    if row['balance'] < 0:
        return 'to_order'
    return 'on_order' if row['on_order'] > 0 else 'available'


_WORST_RANK = {'to_order': 3, 'on_order': 2, 'available': 1}


def _worst_of(statuses):
    """Цвет шапки = худший из присутствующих статусов строк."""
    if not statuses:
        return 'available'
    return max(statuses, key=lambda s: _WORST_RANK[s])


def _best_of(statuses):
    """Бейдж = лучший достигнутый прогресс (для инвертированного цвета прибора)."""
    if not statuses:
        return 'available'
    return min(statuses, key=lambda s: _WORST_RANK[s])


# --------------------------------------------------------------------------- #
#  Разузлование потребности до покупных листьев (Ф5, волна 16)
# --------------------------------------------------------------------------- #
def _explode_demand(item, qty, leaves, incomplete, visiting):
    """Разузловать потребность `item × qty` до покупных листьев (структурно).

    Покупной `item` (`native=False`) — лист-терминал: копим `qty` в `leaves`.
    Производимый узел — рекурсия в детей (`qty × bl.qty`): купить его нельзя, деньги/
    заказ живут на листьях (резисторы/ИС/материалы). Нетинга подсборок здесь НЕТ
    (согласовано В16): покрытие складом/заказом считает `_coverage` на самом листе.
    Производимый узел без BOM оценить нечем → в `incomplete` (вклад 0). Циклы —
    страховка `visiting` (гасит уже `add_bom_line`), как в `_rollup_cost`.
    """
    if not item.native:
        leaves[item] = leaves.get(item, ZERO) + qty
        return
    if item.id in visiting:
        raise ValidationError(f'Цикл в составе: {item.code}.')
    visiting.add(item.id)
    lines = list(item.bom_lines.select_related('component'))
    if not lines:
        incomplete.append(item.code)
    for bl in lines:
        _explode_demand(bl.component, qty * bl.qty, leaves, incomplete, visiting)
    visiting.discard(item.id)


def project_leaf_demand(project):
    """Свод потребности проекта до покупных листьев: `({leaf: qty}, incomplete)`.

    Разузловывает каждую потребность (`target_item × demand.qty`) до листьев и
    складывает по Item через все приборы проекта. `incomplete` — производимые узлы
    без BOM (оценить нечем), для честного флага неполноты.
    """
    leaves, incomplete = {}, []
    for demand in project.demands.select_related('target_item'):
        _explode_demand(demand.target_item, demand.qty, leaves, incomplete, set())
    return leaves, incomplete


def _project_usage_map(project):
    """Обратное разузлование проекта: `{leaf_item_id: [применения]}` — за один обход.

    Прямое разузлование отвечает «что купить», обратное — «зачем»: по сколько штук
    этого Item уходит в каждое производимое изделие проекта. Считается той же
    рекурсией (`_explode_demand`), просто пущенной от **единицы** прибора: разложение
    `target_item × 1` и есть норма на изделие (`per_unit`), а `per_unit × demand.qty` —
    вклад этой потребности в общую нужду. Третьего прохода по BOM не заводим.

    Карта строится на весь проект разом (обходов = потребностей, а не потребностей ×
    строк плана) — аккордеон закупки спрашивает её по каждой строке.
    """
    usage = {}
    for demand in project.demands.select_related('target_item'):
        leaves = {}
        _explode_demand(demand.target_item, Decimal('1'), leaves, [], set())
        for leaf, per_unit in leaves.items():
            usage.setdefault(leaf.id, []).append({
                'target_item_id': demand.target_item_id,
                'target_code': demand.target_item.code,
                'target_description': demand.target_item.description,
                'per_unit': per_unit,               # норма на одно изделие (через все уровни)
                'demand_qty': demand.qty,           # сколько изделий нужно проекту
                'total': per_unit * demand.qty,     # вклад в нужду проекта по этому Item
            })
    for rows in usage.values():
        rows.sort(key=lambda r: r['target_code'])
    return usage


def item_usage_in_project(item, project):
    """«Куда идёт этот Item в проекте»: список применений (изделие, норма, вклад).

    Точечный вход в `_project_usage_map` — для одной строки; массовые витрины берут
    карту целиком, чтобы не гонять разузлование на каждую строку.
    """
    return _project_usage_map(project).get(item.id, [])


def _leaf_row(item, need, project):
    """Строка баланса листа (`_balance` + реквизиты Item) для секций потребности.

    Ключи `component_*` сохранены (фронт/сериализатор их ждут): лист — тот же Item.
    """
    row = _balance(need, item_kitted(item, project), item_available(item, project),
                   item_on_order(item, project))
    row.update({
        'component_id': item.id,
        'component_code': item.code,
        'component_description': item.description,
        'component_native': item.native,       # глиф строки по режиму (Ф3a): native→замок,
        'component_synced': item.synced,        # not native→sync (зел. библ. / оранж. ручной)
        'component_locked': item.locked,
        'uom': item.uom,
        'supply': _supply_status(row),         # ось «как дела» (цвет прибора/проекта)
        'anomaly': item_has_negative_lot(item, project),
    })
    return row


def _demand_tree(item, qty, project, depth, visiting, out):
    """Плоский pre-order список узлов BOM с `depth` — древовидный аккордеон прибора
    (Ф5b, В16): видно вложенность (прибор → подсборка → лист), а не только листья.

    Покупной лист — строка баланса (`_leaf_row`) + `depth`/`is_leaf=True`. Производимый
    узел — структурная строка (купить нельзя): только `need` и дети рекурсией ниже. Цикл
    — страховка `visiting`.

    Статуса у узла-подсборки больше НЕТ (2026-08-05): он существовал ровно ради цветной
    полосы слева («где горит под свёрнутым узлом»), а полосы сняты во всём Plume. Считать
    worst-of поддерева, чтобы никто не смотрел, — мёртвая работа."""
    if not item.native:
        row = _leaf_row(item, qty, project)
        row['need'] = qty
        row['depth'] = depth
        row['is_leaf'] = True
        out.append(row)
        return
    if item.id in visiting:
        raise ValidationError(f'Цикл в составе: {item.code}.')
    visiting.add(item.id)
    row = {
        'component_id': item.id,
        'component_code': item.code,
        'component_description': item.description,
        'component_native': item.native,       # глиф строки по режиму (Ф3a)
        'component_synced': item.synced,
        'component_locked': item.locked,
        'uom': item.uom,
        'need': qty,
        'depth': depth,
        'is_leaf': False,
    }
    out.append(row)          # pre-order: узел раньше детей
    for bl in item.bom_lines.select_related('component'):
        _demand_tree(bl.component, qty * bl.qty, project, depth + 1, visiting, out)
    visiting.discard(item.id)


# --------------------------------------------------------------------------- #
#  Дефицит проекта (главная проекция волны 1)
# --------------------------------------------------------------------------- #
def project_deficit(project, with_tree=True):
    """Дефицит проекта: по каждой потребности — прибор и его состав (Ф5b: аккордеон —
    дерево BOM со всеми уровнями; свод «Потребность» — плоско по покупным листьям).

    «ДО» купить нельзя — деньги/дефицит на листьях; но структуру (прибор → подсборка →
    лист) видно в дереве. Возвращает структуру под сериализацию (Decimal → строки в DRF).

    `with_tree=False` — пропустить построение дерева-аккордеона (нужно только детальной
    форме): `_project_health` считает worst-of без него, чтобы не гонять лишнее в списке.
    """
    demands = []
    # Свод потребности по листьям на весь проект (секция «Потребность»): каждый прибор
    # разузловываем до покупных листьев, копим Σ. Склад/заказано — общие по листу.
    need_by_leaf = {}          # leaf Item → суммарная потребность (Decimal)
    for demand in project.demands.select_related('target_item'):
        target = demand.target_item
        # Свод листьев этой потребности — для секции «Потребность» и цвета прибора.
        leaves, _incomplete = {}, []
        _explode_demand(target, demand.qty, leaves, _incomplete, set())
        statuses = []
        for leaf, need in leaves.items():
            need_by_leaf[leaf] = need_by_leaf.get(leaf, ZERO) + need
            statuses.append(_supply_status(
                _balance(need, item_kitted(leaf, project), item_available(leaf, project),
                         item_on_order(leaf, project))))
        # Дерево BOM для аккордеона (все уровни, pre-order + depth). Для цвета проекта
        # (`_project_health`) дерево не нужно — пропускаем, чтобы не гонять лишнее в списке.
        tree = []
        if with_tree:
            visiting = {target.id}
            for bl in target.bom_lines.select_related('component'):
                _demand_tree(bl.component, demand.qty * bl.qty, project, 0, visiting, tree)

        # триплет прибора: готово (зафиксированные лоты) / делается (в работе) / не начато
        done = models.StockMovement.objects.filter(
            lot__item=target, lot__project=project,
            lot__origin__kind=models.StockDocument.Kind.KITTING,
            lot__origin__locked=True,
        ).aggregate(s=Sum('qty'))['s'] or ZERO
        wip = _manufactured_in_progress(target, project)
        not_started = max(ZERO, demand.qty - done - wip)

        demands.append({
            'demand_id': demand.id,
            'target_id': target.id,
            'target_code': target.code,
            'target_description': target.description,
            'target_native': target.native,     # глиф строки по режиму (Ф3a): прибор → замок
            'target_synced': target.synced,
            'target_locked': target.locked,
            'qty': demand.qty,
            'device': {'done': done, 'wip': wip, 'not_started': not_started},
            # цвет прибора: worst-of листьев (внимание) + бейдж лучшего прогресса
            'status': _worst_of(statuses),
            'badge': _best_of(statuses),
            'tree': tree,
        })

    # Сводная таблица по листьям (полная покупная картина проекта, всегда видна).
    components = [_leaf_row(leaf, need, project) for leaf, need in need_by_leaf.items()]
    # «Горит вперёд» по тону баланса: сначала недобор, затем сошедшиеся в ноль, затем
    # строки с запасом; внутри — по коду.
    components.sort(key=lambda c: (-_WORST_RANK[c['status']], c['component_code']))

    return {
        'project_id': project.id,
        'project_code': project.code,
        'project_name': project.description,
        'demands': demands,
        'components': components,
    }


def project_health(project):
    """Цвет проекта в списке (Ф1b): worst-of здоровья, вычисляемая проекция (как дефицит —
    НЕ храним в БД). Семантика: `to_order` (красный) — хоть что-то не заказано; `on_order`
    (оранжевый) — хоть что-то в пути ИЛИ прибор не собран; `available` (зелёный) — всё
    приехало и собрано. Внутренние проекты и пустые (без потребностей) — `None` (дефицит
    неприменим, глиф нейтрален). Лёгкая версия дефицита (`with_tree=False`)."""
    if project.kind != models.Project.Kind.EXTERNAL:
        return None
    d = project_deficit(project, with_tree=False)
    if not d['demands']:
        return None
    # Здоровье смотрит на ось снабжения (`supply`), а не на тон колонки «Баланс»:
    # сошедшийся в ноль проект здоров, хотя в колонке он оранжевый (запаса нет).
    comp = _worst_of([c['supply'] for c in d['components']]) if d['components'] else 'available'
    assembled = all(dm['device']['not_started'] <= 0 and dm['device']['wip'] <= 0
                    for dm in d['demands'])
    if comp == 'to_order':
        return 'to_order'
    if comp == 'on_order' or not assembled:
        return 'on_order'
    return 'available'


def purchase_coverage(purchase):
    """Цвет заказа в списке (Ф1b): покрытие строк лотами приходов. `to_order` (красный) —
    ни одного лота; `on_order` (оранжевый) — часть пришла; `available` (зелёный) — все
    строки закрыты остатком. Пустой заказ — `to_order` (ничего нет). Из `_line_received`
    (без разузлования — дёшево)."""
    lines = list(purchase.lines.all())
    if not lines:
        return 'to_order'
    total_received = ZERO
    fully = True
    for line in lines:
        received = _line_received(line)
        total_received += received
        if received < line.qty:
            fully = False
    if fully:
        return 'available'
    return 'on_order' if total_received > ZERO else 'to_order'


# --------------------------------------------------------------------------- #
#  Бюджет проекта: два числа денег + себестоимость/экономия (north-star окупаемости)
# --------------------------------------------------------------------------- #
def _project_spent(project):
    """Потрачено (факт) = Σ(unit_cost×qty) приходных (`Receipt`) лотов проекта.

    Точная застывшая цифра «кэша ФЛС». Заёмные/свои бесплатные лоты (origin
    requisition/inventory/kitting) сюда не входят → бесплатны в бюджете (платил
    источник). Только покупные материалы — снимок цены собранного узла (лот-прибор
    из Kitting) в бюджет не складываем (иначе двойной счёт).

    Волна 19, Ф15 (решение Ивана): считаем только лоты **зафиксированных** поставок.
    Это единственное место, где деньги читались с лотов напрямую, мимо движений, —
    поэтому гейт склада его бы не накрыл, а черновой УПД двигал бы бюджет проекта
    всем пользователям. Черновая поставка не в «потрачено» и не на складе: её
    позиции остаются в «в пути / к заказу» — картина связная.
    """
    total = ZERO
    for lot in project.lots.filter(origin__kind=models.StockDocument.Kind.RECEIPT,
                                   origin__locked=True):
        total += lot.qty * lot.unit_cost
    return total


def _project_estimate(project):
    """Прогнозная стоимость ещё-не-полученного покупного материала (по estimated_cost).

    Ф5 (В16): потребность разузловываем до покупных **листьев** (`project_leaf_demand`),
    затем один разбор `_coverage` на лист. Считаем только листья — производимые узлы
    разузловываются насквозь, их роллап-`estimated_cost` в план НЕ складываем (иначе
    двойной счёт: цена узла = Σ его листьев). Оцениваем «в пути + к заказу» (● + ▲) —
    то, что ещё потребует денег; ✓ уже покрыто (склад приходной = в «потрачено», заём =
    бесплатно). Возвращает `(estimate, unestimated_codes)` — сумма оценки и коды
    покупных листьев без `estimated_cost` (план по ним неполон, не молчим 0).
    """
    leaves, _incomplete = project_leaf_demand(project)
    estimate = ZERO
    unestimated = []
    for leaf, need in leaves.items():
        cov = _coverage(need, item_available(leaf, project),
                        item_on_order(leaf, project))
        remaining = cov['on_order'] + cov['to_order']
        if remaining <= 0:
            continue
        if leaf.estimated_cost is None:
            unestimated.append(leaf.code)
            continue
        estimate += remaining * leaf.estimated_cost
    return estimate, unestimated


def _project_cost(project):
    """Себестоимость проекта = Σ(qty×снимок) по лотам-приборам закрытых комплектаций,
    чьё изделие — цель потребности проекта (только верхние приборы, без задвоения
    подсборок: их цена уже в снимке верхнего прибора).

    Снимок `unit_cost` лота-прибора взят на закрытии (`_device_unit_cost`) и включает
    заёмные компоненты по реальной цене (Requisition-лот наследует цену предка) —
    честная цена для КП.
    """
    targets = {d.target_item_id for d in project.demands.all()}
    if not targets:
        return ZERO
    total = ZERO
    lots = project.lots.filter(
        origin__kind=models.StockDocument.Kind.KITTING,
        origin__locked=True, item_id__in=targets,
    )
    for lot in lots:
        total += lot.qty * lot.unit_cost
    return total


def project_budget(project):
    """Проекция бюджета проекта (north-star окупаемости линзы).

    Два числа денег (не путать): **потрачено** (факт по `Receipt`-лотам) и **план**
    (прогноз «факт где есть, оценка где нет»). Компас `budget − план` = запас/
    перерасход. **Себестоимость** (честная цена, заём по реальной цене) и **экономия**
    = себестоимость − потрачено (оцифрованная польза внутреннего заёма = польза PLM).
    """
    spent = _project_spent(project)
    estimate, unestimated = _project_estimate(project)
    plan = spent + estimate
    cost = _project_cost(project)
    budget = project.budget
    return {
        'project_id': project.id,
        'project_code': project.code,
        'project_name': project.description,
        'budget': budget,                       # может быть None
        'spent': spent,                         # потрачено (факт)
        'plan': plan,                           # прогноз полной стоимости
        'compass': (budget - plan) if budget is not None else None,
        'unestimated': unestimated,             # покупные позиции без оценки
        'cost': cost,                           # себестоимость (для КП)
        'economy': cost - spent,                # экономия = польза заёма
    }


# --------------------------------------------------------------------------- #
#  Карта остатков по складам-проектам (north-star)
# --------------------------------------------------------------------------- #
def stock_map(item):
    """Где этот Item лежит по всем складам-проектам, с доступным qty.

    Переносит знание «у кого что есть» из головы в БД. Авто-зачёта между
    проектами нет — это справка для решения «что закупить».
    """
    rows = []
    project_ids = models.Lot.objects.filter(item=item).values_list(
        'project_id', flat=True).distinct()
    for pid in project_ids:
        project = models.Project.objects.get(id=pid)
        available = item_available(item, project)
        if available == 0:
            continue
        # Волна 13, Ф2e (мультисклад): аддитивная разбивка остатка проекта по местам
        # хранения (пары `(лот, локация)` свёрнуты по локации). Ключ новый — фронт его
        # пока игнорирует (вьюхи потом); строки с нулём не показываем.
        loc_agg = (models.StockMovement.objects
                   .filter(lot__item=item, lot__project=project)
                   .values('location', 'location__code', 'location__description')
                   .annotate(q=Sum('qty')).order_by('location'))
        by_location = [
            {'location_id': r['location'], 'code': r['location__code'],
             'description': r['location__description'], 'available': r['q']}
            for r in loc_agg if r['q']
        ]
        rows.append({
            'project_id': project.id,
            'project_code': project.code,
            'project_name': project.description,
            'project_kind': project.kind,
            'available': available,
            'by_location': by_location,
        })
    # подсказка-порядок: белый → серый — мягкая сортировка по виду, потом по коду
    kind_rank = {
        models.Project.Kind.INTERNAL_STOCK: 0,
        models.Project.Kind.EXTERNAL: 1,
        models.Project.Kind.INTERNAL_WRITEOFF: 2,
    }
    rows.sort(key=lambda r: (kind_rank.get(r['project_kind'], 1), r['project_code']))
    return {
        'item_id': item.id,
        'item_code': item.code,
        'item_description': item.description,
        'uom': item.uom,
        'rows': rows,
    }


# --------------------------------------------------------------------------- #
#  Форма комплектации (волна 2): реальные строки + призрачные строки
# --------------------------------------------------------------------------- #
def available_lots(item, project, location=None):
    """Лоты item в проекте с живым остатком > 0 — кандидаты под пайку.

    Волна 13, Ф2e: опциональный `location` сужает до лотов с остатком > 0 **в этом
    месте хранения** (пикер под конкретную локацию). По умолчанию — как раньше
    (остаток по всем локациям), контракт формы комплектации байт-в-байт.
    """
    result = []
    for lot in models.Lot.objects.filter(item=item, project=project).select_related('item'):
        live = lot_live_qty(lot, location)
        if live > 0:
            result.append({
                'lot_id': lot.id, 'live_qty': live, 'unit_cost': lot.unit_cost,
                'part_number': lot.part_number,
                'origin': lot.origin_kind, 'lot_name': lot.lot_name,
            })
    return result


def _require_native_target(item):
    """«Комплектуем только своё» (решение Ивана 2026-07-31): целью комплектации может
    быть лишь наше изделие. Дружелюбный гейт входного слоя; страховка — в
    `StockDocument.clean` (CHECK на чужую таблицу MySQL не умеет, см. модель)."""
    if item is not None and not item.native:
        raise ValidationError(models.KITTING_TARGET_NATIVE)


def create_kitting(project, user, target_item=None, qty=None, date=None):
    """Создать комплектацию — рождение переехало из вьюхи в движок (Ф12e).

    `target_item`/`qty` необязательны при рождении: их требует ФИКСАЦИЯ (CHECK
    `doc_locked_kitting_has_target`), потому что до выбора прибора форма всё равно
    пуста — BOM разузловать не от чего. Номера у комплектации нет
    (`REQUIRED_HEADER_BY_KIND[KITTING] = ()`), но `code` она получает общий."""
    _require_native_target(target_item)
    return _born_order(models.Kitting, project, user, date=date,
                       target_item=target_item, qty=qty)


def kitting_form(kitting):
    """Проекция формы сборки (1 уровень BOM целевого прибора).

    Каждая строка BOM: `надо = bom.qty × kitting.qty`, пробитые `KittingLine`
    (реальные зелёные строки) и остаток → **призрачная строка**, покрашенная по
    доступности компонента в проекте (`_coverage`, тот же словарь ✓/●/▲) с лотами-
    кандидатами под пайку. Ничего не хранит — чистая проекция.
    """
    # Ф12e: комплектация рождается по клику, прибор-цель выбирают уже в форме —
    # до выбора разузловывать нечего, и это не ошибка, а нормальный черновик.
    target = kitting.target_item
    project = kitting.project
    rows = []
    statuses = []
    is_wip = not kitting.locked
    # `qty` черновика может быть пустым (Ф12e: рождение по клику не спрашивает ничего),
    # а состав у цели уже есть — проекция обязана это пережить. Раньше здесь падало
    # `Decimal * None` → 500 ровно на выборе НАШЕГО изделия (у покупного состава нет,
    # цикл не заходил, и баг выглядел вывернутым наизнанку). Пустое кол-во = «потребность
    # ещё не задана» → надо 0, призраков нет; выбор цели её тут же и проставляет
    # (`_set_target_item`), так что состояние мимолётное.
    for bl in (target.bom_lines.select_related('component') if target else ()):
        component = bl.component
        need = bl.qty * (kitting.qty or ZERO)
        real_lines = []
        pierced = ZERO
        # компонент строки выводится из lot.item (StockLine его не хранит);
        # qty знаковый — наружу отдаём магнитуду (положительную), проекция без изменений.
        for kl in kitting.lines.filter(lot__item=component).select_related('lot'):
            mag = -kl.qty
            pierced += mag
            real_lines.append({
                'id': kl.id, 'lot_id': kl.lot_id,
                'lot_label': f'#{kl.lot_id} {kl.lot.lot_name or component.code}',
                'qty': mag, 'date': kl.date,
            })
        remaining = max(ZERO, need - pierced)
        ghost = None
        if remaining > 0 and is_wip:
            cov = _coverage(remaining, item_available(component, project),
                            item_on_order(component, project))
            ghost = {
                'status': cov['status'], 'have': cov['have'],
                'on_order': cov['on_order'], 'to_order': cov['to_order'],
                'candidate_lots': available_lots(component, project),
            }
            statuses.append(cov['status'])
        rows.append({
            'component_id': component.id, 'component_code': component.code,
            'component_description': component.description, 'uom': component.uom,
            # Оси компонента — под глиф строки (§7a), как в заказе и закупке-плане:
            # форма = изделие/компонент, цвет = покрытие строки складом проекта.
            'component_native': component.native, 'component_synced': component.synced,
            'component_locked': component.locked,
            'need': need, 'pierced': pierced, 'remaining': remaining,
            'real_lines': real_lines, 'ghost': ghost,
        })
    born_lots = [
        {'id': lot.id, 'qty': lot.qty, 'unit_cost': lot.unit_cost,
         'lot_name': lot.lot_name, 'part_number': lot.part_number}
        for lot in kitting.lots.all()
    ]
    return {
        'id': kitting.id, **_author(kitting), 'locked': kitting.locked,
        'code': kitting.code, 'description': kitting.description,
        'project_id': project.id, 'project_code': project.code,
        'target_id': target.id if target else None,
        'target_code': target.code if target else '',
        'target_description': target.description if target else '',
        'uom': target.uom if target else '',
        'qty': kitting.qty, 'date': kitting.date,
        'worst_status': _worst_of(statuses),   # worst-of призрачных строк
        'rows': rows,
        'born_lots': born_lots,   # рождённые лоты-приборы (после закрытия)
    }


# --------------------------------------------------------------------------- #
#  Мутации формы (единый источник правил + пересборка проекции склада)
# --------------------------------------------------------------------------- #
def add_kitting_line(kitting, component, lot, qty, location=None, date=None):
    """Пайка: промоушн призрачной строки в реальную `KittingLine` + `-ISSUE`."""
    _require_unlocked(kitting)
    if lot.item_id != component.id:
        raise ValidationError('Лот не соответствует компоненту строки.')
    if lot.project_id != kitting.project_id:
        raise ValidationError('Лот из другого проекта (заём — отдельным требованием).')
    if qty is None or qty <= 0:
        raise ValidationError('Количество пайки должно быть положительным.')
    line = models.StockLine.objects.create(
        document=kitting, lot=lot,
        location=location or _main_location(), qty=-qty, date=date,
    )
    rebuild_movements(lot)
    return line


def update_kitting_line(line, qty):
    """Автосейв количества пайки (правка провизорной строки до замка)."""
    _require_unlocked(line.document)
    if qty is None or qty <= 0:
        raise ValidationError('Количество пайки должно быть положительным.')
    line.qty = -qty                      # знаковая строка (− расход)
    line.save(update_fields=['qty'])
    rebuild_movements(line.lot)


def remove_kitting_line(line):
    """Удалить пробитую строку (коррекция до замка) + пересобрать движения лота."""
    _require_unlocked(line.document)
    lot = line.lot
    line.delete()
    rebuild_movements(lot)


def _device_unit_cost(kitting):
    """Снимок себестоимости прибора на закрытии = Σ(qty×цена лотов) / кол-во."""
    total = ZERO
    for kl in kitting.lines.select_related('lot'):
        total += -kl.qty * kl.lot.unit_cost     # qty знаковый (− расход) → магнитуда
    if kitting.qty and kitting.qty != ZERO:
        return (total / kitting.qty).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    return ZERO


def lock_kitting(kitting):
    """Закрыть комплектацию: рождается лот-прибор (`+RECEIPT`), замок ставится.

    Ф15 сделала фиксацию единым событием склада: вместе с рождением прибора
    материализуется и **пайка** (`−ISSUE` компонентов) — до замка её строки склад не
    двигали. Раньше комплектация была наполовину по правилу: лот ждал замка, а
    компоненты списывались сразу.
    """
    _require_unlocked(kitting)
    # Ф12e: комплектация рождается без прибора-цели — гейт здесь (своей дорогой
    # мимо `lock_document`, поэтому зовём явно).
    _require_header(kitting)
    if kitting.lots.exists():
        raise ValidationError('У комплектации уже есть рождённый лот-прибор.')
    models.Lot.objects.create(
        item=kitting.target_item, project=kitting.project, origin=kitting,
        qty=kitting.qty, unit_cost=_device_unit_cost(kitting),
    )
    kitting.locked = True
    kitting.save(update_fields=['locked'])
    rebuild_document_movements(kitting)     # лот-прибор + пайка компонентов
    return kitting.lots.first()


def unlock_kitting(kitting):
    """Расфиксировать комплектацию: снять лот-прибор + отпустить пайку (Ф15).

    Guard: лот-прибор не должен быть потреблён/передан/отпочкован ниже.
    """
    if not kitting.locked:
        raise ValidationError('Расфиксировать можно только зафиксированную комплектацию.')
    for lot in kitting.lots.all():
        if _lot_consumed_downstream(lot):
            raise ValidationError(
                'Лот-прибор уже потреблён/передан ниже — переоткрытие заблокировано.')
    for lot in kitting.lots.all():
        lot.movements.all().delete()
        lot.delete()
    kitting.locked = False
    kitting.save(update_fields=['locked'])
    rebuild_document_movements(kitting)     # снять `−ISSUE` компонентов


# --------------------------------------------------------------------------- #
#  Форма прихода / УПД (волна 3): строки-лоты, рождение +RECEIPT, мягкий замок
# --------------------------------------------------------------------------- #
def _lot_consumed_downstream(lot):
    """Потреблён ли лот ниже: расход/пайка/передача/списание/отпочкование/успех.

    Общий guard: до `PROTECT`-ошибки БД даём дружелюбный отказ на разрушающую
    правку (удаление строки прихода, переоткрытие комплектации).
    """
    return (lot.movements.filter(qty__lt=0).exists() or lot.successors.exists()
            or lot.stock_lines.exists())


def create_receipt(project, user, number='', date=None, contractor=None):
    """Создать поставку (УПД) — рождение переехало из вьюхи в движок (Ф12e).

    Раньше `views.receipts` строила `Receipt` сырым `objects.create` и требовала
    `contractor_id` + непустой `number`: артефакт формы создания. Оба поля теперь
    обязательны к ФИКСАЦИИ (`REQUIRED_HEADER_BY_KIND` + CHECK
    `doc_locked_receipt_has_contractor`), а не к рождению."""
    return _born_order(models.Receipt, project, user, number, date,
                       contractor=contractor)


def receipt_form(receipt):
    """Проекция формы прихода: шапка УПД + строки-лоты (каждая строка = Lot).

    В модели отдельной `ReceiptLine` нет — строки прихода это его лоты. Каждый лот
    показывает рождённое кол-во, живой остаток (просел ли под пайку), цену и
    название из УПД. Ничего не хранит — чистая проекция.
    """
    lots = []
    total = ZERO
    for lot in receipt.lots.select_related('item').order_by('id'):
        total += lot.qty * lot.unit_cost
        lots.append({
            'id': lot.id, 'item_id': lot.item_id, 'item_code': lot.item.code,
            'item_description': lot.item.description, 'uom': lot.item.uom,
            'qty': lot.qty, 'live_qty': lot_live_qty(lot),
            'unit_cost': lot.unit_cost, 'lot_name': lot.lot_name,
            'part_number': lot.part_number,
            'consumed': _lot_consumed_downstream(lot),
        })
    return {
        'id': receipt.id, **_author(receipt), 'number': receipt.number, 'date': receipt.date,
        'code': receipt.code, 'description': receipt.description,
        # Ф14: колонка стала nullable (обязательность — на фиксации, не на рождении),
        # поэтому проекция больше не опирается на прежний NOT NULL.
        **_contractor_view(receipt.contractor),
        # Ф17: «кто привёз» ≠ «у кого купили» — расхождение с контрагентом заказа
        # считает движок, форма только рисует оранжевый глиф (не гейт).
        'contractor_mismatch': _contractor_mismatch(receipt.contractor_id,
                                                    receipt.purchase),
        'project_id': receipt.project_id, 'project_code': receipt.project.code,
        'project_name': receipt.project.description,
        'purchase_id': receipt.purchase_id,   # связанный заказ (закрытие строк)
        'locked': receipt.locked, 'total_cost': total,
        'lots': lots,
    }


def add_receipt_lot(receipt, item, qty, unit_cost=ZERO, lot_name='',
                    part_number=''):
    """Добавить строку УПД: рождается партия (`+RECEIPT`) в проекте прихода."""
    _require_unlocked(receipt)
    if qty is None or qty <= 0:
        raise ValidationError('Количество прихода должно быть положительным.')
    if unit_cost is not None and unit_cost < 0:
        raise ValidationError('Цена не может быть отрицательной.')
    lot = models.Lot.objects.create(
        item=item, project=receipt.project, origin=receipt, qty=qty,
        unit_cost=unit_cost or ZERO, lot_name=lot_name or '',
        part_number=part_number or '',
    )
    rebuild_movements(lot)
    return lot


def update_receipt_lot(lot, qty=None, unit_cost=None, lot_name=None,
                       part_number=None):
    """Автосейв строки УПД (кол-во/цена/название/PN). Правка до замка.

    Кол-во не клампим по потреблению: уронить ниже списанного можно — живой остаток
    уйдёт в минус (недостача информативнее, в духе мутабельной ДНК).
    """
    _require_unlocked(lot.origin)
    fields = []
    if qty is not None:
        if qty <= 0:
            raise ValidationError('Количество прихода должно быть положительным.')
        lot.qty = qty
        fields.append('qty')
    if unit_cost is not None:
        if unit_cost < 0:
            raise ValidationError('Цена не может быть отрицательной.')
        lot.unit_cost = unit_cost
        fields.append('unit_cost')
    if lot_name is not None:
        lot.lot_name = lot_name
        fields.append('lot_name')
    if part_number is not None:
        lot.part_number = part_number
        fields.append('part_number')
    if fields:
        lot.save(update_fields=fields)
        rebuild_movements(lot)
    return lot


def remove_receipt_lot(lot):
    """Удалить строку УПД (до замка). Guard: лот не потреблён ниже."""
    _require_unlocked(lot.origin)
    if _lot_consumed_downstream(lot):
        raise ValidationError(
            'Партия уже потреблена ниже (пайка/передача) — удаление заблокировано.')
    lot.movements.all().delete()
    lot.delete()


def lock_receipt(receipt):
    """Поставить замок «сверено со сканом» — форма прихода read-only, партии УПД
    ложатся на склад (Ф15). Ф15 заодно сняла рукописную копию `lock_document`:
    поставка ходит общей дорогой, как остальные пять ордеров."""
    return lock_document(receipt, receipt.lots,
                         'Нельзя сверить пустой приход — добавьте строку.')


def unlock_receipt(receipt):
    """Снять замок — снова разрешить правку (партии УПД уходят со склада, Ф15).
    Ничего не разрушает (в отличие от переоткрытия комплектации), поэтому guard по
    потомкам не нужен."""
    return unlock_document(receipt)


def set_receipt_purchase(receipt, purchase):
    """Связать приход с заказом (гасит строки заказа) или отвязать (`None`).

    Один заказ закрывается одним/несколькими приходами через `Receipt.purchase`.
    Приход закрывает только заказ **своего проекта** (чистота «один УПД ↔ один проект»).
    Лоты не двигает — `rebuild_movements` не нужен; на «заказано» влияет через
    `_line_received` (проекция).
    """
    if purchase is not None and purchase.project_id != receipt.project_id:
        raise ValidationError(
            'Заказ другого проекта — приход закрывает только заказ своего проекта.')
    receipt.purchase = purchase
    receipt.save(update_fields=['purchase'])
    return receipt


PURCHASE_DRAFT_NO_RECEIPT = (
    'УПД создаётся по зафиксированному заказу — сперва зафиксируйте заказ.')
PURCHASE_CLOSED_NO_RECEIPT = (
    'Заказ закрыт полностью — по нему нечего принимать.')


def create_receipt_from_purchase(purchase, user, number='', date=None):
    """Родить черновик УПД по заказу: строки — ОСТАТКОМ заказа (волна 19, Ф6).

    В 98% случаев накладная повторяет заказ 1:1, и набивать строки заново незачем.
    Остаток (`qty − уже получено`, тот же `_line_received`, что красит форму заказа)
    вместо полного количества даёт «поставку частями» бесплатно: повторный вызов на
    частично закрытом заказе рождает вторую накладную ровно на недостающее.

    Поставщик — `purchase.contractor` (волна 19, Ф17): «кто привёз» наследуется от «у
    кого купили», копией при рождении. До Ф17 он читался сквозь цепочку
    `purchase.procurement.contractor` (Р3, отменена), и заказ без плана поставщика не
    знал вовсе. Заказ проставляется сразу (`purchase`), поэтому руками связывать в форме
    УПД больше не нужно.

    **Цена ставится в 0, а не в `item.estimated_cost`** (решение Ивана 2026-07-29):
    `Lot` по канону хранит цену ИЗ УПД, а `_project_spent` считает факт трат по
    `unit_cost` зафиксированных приходных лотов. Правдоподобная оценка, забытая при
    фиксации, молча стала бы фактом бюджета; ноль виден глазом (итог поставки 0 ₽).
    Имя лота преднабиваем описанием изделия — его пользователь и правит по УПД.

    Всё рождённое — черновик: до фиксации на складе ничего не лежит (Ф15).
    """
    if not purchase.locked:
        raise ValidationError(PURCHASE_DRAFT_NO_RECEIPT)
    # Черновая накладная заказ не гасит (Ф6-правка `_line_received`), поэтому повторный
    # клик молча плодил бы её близнеца. Отправляем в уже заведённую: «частями» — это
    # ЗАФИКСИРОВАТЬ первую, а потом жать снова на остаток.
    open_receipt = models.Receipt.objects.filter(purchase=purchase, locked=False).first()
    if open_receipt is not None:
        raise ValidationError(
            f'По заказу уже заведена черновая поставка {open_receipt.code or open_receipt.number}'
            ' — заполните и зафиксируйте её; за остатком вернётесь сюда.')
    remaining = [
        (line, line.qty - _line_received(line))
        for line in purchase.lines.select_related('item').order_by('id')
    ]
    remaining = [(line, rest) for line, rest in remaining if rest > 0]
    if not remaining:
        raise ValidationError(PURCHASE_CLOSED_NO_RECEIPT)

    receipt = _born_order(
        models.Receipt, purchase.project, user, number, date,
        contractor=purchase.contractor, purchase=purchase)
    for line, rest in remaining:
        add_receipt_lot(receipt, line.item, rest, unit_cost=ZERO,
                        lot_name=line.item.description)
    return receipt


# --------------------------------------------------------------------------- #
#  Форма заказа / Purchase (волна 4): строки-обязательства + гашение приходом
# --------------------------------------------------------------------------- #
def create_purchase(project, user, date=None, code=None, description=''):
    """Создать заказ проекта (черновик) — **без закупки-плана**. Пустой код — фолбэком
    «Заказ 12» (Ф12e).

    Волна 19, Ф17: `procurement` стал nullable, и заказ рождается с `None`. Прежний
    `_solo_procurement` плодил закупку-пустышку на каждый одиночный заказ только чтобы
    удовлетворить NOT NULL, а список закупок прятал эти пустышки эвристикой
    (`_plan_procurements`) — обе конструкции удалены вместе. План выбирают в форме, если
    он есть; контрагента — тоже (он обязателен к фиксации, `lock_purchase`).
    """
    require_unique_code(models.Purchase, code)
    p = models.Purchase.objects.create(
        project=project, user=user,
        locked=False, date=date, code=code, description=description or '')
    return p if code else fallback_code(p, 'Заказ')


def _purchase_closure(purchase):
    """`{item_id: [{receipt_id, number, date, qty}]}` — чем закрыта каждая строка (Ф6).

    Источник — лоты приходов заказа (`Lot.origin → Receipt.purchase`); сопоставление
    по изделию, отдельного FK «строка заказа ↔ лот» не заводим: `PurchaseLine` уникален
    по `(purchase, item)`, поэтому атрибуция однозначна. Несколько лотов одного изделия
    в одной накладной (разные партии одной позиции УПД) сворачиваются в одну строку —
    пользователь спрашивает «какой накладной закрыто», а не «сколькими партиями».

    Заодно это ИСТОЧНИК `received` для формы: та же выборка, что у `_line_received`
    (включая гейт `origin__locked` — закрывают заказ только зафиксированные накладные),
    но одним запросом на заказ вместо запроса на строку.
    """
    out = {}
    lots = (models.Lot.objects.filter(origin__purchase=purchase, origin__locked=True)
            .select_related('origin').order_by('origin__date', 'origin_id', 'id'))
    for lot in lots:
        rows = out.setdefault(lot.item_id, [])
        for r in rows:
            if r['receipt_id'] == lot.origin_id:
                r['qty'] += lot.qty
                break
        else:
            rows.append({'receipt_id': lot.origin_id, 'number': lot.origin.number,
                         'date': lot.origin.date, 'qty': lot.qty})
    return out


def _contractor_mismatch(own_contractor_id, parent):
    """Расхождение контрагента с уровнем НАД ним (волна 19, Ф17) — флаг для формы.

    Контрагент есть у всех трёх уровней контура (закупка = ЧТО купить, заказ = У КОГО,
    поставка = КТО привёз) и наследуется копией при рождении, после чего живёт своей
    жизнью. Расхождение — законная ситуация («одна закупка точно может пойти от разных
    поставщиков»), поэтому это **не гейт фиксации**, а оранжевый глиф-warning в форме:
    в случае ошибки она кричит и подсказывает, в редком законном случае не мешает.

    Считаем только когда **оба** контрагента заданы и различаются: пустой снизу — это
    «не выбран», а не расхождение. Флаг считает движок, знак выбирает вью
    ([[engine-view-seam]]) — иначе форма сравнивала бы два id сама.
    """
    parent_id = parent.contractor_id if parent is not None else None
    return bool(own_contractor_id and parent_id and own_contractor_id != parent_id)


def lines_estimate(lines):
    """Оценка документа-намерения: `Σ(qty × item.estimated_cost)` + позиции без оценки.

    Прямое перемножение по строкам, без разузлования: строка закупки/заказа — это то,
    что покупают КАК ЕСТЬ, и цена берётся с самого изделия (у производимого там лежит
    роллап его состава — тоже валидная оценка покупки узла на стороне).

    Считаем по тем, у кого оценка есть, а коды остальных возвращаем списком: молча
    сложить их как 0 значило бы выдать неполную сумму за полную. Тот же приём, что у
    `_project_estimate` (`unestimated`) — вью показывает счётчик, а коды кладёт в title.

    Это НЕ деньги документа: факт живёт в поставке (`Lot.unit_cost`). Здесь — прогноз,
    «во сколько это обойдётся», и цена может ещё десять раз поменяться.
    """
    estimate = ZERO
    unestimated = []
    for line in lines:
        if line.item.estimated_cost is None:
            unestimated.append(line.item.code)
            continue
        estimate += line.qty * line.item.estimated_cost
    return estimate, unestimated


def demand_map(projects):
    """`{item_id: нужда по листьям}` через набор проектов — знаменатель денег намерения.

    Та же нужда, что в «Потребности» и в своде «К закупке» (`project_leaf_demand`,
    разузлование насквозь), только сложенная по Item и **без покрытия**: панель бюджета
    отвечает «сколько деталей нужно по-настоящему», а не «сколько осталось докупить».
    Между проектами не перенеттим — как и `scope_deficit`.
    """
    total = {}
    for project in projects:
        leaves, _incomplete = project_leaf_demand(project)
        for leaf, qty in leaves.items():
            total[leaf.id] = total.get(leaf.id, ZERO) + qty
    return total


def intent_money(lines, need_by_item):
    """Три числа денег документа-намерения (2026-08-07): нужда / сумма строк / разница.

    - `demand` — **потребность**: `Σ` по строкам `нужда_проекта(item) × estimated_cost`.
      Сколько денег на эти позиции нужно по-настоящему, без запаса. Количество берётся
      НЕ из строки, а из потребности проекта — строка задаёт только номенклатуру.
    - `estimate` — сумма самого документа (`lines_estimate`, «во сколько обойдётся»).
    - `overpay` — их разница со знаком: `+` переплата (взяли с запасом), `−` недозаказ
      (взяли меньше нужды). Знак считает движок, подпись и цвет выбирает вью
      ([[engine-view-seam]]).

    Цена у обоих чисел одна — `estimated_cost` изделия (в строках заказа/закупки цены
    нет вовсе, факт рождается только в поставке). Поэтому позиция без оценки выпадает
    из **обеих** сумм разом и знак разницы не врёт; её код — в `unestimated`.

    Числа — витрина, а не снимок: замок морозит строки, но не цену и не BOM, поэтому у
    зафиксированного документа они законно едут вслед за справочником.

    Потребность проекта не поделена между его заказами: два заказа на один Item покажут
    каждый полную нужду. Число отвечает «сколько из ЭТОГО документа оправдано нуждой», и
    по проекту не складывается.
    """
    estimate, unestimated = lines_estimate(lines)
    demand = ZERO
    for line in lines:
        if line.item.estimated_cost is None:
            continue                      # уже учтён в `unestimated` — молчим в обеих суммах
        demand += need_by_item.get(line.item_id, ZERO) * line.item.estimated_cost
    return {'demand': demand, 'estimate': estimate, 'overpay': estimate - demand,
            'unestimated': unestimated}


# Строка взята с перебором больше чем в ПОЛТОРА раза от нужды — красный член. Порог
# задан в деньгах («переплата больше половины потребности»), но цена в неравенстве
# сокращается: `(qty − need)·цена > ½·need·цена` ⇔ `qty > 1.5·need`. Поэтому правило
# работает и там, где `estimated_cost` нет вовсе (красить, впрочем, будет нечего).
OVERPAY_LIMIT = Decimal('1.5')
# Шкала округления порога (2026-08-07): до 5 → до 10 → до 50 → до 100, ступень выбирается
# по величине самого порога. Округляем ПОРОГ, а не потребность: иначе «нужно 13, беру
# круглые 20» ловилось красным (порог 19.5), хотя это ровно то решение, к которому
# приходят руками — некратные числа не заказывают.
ROUND_STEPS = ((Decimal(50), Decimal(5)), (Decimal(100), Decimal(10)),
               (Decimal(500), Decimal(50)))
ROUND_STEP_MAX = Decimal(100)
# Доля строк, помеченных «дорогая»: верхняя четверть по стоимости, округление ВВЕРХ
# (в списке из двух строк дорогая — одна; решение Ивана 2026-08-07).
COSTLY_SHARE = 4


def line_cost(item, qty):
    """Стоимость строки: `qty × estimated_cost`. Без оценки — `None`, а не ноль.

    Тот же приём, что у `lines_estimate`/`intent_money`: неизвестную цену нельзя
    показывать нулём — вью рисует прочерк, и строка не участвует в ранжировании.
    """
    return None if item.estimated_cost is None else qty * item.estimated_cost


def overpay_threshold(need):
    """Сколько можно взять при нужде `need`, чтобы это ещё не считалось перебором:
    **полторы нормы, округлённые вверх до круглого числа**.

    Одна фраза — и её видно в подсказке строки («без перебора можно до 20»). Понятность
    правила тут важнее точности (решение Ивана 2026-08-07): закупщик держит его в голове
    и заранее знает, загорится строка или нет.

    Округление вверх убирает жёсткость на малых количествах, где процент режет по живому:
    нужно 13 → полторы нормы 19.5 → **20**, то есть ровно то круглое число, которым и
    закупают. На больших партиях ступень грубее (5 → 10 → 50 → 100), и процент снова
    берёт своё: 100 → 150, 333 → 500.

    `need = 0` даёт порог 0: позиции нет в составе проекта, любое количество — перебор.
    """
    raw = need * OVERPAY_LIMIT
    step = next((s for limit, s in ROUND_STEPS if raw < limit), ROUND_STEP_MAX)
    return (raw / step).to_integral_value(rounding=ROUND_CEILING) * step


def is_overpaid(qty, need):
    """Строка взята сверх порога (красный). Пустая ячейка (`qty = 0`) не горит."""
    return qty > overpay_threshold(need) and qty > ZERO


def mark_costs(rows):
    """Проставить `cost_status` строкам: `overpaid` (красный) > `costly` (оранжевый).

    Смысл считает движок, знак выбирает вью ([[engine-view-seam]]): здесь — «эта строка
    дорогая» и «эта взята с перебором», а не «оранжевая»/«красная».

    «Дорогая» — верхняя четверть НАБОРА (ранжирование внутри одного списка, поэтому
    считается тут, а не построчно): в разных документах дорого разное. Строки без цены и
    нулевые в ранжировании не участвуют — «самый дорогой ноль» не бывает.

    Красное сильнее оранжевого (worst-of, как везде в продукте): перебор — повод
    вмешаться, дороговизна — повод посмотреть.

    Ждёт в строках `cost` (может быть `None`), `qty` и `need`; правит их на месте.
    """
    ranked = sorted((r for r in rows if r['cost'] is not None and r['cost'] > ZERO),
                    key=lambda r: r['cost'], reverse=True)
    costly = -(-len(ranked) // COSTLY_SHARE)          # ceil: 4→1, 6→2, 10→3
    for i, row in enumerate(ranked):
        row['cost_status'] = 'costly' if i < costly else None
    for row in rows:
        row.setdefault('cost_status', None)
        # Порог едет в строку: им вью объясняет цвет («без перебора можно до 20»).
        # Правило, которое видно, закупщик держит в голове — а держа, не спорит с ним.
        row['overpay_at'] = overpay_threshold(row['need'])
        if row['cost'] is not None and row['qty'] > row['overpay_at'] and row['qty'] > ZERO:
            row['cost_status'] = 'overpaid'
    return rows


def purchase_form(purchase):
    """Проекция формы заказа: шапка + строки (заказано/поступило/остаток) + приходы.

    Закрытость строки красится тем же словарём ✓/●/▲, что дефицит/формы:
    получено полностью → ✓ (available), частично → ● (on_order), ничего → ▲ (to_order).
    Статусы `partial`/`received` не храним — это вычисляемая закрытость. Ничего не
    хранит (чистая проекция).

    Ф6: строка несёт ещё и **чем** она закрыта (`receipts`) — обратная связь потока
    «Заказ → УПД». Список приходов заказа целиком остаётся отдельным табом.

    2026-08-07: строка несёт **баланс проекта** по своему изделию (`_balance`, четвёрка
    слагаемых + невязка) — третий экран, показывающий одно и то же число: в проекте для
    общей оценки, в «Привязке» для раскладки, здесь для решения «сколько заказывать».
    Две закрытости не путать: `status` — приехало ли по ЭТОЙ строке, `balance_status` —
    как дела у проекта с этим изделием вообще.
    """
    editable = not purchase.locked
    closure = _purchase_closure(purchase)
    lines = list(purchase.lines.select_related('item').order_by('id'))
    # Нужда своего проекта — знаменатель и панели бюджета, и цвета строк (одно
    # разузлование на форму).
    need_by_item = demand_map([purchase.project])
    rows = []
    statuses = []
    total_ordered = ZERO
    total_received = ZERO
    for line in lines:
        closed_by = closure.get(line.item_id, [])
        received = sum((r['qty'] for r in closed_by), ZERO)
        remaining = max(ZERO, line.qty - received)
        total_ordered += line.qty
        total_received += received
        if line.qty > 0 and received >= line.qty:
            st = 'available'      # ✓ получено полностью
        elif received > 0:
            st = 'on_order'       # ● частично получено
        else:
            st = 'to_order'       # ▲ ждём поставки
        statuses.append(st)
        # Баланс считаем ЗДЕСЬ, а не в вью: три обращения к складу на строку — та же
        # цена, что платит ячейка «Привязки», и та же арифметика (`_balance`).
        bal = _balance(need_by_item.get(line.item_id, ZERO),
                       item_kitted(line.item, purchase.project),
                       item_available(line.item, purchase.project),
                       item_on_order(line.item, purchase.project))
        rows.append({
            'id': line.id, 'item_id': line.item_id, 'item_code': line.item.code,
            'item_description': line.item.description, 'uom': line.item.uom,
            # Оси изделия — под глиф строки (§7a): форма = изделие/компонент,
            # цвет = закрытость строки. Так же устроена форма закупки-плана (Ф3a).
            'item_native': line.item.native, 'item_synced': line.item.synced,
            'item_locked': line.item.locked,
            'qty': line.qty, 'received': received, 'remaining': remaining,
            'status': st,
            # Баланс проекта по этому изделию (2026-08-07) — то же число и та же
            # четвёрка слагаемых, что в «Потребности» проекта и в ячейке «Привязки»:
            # заказ без него не отвечает на вопрос «а сколько вообще надо». Считает
            # только зафиксированное, поэтому свой же черновик его не двигает — гаснет
            # на фиксации заказа, в осмысленный момент.
            **{k: v for k, v in bal.items() if k != 'status'},
            'balance_status': bal['status'],
            # Деньги строки (2026-08-07): сколько она стоит и оправдана ли нуждой.
            # `need` приезжает выше вместе с балансом — он и есть его первый член.
            'cost': line_cost(line.item, line.qty),
            'receipts': closed_by,     # Ф6: какими накладными строка закрыта
        })
    mark_costs(rows)                   # `cost_status`: overpaid (красный) > costly (оранж.)
    receipts = [
        # Ф6: `code` — первичная идентичность накладной (у рождённой из заказа
        # `number` пуст до заполнения по бумаге, и ссылка на него рисовала пустоту).
        {'id': r.id, 'code': r.code, 'number': r.number, 'date': r.date,
         'locked': r.locked,
         **_contractor_view(r.contractor),
         'lines': r.lots.count()}
        for r in purchase.receipts.select_related('contractor').order_by('id')
    ]
    # Деньги намерения (2026-08-07): потребность / заказ / переплата. Знаменатель —
    # нужда СВОЕГО проекта: он у заказа один и назван явно, отсева тут нет (замок или
    # вид проекта потребность не отменяют — заказ уже заведён именно под него).
    money = intent_money(lines, need_by_item)
    return {
        'id': purchase.id, **_author(purchase), 'locked': purchase.locked,
        'project_id': purchase.project_id, 'project_code': purchase.project.code,
        'project_name': purchase.project.description,
        # Ф17: закупка-план опциональна (`None` = заказ без плана — законное состояние).
        'procurement_id': purchase.procurement_id,   # якорь #A: закупка-план (Ф2k)
        **_contractor_view(purchase.contractor),      # Ф17: у кого купили
        # Расхождение с намерением плана считает ДВИЖОК, вью только выбирает знак
        # ([[engine-view-seam]]): отдаём готовый флаг, а не два id на сравнение во фронте.
        'contractor_mismatch': _contractor_mismatch(purchase.contractor_id,
                                                    purchase.procurement),
        'code': purchase.code, 'description': purchase.description,
        'date': purchase.date,
        'editable': editable,                       # строки правятся только пока не зафиксировано
        'worst_status': _worst_of(statuses),      # worst-of закрытости строк
        'total_ordered': total_ordered, 'total_received': total_received,
        **money,          # demand / estimate / overpay / unestimated — панель бюджета
        'rows': rows, 'receipts': receipts,
    }


PURCHASE_LOCKED = 'Строки правятся только в черновике заказа — снимите замок (unlock).'


def add_purchase_line(purchase, item, qty):
    """Добавить строку заказа (только в черновике). `(purchase, item)` уникальна."""
    _require_unlocked(purchase, PURCHASE_LOCKED)
    if qty is None or qty <= 0:
        raise ValidationError('Количество заказа должно быть положительным.')
    if purchase.lines.filter(item=item).exists():
        raise ValidationError(
            f'Изделие {item.code} уже в заказе — правьте существующую строку.')
    return models.PurchaseLine.objects.create(purchase=purchase, item=item, qty=qty)


def update_purchase_line(line, qty):
    """Автосейв количества строки заказа (только в черновике)."""
    _require_unlocked(line.purchase, PURCHASE_LOCKED)
    if qty is None or qty <= 0:
        raise ValidationError('Количество заказа должно быть положительным.')
    line.qty = qty
    line.save(update_fields=['qty'])
    return line


def remove_purchase_line(line):
    """Удалить строку заказа (только в черновике)."""
    _require_unlocked(line.purchase, PURCHASE_LOCKED)
    line.delete()


def lock_purchase(purchase):
    """Поставить замок заказа — теперь считается в «заказано»,
    строки становятся read-only. Снятие (`unpost`) ничего не разрушает.

    Волна 19 (Ф1): бывший `send_purchase`. Отмены больше нет — отмена = удаление
    (развилка Р1), поэтому проверять «отменённый нельзя отправить» стало нечего.

    Ф17: контрагент обязателен к ФИКСАЦИИ (CHECK `purchase_locked_has_contractor`).
    Прикладной гейт дублирует инвариант БД, чтобы пользователь получил внятный отказ,
    а не `IntegrityError 3819` (урок Ф12e: CHECK без гейта — это 500 в лицо).
    """
    if not purchase.lines.exists():
        raise ValidationError('Нельзя утвердить пустой заказ — добавьте строку.')
    if purchase.contractor_id is None:
        raise ValidationError(
            'Перед фиксацией выберите контрагента — у кого заказано.')
    purchase.locked = True
    purchase.save(update_fields=['locked'])
    return purchase


def unlock_purchase(purchase):
    """Снять замок заказа. Purchase лотов не рождает — снятие
    обязательства ничего не разрушает (связанные приходы остаются, заказ просто
    выходит из счёта «заказано»), guard по потомкам не нужен."""
    purchase.locked = False
    purchase.save(update_fields=['locked'])
    return purchase


def delete_purchase(purchase):
    """Удалить заказ (WAVE14 Ф2). Мягкий замок как у ордеров: утверждённый сперва
    вернуть в черновик (снять замок); привязанный приход (`Receipt.purchase`,
    SET_NULL) держит — иначе удаление молча обнулило бы ссылку у поставок. Строки
    заказа (`PurchaseLine`) — каскад."""
    if purchase.locked:
        raise ValidationError(
            'Заказ утверждён — сперва верните его в черновик (снимите замок), затем удаляйте.')
    if purchase.receipts.exists():
        raise ValidationError('К заказу привязаны поставки (приход) — удаление заблокировано.')
    for att in purchase.attachments.all():         # физические файлы (каскад их сиротит)
        delete_attachment(att)
    try:
        purchase.delete()                          # каскад: строки заказа
    except ProtectedError:
        raise ValidationError('Заказ связан с другими записями — удаление заблокировано.')


# --------------------------------------------------------------------------- #
#  Форма передачи / Transfer (волна 5): отдаём партию заказчику (−ISSUE)
# --------------------------------------------------------------------------- #
def project_available_lots(project):
    """Лоты проекта с живым остатком > 0 — кандидаты на отгрузку заказчику.

    Пикер строки передачи: любой лот проекта, где ещё что-то лежит (обычно
    готовое железо — приборы из комплектации, но модель не ограничивает).
    """
    result = []
    for lot in (models.Lot.objects.filter(project=project)
                .select_related('item').order_by('item__code', 'id')):
        live = lot_live_qty(lot)
        if live > 0:
            result.append({
                'lot_id': lot.id, 'item_id': lot.item_id,
                'item_code': lot.item.code, 'item_description': lot.item.description,
                'uom': lot.item.uom, 'live_qty': live, 'origin': lot.origin_kind,
                'part_number': lot.part_number,
                'lot_name': lot.lot_name,
            })
    return result


def _lot_label(lot):
    """Человекочитаемая метка лота для накладной/строки (название / PN / артикул)."""
    tail = lot.lot_name or lot.part_number or lot.item.code
    return f'#{lot.id} {tail}'


def _author(doc):
    """Автор документа для проекции формы (Ф2j): id + человеческое имя (для
    пикера авторства в шапке формы). `user` NOT NULL на всех ордерах/закупках."""
    u = doc.user
    if u is None:                        # страховка на легаси-строки
        return {'user_id': None, 'user_name': ''}
    return {'user_id': u.id, 'user_name': u.get_full_name() or u.get_username()}


def transfer_form(transfer):
    """Проекция формы передачи: шапка накладной + строки-лоты + итог.

    Каждая строка отдаёт партию заказчику (`−ISSUE`); показываем живой остаток
    источника (просел ли под передачу, не ушёл ли в минус). Ничего не хранит.
    """
    lines = []
    total_qty = ZERO
    for line in transfer.lines.select_related('lot__item', 'lot__origin').order_by('id'):
        lot = line.lot
        mag = -line.qty                       # знаковая строка (− расход) → магнитуда
        total_qty += mag
        lines.append({
            'id': line.id, 'lot_id': lot.id,
            'lot_label': _lot_label(lot),
            'origin': lot.origin_kind,        # глиф партии (§7a): форма = откуда родилась
            'item_id': lot.item_id, 'item_code': lot.item.code,
            'item_description': lot.item.description, 'uom': lot.item.uom,
            'qty': mag, 'display_name': line.display_name,
            'lot_live_qty': lot_live_qty(lot),   # остаток источника после отгрузки
            'lot_name': lot.lot_name,
        })
    return {
        'id': transfer.id, **_author(transfer), 'number': transfer.number, 'date': transfer.date,
        'code': transfer.code, 'description': transfer.description,
        **_contractor_view(transfer.contractor),
        'project_id': transfer.project_id, 'project_code': transfer.project.code,
        'project_name': transfer.project.description, 'locked': transfer.locked,
        'total_qty': total_qty, 'lines': lines,
    }


def _movement_row(document, lot, qty, event):
    """Одна строка ленты движений: ордер + партия + знаковое кол-во."""
    return {
        'event': event,                 # 'born' — партия родилась | 'move' — движение
        'kind': document.kind, 'document_id': document.id,
        'code': document.code, 'number': document.number, 'date': document.date,
        'locked': document.locked, 'project_code': document.project.code,
        'lot_id': lot.id, 'lot_name': lot.lot_name, 'qty': qty,
    }


def item_movements(item):
    """Все ордера, коснувшиеся изделия, одной лентой — для его экрана (волна 19, Ф12a).

    Заменила узкую `item_shipments` (только передачи): изделие живёт не одними
    отгрузками, и «куда делось» читается лишь по всей ленте. Два источника, ровно
    как в модели:

    * **рождение партии** (born-direct) — `Lot.origin` (поставка / комплектация /
      инвентаризация / требование), знак `+`;
    * **движение существующей партии** — `StockLine` (пайка в комплектацию, передача,
      списание, требование, перемещение), знак уже в `qty` (− расход).

    `StockMovement` намеренно НЕ читаем: это проекция-остаток, её `source_type`/
    `source_id` пришлось бы резолвить обратно в документы. Порядок — свежие сверху;
    дата у ордера nullable (черновик комплектации), такие уходят в конец.
    """
    rows = [
        _movement_row(lot.origin, lot, lot.qty, 'born')
        for lot in (models.Lot.objects.filter(item=item)
                    .select_related('origin', 'origin__project'))
    ]
    rows += [
        _movement_row(line.document, line.lot, line.qty, 'move')
        for line in (models.StockLine.objects.filter(lot__item=item)
                     .select_related('document', 'document__project', 'lot'))
    ]
    rows.sort(key=lambda r: (r['date'] is not None, r['date'], r['document_id']),
              reverse=True)
    return rows


def create_transfer(project, user, number='', date=None, contractor=None):
    """Создать передачу (накладную) проекта. Строки добавляются в форме.

    `contractor` — контрагент-заказчик (опционален: получатель может быть проставлен
    позже в форме). Ф12e: номер тоже опционален — он обязателен к ФИКСАЦИИ
    (`REQUIRED_HEADER_BY_KIND`), а не к рождению."""
    return _born_order(models.Transfer, project, user, number, date,
                       contractor=contractor)


def add_transfer_line(transfer, lot, qty, display_name=''):
    """Отгрузить партию заказчику: строка передачи (`−ISSUE` на лоте).

    Лот — того же проекта (передаём своё, чужое — через требование). Кол-во не
    клампим по остатку: переотдать можно, лот уйдёт в минус (недостача информативна,
    в духе мутабельной ДНК).
    """
    _require_unlocked(transfer)
    if lot.project_id != transfer.project_id:
        raise ValidationError('Лот из другого проекта — передаём только своё.')
    if qty is None or qty <= 0:
        raise ValidationError('Количество передачи должно быть положительным.')
    line = models.StockLine.objects.create(
        document=transfer, lot=lot, location=_main_location(), qty=-qty,
        display_name=(display_name or '').strip() or _lot_label(lot))
    rebuild_movements(lot)
    return line


def update_transfer_line(line, qty=None, display_name=None):
    """Автосейв строки передачи (кол-во / отображаемое имя для накладной)."""
    _require_unlocked(line.document)
    fields = []
    if qty is not None:
        if qty <= 0:
            raise ValidationError('Количество передачи должно быть положительным.')
        line.qty = -qty                   # знаковая строка (− расход)
        fields.append('qty')
    if display_name is not None:
        line.display_name = display_name
        fields.append('display_name')
    if fields:
        line.save(update_fields=fields)
        if 'qty' in fields:
            rebuild_movements(line.lot)
    return line


def remove_transfer_line(line):
    """Убрать строку передачи (коррекция) — источник возвращает остаток."""
    _require_unlocked(line.document)
    lot = line.lot
    line.delete()
    rebuild_movements(lot)


def lock_transfer(transfer):
    """Поставить замок «отгружено»: накладная read-only, партии уходят со склада
    (`−ISSUE` материализуется, Ф15). Сюда позже ляжет подписанная накладная
    (Attachment). Ф15 сняла рукописную копию `lock_document` — общая дорога."""
    return lock_document(transfer, transfer.lines,
                         'Нельзя отгрузить пустую накладную — добавьте строку.')


def unlock_transfer(transfer):
    """Снять замок — снова разрешить правку (`−ISSUE` отпускается, партии
    возвращаются на склад, Ф15). Строки остаются, guard по потомкам не нужен."""
    return unlock_document(transfer)


# --------------------------------------------------------------------------- #
#  Закрытие проекта (волна 6): списание / требование + панель + мягкий замок
# --------------------------------------------------------------------------- #
def default_document_project():
    """Проект-якорь для ордера, рождённого по клику (Ф12e).

    `StockDocument.project` — NOT NULL (якорь Ф2k), фолбэк нужен обязательно.
    Берём белый «Собственный склад»: он всегда существует (синглтон сида), и это
    не ложные данные — документ и правда пока ничей, а не приписан чужому НИР.
    Меняется в шапке, пока ордер пуст (`_set_project`)."""
    return _internal_project(models.Project.Kind.INTERNAL_STOCK)


def _internal_project(kind):
    """Найти-или-создать внутренний проект-склад (белый/серый) — синглтон.

    Целевая «куча» для постановки на баланс (белый «Собственный склад») —
    `INTERNAL_KINDS` синглтоны (см. `Project.clean`), сид их заводит; здесь мягко
    добираем, чтобы движок был робастен и в голой БД/тестах.
    """
    proj = models.Project.objects.filter(kind=kind).first()
    if proj is not None:
        return proj
    code, description = {
        models.Project.Kind.INTERNAL_STOCK: ('WHITE', 'Собственный склад'),
        models.Project.Kind.INTERNAL_WRITEOFF: ('GREY', 'Свободные неучтённые'),
    }[kind]
    return models.Project.objects.create(
        code=code, description=description, kind=kind)


def _auto_number(prefix, project):
    """Авто-№ документа для мостов панели (акт списания/требование одним кликом)."""
    return f'{prefix}-{project.code}-{timezone.localdate():%Y%m%d}'


def all_available_lots():
    """Лоты всех проектов с живым остатком > 0 — пикер источника требования.

    Требование тянет из любого проекта (постановка своего на баланс, заём у
    соседнего активного) — поэтому пикер сквозной, с кодом проекта-источника.
    """
    result = []
    for lot in (models.Lot.objects.select_related('item', 'project')
                .order_by('project__code', 'item__code', 'id')):
        live = lot_live_qty(lot)
        if live > 0:
            result.append({
                'lot_id': lot.id, 'item_id': lot.item_id,
                'item_code': lot.item.code, 'item_description': lot.item.description,
                'uom': lot.item.uom, 'live_qty': live, 'origin': lot.origin_kind,
                'project_id': lot.project_id, 'project_code': lot.project.code,
                'part_number': lot.part_number,
                'lot_name': lot.lot_name,
            })
    return result


# ── Списание / Writeoff (серый путь: чистый −ISSUE, лот покидает учёт) ──
def writeoff_form(writeoff):
    """Проекция формы списания: шапка акта + строки-лоты (`−ISSUE`) + итог.

    Списание — чистое выбытие из проекта (born-лота нет, `Writeoff` не origin);
    «в серый склад» — конвенция учёта, ре-материализация серых остатков — актом
    `Inventory` (следующая волна). Живой остаток источника показываем — просел ли.
    """
    lines = []
    total_qty = ZERO
    for line in writeoff.lines.select_related('lot__item', 'lot__origin').order_by('id'):
        lot = line.lot
        mag = -line.qty                       # знаковая строка (− расход) → магнитуда
        total_qty += mag
        lines.append({
            'id': line.id, 'lot_id': lot.id, 'lot_label': _lot_label(lot),
            'origin': lot.origin_kind,        # глиф партии (§7a): форма = откуда родилась
            'item_id': lot.item_id, 'item_code': lot.item.code,
            'item_description': lot.item.description, 'uom': lot.item.uom,
            'qty': mag, 'lot_live_qty': lot_live_qty(lot),
            'lot_name': lot.lot_name,
        })
    return {
        'id': writeoff.id, **_author(writeoff), 'number': writeoff.number, 'date': writeoff.date,
        'code': writeoff.code, 'description': writeoff.description,
        'reason': writeoff.reason,
        'project_id': writeoff.project_id, 'project_code': writeoff.project.code,
        'project_name': writeoff.project.description, 'locked': writeoff.locked,
        'total_qty': total_qty, 'lines': lines,
    }


def create_writeoff(project, user, number='', date=None, reason=''):
    """Создать акт списания проекта. Строки добавляются в форме."""
    return _born_order(models.Writeoff, project, user, number, date,
                       reason=(reason or '').strip())


def add_writeoff_line(writeoff, lot, qty, location=None):
    """Списать партию из проекта: строка списания (`−ISSUE` на лоте).

    Списываем только своё (`lot.project == writeoff.project`). Кол-во не клампим по
    остатку (как приход/передача): пересписать можно, лот уйдёт в минус — недостача
    информативнее нуля (мутабельная ДНК).
    """
    _require_unlocked(writeoff)
    if lot.project_id != writeoff.project_id:
        raise ValidationError('Лот из другого проекта — списываем только своё.')
    if qty is None or qty <= 0:
        raise ValidationError('Количество списания должно быть положительным.')
    line = models.StockLine.objects.create(
        document=writeoff, lot=lot, location=location or _main_location(), qty=-qty)
    rebuild_movements(lot)
    return line


def update_writeoff_line(line, qty):
    """Автосейв количества строки списания. Только черновик (замок)."""
    _require_unlocked(line.document)
    if qty is None or qty <= 0:
        raise ValidationError('Количество списания должно быть положительным.')
    line.qty = -qty                      # знаковая строка (− расход)
    line.save(update_fields=['qty'])
    rebuild_movements(line.lot)
    return line


def remove_writeoff_line(line):
    """Убрать строку списания (коррекция) — источник возвращает остаток."""
    _require_unlocked(line.document)
    lot = line.lot
    line.delete()
    rebuild_movements(lot)


def lock_writeoff(writeoff):
    """Провести списание (замок «списано», форма read-only)."""
    return lock_document(writeoff, writeoff.lines,
                         'Нельзя провести пустой акт списания — добавьте строку.')


def unlock_writeoff(writeoff):
    """Снять замок списания — снова разрешить правку."""
    return unlock_document(writeoff)


# ── Требование / Requisition (белый путь: −ISSUE источника + рождение потомка) ──
def _requisition_born_lot(requisition, source_lot):
    """Порождённый требованием лот-потомок источника (пара строки).

    Один источник = одна строка (guard в `add_requisition_line`), поэтому пара
    `(requisition, predecessor)` однозначна.
    """
    return requisition.lots.filter(predecessor=source_lot,
                                   item=source_lot.item).first()


def requisition_form(requisition):
    """Проекция формы требования: шапка (проект-получатель) + строки + итог.

    Каждая строка тянет из лота-источника (`−ISSUE`) и рождает лот-потомок в
    проекте-получателе (`+RECEIPT`, наследует item/цену/провенанс через
    `predecessor`). Показываем живой остаток источника (просел ли).
    """
    lines = []
    total_qty = ZERO
    for line in (requisition.lines
                 .select_related('lot__item', 'lot__project', 'lot__origin')
                 .order_by('id')):
        src = line.lot                        # StockLine.lot = лот-источник расхода
        mag = -line.qty                       # знаковая строка (− расход) → магнитуда
        total_qty += mag
        born = _requisition_born_lot(requisition, src)
        lines.append({
            'id': line.id, 'source_lot_id': src.id, 'lot_label': _lot_label(src),
            'origin': src.origin_kind,        # глиф партии (§7a): форма = откуда родилась
            'source_project_code': src.project.code,
            'item_id': src.item_id, 'item_code': src.item.code,
            'item_description': src.item.description, 'uom': src.item.uom,
            'qty': mag, 'source_live_qty': lot_live_qty(src),
            'born_lot_id': born.id if born else None,
            'lot_name': src.lot_name,
        })
    return {
        'id': requisition.id, **_author(requisition), 'number': requisition.number, 'date': requisition.date,
        'code': requisition.code, 'description': requisition.description,
        'project_id': requisition.project_id, 'project_code': requisition.project.code,
        'project_name': requisition.project.description, 'locked': requisition.locked,
        'total_qty': total_qty, 'lines': lines,
    }


def create_requisition(project, user, number='', date=None):
    """Создать требование в проект-получатель (`project` = куда кладём потомков)."""
    return _born_order(models.Requisition, project, user, number, date)


def add_requisition_line(requisition, source_lot, qty, location=None):
    """Отпочкование: `−ISSUE` от источника + рождение лота-потомка у получателя.

    Источник — из любого проекта (постановка своего на баланс → белый, заём у
    соседнего активного B→A). Потомок наследует item/цену/название/PN источника,
    `predecessor` → источник (генеалогия/провенанс для kitting из остатков). Один
    источник = одна строка (пара строки↔потомок однозначна). Кол-во не клампим.
    """
    _require_unlocked(requisition)
    if qty is None or qty <= 0:
        raise ValidationError('Количество требования должно быть положительным.')
    if source_lot.project_id == requisition.project_id:
        raise ValidationError('Источник и получатель — один проект (перекладывать некуда).')
    if requisition.lines.filter(lot=source_lot).exists():
        raise ValidationError('Этот лот уже в требовании — правьте существующую строку.')
    line = models.StockLine.objects.create(
        document=requisition, lot=source_lot,
        location=location or _main_location(), qty=-qty)
    born = models.Lot.objects.create(
        item=source_lot.item, project=requisition.project, origin=requisition,
        predecessor=source_lot, qty=qty, unit_cost=source_lot.unit_cost,
        lot_name=source_lot.lot_name, part_number=source_lot.part_number)
    rebuild_movements(source_lot)   # −ISSUE у источника
    rebuild_movements(born)         # +RECEIPT у потомка
    return line


def update_requisition_line(line, qty):
    """Автосейв количества: правит и строку-источник (`−ISSUE`), и потомок (`+RECEIPT`).
    Только черновик (замок)."""
    _require_unlocked(line.document)
    if qty is None or qty <= 0:
        raise ValidationError('Количество требования должно быть положительным.')
    line.qty = -qty                      # знаковая строка (− расход) у источника
    line.save(update_fields=['qty'])
    born = _requisition_born_lot(line.document, line.lot)
    if born is not None:
        born.qty = qty                   # рождённый лот-потомок — положительное кол-во
        born.save(update_fields=['qty'])
        rebuild_movements(born)
    rebuild_movements(line.lot)
    return line


def remove_requisition_line(line):
    """Убрать строку требования: снять потомок + вернуть остаток источнику.

    Guard: потомок не должен быть потреблён ниже (спаян/передан/списан из белого).
    """
    _require_unlocked(line.document)
    src = line.lot
    born = _requisition_born_lot(line.document, src)
    if born is not None and _lot_consumed_downstream(born):
        raise ValidationError(
            'Поставленный на баланс лот уже потреблён ниже — удаление заблокировано.')
    line.delete()
    if born is not None:
        born.movements.all().delete()
        born.delete()
    rebuild_movements(src)


def lock_requisition(requisition):
    """Провести требование (замок, форма read-only)."""
    return lock_document(requisition, requisition.lines,
                         'Нельзя провести пустое требование — добавьте строку.')


def unlock_requisition(requisition):
    """Снять замок требования — снова разрешить правку."""
    return unlock_document(requisition)


# ── Перемещение / Relocation (мультисклад: лот между локациями внутри проекта) ──
def _relocation_pair(relocation, lot):
    """Пара знаковых строк одного хода: (источник `−q`, приёмник `+q`). Один лот =
    один ход в перемещении (guard в `add_relocation_line`), поэтому пара однозначна."""
    lines = list(relocation.lines.filter(lot=lot).select_related('location'))
    src = next((l for l in lines if l.qty < 0), None)
    dst = next((l for l in lines if l.qty > 0), None)
    return src, dst


def relocation_form(relocation):
    """Проекция формы перемещения: шапка + ходы (лот, откуда→куда, кол-во) + итог.

    Каждый ход — пара строк (`−q`@источник, `+q`@приёмник, волна 13 Ф2e). Показываем
    остаток лота в источнике и приёмнике (пары `(лот,локация)`) — куда и сколько ушло.
    Перемещение не меняет тотал лота/проекта, только распределение по местам.
    """
    moves = []
    total_qty = ZERO
    seen = set()
    for line in relocation.lines.select_related('lot__item').order_by('id'):
        if line.lot_id in seen:
            continue
        seen.add(line.lot_id)
        lot = line.lot
        src, dst = _relocation_pair(relocation, lot)
        mag = (-src.qty) if src else (dst.qty if dst else ZERO)
        total_qty += mag
        moves.append({
            'lot_id': lot.id, 'lot_label': _lot_label(lot),
            'origin': lot.origin_kind,        # глиф партии (§7a): форма = откуда родилась
            'item_id': lot.item_id, 'item_code': lot.item.code,
            'item_description': lot.item.description, 'uom': lot.item.uom, 'qty': mag,
            'from_location_id': src.location_id if src else None,
            'from_location': src.location.code if src else '',
            'to_location_id': dst.location_id if dst else None,
            'to_location': dst.location.code if dst else '',
            'from_live_qty': lot_live_qty(lot, src.location) if src else ZERO,
            'to_live_qty': lot_live_qty(lot, dst.location) if dst else ZERO,
        })
    return {
        'id': relocation.id, **_author(relocation), 'number': relocation.number, 'date': relocation.date,
        'code': relocation.code, 'description': relocation.description,
        'project_id': relocation.project_id, 'project_code': relocation.project.code,
        'project_name': relocation.project.description, 'locked': relocation.locked,
        'total_qty': total_qty, 'moves': moves,
    }


def create_relocation(project, user, number='', date=None):
    """Создать перемещение внутри проекта (`project` — где двигаем лоты по местам)."""
    return _born_order(models.Relocation, project, user, number, date)


def add_relocation_line(relocation, lot, qty, from_location, to_location):
    """Ход перемещения: пара знаковых строк (`−q` на источнике, `+q` на приёмнике).

    Двигаем только свой лот (`lot.project == relocation.project`) между двумя РАЗНЫМИ
    местами хранения. Один лот = один ход (правьте существующий). Кол-во не клампим по
    остатку источника (как передача/списание): пересместить можно, источник уйдёт в
    минус — недостача информативнее нуля (мутабельная ДНК). Тотал лота сохранён
    (`−q+q=0`) — двигаем распределение, не остаток.
    """
    _require_unlocked(relocation)
    if lot.project_id != relocation.project_id:
        raise ValidationError('Лот из другого проекта — перемещаем только своё.')
    if qty is None or qty <= 0:
        raise ValidationError('Количество перемещения должно быть положительным.')
    if from_location is None or to_location is None:
        raise ValidationError('Нужны место-источник и место-приёмник.')
    if from_location.id == to_location.id:
        raise ValidationError('Источник и приёмник — одно место (перемещать некуда).')
    if relocation.lines.filter(lot=lot).exists():
        raise ValidationError('Этот лот уже в перемещении — правьте существующий ход.')
    src = models.StockLine.objects.create(
        document=relocation, lot=lot, location=from_location, qty=-qty)
    dst = models.StockLine.objects.create(
        document=relocation, lot=lot, location=to_location, qty=qty)
    rebuild_movements(lot)
    return src, dst


def update_relocation_line(relocation, lot, qty=None, from_location=None,
                           to_location=None):
    """Автосейв хода перемещения (кол-во/места). Только черновик (замок)."""
    _require_unlocked(relocation)
    src, dst = _relocation_pair(relocation, lot)
    if src is None or dst is None:
        raise ValidationError('Ход перемещения не найден.')
    if qty is not None:
        if qty <= 0:
            raise ValidationError('Количество перемещения должно быть положительным.')
        src.qty = -qty
        dst.qty = qty
    if from_location is not None:
        src.location = from_location
    if to_location is not None:
        dst.location = to_location
    if src.location_id == dst.location_id:
        raise ValidationError('Источник и приёмник — одно место (перемещать некуда).')
    src.save(update_fields=['qty', 'location'])
    dst.save(update_fields=['qty', 'location'])
    rebuild_movements(lot)
    return src, dst


def remove_relocation_line(relocation, lot):
    """Убрать ход перемещения (обе строки пары) + пересобрать движения лота."""
    _require_unlocked(relocation)
    relocation.lines.filter(lot=lot).delete()
    rebuild_movements(lot)


def lock_relocation(relocation):
    """Провести перемещение (замок, форма read-only)."""
    return lock_document(relocation, relocation.lines,
                         'Нельзя провести пустое перемещение — добавьте ход.')


def unlock_relocation(relocation):
    """Снять замок перемещения — снова разрешить правку."""
    return unlock_document(relocation)


def relocation_source_lots(project):
    """Лоты проекта с живым остатком > 0 — кандидаты на перемещение, с разбивкой по
    местам хранения (`lot_locations`): пикер видит, где лот лежит и сколько."""
    result = []
    for lot in (models.Lot.objects.filter(project=project)
                .select_related('item').order_by('item__code', 'id')):
        live = lot_live_qty(lot)
        if live > 0:
            result.append({
                'lot_id': lot.id, 'item_id': lot.item_id,
                'item_code': lot.item.code, 'item_description': lot.item.description,
                'uom': lot.item.uom, 'live_qty': live,
                'part_number': lot.part_number,
                'lot_name': lot.lot_name,
                'by_location': lot_locations(lot),
            })
    return result


# ── Панель закрытия проекта + мягкий замок статуса ──
def project_closure(project):
    """Панель сведения остатков проекта к 0 + готовность к закрытию.

    Остаточные лоты (live≠0) — то, что мешает закрытию: положительные сводим в 0
    выходами (передача/списание/на баланс), отрицательные — аномалия «подбей лоты»
    (недостача, чинится правкой документа-потребителя). Закрыть можно **внешний**
    проект, когда остатков нет (внутренние склады постоянны — не закрываются).

    Волна 19, Ф15: раз черновик склад не двигает, мост «списать/на баланс» больше не
    уводит остаток мгновенно — он кладёт его в **черновой** закрывающий документ.
    Чтобы панель не выглядела «кнопка не сработала», отдаём `closing_drafts` —
    расфиксированные документы, уже разобравшие остатки этого проекта (в т.ч. чужого
    проекта-получателя, как требование в «Собственный склад»): их видно строкой
    «ждут фиксации», и они же становятся причиной отказа в `blocker`.
    """
    residuals = []
    positive = ZERO
    anomaly_count = 0
    for lot in project.lots.select_related('item', 'origin').order_by('item__code', 'id'):
        live = lot_live_qty(lot)
        if live == 0:
            continue
        residuals.append({
            'lot_id': lot.id, 'lot_label': _lot_label(lot),
            'origin': lot.origin_kind,          # глиф партии (§7a): форма = откуда родилась
            'item_id': lot.item_id, 'item_code': lot.item.code,
            'item_description': lot.item.description, 'uom': lot.item.uom,
            'live_qty': live, 'anomaly': live < 0,
        })
        if live > 0:
            positive += live
        else:
            anomaly_count += 1
    drafts = _closing_drafts(project)
    # Аудит-1, Б2б-5: при смеси показываем ОБЕ половины работы. Черновики разобрали
    # часть остатка, но `live` до фиксации не гаснет — «зафиксируйте черновики» в
    # одиночку врало бы, что дальше делать нечего. Разбор считаем от магнитуды
    # расхода черновиков (`_closing_drafts.qty`), нижний зажим — от возможной
    # пере-разборки (два черновика на один лот).
    in_drafts = sum((row['qty'] for row in drafts), ZERO)
    unsorted = max(positive - in_drafts, ZERO)
    is_external = project.kind == models.Project.Kind.EXTERNAL
    is_closed = project.locked
    can_close = is_external and not is_closed and not residuals
    if not is_external:
        blocker = 'Внутренний склад постоянный — не закрывается.'
    elif is_closed:
        blocker = ''
    elif residuals and drafts and unsorted:
        blocker = (f'В черновиках {_num(in_drafts)}, не разобрано {_num(unsorted)} — '
                   'зафиксируйте черновики и сведите остаток в 0.')
    elif residuals and drafts:
        blocker = 'Закрывающие документы ещё черновики — зафиксируйте их.'
    elif residuals:
        blocker = 'Есть остаточные лоты — сведите их в 0.'
    else:
        blocker = ''
    return {
        'project_id': project.id, 'project_code': project.code,
        'project_name': project.description, 'kind': project.kind,
        'locked': project.locked, 'closed': project.closed,
        'is_external': is_external,
        'residuals': residuals, 'residual_positive': positive,
        'residual_in_drafts': in_drafts, 'residual_unsorted': unsorted,
        'anomaly_count': anomaly_count, 'closing_drafts': drafts,
        'can_close': can_close, 'blocker': blocker,
    }


def _closing_drafts(project):
    """Расфиксированные документы, уже разобравшие остатки проекта (Ф15).

    Ищем по строкам: любой `locked=False` документ, чья `StockLine` ссылается на лот
    этого проекта. Документ может жить в другом проекте (требование в «Собственный
    склад» тянет наш лот) — поэтому фильтр по строкам, а не по `document.project`.
    `qty` — сколько остатка ждёт фиксации (магнитуда расхода по нашим лотам).
    """
    rows = []
    docs = (models.StockDocument.objects
            .filter(locked=False, lines__lot__project=project)
            .distinct().order_by('id'))
    for doc in docs:
        qty = -(doc.lines.filter(lot__project=project, qty__lt=0)
                .aggregate(s=Sum('qty'))['s'] or ZERO)
        rows.append({
            'document_id': doc.id, 'kind': doc.kind,
            'code': doc.code or '', 'number': doc.number, 'qty': qty,
        })
    return rows


def lock_project(project):
    """Зафиксировать проект — мягкий замок-веха.

    Gate: внешний проект без остаточных лотов (всё сведено в 0). Ничего не
    разрушает — веха «проект отработан», реюз/расфиксация свободны.

    Волна 19, Ф1c: дату `closed` здесь больше НЕ штампуем — она информационная,
    Иван проставляет реальный срок руками (сколько работа заняла на самом деле,
    независимо от формальностей закрытия).
    """
    if project.kind != models.Project.Kind.EXTERNAL:
        raise ValidationError('Фиксировать можно только внешний проект (НИР/контракт).')
    if project.locked:
        raise ValidationError('Проект уже зафиксирован.')
    if project_closure(project)['residuals']:
        raise ValidationError(
            'Нельзя зафиксировать: есть остаточные лоты — сведите их в 0.')
    project.locked = True
    project.save(update_fields=['locked'])
    return project


def unlock_project(project):
    """Расфиксировать проект. Замок ничего не разрушал."""
    if not project.locked:
        raise ValidationError('Расфиксировать можно только зафиксированный проект.')
    project.locked = False
    project.save(update_fields=['locked'])
    return project


# ── Мосты панели закрытия (один клик = свести остаточный лот в 0) ──
def _require_no_closing_draft(lot):
    """Отказать, если остаток лота уже разобран черновым закрывающим актом (Ф15).

    Мосты панели «найти-или-создать»: лот, уже лежащий в черновике, раньше уводил
    их в ветку «создать новый акт» — и повторный клик плодил ВТОРОЙ документ на тот
    же остаток (после фиксации обоих лот ушёл бы в минус). Спровоцировать это легко:
    после Ф15 остаток на панели не гаснет до фиксации, и клик выглядит «не
    сработавшим». Правильный ответ — не второй документ, а напоминание, где лежит
    первый: работа уже сделана, осталось её зафиксировать.
    """
    kinds = (models.StockDocument.Kind.WRITEOFF, models.StockDocument.Kind.REQUISITION)
    doc = (models.StockDocument.objects
           .filter(locked=False, kind__in=kinds, lines__lot=lot, lines__qty__lt=0)
           .distinct().order_by('id').first())
    if doc is not None:
        raise ValidationError(
            f'Остаток лота уже разобран черновиком «{doc.code}» — '
            f'зафиксируйте его.')


def writeoff_lot(project, lot, qty, user):
    """Мост «списать остаток»: найти-или-создать акт списания проекта + строка.

    Оживляет действие панели: один клик кладёт остаток лота в акт списания. Ф15:
    остаток уйдёт в 0 не сейчас, а на фиксации акта — панель показывает лот
    остаточным, пока акт черновик (это и есть «ответственно нажал Зафиксировать»).
    Переиспользует последний **черновой** акт проекта; зафиксированный не трогаем —
    правка под замком запрещена.
    """
    if lot.project_id != project.id:
        raise ValidationError('Лот из другого проекта.')
    _require_no_closing_draft(lot)
    # Ф2c: `project` поднят в StockDocument (реверс — `project.documents`); типизированный
    # доступ через дочерний менеджер (прозрачно фильтрует по родительскому полю).
    writeoff = (models.Writeoff.objects.filter(project=project, locked=False)
                .order_by('-id').first())
    if writeoff is None:
        writeoff = create_writeoff(
            project, user, _auto_number('СПИС', project), reason='закрытие проекта')
    add_writeoff_line(writeoff, lot, qty)
    return writeoff


def requisition_lot(project, lot, qty, user, dest_kind=None):
    """Мост «на баланс»: отпочковать остаток проекта в белый «Собственный склад».

    Один клик панели кладёт остаток в требование: на фиксации он уйдёт в 0 у проекта
    (`−ISSUE`) и появится лотом-потомком на балансе (`+RECEIPT`) — до неё склад не
    двигается (Ф15). Переиспользует последнее **черновое** требование в целевой склад;
    зафиксированное не трогаем.
    """
    if lot.project_id != project.id:
        raise ValidationError('Лот из другого проекта.')
    _require_no_closing_draft(lot)
    dest = _internal_project(dest_kind or models.Project.Kind.INTERNAL_STOCK)
    requisition = (models.Requisition.objects.filter(project=dest, locked=False)
                   .order_by('-id').first())
    if requisition is None:
        requisition = create_requisition(dest, user, _auto_number('ТРБ', dest))
    add_requisition_line(requisition, lot, qty)
    return requisition


# --------------------------------------------------------------------------- #
#  Правка шапки форм (сквозная, все документы): номер/дата/мягкие поля
# --------------------------------------------------------------------------- #
# Инлайн-правка несруктурных полей шапки прямо в форме (автосейв по полю,
# read-only под замком). Структурные якоря (проект/поставщик — дом лотов) не
# трогаем: их смена переселяет лоты, это отдельная операция, не инлайн.
def _apply(instance, updates):
    """Присвоить непустые поля (None → не трогаем) и сохранить изменённые."""
    fields = []
    for name, value in updates.items():
        if value is not None:
            setattr(instance, name, value)
            fields.append(name)
    if fields:
        instance.save(update_fields=fields)
    return instance


# Номер и дата шапки очищаются в черновике свободно (аудит-1, Б2б-4). Прежние
# `_require_number`/`_require_date` отклоняли пустую строку в PATCH — «заполнить можно,
# передумать нельзя», что расходилось с Ф12e: черновик имеет право быть неполным.
# Авторитетная per-kind политика обязательности живёт на модели
# (`StockDocument.REQUIRED_HEADER_BY_KIND`/`clean`, волна 13, Ф2d) и гейтится на
# ФИКСАЦИИ (`_require_header`) — там ошибиться документом уже поздно, а в черновике
# «стереть чужой номер» — обычный ход.
_UNSET = object()   # часовой «поле не передано» (отличает от «выставить None»)


def require_unique_code(model, code, pk=None):
    """Мягкая уникальность `code` (волна 19, Ф10, правило Ивана): плохой/занятый код
    ловим дружелюбно ДО IntegrityError, чтобы трение не уходило команде в оффлайн.

    `code` — уникальный короткий жаргон; занят в пределах своей сущности → отказ с
    внятным текстом (не 500). Пустой/`None` пропускаем (несколько NULL легальны). Для
    документов `model` = `StockDocument` (единое пространство кода на все ордера)."""
    code = (code or '').strip()
    if code and model.objects.filter(code=code).exclude(pk=pk).exists():
        raise ValidationError(f'Код «{code}» уже занят — выберите уникальный.')


def fallback_code(instance, label):
    """Проставить код-фолбэк «Поставка 12» только что рождённой сущности (Ф12e).

    «＋ Новый» больше не открывает форму создания — сущность рождается по клику, а
    `code` у половины сущностей `unique NOT NULL`: пустым его оставить нельзя.
    Решение Ивана 2026-07-28 — **единый шаблон везде** (`«{вид} {id}»`), включая
    ордера, где колонка допустила бы NULL: правило одно на всех, и титул формы
    (`code`, Ф11) нигде не вычисляется в обход данных.

    Код заведомо временный — человек перебьёт его своим жаргоном
    ([[code-identity-principle]]), поэтому уникальность добираем суффиксом, а не
    ошибкой: `id` уникален внутри сущности, но человек мог руками занять ровно
    «Изделие 12» раньше. Зовётся ПОСЛЕ `create` — до вставки `id` не существует.
    """
    base = f'{label} {instance.pk}'
    code, n = base, 1
    # Через `concrete_model` — у ордера это `StockDocument` (единое пространство
    # кода на семь видов); `Receipt.objects` фильтрует по `kind` и «Списание 12»
    # не увидел бы. У остальных сущностей proxy нет, и это тот же класс.
    model = instance._meta.concrete_model
    while model._default_manager.filter(code=code).exclude(pk=instance.pk).exists():
        n += 1
        code = f'{base}-{n}'
    instance.code = code
    instance.save(update_fields=['code'])
    return instance


def create_with_fallback_code(model, label, **fields):
    """Родить сущность, чей `code` — `unique NOT NULL`, без кода от человека (Ф12e).

    Шаблон фолбэка опирается на `id`, а `id` появляется только после вставки —
    отсюда два шага. Временный код — случайный: транзакция не даёт ему стать
    видимым снаружи, но защищает от коллизии двух одновременных рождений (пустая
    строка на этом месте оставила бы дыру шириной в один `INSERT`).
    """
    with transaction.atomic():
        obj = model.objects.create(code=uuid4().hex, **fields)
        return fallback_code(obj, label)


def _born_order(model, project, user, number='', date=None, **specifics):
    """Родить ордер любого вида (Ф12e) — одна точка на семь `create_*`.

    Что раньше отличало семь рождений — только вид и своя строка «Нужен № акта».
    Ф12e эти проверки снимает (номер обязателен к ФИКСАЦИИ, `REQUIRED_HEADER_BY_KIND`,
    а не к рождению), и различий не остаётся вовсе. `date` — не фолбэк-ложь, а факт:
    сегодня документ и правда заводят; к фиксации человек поправит его на дату УПД.
    Код — общий шаблон («Поставка 12»), пространство кода одно на все виды.
    """
    doc = model.objects.create(
        project=project, user=user, number=(number or '').strip(),
        date=date or timezone.localdate(), **specifics)
    return fallback_code(doc, models.DocumentKind(doc.kind).label)


def _set_code(instance, code):
    """Выставить `code` документа/сущности (волна 19, Ф10) под замком формы.

    `code` — часовой: `_UNSET` → не трогаем; иначе пустой → `NULL` (в MySQL несколько
    NULL не конфликтуют по unique, так очистка кода не ловит IntegrityError), непустой
    → строка. Уникальность — мягко (`require_unique_code`) на `StockDocument` (единое
    пространство кода ордеров). Отдельным `save` (как контрагент у передачи) — вне
    `_apply`, который пропускает None и не дал бы очистить код в NULL."""
    if code is not _UNSET:
        require_unique_code(models.StockDocument, code, instance.pk)
        instance.code = (code or '').strip() or None
        instance.save(update_fields=['code'])


def _set_author(doc, user):
    """Сменить автора документа (Ф2j) — сквозная правка шапки под замком.

    `user` — часовой: `_UNSET` → не трогаем; `User` → выставить. Автор обязателен
    (FK `StockDocument.user` NOT NULL), поэтому `None` отклоняем. Замок проверяет
    вызывающий (`update_*` уже гейтит `_require_unlocked`)."""
    if user is _UNSET:
        return
    if user is None:
        raise ValidationError('Автор документа обязателен.')
    doc.user = user
    doc.save(update_fields=['user'])


def _set_project(doc, project):
    """Сменить проект-якорь ордера (Ф2k) — вторая связка «Свода расхождений #A».

    Проект — **якорь**: `Lot.project` выводится из ордера-origin, а строки движения
    (`StockLine`) ссылаются на лоты того же проекта (движок гейтит эту чистоту при
    добавлении). Поэтому менять якорь можно только у **пустого** ордера — без
    born-лотов (`lots`) и строк (`lines`); иначе дружелюбный отказ: сперва удалить
    зависимые. `_UNSET` → не трогаем; `None` → отказ (FK NOT NULL); тот же проект →
    ноль-оп. Замок проверяет вызывающий (`update_*` уже гейтит `_require_unlocked`)."""
    if project is _UNSET:
        return
    if project is None:
        raise ValidationError('Проект ордера обязателен.')
    if project.pk == doc.project_id:
        return
    if doc.lots.exists() or doc.lines.exists():
        raise ValidationError(
            'Проект — якорь ордера: лоты и строки следуют за ним. Сначала удалите '
            'строки/лоты ордера, затем меняйте проект.')
    doc.project = project
    doc.save(update_fields=['project'])


def _set_target_item(kitting, item):
    """Сменить целевое изделие комплектации (Ф2k) — якорь #A, специфичный для kitting.

    Целевое изделие определяет состав (призрачные строки BOM-потребности) и рождаемый
    прибор, поэтому менять его можно только пока у комплектации нет строк пайки
    (`lines`) и рождённого прибора (`lots`). Иначе дружелюбный отказ."""
    if item is _UNSET:
        return
    if item is None:
        raise ValidationError('Целевое изделие комплектации обязательно.')
    _require_native_target(item)
    if item.pk == kitting.target_item_id:
        return
    if kitting.lines.exists() or kitting.lots.exists():
        raise ValidationError(
            'Целевое изделие определяет состав — сначала удалите строки пайки '
            'и рождённый прибор.')
    kitting.target_item = item
    fields = ['target_item']
    # Цель появилась → появилась и потребность: без кол-ва образцов состав считать не
    # из чего. Ставим 1, ровно как рождение СРАЗУ с целью (`kittings` POST:
    # `qty=… or (1 if target else None)`) — там это правило уже жило, а путь «родил по
    # клику, цель выбрал в форме» его не проходил и оставлял `qty` пустым.
    if kitting.qty is None:
        kitting.qty = Decimal('1')
        fields.append('qty')
    kitting.save(update_fields=fields)


def update_document(doc, number=None, date=None, code=_UNSET, description=None,
                    user=_UNSET, project=_UNSET, contractor=_UNSET, reason=None):
    """**Единая** правка шапки ордера (волна 19, Ф14).

    До Ф14 шесть видов несли шесть одинаковых `update_*`, различавшихся ровно двумя
    строками специфики (`contractor` у передачи, `reason` у списания) — различие было
    структурным (у каждого своя таблица), а не поведенческим. MTI снят, специфика
    живёт колонками одной таблицы — вместе с ним схлопывается и эта шестерня.
    Per-kind имена остаются тонкими делегатами ниже: на них завязан словарь вызова
    (`update_writeoff(w, reason=…)` читается лучше общего), но реализация одна.

    `contractor` — часовой: не передан → не трогаем; `Counterparty` → выставить;
    `None` → снять. Применим только к своим видам (`CONTRACTOR_KINDS`) — то же, что
    стережёт CHECK `doc_contractor_only_own_kinds`, но дружелюбной ошибкой вместо
    `IntegrityError`.

    `number`/`date` — не переданы (`None`) → не трогаем; пустая строка → **очистка**
    (`''` / `NULL`): в черновике можно передумать, обязательность стережёт фиксация
    (аудит-1, Б2б-4). Дата идёт мимо `_apply` — там `None` занят под «не передано».
    """
    _require_unlocked(doc)
    _set_author(doc, user)
    _set_project(doc, project)
    _set_code(doc, code)
    if contractor is not _UNSET:
        if doc.kind not in models.CONTRACTOR_KINDS:
            raise ValidationError(
                'Контрагент есть только у поставки (поставщик) и передачи (заказчик).')
        doc.contractor = contractor
        doc.save(update_fields=['contractor'])
    if date is not None:
        doc.date = date or None         # '' → NULL (как у заказа, `update_purchase`)
        doc.save(update_fields=['date'])
    return _apply(doc, {
        'number': None if number is None else str(number).strip(),
        'reason': None if reason is None else reason.strip(),
        'description': None if description is None else description.strip(),
    })


def update_receipt(receipt, **kw):
    """Шапка поставки: № УПД / дата / код / описание / автор / проект-якорь."""
    return update_document(receipt, **kw)


def update_purchase(purchase, date=None, code=_UNSET, description=None, user=_UNSET,
                    project=_UNSET, procurement=_UNSET, contractor=_UNSET):
    """Правка шапки заказа (дата / код / описание / автор / проект / закупка /
    контрагент). Только в черновике.

    Дата заказа nullable — пустая строка очищает её в NULL (в отличие от
    документов с обязательной датой). `code` — часовой `_UNSET` (пустой → NULL, чтобы
    несколько заказов без кода не конфликтовали по unique).

    Якоря #A (Ф2k): `project` обязателен — заказ проектный по определению, `None`
    отклоняем. `procurement` и `contractor` — часовые (Ф17): не переданы → не трогаем,
    сущность → выставить, `None` → снять. Закупка-план опциональна (заказ бывает и без
    плана), контрагент обязателен к ФИКСАЦИИ, а не к правке черновика. Смена проекта у
    заказа со связанными приходами ломает инвариант «УПД ↔ проект заказа» → дружелюбный
    отказ (сперва отвязать приходы).
    """
    _require_unlocked(purchase, PURCHASE_LOCKED)
    _set_author(purchase, user)
    fields = []
    if project is not _UNSET:
        if project is None:
            raise ValidationError('Проект заказа обязателен.')
        if project.pk != purchase.project_id and purchase.receipts.exists():
            raise ValidationError(
                'К заказу привязаны приходы (УПД ↔ проект) — сначала отвяжите их, '
                'затем меняйте проект.')
        purchase.project = project
        fields.append('project')
    if procurement is not _UNSET:
        purchase.procurement = procurement
        fields.append('procurement')
    if contractor is not _UNSET:
        purchase.contractor = contractor
        fields.append('contractor')
    if date is not None:
        purchase.date = date or None
        fields.append('date')
    if code is not _UNSET:
        require_unique_code(models.Purchase, code, purchase.pk)
        purchase.code = (code or '').strip() or None
        fields.append('code')
    if description is not None:
        purchase.description = description.strip()
        fields.append('description')
    if fields:
        purchase.save(update_fields=fields)
    return purchase


def update_transfer(transfer, **kw):
    """Шапка передачи: № накладной / дата / код / описание / **заказчик** / автор / проект."""
    return update_document(transfer, **kw)


def update_writeoff(writeoff, **kw):
    """Шапка списания: № акта / дата / **причина** / код / описание / автор / проект."""
    return update_document(writeoff, **kw)


def update_requisition(requisition, **kw):
    """Шапка требования: № / дата / код / описание / автор / проект-получатель."""
    return update_document(requisition, **kw)


def update_relocation(relocation, **kw):
    """Шапка перемещения: № / дата / код / описание / автор / проект-якорь.

    Проект — якорь (`_set_project`): у перемещения строки-ходы ссылаются на лоты того
    же проекта, поэтому сменить его можно лишь у пустого ордера."""
    return update_document(relocation, **kw)


def update_kitting(kitting, qty=None, date=None, code=_UNSET, description=None,
                   user=_UNSET, project=_UNSET, target_item=_UNSET):
    """Правка шапки комплектации (кол-во образцов / дата / код / описание / автор / проект / цель).
    Только «в работе».

    Кол-во образцов пересчитывает потребности BOM — правится, пока `wip`. `project`/
    `target_item` — якоря #A (Ф2k): меняются только у пустой комплектации.
    """
    _require_unlocked(kitting)
    _set_author(kitting, user)
    _set_project(kitting, project)
    _set_target_item(kitting, target_item)
    _set_code(kitting, code)
    if qty is not None and qty <= 0:
        raise ValidationError('Количество образцов должно быть положительным.')
    fields = []
    if qty is not None:
        kitting.qty = qty
        fields.append('qty')
    if date is not None:                 # дата комплектации nullable
        kitting.date = date or None
        fields.append('date')
    if description is not None:
        kitting.description = description.strip()
        fields.append('description')
    if fields:
        kitting.save(update_fields=fields)
    return kitting


# --------------------------------------------------------------------------- #
#  Волна 7 — планирование закупок: командный свод + записываемый Procurement
# --------------------------------------------------------------------------- #
def counts_in_scope(project):
    """Проект участвует в арифметике охвата: **активный внешний**.

    Внутренние склады — источник покрытия, а не потребитель; закрытый проект не
    закупают. Правило одно на весь охват (свод «К закупке» и деньги закупки считают
    один и тот же набор — иначе панель и витрина разошлись бы молча).
    """
    return project.kind == models.Project.Kind.EXTERNAL and not project.locked


def scope_deficit(projects):
    """Дефицит по **охвату**: суммарная нужда по оси Item через заданный набор проектов.

    Консолидация-проекция (не таблица): для каждого проекта считаем потребность по
    покупным **листьям** (Ф5 В16: разузлование насквозь, а не 1 уровень) и покрываем её
    складом/заказами **этого** проекта (`_coverage`, как дефицит проекта — покрытие на
    уровне Item в проекте, агрегат), затем складываем сегменты по Item через проекты.
    Между проектами **не** перенеттим (чужие ФЛС/склады не смешиваем): профицит проекта
    A не гасит нужду проекта B. Итог по Item: `to_order` = сколько всего докупить (▲
    красный член), `have`/`on_order` — контекст.

    Волна 19, Ф13: набор проектов приходит **снаружи** — это охват конкретной закупки
    (`procurement_scope`: проекты её заказов), а не «все активные внешние». Пустой охват
    → пустой результат: «общего через отсутствие» в продукте больше нет, и слепая
    закупка честнее закупки, молча считающей за всю организацию. Read-only витрина.

    Считаем только по **активным внешним** проектам охвата: внутренние склады —
    источник покрытия, а не потребитель, а закрытый проект не закупают (пегать на него
    движок и так отказывается — наводка на непегаемое была бы враньём). Отсев — общий
    предикат `counts_in_scope`: это инвариант самой арифметики охвата, и деньги закупки
    считают по тому же набору.
    """
    acc = {}  # item_id → агрегат по Item через проекты
    for project in projects:
        if not counts_in_scope(project):
            continue
        # потребность проекта по покупным листьям (разузлование насквозь)
        need_by_item, _incomplete = project_leaf_demand(project)
        for component, need in need_by_item.items():
            cov = _coverage(need, item_available(component, project),
                            item_on_order(component, project))
            row = acc.setdefault(component.id, {
                'item_id': component.id, 'item_code': component.code,
                'item_description': component.description, 'uom': component.uom,
                'native': component.native,
                'need': ZERO, 'have': ZERO, 'on_order': ZERO, 'to_order': ZERO,
                'by_project': [],
            })
            row['need'] += cov['need']
            row['have'] += cov['have']
            row['on_order'] += cov['on_order']
            row['to_order'] += cov['to_order']
            row['by_project'].append({
                'project_id': project.id, 'project_code': project.code,
                'project_name': project.description, 'need': cov['need'],
                'have': cov['have'], 'on_order': cov['on_order'],
                'to_order': cov['to_order'], 'status': cov['status'],
            })
    rows = []
    for row in acc.values():
        # статус Item = тот же словарь по сегментам итога (worst-of)
        if row['to_order'] > 0:
            row['status'] = 'to_order'
        elif row['on_order'] > 0:
            row['status'] = 'on_order'
        else:
            row['status'] = 'available'
        rows.append(row)
    # худшее наверх (красное просит внимания), потом по артикулу
    rows.sort(key=lambda r: (-_WORST_RANK[r['status']], r['item_code']))
    return {'rows': rows}


def procurement_scope(procurement):
    """Охват закупки — проекты её заказов, по коду. **Вычисляемый** (2026-08-05).

    Раньше охват был M2M-галочками (Ф13) и жил параллельно с тем же фактом, записанным
    через `Purchase.project`. Две дороги к одному отношению — верный способ разойтись, и
    они расходились молча. Теперь дорога одна: завёл заказ под проект, привязал его к
    закупке — проект в охвате. Хранить нечего, и «пусто = пусто» осталось само собой:
    нет заказов → нет охвата → закупка честно ничего не считает.

    Единственный вход в охват для всего движка (витрина «К закупке»), поэтому «что
    считаем» по-прежнему задано в одном месте.
    """
    return models.Project.objects.filter(
        purchases__procurement=procurement).distinct().order_by('code')


def procurement_deficit(procurement):
    """Дефицит по охвату закупки — таб «К закупке» её формы (бывший «командный свод»).

    Свод перестал быть отдельным экраном-фантомом (без `code`, без замка, вечно один):
    он стал витриной конкретной закупки, суженной её охватом. Отметить все проекты =
    прежний общий свод, отметить один = закупка под проект без шума остальных.

    К каждой строке добавляется `planned` — сколько этого Item **уже в строках плана**:
    витрина внутри закупки обязана показывать, что из наводки уже взято, иначе мост
    «＋ в план» вслепую.
    """
    data = scope_deficit(procurement_scope(procurement))
    planned = dict(procurement.lines.values_list('item_id', 'qty'))
    for row in data['rows']:
        row['planned'] = planned.get(row['item_id'], ZERO)
    return data


def procurement_form(procurement):
    """Проекция формы закупки-плана: шапка + охват проектов + строки (`item`, `qty`) + итог.

    `Procurement` — план **под охват проектов** (волна 19, Ф13; до неё — «без проекта,
    командная высота»); нарезка на проектные `Purchase` (pegging) — волна 8. Мягкий замок
    `status` зеркалит заказ: строки правятся только пока не зафиксировано. Чистая проекция.
    """
    editable = not procurement.locked
    lines = list(procurement.lines.select_related('item').order_by('id'))
    rows = []
    total_qty = ZERO
    for line in lines:
        total_qty += line.qty
        rows.append({
            'id': line.id, 'item_id': line.item_id, 'item_code': line.item.code,
            'item_description': line.item.description, 'uom': line.item.uom, 'qty': line.qty,
            'item_native': line.item.native, 'item_synced': line.item.synced,
            'item_locked': line.item.locked,   # глиф строки по режиму (Ф3a)
        })
    scope = list(procurement_scope(procurement))
    # Деньги намерения (2026-08-07): потребность / закупка / переплата. Знаменатель —
    # нужда ОХВАТА (у закупки своего проекта нет), по тому же набору, что свод «К
    # закупке». Охват пуст (заказов ещё нет) → нужда 0, и вся сумма плана читается
    # переплатой: спросить «сколько надо» закупке пока не у кого, и она это показывает.
    money = intent_money(lines, demand_map(p for p in scope if counts_in_scope(p)))
    return {
        'id': procurement.id, **_author(procurement), 'locked': procurement.locked,
        'code': procurement.code, 'description': procurement.description,
        'date': procurement.date,
        **_contractor_view(procurement.contractor),          # Ф4: поставщик (Р3)
        # Охват — область расчёта витрины «К закупке». Вычисляемый (проекты заказов),
        # правке не подлежит: поле шапки показывает его ссылками и только.
        'projects': [{'id': p.id, 'code': p.code, 'description': p.description}
                     for p in scope],
        'editable': editable,                       # строки правятся только пока не зафиксировано
        'total_qty': total_qty,
        **money,          # demand / estimate / overpay / unestimated — панель бюджета
        'lines': rows,
    }


def create_procurement(user, date=None, code=None, description=''):
    """Создать закупку-план (черновик) без проекта.
    Пустой код — фолбэком «Закупка 12» (Ф12e)."""
    require_unique_code(models.Procurement, code)
    p = models.Procurement.objects.create(
        user=user, locked=False,
        date=date, code=code, description=(description or '').strip())
    return p if code else fallback_code(p, 'Закупка')


PROCUREMENT_LOCKED = (
    'Строки правятся только в черновике закупки — снимите замок (unlock).')


def add_procurement_line(procurement, item, qty):
    """Добавить строку закупки-плана (только в черновике). `(procurement, item)` — одна строка."""
    _require_unlocked(procurement, PROCUREMENT_LOCKED)
    if qty is None or qty <= 0:
        raise ValidationError('Количество закупки должно быть положительным.')
    if procurement.lines.filter(item=item).exists():
        raise ValidationError(
            f'Изделие {item.code} уже в закупке — правьте существующую строку.')
    return models.ProcurementLine.objects.create(
        procurement=procurement, item=item, qty=qty)


def update_procurement_line(line, qty):
    """Автосейв количества строки закупки-плана (только в черновике)."""
    _require_unlocked(line.procurement, PROCUREMENT_LOCKED)
    if qty is None or qty <= 0:
        raise ValidationError('Количество закупки должно быть положительным.')
    line.qty = qty
    line.save(update_fields=['qty'])
    return line


def remove_procurement_line(line):
    """Удалить строку закупки-плана (только в черновике)."""
    _require_unlocked(line.procurement, PROCUREMENT_LOCKED)
    line.delete()


def lock_procurement(procurement):
    """Поставить замок закупки-плана — строки read-only.

    Волна 19 (Ф1): бывший `send_procurement`; отмена = удаление (Р1).
    """
    if not procurement.lines.exists():
        raise ValidationError('Нельзя утвердить пустую закупку — добавьте строку.')
    procurement.locked = True
    procurement.save(update_fields=['locked'])
    return procurement


def unlock_procurement(procurement):
    """Снять замок закупки-плана — ничего не разрушает."""
    procurement.locked = False
    procurement.save(update_fields=['locked'])
    return procurement


def update_procurement(procurement, date=None, code=_UNSET, description=None, user=_UNSET,
                       contractor=_UNSET):
    """Правка шапки закупки-плана (дата / код / описание / автор / контрагент). Только в черновике.

    Дата закупки nullable — пустая строка очищает её в NULL (как заказ). `code` —
    часовой `_UNSET` (пустой → NULL, чтобы несколько закупок без кода не конфликтовали
    по unique). `contractor` — часовой (волна 19, Ф4): не передан → не трогаем;
    `Counterparty` → выставить; `None` → снять (nullable). Охвата тут нет вовсе — он
    вычисляемый (проекты заказов закупки), править нечего.
    """
    _require_unlocked(procurement, PROCUREMENT_LOCKED)
    _set_author(procurement, user)
    fields = []
    if date is not None:
        procurement.date = date or None
        fields.append('date')
    if code is not _UNSET:
        require_unique_code(models.Procurement, code, procurement.pk)
        procurement.code = (code or '').strip() or None
        fields.append('code')
    if description is not None:
        procurement.description = description.strip()
        fields.append('description')
    if contractor is not _UNSET:
        procurement.contractor = contractor
        fields.append('contractor')
    if fields:
        procurement.save(update_fields=fields)
    return procurement


def delete_procurement(procurement):
    """Удалить закупку-план (WAVE14 Ф2). Мягкий замок: утверждённую сперва вернуть в
    черновик (снять замок); привязанные заказы (`Purchase.procurement`, PROTECT)
    держат — их сперва открепить/удалить. Строки плана (`ProcurementLine`) — каскад."""
    if procurement.locked:
        raise ValidationError(
            'Закупка утверждена — сперва верните её в черновик (снимите замок), затем удаляйте.')
    if procurement.purchases.exists():
        raise ValidationError('К закупке привязаны заказы — удаление заблокировано.')
    for att in procurement.attachments.all():      # физические файлы (каскад их сиротит)
        delete_attachment(att)
    try:
        procurement.delete()                       # каскад: строки закупки
    except ProtectedError:
        raise ValidationError('Закупка связана с другими записями — удаление заблокировано.')


# Волна 19, Ф17: `_plan_procurements` удалён — список закупок = ВСЕ закупки.
# Эвристика «есть заказы, но нет строк плана» существовала ровно затем, чтобы прятать
# закупки-пустышки, которые `_solo_procurement` плодил на каждый одиночный заказ (плата
# за NOT NULL у `Purchase.procurement`). Стал nullable — пустышек не бывает, и прятать
# нечего: фильтр и заглушка умерли одной правкой.


def add_to_procurement(procurement, item, qty):
    """Мост «витрина К закупке → строки плана»: довести строку плана до наводки.

    Волна 19, Ф13: мост стал **внутренним**. Раньше свод был отдельным экраном и не
    знал, в какую закупку класть, — отсюда магия «найти-или-создать последний
    черновик»; теперь витрина живёт внутри конкретной закупки, и класть надо в неё.
    Топ-ап, а не инкремент: повторный клик по той же строке ничего не удваивает, а
    руками набранное сверх наводки не срезается.
    """
    _require_unlocked(procurement, PROCUREMENT_LOCKED)
    if qty is None or qty <= 0:
        raise ValidationError('Количество должно быть положительным.')
    line = procurement.lines.filter(item=item).first()
    if line is None:
        models.ProcurementLine.objects.create(
            procurement=procurement, item=item, qty=qty)
    elif line.qty < qty:
        line.qty = qty
        line.save(update_fields=['qty'])
    return procurement


def our_organization():
    """Наша сторона — **обычный контрагент**, чей ID знает окружение (волна 19, Ф4b).

    Решение Ивана 2026-07-30: не константа в коде и не отдельная таблица настроек, а
    запись `Counterparty`. Тогда себя правят той же формой, что всех, у нас те же
    `description`/`inn` и та же «карточка предприятия» во вложениях — а окружение лишь
    знает, кто из справочника «мы» (`settings.ORG_COUNTERPARTY_ID`). Название
    организации в репозиторий не попадает, оно живёт в данных и в `.env`.

    `None` — настройка не задана или указывает на удалённую запись. Это не ошибка:
    выгрузка обязана работать и без неё (см. `procurement_xlsx`).
    """
    oid = getattr(settings, 'ORG_COUNTERPARTY_ID', None)
    return models.Counterparty.objects.filter(pk=oid).first() if oid else None


def _party_line(cp):
    """Сторона одной строкой для бланка: наименование + ИНН, если он известен."""
    return f'{cp.description} · ИНН {cp.inn}' if cp.inn else cp.description


def _doc_date(value):
    """Дата для документа НАРУЖУ — `дд.мм.гггг` (тот же формат, что в просмотре формы).

    Принимает и `date`, и ISO-строку: сущность, только что созданная или поправленная в
    этом же запросе, держит в памяти ровно то, что пришло из JSON (`'2026-07-30'`), пока
    её не перечитали из БД. Нераспознанное значение даёт пустую строку — сырой ISO или
    мусор в бланке у поставщика хуже отсутствующей графы.
    """
    if isinstance(value, str):
        try:
            value = dt_date.fromisoformat(value)
        except ValueError:
            return ''
    return value.strftime('%d.%m.%Y') if value else ''


def _xlsx_sheet(ws, title, head, headers, rows, widths):
    """Один лист выгрузки: шапка (B/C) → пустая строка → нумерованная таблица.

    Форма листа поднята из `procurement_xlsx` (2026-07-30, выгрузка изделия): бланк
    закупки и выгрузка изделия рисуют одно и то же — подписи шапки в колонке **B**,
    значения в **C** (колонка A узкая, она под «№», подпись в ней Excel обрезал бы по
    занятой соседке), жирные заголовки таблицы и сквозная нумерация строк.

    `head` — пары `(подпись, значение)`; **пустые строки шапки не рисуются вовсе**:
    пустая графа во внешнем документе читается как брак, а не как «не задано» (внутри
    продукта правило обратное, §13 — там прочерк честнее). `rows` идут БЕЗ номера —
    его проставляет лист.
    """
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    bold = Font(bold=True)
    ws.title = title
    for label, value in head:
        if value is None or value == '':
            continue
        ws.append([None, label, value])
        ws.cell(row=ws.max_row, column=2).font = bold
    if head:
        ws.append([])                  # пустая строка отбивает шапку от таблицы
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = bold
    for n, row in enumerate(rows, start=1):
        ws.append([n, *row])
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    return ws


def procurement_xlsx(procurement):
    """Сгенерировать xlsx-бланк закупки-плана (bytes) — файл поставщику.

    **Документ наружу** (с 2026-07-20 уходит реальным контрагентам), поэтому состав
    меняется осознанно и разом, а не побочным эффектом правки формы. Волна 19, Ф4b —
    первое такое изменение: над таблицей появилась шапка запроса (кто просит, у кого,
    когда, кто автор), а у позиций — порядковый номер.

    Шапка: **заказчик** — наша сторона (`our_organization`, обычный контрагент);
    **контрагент** — `Procurement.contractor` (намерение плана: у кого собираемся
    купить); **дата запроса** — `Procurement.date`; **автор** — `Procurement.user`.
    Незаданные строки шапки **не рисуются вовсе**: пустая графа во внешнем документе
    читается как брак, а не как «значение не задано» (внутри продукта правило обратное,
    §13 — там прочерк честнее).

    **Порядок позиций = порядок на экране** (`order_by('id')`, тот же, что у
    `procurement_form`): нумерация в файле обязана совпадать с тем, что видит человек,
    иначе «позиция 3» в переписке означает разное у нас и у поставщика.

    Синхронно в запросе (файл небольшой, тяжёлых рантаймов нет). openpyxl — импорт
    ленивый (зависимость только ради экспорта).
    """
    from openpyxl import Workbook

    us = our_organization()
    wb = Workbook()
    _xlsx_sheet(
        wb.active, 'Заказ',
        head=[
            ('Заказчик', _party_line(us) if us else ''),
            ('Контрагент', _party_line(procurement.contractor)
                           if procurement.contractor_id else ''),
            ('Дата запроса', _doc_date(procurement.date)),
            ('Автор', procurement.user.get_full_name() or procurement.user.get_username()),
        ],
        headers=['№', 'Артикул', 'Наименование', 'Кол-во', 'Ед.'],
        rows=[(line.item.code, line.item.description, float(line.qty), line.item.uom)
              for line in procurement.lines.select_related('item').order_by('id')],
        widths=[5, 22, 48, 12, 8])       # «№» узкая; остальные — по содержимому
    return _xlsx_bytes(wb)


def _xlsx_bytes(wb):
    """Книга → bytes (в HTTP-ответ отдаём тело, файлов на диске не заводим)."""
    from io import BytesIO

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Листы выгрузки изделия = ВКЛАДКИ его формы (кроме «Файлов» — вложения в книгу не
# кладём). Порядок тот же, что на экране: состав → применение → склад → движения.
ITEM_XLSX_BOM = 'bom'          # «Только состав» — одна вкладка
ITEM_XLSX_ALL = 'all'          # «Все вкладки»


def item_xlsx(item, scope=ITEM_XLSX_BOM):
    """Сгенерировать xlsx-выгрузку изделия (bytes) — снимок его вкладок.

    Решение Ивана 2026-07-30. Файл повторяет ЭКРАН: колонки названы так же, как в
    табах формы, порядок листов — порядок вкладок. Два режима (пункты меню кнопки
    «Скачать»): `bom` — один лист «Состав»; `all` — все вкладки, кроме «Файлов».

    Лист «Состав» есть ВСЕГДА, даже у покупного компонента без BOM (у него и вкладки
    такой нет): пустой лист с заголовками честнее отказа скачать — человек видит, что
    состава нет, а не гадает, почему кнопка молчит. Он же первый — и потому несёт
    **шапку изделия** (код, описание, категория, единицы, оценка); остальные листы
    чистые таблицы, идентичность им даёт имя файла (= `code`).

    **Состав — один уровень**, ровно как на вкладке: разузлование до листьев — другой
    документ (сводная спецификация), и мешать их в одном файле нельзя.

    Синхронно в запросе, openpyxl — ленивым импортом (как `procurement_xlsx`).
    """
    from openpyxl import Workbook

    wb = Workbook()
    _xlsx_sheet(
        wb.active, 'Состав',
        head=[
            ('Код', item.code),
            ('Описание', item.description),
            ('Категория', item.category.code if item.category_id else ''),
            ('Единицы', item.uom),
            ('Оценка', float(item.estimated_cost)
                       if item.estimated_cost is not None else ''),
        ],
        headers=['№', 'Компонент', 'Описание', 'Кол-во', 'Ед.'],
        rows=[(bl.component.code, bl.component.description,
               float(bl.qty), bl.component.uom)
              for bl in item.bom_lines.select_related('component')],
        widths=[5, 22, 48, 12, 8])
    if scope != ITEM_XLSX_ALL:
        return _xlsx_bytes(wb)

    _xlsx_sheet(
        wb.create_sheet(), 'Применение',
        head=[],
        headers=['№', 'Изделие', 'Описание', 'Кол-во', 'Ед.'],
        # Кол-во — вхождение изделия в родителя, единица поэтому НАША (как в табе).
        rows=[(bl.parent.code, bl.parent.description, float(bl.qty), item.uom)
              for bl in item.used_in.select_related('parent')],
        widths=[5, 22, 48, 12, 8])
    _xlsx_sheet(
        wb.create_sheet(), 'Склад',
        head=[],
        headers=['№', 'Партия', 'Проект', 'Рожд.', 'Остаток', 'Ед.',
                 'Part number', 'Название'],
        # Остаток партии черновика — прочерк, а не 0 (Ф15: «едет, ещё не принято»).
        rows=[(f'#{lot.id}', lot.project.code, float(lot.qty),
               float(lot_live_qty(lot)) if lot.origin.locked else '—',
               item.uom, lot.part_number, lot.lot_name)
              for lot in item.lots.select_related('project', 'origin')],
        widths=[5, 10, 16, 12, 12, 8, 22, 32])
    _xlsx_sheet(
        wb.create_sheet(), 'Движения',
        head=[],
        headers=['№', 'Ордер', 'Вид', 'Дата', 'Проект', 'Партия', 'Кол-во', 'Ед.'],
        rows=[(_movement_order_label(m), _movement_kind_label(m), _doc_date(m['date']),
               m['project_code'], _movement_lot_label(m), float(m['qty']), item.uom)
              for m in item_movements(item)],
        widths=[5, 22, 24, 12, 16, 28, 12, 8])
    return _xlsx_bytes(wb)


def _movement_order_label(m):
    """Ярлык ордера в ленте — тот же фолбэк, что на экране: код → номер → «Вид #id»."""
    return (m['code'] or m['number']
            or f'{models.DocumentKind(m["kind"]).label} #{m["document_id"]}')


def _movement_kind_label(m):
    """Вид ордера; у рождения партии — с пометкой (на экране она идёт подсказкой)."""
    label = models.DocumentKind(m['kind']).label
    return f'{label} · партия рождена' if m['event'] == 'born' else label


def _movement_lot_label(m):
    """Партия в ленте: `#id`, плюс имя из УПД, когда оно есть."""
    return f'#{m["lot_id"]} · {m["lot_name"]}' if m['lot_name'] else f'#{m["lot_id"]}'


# --------------------------------------------------------------------------- #
#  Волна 8 — pegging: нарезка плана (Procurement) на проектные заказы (Purchase)
# --------------------------------------------------------------------------- #
def _plan_allocations(procurement):
    """Разложенное под этой закупкой: `{(purchase_id, item_id): qty}` — одним запросом.

    Ячейка раскладки = строка проектного заказа (`PurchaseLine`), пара
    `(purchase, item)` уникальна по схеме, поэтому ячейка ровно одна.
    """
    return {(pid, iid): qty for pid, iid, qty in (
        models.PurchaseLine.objects
        .filter(purchase__procurement=procurement)
        .values_list('purchase_id', 'item_id', 'qty'))}


def _allocation_status(planned, allocated):
    """Глиф строки плана: разложена ли она (правило Ивана 2026-08-05).

    Ровно в ноль → зелёный check. Не тронута (0) → красный warning. Перепег (разложено
    больше плана — движок его разрешает) → тоже красный: это расхождение, а не прогресс.
    Всё между → оранжевый warning: тронуто, но не подбито.
    """
    if allocated == 0:
        return 'to_order'
    if allocated == planned:
        return 'available'
    if allocated > planned:
        return 'to_order'
    return 'on_order'


def procurement_allocation(procurement):
    """Проекция «Привязка»: строка плана × заказы закупки + баланс проекта в ячейке.

    Раскрытие строки даёт **заказы** (2026-08-05), а не проекты. Проект — свойство
    заказа, а не уровень закупочного контура (Закупка = ЧТО, Заказ = У КОГО, Поставка =
    КТО привёз), поэтому раскладка «из ЧТО в У КОГО» идёт по заказам напрямую. Матрица
    получается прямоугольной и без скрытого состояния: ячейка ↔ `PurchaseLine`, пустая
    ячейка — тоже факт («в этот заказ не кладём»).

    Что ушло вместе с проектной осью: выбор заказа дропдауном, ленивое рождение заказа
    под пег, «найти-или-создать черновик», autopeg и обратное разузлование внутрь
    проекта (шумело, не пригодилось). Заказы заводятся руками в своей форме и
    привязываются к закупке — намерение сопровождается действием и следом в БД.

    `balance` ячейки — тот же баланс, что в «Потребности» проекта (`_balance`): считает
    только зафиксированное, поэтому черновая раскладка его НЕ двигает и он не гаснет,
    пока набиваешь. Гаснет на фиксации заказа — в осмысленный момент. Два экрана
    показывают одно число: в проекте для общей оценки, здесь для конкретного решения.

    `cost` (2026-08-07) есть на ОБОИХ уровнях — `qty × estimated_cost` того документа,
    чья это строка (у плана — в закупке, у ячейки — в заказе; та же двухуровневость, что
    у «Кол-ва»). Знаменатель для цвета у каждого уровня свой и честный: строка плана
    меряется нуждой ОХВАТА, ячейка — нуждой своего проекта.
    """
    allocations = _plan_allocations(procurement)
    orders = list(procurement.purchases.select_related('project')
                  .order_by('project__code', 'id'))
    # Потребность проекта по листьям — по одному разузлованию на проект, не на ячейку.
    demand_by_project = {}
    for pu in orders:
        if pu.project_id not in demand_by_project:
            leaves, _incomplete = project_leaf_demand(pu.project)
            demand_by_project[pu.project_id] = {leaf.id: qty for leaf, qty in leaves.items()}

    # Нужда ОХВАТА для строк плана: у строки плана своего проекта нет, и её деньги
    # меряются тем же знаменателем, что панель бюджета закупки (отсев `counts_in_scope`,
    # иначе панель и таб разошлись бы). Складываем уже посчитанные разузлования — второй
    # раз по BOM не ходим.
    scope_need = {}
    counted = set()
    for pu in orders:
        if pu.project_id in counted or not counts_in_scope(pu.project):
            continue
        counted.add(pu.project_id)
        for item_id, need in demand_by_project[pu.project_id].items():
            scope_need[item_id] = scope_need.get(item_id, ZERO) + need

    rows = []
    for line in procurement.lines.select_related('item').order_by('id'):
        cells = []
        allocated = ZERO
        for pu in orders:
            qty = allocations.get((pu.id, line.item_id), ZERO)
            allocated += qty
            need = demand_by_project[pu.project_id].get(line.item_id, ZERO)
            bal = _balance(need, item_kitted(line.item, pu.project),
                           item_available(line.item, pu.project),
                           item_on_order(line.item, pu.project))
            cells.append({
                'purchase_id': pu.id, 'purchase_code': pu.code or f'Заказ {pu.id}',
                'locked': pu.locked,
                'project_id': pu.project_id, 'project_code': pu.project.code,
                'project_name': pu.project.description,
                'qty': qty,
                # баланс проекта по этому Item + слагаемые (расшифровка под курсором)
                'need': bal['need'], 'kitted': bal['kitted'], 'in_stock': bal['in_stock'],
                'on_order': bal['on_order'], 'balance': bal['balance'],
                'balance_status': bal['status'],
                # Деньги ячейки (2026-08-07): её стоимость и перебор относительно нужды
                # ЭТОГО проекта. Оранжевого «дорогая» тут нет — верхнюю четверть ищем
                # среди строк плана, где строки сопоставимы (решение Ивана).
                'cost': line_cost(line.item, qty),
                'overpay_at': overpay_threshold(bal['need']),
                'cost_status': ('overpaid'
                                if line.item.estimated_cost is not None
                                and is_overpaid(qty, bal['need']) else None),
            })
        rows.append({
            'line_id': line.id, 'item_id': line.item_id,
            'item_code': line.item.code, 'item_description': line.item.description,
            'uom': line.item.uom, 'qty': line.qty,
            'allocated': allocated, 'remaining': line.qty - allocated,
            'status': _allocation_status(line.qty, allocated),
            # Деньги строки плана: знаменатель — нужда ОХВАТА (см. `scope_need` выше).
            'cost': line_cost(line.item, line.qty),
            'need': scope_need.get(line.item_id, ZERO),
            'orders': cells,
        })
    mark_costs(rows)                   # `cost_status` строк плана: overpaid > costly
    fan = []
    for pu in orders:
        fan.append({
            'purchase_id': pu.id, 'purchase_code': pu.code or f'Заказ {pu.id}',
            'locked': pu.locked,
            'project_id': pu.project_id, 'project_code': pu.project.code,
            'project_name': pu.project.description, 'lines': pu.lines.count(),
            'total': pu.lines.aggregate(s=Sum('qty'))['s'] or ZERO,
        })
    return {
        'id': procurement.id, 'locked': procurement.locked,
        'rows': rows, 'fan': fan,
    }


def set_allocation(procurement, purchase, item, qty):
    """Положить в ячейку раскладки ровно `qty` — присвоение, не добавление.

    Ввод в поле = состояние строки заказа, а не дельта к ней: `0` (или пусто) снимает
    строку, число — ставит его. Так «разложить план по заказам» становится правкой
    матрицы, а не серией жестов ✓/＋/корзина, каждый из которых надо было подтверждать.

    Замок заказа гейтит правку: зафиксированный заказ — обязательство, его строки
    меняют в его собственной форме, расфиксировав. Количество не клампим по остатку
    плана (перепег законен и информативен — в духе мутабельной ДНК), но строка плана
    покажет его красным.
    """
    if qty is None or qty < 0:
        raise ValidationError('Количество не может быть отрицательным.')
    if purchase.procurement_id != procurement.id:
        raise ValidationError('Заказ заведён под другой закупкой.')
    if purchase.locked:
        raise ValidationError(
            'Заказ зафиксирован — расфиксируйте его в форме заказа, потом правьте.')
    if not procurement.lines.filter(item=item).exists():
        raise ValidationError(
            f'Изделие {item.code} не в плане закупки — сначала добавьте строку плана.')
    line = purchase.lines.filter(item=item).first()
    if qty == 0:
        if line:
            line.delete()
        return procurement
    if line:
        line.qty = qty
        line.save(update_fields=['qty'])
    else:
        models.PurchaseLine.objects.create(purchase=purchase, item=item, qty=qty)
    return procurement


# --------------------------------------------------------------------------- #
#  Волна 20 — контрагент: справочник + витрина двух сторон документооборота
# --------------------------------------------------------------------------- #
# `Counterparty` — единая внешняя сторона документооборота, и витрина устроена по её
# оси: **две стороны**, а не одна лента. Закупочная сторона (мы платим, к нам едет) —
# закупки → заказы → поставки; передачная (мы отдаём) — передачи заказчику. Стороны не
# взаимоисключающие: одно юрлицо законно и привозит нам, и принимает от нас.
#
# Каждая сторона отдаётся как **интеграл или `None`** (решение Ивана 2026-07-30):
# «ноль движений — стороны нет». Тот же приём, что `project_health` (внутренний склад
# → `None`, глиф нейтрален): пустоту решает ДВИЖОК, вью только не рисует панель.
#
# **Ролей-флагов не существует** (снесены 2026-07-30, Ф3): сторона — не свойство
# справочника, а факт документооборота. Два bool были декларацией о намерениях,
# которую человек поддерживал руками, тогда как правда всегда лежала в документах;
# хуже того, пикер по ним ФИЛЬТРОВАЛ и прятал нужную запись. Свежезаведённый
# контрагент честно показывает пустую форму (ни панелей, ни табов контура) — он
# заводится ровно тогда, когда ему собираются оформить заказ, и пустым не живёт.


def counterparty_sides(queryset):
    """Аннотировать контрагентов их сторонами ПО ФАКТАМ: `has_supply` / `has_shipment`.

    Заменила пару флагов-ролей: «поставщик» = у него что-то покупали (закупка-план,
    заказ или поставка), «заказчик» = ему что-то передавали. Один запрос на весь
    список (`Exists`, не счёт), потому что это ось СПИСКА режима — глиф строки
    (`fold-*`) и порядок в пикере («свои» для этого вида ордера — наверх).
    """
    receipts = models.StockDocument.objects.filter(
        contractor=OuterRef('pk'), kind=models.StockDocument.Kind.RECEIPT)
    transfers = models.StockDocument.objects.filter(
        contractor=OuterRef('pk'), kind=models.StockDocument.Kind.TRANSFER)
    return queryset.annotate(
        has_supply=(Exists(models.Procurement.objects.filter(contractor=OuterRef('pk')))
                    | Exists(models.Purchase.objects.filter(contractor=OuterRef('pk')))
                    | Exists(receipts)),
        has_shipment=Exists(transfers))


def _qty_by_uom(pairs):
    """`[(uom, qty)]` → `[{uom, qty}]` со сложением внутри единицы (по алфавиту).

    Складывать штуки с метрами в одно число — ложь, поэтому «сколько привезли» это
    вектор по единицам, а не скаляр. Тот же приём, что в мете формы склада.
    """
    total = {}
    for uom, qty in pairs:
        total[uom] = total.get(uom, ZERO) + qty
    return [{'uom': uom, 'qty': total[uom]} for uom in sorted(total)]


def _counterparty_receipt_lots(counterparty):
    """Лоты, привезённые контрагентом: партии его **зафиксированных** поставок.

    Гейт `origin__locked=True` — та же граница, что у «потрачено» проекта
    (`_project_spent`, волна 19 Ф15): черновой УПД ещё не факт, и в интеграл он не
    идёт. Список поставок в табе при этом показывает и черновики — интеграл считает
    состоявшееся, таб показывает всё.
    """
    return (models.Lot.objects
            .filter(origin__kind=models.StockDocument.Kind.RECEIPT,
                    origin__contractor=counterparty, origin__locked=True)
            .select_related('item'))


def _counterparty_supply(counterparty):
    """Интеграл закупочной стороны — «сколько у него куплено и что привёз».

    Три счёта контура (закупок-планов / заказов / поставок) + материальный итог
    (партий, штук по единицам, сумма). `open_purchases` — заказы, не закрытые
    поставками целиком (`purchase_coverage` ≠ ✓): единственное число панели, которое
    смотрит вперёд, а не назад. `None` — на этой стороне ни одного документа.
    """
    procurements = counterparty.procurements.count()
    purchases = list(counterparty.purchases.all())
    receipts = models.Receipt.objects.filter(contractor=counterparty)
    receipt_count = receipts.count()
    if not (procurements or purchases or receipt_count):
        return None
    lots = list(_counterparty_receipt_lots(counterparty))
    return {
        'procurements': procurements,
        'purchases': len(purchases),
        'open_purchases': sum(1 for p in purchases
                              if purchase_coverage(p) != 'available'),
        'receipts': receipt_count,
        # Расхождение «поставок 4, а привёз 0 партий» законно и требует объяснения:
        # черновые УПД считаются документами, но не фактом (Ф15). Отдаём их числом —
        # вью подписывает панель, вместо того чтобы пользователь читал это как баг.
        'draft_receipts': receipts.filter(locked=False).count(),
        'lots': len(lots),
        'qty_by_uom': _qty_by_uom((lot.item.uom, lot.qty) for lot in lots),
        'total': sum((lot.qty * lot.unit_cost for lot in lots), ZERO),
    }


def _counterparty_shipment(counterparty):
    """Интеграл передачной стороны — «сколько ему отдано» (роль заказчика).

    Считается по строкам **зафиксированных** передач (тот же гейт Ф15, что у
    закупочной стороны). `qty` строки знаковая (− расход) → берём магнитуду; деньги —
    по цене лота-источника (единственная известная цена изделия: своей цены у передачи
    нет, документооборот с заказчиком живёт вне PLM). `None` — передач не было.
    """
    lines = list(models.StockLine.objects
                 .filter(document__kind=models.StockDocument.Kind.TRANSFER,
                         document__contractor=counterparty, document__locked=True)
                 .select_related('lot__item'))
    transfers = models.Transfer.objects.filter(contractor=counterparty)
    if not transfers.exists():
        return None
    return {
        'transfers': transfers.count(),
        'draft_transfers': transfers.filter(locked=False).count(),
        'lots': len({line.lot_id for line in lines}),
        'qty_by_uom': _qty_by_uom((line.lot.item.uom, -line.qty) for line in lines),
        'total': sum((-line.qty * line.lot.unit_cost for line in lines), ZERO),
    }


def _procurement_rows(queryset):
    """Строки ленты закупок-планов — таб «Закупки» ЛЮБОЙ формы, где эта лента нужна.

    Волна 21: функция берёт **queryset**, а не контрагента. У формы контрагента лента
    отвечает «у кого собираемся купить», у формы аккаунта — «что я планировал»; строка
    одна и та же, и второго словаря под тот же смысл заводить незачем.
    """
    rows = []
    for p in queryset.prefetch_related('lines').order_by('-id'):
        lines = list(p.lines.all())
        rows.append({
            'id': p.id, 'code': p.code, 'description': p.description,
            'date': p.date, 'locked': p.locked,
            'lines': len(lines), 'qty': sum((ln.qty for ln in lines), ZERO),
        })
    return rows


def _purchase_rows(queryset):
    """Строки ленты заказов — таб «Заказы» любой формы (контрагент, аккаунт).

    Закрытость строки красит тем же словарём ✓/●/▲, что список режима «Заказы»
    (`purchase_coverage`) — знак выбирает вью.
    """
    rows = []
    qs = (queryset.select_related('project')
          .prefetch_related('lines').order_by('-id'))
    for p in qs:
        lines = list(p.lines.all())
        rows.append({
            'id': p.id, 'code': p.code, 'description': p.description,
            'date': p.date, 'locked': p.locked,
            'project_code': p.project.code,
            'lines': len(lines), 'qty': sum((ln.qty for ln in lines), ZERO),
            'coverage': purchase_coverage(p),
        })
    return rows


def _cp_receipt_rows(counterparty):
    """Таб «Поставки»: УПД, которые привёз этот контрагент («кто привёз», Ф17)."""
    rows = []
    qs = (models.Receipt.objects.filter(contractor=counterparty)
          .select_related('project').prefetch_related('lots__item').order_by('-id'))
    for r in qs:
        lots = list(r.lots.all())
        rows.append({
            'id': r.id, 'code': r.code, 'number': r.number, 'date': r.date,
            'locked': r.locked, 'project_code': r.project.code,
            'purchase_id': r.purchase_id,
            'lots': len(lots),
            'total': sum((lot.qty * lot.unit_cost for lot in lots), ZERO),
        })
    return rows


def _cp_transfer_rows(counterparty):
    """Таб «Передачи»: накладные, которыми ему отдавали (роль заказчика)."""
    rows = []
    qs = (models.Transfer.objects.filter(contractor=counterparty)
          .select_related('project').prefetch_related('lines__lot__item').order_by('-id'))
    for t in qs:
        lines = list(t.lines.all())
        rows.append({
            'id': t.id, 'code': t.code, 'number': t.number, 'date': t.date,
            'locked': t.locked, 'project_code': t.project.code,
            'lines': len(lines),
            'qty': sum((-ln.qty for ln in lines), ZERO),
            'total': sum((-ln.qty * ln.lot.unit_cost for ln in lines), ZERO),
        })
    return rows


def counterparty_form(counterparty):
    """Проекция формы контрагента (волна 20): ДНК + два интеграла + четыре списка.

    Аккордеона «закупка → накладные» здесь нет (решение Ивана 2026-07-30, отмена
    формулировки семени): контрагент — не документ контура, а его сторона, и вложение
    уровней врало бы про путь. Заказ живёт без плана (Ф17), поставка — без заказа,
    поэтому уровни идут **тремя равными табами**, а связь «чем закрыто» остаётся там,
    где она однозначна — в форме заказа (Ф6).
    """
    return {
        'id': counterparty.id, 'code': counterparty.code,
        'description': counterparty.description, 'inn': counterparty.inn,
        # Стороны: интеграл или `None` («движений нет» решает движок, не вью). Отдельных
        # `has_*` здесь нет намеренно — «сторона есть» это и есть «интеграл не `None`»,
        # и два источника одной правды в одной проекции разошлись бы (Ф3).
        'supply': _counterparty_supply(counterparty),
        'shipment': _counterparty_shipment(counterparty),
        'procurements': _procurement_rows(counterparty.procurements),
        'purchases': _purchase_rows(counterparty.purchases),
        'receipts': _cp_receipt_rows(counterparty),
        'transfers': _cp_transfer_rows(counterparty),
    }


def create_counterparty(code=None, description='', inn=''):
    """Завести контрагента. Пустой код — фолбэком «Контрагент 12» (Ф12e).

    Волна 19 (Ф10) оставила `code` контрагента без авто-фолбэка: заводили его только
    из пикера, где человек вводил имя. Волна 20 дала сущности форму, а форма рождается
    по клику (Ф12e) — титулу нужен код, и правило снова одно на всех.

    Роли при рождении больше нет (Ф3): контрагента заводят из пикера конкретного
    документа, и «кто он» этот документ и скажет — первым же фактом.
    """
    code = (code or '').strip() or None
    require_unique_code(models.Counterparty, code)
    fields = dict(description=(description or '').strip(), inn=(inn or '').strip())
    if code:
        return models.Counterparty.objects.create(code=code, **fields)
    return create_with_fallback_code(models.Counterparty, 'Контрагент', **fields)


def update_counterparty(counterparty, code=_UNSET, description=None, inn=None):
    """Правка ДНК контрагента под интерфейсным замком формы (волна 20).

    ДНК теперь ровно три поля — код, описание, ИНН (роль снесена, Ф3): всё остальное,
    что справочник «знает» о контрагенте, он знает из документов и правке не подлежит.

    `code` — часовой `_UNSET` (не прислали → не трогаем; пустой → NULL, как у
    документов: колонка nullable, и очистка кода не должна ловить IntegrityError).
    `description`/`inn` пустыми быть вправе (Ф12e: рождённое по клику незаполнено, и
    запрет очистки означал бы «заполнить можно, передумать нельзя»).
    """
    if code is not _UNSET:
        require_unique_code(models.Counterparty, code, counterparty.pk)
        counterparty.code = (code or '').strip() or None
    if description is not None:
        counterparty.description = description.strip()
    if inn is not None:
        counterparty.inn = inn.strip()
    counterparty.save()
    return counterparty


def delete_counterparty(counterparty):
    """Удалить контрагента (долг, приехавший вместе с формой; волна 20).

    До этой волны у контрагента не было пути удаления **вообще** — ни эндпойнта, ни
    функции движка, только админка. Друзья-guard'ы:
    — поставки/передачи (`StockDocument.contractor`, PROTECT) и заказы
      (`Purchase.contractor`, PROTECT) держат наглухо: документ без стороны не бывает;
    — закупки-планы (`Procurement.contractor`, SET_NULL) НЕ держат — так решено в Ф17
      (план-черновик не должен падать из-за справочника), поле просто опустеет;
    — вложения («карточка предприятия», Ф12b) сносим ЯВНО: каскад БД унёс бы строки и
      оставил файлы сиротами на диске (та же беда, что у `delete_project`).
    """
    if counterparty.documents.exists():
        raise ValidationError(
            'У контрагента есть поставки или передачи — удаление заблокировано.')
    if counterparty.purchases.exists():
        raise ValidationError('На контрагента оформлены заказы — удаление заблокировано.')
    for att in counterparty.attachments.all():   # физические файлы (каскад их сиротит)
        delete_attachment(att)
    try:
        counterparty.delete()
    except ProtectedError:
        raise ValidationError('Контрагент связан с документами — удаление заблокировано.')


# --------------------------------------------------------------------------- #
#  Волна 9 — инвентаризация (Inventory): 4-й origin партии + серая ре-материализация
# --------------------------------------------------------------------------- #
# `Inventory` рождает «найденные» партии — излишки, всплывшие при пересчёте, и
# ре-материализацию серых остатков (списанное → −ISSUE «в серый»; найдено физически
# → возвращаем на баланс новым лотом с `predecessor` → списанный, наследуя
# item/цену/название/зав.№). Отдельной `InventoryLine` в модели нет: строки акта =
# его лоты (`inventory.lots`, как приход/УПД). Origin `inventory` несёт единый
# `Lot.origin` (Ф2b) и знает `rebuild_movements` — волна добавила записываемую надстройку.
# Замка нет (у модели нет поля-статуса, как у Writeoff/Requisition): правимо всегда,
# корректность держат guard'ы + PROTECT.
def inventory_form(inventory):
    """Проекция формы инвентаризации: шапка акта + строки-лоты (`+RECEIPT`) + итог.

    Каждая строка — рождённый актом лот («найденная» партия): кол-во, живой остаток
    (просел ли под последующий расход), цена/название, зав.№ и провенанс
    (`predecessor` — из какого списанного лота ре-материализован). Чистая проекция.
    """
    lots = []
    total = ZERO
    for lot in (inventory.lots.select_related('item', 'predecessor__project')
                .order_by('id')):
        total += lot.qty * lot.unit_cost
        pred = lot.predecessor
        lots.append({
            'id': lot.id, 'item_id': lot.item_id, 'item_code': lot.item.code,
            'item_description': lot.item.description, 'uom': lot.item.uom,
            'qty': lot.qty, 'live_qty': lot_live_qty(lot),
            'unit_cost': lot.unit_cost, 'lot_name': lot.lot_name,
            'part_number': lot.part_number,
            'predecessor_id': lot.predecessor_id,
            'predecessor_label': _lot_label(pred) if pred else '',
            'consumed': _lot_consumed_downstream(lot),
        })
    return {
        'id': inventory.id, **_author(inventory), 'number': inventory.number, 'date': inventory.date,
        'code': inventory.code, 'description': inventory.description,
        'project_id': inventory.project_id, 'project_code': inventory.project.code,
        'project_name': inventory.project.description, 'locked': inventory.locked,
        'total_cost': total, 'lots': lots,
    }


def create_inventory(project, user, number='', date=None):
    """Создать акт инвентаризации в проект-дом (куда рождаются найденные лоты).
    `description` (Ф10) заполняется в детальной форме под замком, не при создании."""
    return _born_order(models.Inventory, project, user, number, date)


def add_inventory_lot(inventory, item, qty, unit_cost=ZERO, lot_name='',
                      part_number='', predecessor=None):
    """Добавить строку акта: рождается «найденная» партия (`+RECEIPT`) в его проекте.

    `predecessor` (опц.) связывает найденный лот со списанным-источником
    (ре-материализация серого остатка — провенанс/генеалогия). Кол-во не клампим.
    """
    _require_unlocked(inventory)
    if qty is None or qty <= 0:
        raise ValidationError('Количество должно быть положительным.')
    if unit_cost is not None and unit_cost < 0:
        raise ValidationError('Цена не может быть отрицательной.')
    lot = models.Lot.objects.create(
        item=item, project=inventory.project, origin=inventory, qty=qty,
        unit_cost=unit_cost or ZERO, lot_name=lot_name or '',
        part_number=part_number or '', predecessor=predecessor)
    rebuild_movements(lot)
    return lot


def update_inventory_lot(lot, qty=None, unit_cost=None, lot_name=None,
                         part_number=None):
    """Автосейв строки акта (кол-во/цена/название/PN). Кол-во не клампим по расходу.
    Только черновик (замок)."""
    _require_unlocked(lot.origin)
    fields = []
    if qty is not None:
        if qty <= 0:
            raise ValidationError('Количество должно быть положительным.')
        lot.qty = qty
        fields.append('qty')
    if unit_cost is not None:
        if unit_cost < 0:
            raise ValidationError('Цена не может быть отрицательной.')
        lot.unit_cost = unit_cost
        fields.append('unit_cost')
    if lot_name is not None:
        lot.lot_name = lot_name
        fields.append('lot_name')
    if part_number is not None:
        lot.part_number = part_number
        fields.append('part_number')
    if fields:
        lot.save(update_fields=fields)
        rebuild_movements(lot)
    return lot


def remove_inventory_lot(lot):
    """Удалить строку акта (коррекция). Guard: черновик + найденный лот не потреблён ниже."""
    _require_unlocked(lot.origin)
    if _lot_consumed_downstream(lot):
        raise ValidationError(
            'Найденная партия уже потреблена ниже — удаление заблокировано.')
    lot.movements.all().delete()
    lot.delete()


def lock_inventory(inventory):
    """Провести инвентаризацию (замок, форма read-only)."""
    return lock_document(inventory, inventory.lots,
                         'Нельзя провести пустой акт инвентаризации — добавьте строку.')


def unlock_inventory(inventory):
    """Снять замок инвентаризации — снова разрешить правку."""
    return unlock_document(inventory)


def update_inventory(inventory, **kw):
    """Шапка инвентаризации: № акта / дата / код / описание / автор / проект.

    Ф14: заодно приведена к общей дороге — своя рукописная обработка `code` (мимо
    `require_unique_code`, единственная такая среди шести) была недосмотром, а не
    решением; теперь код инвентаризации стережёт та же мягкая уникальность, что у
    остальных ордеров."""
    return update_document(inventory, **kw)


def written_off_lots():
    """Списанные лоты (серый путь) — кандидаты ре-материализации инвентаризацией.

    Списание — чистый `−ISSUE`: лот покинул учёт «в серый». Если серую партию нашли
    физически, инвентаризация возвращает её на баланс лотом-потомком (`predecessor` →
    списанный, наследование item/цены/названия/зав.№). Показываем суммарно списанное
    с лота (сколько «серого» доступно вернуть).

    Ф15: считаем только по **зафиксированным** актам — черновик списания это
    намерение, а не факт, и предлагать вернуть ещё не списанное значило бы делать
    склад из воздуха. Та же граница, что у остальных интегралов движка.
    """
    result = []
    wo = models.StockDocument.Kind.WRITEOFF
    for lot in (models.Lot.objects
                .filter(stock_lines__document__kind=wo,
                        stock_lines__document__locked=True).distinct()
                .select_related('item', 'project').order_by('project__code',
                                                            'item__code', 'id')):
        # qty знаковый (− расход) → магнитуда списанного = −Σ
        written = -(lot.stock_lines.filter(document__kind=wo, document__locked=True)
                    .aggregate(s=Sum('qty'))['s'] or ZERO)
        result.append({
            'lot_id': lot.id, 'item_id': lot.item_id,
            'item_code': lot.item.code, 'item_description': lot.item.description,
            'uom': lot.item.uom, 'written_qty': written,
            'project_code': lot.project.code, 'unit_cost': lot.unit_cost,
            'lot_name': lot.lot_name, 'part_number': lot.part_number,
        })
    return result


# --------------------------------------------------------------------------- #
#  Справочники: создание изделий и проектов (канон «＋ Новая», 2026-07-03)
# --------------------------------------------------------------------------- #
def _resolve_category(category_id):
    """Категория изделия из справочника по PK (обязательна). Волна 15: `kind`-enum
    сменён FK-справочником `Category`."""
    if not category_id:
        raise ValidationError('Нужно выбрать категорию изделия.')
    try:
        return models.Category.objects.get(pk=category_id)
    except (models.Category.DoesNotExist, ValueError, TypeError):
        raise ValidationError('Неизвестная категория изделия.')


def item_is_used(item):
    """Изделие «используется» = есть хотя бы одна живая ссылка на него. Волна 15:
    заменяет снятое хранимое `active` вычисляемым признаком (спящий = 0 ссылок =
    кандидат на удаление). Зеркалит guard'ы `delete_item` (used ⇔ неудаляемо):
    вхождение в чужой BOM, лоты, строки заказа/закупки-плана, потребность проекта,
    цель комплектации. Дёшево — набор `Exists` с коротким замыканием на `or`."""
    return (item.used_in.exists() or item.lots.exists()
            or item.purchase_lines.exists() or item.demanded_in.exists()
            or item.kittings.exists()
            or models.ProcurementLine.objects.filter(item=item).exists())


def create_item(code=None, description='', category_id=None, uom='шт',
                native=False, estimated_cost=None, temperature=''):
    """Создать изделие справочника. `code` (заказной PN, канон библиотеки — колонка
    CSV «Design Item Id») уникален.

    Волна 19, Ф12e: рождается по клику «＋ Новое», поэтому пустым может быть всё —
    код добирается фолбэком «Изделие 12», описание остаётся пустым, категория
    остаётся `NULL` и требуется только к фиксации (`lock_item`). Ручное изделие
    рождается `synced=False` (руками, не из библиотеки)."""
    code = (code or '').strip()
    require_unique_code(models.Item, code)
    fields = dict(
        description=(description or '').strip(),
        category=_resolve_category(category_id) if category_id else None,
        uom=(uom or '').strip() or 'шт',
        temperature=(temperature or '').strip(),
        native=bool(native),
        estimated_cost=estimated_cost)
    if code:
        return models.Item.objects.create(code=code, **fields)
    return create_with_fallback_code(models.Item, 'Изделие', **fields)


def _require_item_unlocked(item):
    """Гейт фиксации (волна 17): мутации свойств/состава запрещены у зафиксированного
    изделия — сперва расфиксировать. Защищает библиотечные изделия от ручного
    дрейфа (библиотека = источник правды). Ловит и прямой запрос к API, не только
    UI. Синк библиотеки правит в обход (свои прямые ORM-операции)."""
    if item.locked:
        raise ValidationError(
            f'Изделие {item.code} зафиксировано — сперва расфиксируйте.')


def lock_item(item):
    """Зафиксировать изделие: форма становится read-only, мутации
    гейтятся. Идемпотентно.

    Ф12e: категория обязательна ЗДЕСЬ, а не при рождении (изделие заводится по
    клику пустым). Дружелюбный отказ до CHECK `item_locked_has_category` — тот
    остаётся страховкой от прямого ORM/админки."""
    if item.category_id is None:
        raise ValidationError('Перед фиксацией выберите категорию изделия.')
    if not item.locked:
        item.locked = True
        item.save(update_fields=['locked'])
    return item


def unlock_item(item):
    """Расфиксировать изделие: снова редактируемо. Идемпотентно.
    У библиотечного изделия следующий синк вернёт замок."""
    if item.locked:
        item.locked = False
        item.save(update_fields=['locked'])
    return item


def update_item(item, changes):
    """Правка свойств изделия под замком формы (§6). `changes` — только присланные
    поля (частичный PATCH). `code` уникален; категория из справочника
    (ключ `category_id`); описание непустое.

    Матрица `synced × locked` (Ф3a): зафиксированное (`locked`) — read-only целиком;
    библиотечное расфиксированное (`synced`) — правится ТОЛЬКО оценочная стоимость
    (остальные поля приходят из библиотеки, руками не трогаем); ручное расфиксированное
    — правится всё. `synced` руками не переключается (ставит только синк)."""
    _require_item_unlocked(item)
    if item.synced:
        illegal = set(changes) - {'estimated_cost'}
        if illegal:
            raise ValidationError(
                f'Изделие {item.code} — из библиотеки: правится только '
                f'оценочная стоимость (остальные поля берутся из библиотеки).')
    fields = []
    if 'code' in changes:
        v = (changes['code'] or '').strip()
        if not v:
            raise ValidationError('Нужен код изделия.')
        require_unique_code(models.Item, v, item.pk)
        item.code = v
        fields.append('code')
    if 'description' in changes:
        # Ф12e: пустое описание легально (см. `update_location`) — идентичность
        # держит `code`, а изделие рождается по клику незаполненным.
        item.description = (changes['description'] or '').strip()
        fields.append('description')
    if 'category_id' in changes:
        # Пустая категория легальна в черновике; гейт — на фиксации (`lock_item`).
        item.category = (_resolve_category(changes['category_id'])
                         if changes['category_id'] else None)
        fields.append('category')
    if 'uom' in changes:
        item.uom = (changes['uom'] or '').strip() or 'шт'
        fields.append('uom')
    if 'temperature' in changes:
        item.temperature = (changes['temperature'] or '').strip()
        fields.append('temperature')
    if 'estimated_cost' in changes:
        item.estimated_cost = changes['estimated_cost']    # Decimal или None (сброс)
        fields.append('estimated_cost')
    if 'native' in changes:
        item.native = bool(changes['native'])
        fields.append('native')
    if fields:
        item.save(update_fields=fields)
    return item


def delete_item(item):
    """Удалить изделие из справочника (WAVE14 Ф2). Friendly-guard переводит `PROTECT`
    в человеческий отказ вместо 500: изделие держат партии, вхождение в чужой BOM,
    потребность проекта, строки заказа/закупки, цель комплектации. Свой состав (строки
    BOM, где изделие — parent) и вложения — каскад; файлы вложений сносим явно, иначе
    каскад БД осиротит их на диске (как в `delete_stock_document`).

    Гейт фиксации (волна 17): зафиксированное изделие руками не удаляют — сперва
    расфиксировать (UI прячет удаление у зафиксированного). Синк библиотеки удаляет
    ушедшие из библиотеки изделия в обход — расфиксирует их перед `delete_item`."""
    _require_item_unlocked(item)
    if item.lots.exists():
        raise ValidationError('У изделия есть партии на складе — удаление заблокировано.')
    if item.used_in.exists():
        raise ValidationError('Изделие входит в состав других изделий — удаление заблокировано.')
    if item.demanded_in.exists():
        raise ValidationError('На изделие есть потребность проекта — удаление заблокировано.')
    if item.purchase_lines.exists():
        raise ValidationError('Изделие есть в заказах — удаление заблокировано.')
    if item.kittings.exists():
        raise ValidationError('Изделие — цель комплектации — удаление заблокировано.')
    if models.ProcurementLine.objects.filter(item=item).exists():
        raise ValidationError('Изделие есть в закупках-планах — удаление заблокировано.')
    for att in item.attachments.all():             # физические файлы (каскад их сиротит)
        delete_attachment(att)
    try:
        item.delete()                              # каскад: свои строки BOM (parent)
    except ProtectedError:
        raise ValidationError('Изделие связано с другими записями — удаление заблокировано.')


# --------------------------------------------------------------------------- #
#  Синхронизация справочника с библиотекой компонентов Altium (волна 15)
# --------------------------------------------------------------------------- #
# Внешняя библиотека — источник правды по покупным изделиям. Форма грузит её
# CSV-таблицы (мульти-файл = вся библиотека за раз); движок парсит → диф против БД
# по ключу `code` → применение подтверждённых строк. Цену библиотека не
# хранит (`estimated_cost` — собственность Plume, синк её не трогает).
LIBRARY_ENCODING = 'cp1251'      # CP1251, разделитель ';' без экранирования, LF
LIBRARY_KEY_COL = 'Design Item Id'
LIBRARY_DESC_COL = 'Description'
LIBRARY_TEMP_COL = 'Temperature'


def _category_code_from_filename(name):
    """Класс компонента = стем имени файла (`csv/capacitors.csv` → `capacitors`)."""
    base = (name or '').rsplit('/', 1)[-1].rsplit('\\', 1)[-1]
    return base.rsplit('.', 1)[0].strip().lower()


def parse_library_file(filename, raw):
    """Разобрать один CSV библиотеки → (category_code, [row...]). `raw` — bytes
    (CP1251). Заголовок = эталон: нужные колонки ищем по имени (терпимо к порядку),
    отсутствие любой из трёх — ошибка. Строки нормализуем в
    `{code, description, temperature, category}`; дубль ключа в файле,
    пустой ключ или недобор колонок — человеческий отказ (не молчим)."""
    category = _category_code_from_filename(filename)
    if not category:
        raise ValidationError(f'{filename}: не удалось определить категорию по имени файла.')
    try:
        text = bytes(raw).decode(LIBRARY_ENCODING)
    except (UnicodeDecodeError, TypeError):
        raise ValidationError(f'{filename}: ожидалась кодировка CP1251.')
    reader = csv.reader(io.StringIO(text), delimiter=';')
    try:
        header = next(reader)
    except StopIteration:
        raise ValidationError(f'{filename}: пустой файл.')
    cols = {h.strip(): i for i, h in enumerate(header)}
    for col in (LIBRARY_KEY_COL, LIBRARY_DESC_COL, LIBRARY_TEMP_COL):
        if col not in cols:
            raise ValidationError(f'{filename}: нет колонки «{col}».')
    ki, di, ti = cols[LIBRARY_KEY_COL], cols[LIBRARY_DESC_COL], cols[LIBRARY_TEMP_COL]
    rows, seen = [], set()
    for lineno, rec in enumerate(reader, start=2):
        if not rec or all(not c.strip() for c in rec):
            continue                       # пустая строка (в т.ч. финальный LF)
        if len(rec) <= max(ki, di, ti):
            raise ValidationError(f'{filename}, строка {lineno}: мало колонок.')
        key = rec[ki].strip()
        if not key:
            raise ValidationError(f'{filename}, строка {lineno}: пустой Design Item Id.')
        if key in seen:
            raise ValidationError(f'{filename}: дубль Design Item Id «{key}».')
        seen.add(key)
        rows.append({'code': key, 'description': rec[di].strip(),
                     'temperature': rec[ti].strip(), 'category': category})
    return category, rows


def ensure_library_categories(codes):
    """Завести недостающие категории загруженной библиотеки → список `Category`
    всех перечисленных кодов (волна 22).

    **Сверка ЗАПИСЫВАЕТ ровно это и ничего больше.** До волны 22 категория рождалась
    только внутри `apply_library_diff` (побочно, при создании первого изделия класса),
    и до применения править её описание было негде — а таб «Категории» экрана
    синхронизации существует именно ради этого: описание класса задаётся из продукта,
    а не только из админки. Класс без изделий безобиден: это строка справочника, не
    документ и не движение, — поэтому «диф без записи» ослаблен здесь осознанно."""
    return [ensure_category(code) for code in codes]


def parse_library(files):
    """Разобрать мульти-файл загрузку. `files` — список `(filename, raw_bytes)`.
    Возвращает `{'categories': [code...], 'rows': [row...]}`. Одна категория дважды
    или один `code` в двух файлах — ошибка (в библиотеке ключ глобально
    уникален). Категории нужны для scoping «пропавших» (сверяем только загруженные
    классы)."""
    if not files:
        raise ValidationError('Не выбрано ни одного файла.')
    categories, rows, owner = [], [], {}
    for filename, raw in files:
        cat, file_rows = parse_library_file(filename, raw)
        if cat in categories:
            raise ValidationError(f'Категория «{cat}» пришла дважды (файл {filename}).')
        categories.append(cat)
        for r in file_rows:
            key = r['code']
            if key in owner:
                raise ValidationError(
                    f'Design Item Id «{key}» есть в двух файлах ({owner[key]} и {cat}).')
            owner[key] = cat
        rows.extend(file_rows)
    return {'categories': categories, 'rows': rows}


# Порядок статусов в диф-вью: сперва требующие действия, потом флаги, потом «совпало».
_DIFF_ORDER = {'new': 0, 'changed': 1, 'mark': 2, 'gone': 3, 'orphan': 4, 'same': 5}


def _category_code(item):
    """Код категории изделия или `None` (Ф12e: у черновика её может не быть)."""
    return item.category.code if item.category_id else None


def _library_changes(item, row):
    """Что изменилось у существующего изделия против библиотеки: сравниваем только
    синкаемые поля (`description`/`category`/`temperature`). Категорию — по `code`
    (стем файла). `estimated_cost`/`uom`/`native` — собственность Plume, не сверяем."""
    changes = {}
    if item.description != row['description']:
        changes['description'] = {'old': item.description, 'new': row['description']}
    # Ф12e: у изделия-черновика категории может не быть — тогда это тоже
    # изменение («было пусто → станет `capacitors`»), а не падение.
    if _category_code(item) != row['category']:
        changes['category'] = {'old': _category_code(item), 'new': row['category']}
    if item.temperature != row['temperature']:
        changes['temperature'] = {'old': item.temperature, 'new': row['temperature']}
    return changes


def _diff_row(status, row, item, changes=None):
    out = {'status': status,
           'code': (row or {}).get('code') or item.code,
           'item_id': item.id if item else None}
    if row is not None:
        out['incoming'] = {'description': row['description'],
                           'temperature': row['temperature'], 'category': row['category']}
    if item is not None:
        out['current'] = {'description': item.description,
                          'temperature': item.temperature,
                          'category': _category_code(item),
                          'native': item.native, 'synced': item.synced,
                          'locked': item.locked}
    if changes:
        out['changes'] = changes
    return out


def library_diff(parsed):
    """Полная сверка загруженной библиотеки против БД по ключу `code`.
    `parsed` — из `parse_library`. Возвращает список диф-строк со статусом:

    - `new`     — ключа нет в БД → создать (`native=false`, `synced=true`, `locked=false`);
    - `changed` — есть, отличается description/category/temperature → обновить
      (+ пометить `synced`, снять замок);
    - `mark`    — есть, содержимое совпадает, но ещё не помечено `synced` → пометить
      библиотечным (`synced=true`, замок снять). Путь бэкфилла после Ф3a: все
      существующие библиотечные приходят `synced=false` и помечаются первым же синком;
    - `same`    — есть, совпадает И уже `synced` → ничего;
    - `gone`    — в БД (в одном из загруженных классов), нет в загрузке, не
      используется → кандидат на удаление;
    - `orphan`  — то же, но используется (живые ссылки) → флаг «сирота, нет в
      библиотеке» (не действие: удалить нельзя, обновлять нечем).

    Замок синк снимает всюду, где метит `synced` — этого требует инвариант 0008
    (`synced ⟹ not locked`): библиотечное защищено матрицей «правь только цену», а
    не фиксацией, и стухший замок на нём был бы вторым, противоречащим словарём.
    (Раньше здесь стояло «`locked` не трогаем» — докстринг отстал от инварианта,
    аудит-1 Б1а-6.)

    Scoping «пропавших» — по категории: изделие из класса, которого нет среди
    загруженных файлов, не считается пропавшим (его библиотеку просто не грузили)."""
    by_key = {r['code']: r for r in parsed['rows']}
    categories = set(parsed['categories'])
    existing = {i.code: i
                for i in models.Item.objects.select_related('category')}
    result = []
    for key, row in by_key.items():
        item = existing.get(key)
        if item is None:
            result.append(_diff_row('new', row, None))
            continue
        changes = _library_changes(item, row)
        if changes:
            status = 'changed'
        elif not item.synced:
            status = 'mark'       # совпадает по содержимому, но ещё не помечено библиотечным
        else:
            status = 'same'
        result.append(_diff_row(status, row, item, changes))
    for key, item in existing.items():
        # Изделие без категории (Ф12e, черновик) не принадлежит ни одному
        # синкаемому классу — сиротой библиотеки объявлять его не за что.
        if key in by_key or _category_code(item) not in categories:
            continue
        result.append(_diff_row('orphan' if item_is_used(item) else 'gone', None, item))
    result.sort(key=lambda d: (_DIFF_ORDER[d['status']], d['code']))
    return result


def apply_library_diff(parsed, confirmed):
    """Применить подтверждённые строки дифа. `confirmed` — множество
    `code`, отмеченных галочкой. Диф пересчитываем здесь заново (не
    доверяем присланным клиентом значениям): действие берётся из свежего статуса,
    поля — из `parsed`. Всё в одной транзакции (bulk = всё-или-ничего).

    - `new`     → создать `Item` (`native=false`, `synced=true`, `locked=false`,
      категория `ensure_category`);
    - `changed` → обновить description/category/temperature (+ пометить `synced`,
      снять замок — инвариант 0008 `synced ⟹ not locked`);
    - `mark`    → пометить библиотечным (`synced=true`, замок снять);
    - `gone`    → удалить (`delete_item` — guard добьёт, если стало используемым);
    - `orphan`/`same` → no-op даже если ключ подтверждён.

    Возвращает сводку `{created, updated, marked, deleted}`."""
    confirmed = set(confirmed or [])
    by_key = {r['code']: r for r in parsed['rows']}
    summary = {'created': 0, 'updated': 0, 'marked': 0, 'deleted': 0}
    with transaction.atomic():
        for diff in library_diff(parsed):
            key = diff['code']
            if key not in confirmed:
                continue
            status = diff['status']
            if status == 'new':
                src = by_key[key]
                # Библиотека = источник правды → изделие рождается библиотечным
                # (`synced`), но НЕ запертым (`locked=false`) — готово под ввод цены
                # (Ф3a, решение Ивана 2026-07-24). Защита остального — матрицей `synced`.
                models.Item.objects.create(
                    code=key, description=src['description'],
                    category=ensure_category(src['category']),
                    temperature=src['temperature'], native=False,
                    synced=True, locked=False)
                summary['created'] += 1
            elif status == 'changed':
                src = by_key[key]
                item = models.Item.objects.get(code=key)
                item.description = src['description']
                item.category = ensure_category(src['category'])
                item.temperature = src['temperature']
                # Синк подтверждает библиотечное происхождение → метим `synced` и
                # снимаем стухший замок (инвариант `synced ⟹ not locked`; библиотечное
                # защищено матрицей «правь только цену», а не фиксацией).
                item.synced = True
                item.locked = False
                item.save(update_fields=['description', 'category', 'temperature',
                                         'synced', 'locked'])
                summary['updated'] += 1
            elif status == 'mark':
                # Содержимое совпадает — метим библиотечным и снимаем стухший замок
                # (инвариант `synced ⟹ not locked`).
                item = models.Item.objects.get(code=key)
                item.synced = True
                item.locked = False
                item.save(update_fields=['synced', 'locked'])
                summary['marked'] += 1
            elif status == 'gone':
                item = models.Item.objects.get(code=key)
                unlock_item(item)          # расфиксировать перед удалением (гейт)
                delete_item(item)
                summary['deleted'] += 1
    return summary


# --------------------------------------------------------------------------- #
#  Роллап оценочной стоимости по BOM (волна 15)
# --------------------------------------------------------------------------- #
def _rollup_cost(item, cache, visiting, updated, incomplete):
    """Пост-order обход BOM: стоимость узла. Покупной лист → его `estimated_cost`
    (None → неполнота, вверх идёт как 0). Производимый узел → Σ(стоимость компонента
    × qty); результат пишем в `estimated_cost` ВСЕХ производимых узлов поддерева.
    `cache` мемоизирует общие поддеревья, `visiting` ловит циклы (страховка — их
    гасит уже `add_bom_line`)."""
    if item.id in cache:
        return cache[item.id]
    if not item.native:
        cost = item.estimated_cost
        if cost is None:
            incomplete.append(item.code)
        cache[item.id] = cost
        return cost
    if item.id in visiting:
        raise ValidationError(f'Цикл в составе: {item.code}.')
    visiting.add(item.id)
    lines = list(item.bom_lines.select_related('component'))
    if not lines:
        incomplete.append(item.code)   # производимый без состава — оценить нечем
    total = ZERO
    for bl in lines:
        c = _rollup_cost(bl.component, cache, visiting, updated, incomplete)
        total += (c if c is not None else ZERO) * bl.qty
    visiting.discard(item.id)
    if item.estimated_cost != total:
        item.estimated_cost = total
        item.save(update_fields=['estimated_cost'])
        updated.append(item.code)
    cache[item.id] = total
    return total


def rollup_estimated_cost(item):
    """Пересчитать оценочную стоимость производимого изделия роллапом по BOM
    (рекурсивно до листьев). Пишет оценку во все производимые узлы поддерева, не
    только в вершину (сменили цену листа → обновились и промежуточные платы, и
    прибор). Возвращает `{estimated_cost, updated, incomplete}`: `updated` — какие
    узлы переоценены, `incomplete` — листья/узлы без известной стоимости (учтены
    как 0, но помечены). Только для `native` (у покупного оценка — ручная)."""
    if not item.native:
        raise ValidationError('Пересчёт стоимости — только для производимого изделия.')
    updated, incomplete = [], []
    with transaction.atomic():
        cost = _rollup_cost(item, {}, set(), updated, incomplete)
    return {'estimated_cost': cost, 'updated': updated, 'incomplete': incomplete}


def create_project(code=None, description='', budget=None, started=None):
    """Создать внешний проект (НИР/контракт). Код уникален; пустой — фолбэком
    «Проект 12» (Ф12e: сущность рождается по клику, а не из формы создания).

    Только `kind=external`: внутренние склады (WHITE/GREY) — синглтоны из сида
    (`Project.clean`), кнопкой «＋ Новый» не заводятся.
    """
    code = (code or '').strip()
    if code and models.Project.objects.filter(code=code).exists():
        raise ValidationError(f'Проект с кодом {code} уже есть.')
    fields = dict(description=(description or '').strip(),
                  kind=models.Project.Kind.EXTERNAL,
                  budget=budget, started=started or None)
    if code:
        return models.Project.objects.create(code=code, **fields)
    return create_with_fallback_code(models.Project, 'Проект', **fields)


def update_project(project, changes):
    """Правка реквизитов проекта под замком формы (§6): код, описание, бюджет, дата начала.
    Статус (закрытие/переоткрытие) — отдельным путём, здесь не трогаем. Код правим всем
    проектам (WAVE14 Ф1): он не PK, переименование безопасно; guard как в update_item."""
    fields = []
    if 'code' in changes:
        code = (changes['code'] or '').strip()
        if not code:
            raise ValidationError('Нужен код проекта.')
        if models.Project.objects.filter(code=code).exclude(pk=project.pk).exists():
            raise ValidationError(f'Проект с кодом {code} уже есть.')
        project.code = code
        fields.append('code')
    if 'description' in changes:
        # Ф12e: пустое описание легально (см. `update_location`).
        project.description = (changes['description'] or '').strip()
        fields.append('description')
    if 'budget' in changes:
        project.budget = changes['budget']                 # Decimal или None (сброс)
        fields.append('budget')
    if 'started' in changes:
        project.started = changes['started'] or None
        fields.append('started')
    if fields:
        project.save(update_fields=fields)
    return project


def delete_project(project):
    """Удалить проект (WAVE14 Ф2) — только пустой (решение Ивана): внутренние склады
    неудаляемы (системные синглтоны); непустой проект уходит из жизни закрытием, не
    удалением. Держат: лоты, заказы, потребности; ссылку из ордеров (StockDocument.
    project PROTECT) ловит catch-all — переводим в человеческий отказ вместо 500."""
    if project.kind in models.Project.INTERNAL_KINDS:
        raise ValidationError('Внутренний склад удалять нельзя — это системный проект.')
    if project.lots.exists():
        raise ValidationError(
            'В проекте есть партии — удаление заблокировано; закройте проект закрывающими документами.')
    if project.purchases.exists():
        raise ValidationError('К проекту привязаны заказы — удаление заблокировано.')
    if project.demands.exists():
        raise ValidationError('В проекте есть потребности (приборы) — сперва уберите их.')
    for att in project.attachments.all():          # физические файлы (каскад их сиротит)
        delete_attachment(att)
    try:
        project.delete()
    except ProtectedError:
        raise ValidationError('Проект связан с документами — удаление заблокировано; закройте проект.')


# --------------------------------------------------------------------------- #
#  Потребность проекта (секция «Приборы» формы проекта): что и сколько делаем
# --------------------------------------------------------------------------- #
def _editable_project(project):
    """Потребность правится только у активного внешнего проекта (не склад, не закрыт)."""
    if project.kind in models.Project.INTERNAL_KINDS:
        raise ValidationError('У внутреннего склада нет потребностей.')
    if project.locked:
        raise ValidationError('Проект закрыт — переоткройте, чтобы править потребность.')


def add_project_demand(project, item, qty):
    """Добавить прибор в потребность проекта. Пара (проект, изделие) уникальна."""
    _editable_project(project)
    if qty is None or qty <= ZERO:
        raise ValidationError('Кол-во приборов должно быть больше нуля.')
    if models.ProjectDemand.objects.filter(project=project, target_item=item).exists():
        raise ValidationError(f'Прибор {item.code} уже в потребности проекта.')
    return models.ProjectDemand.objects.create(
        project=project, target_item=item, qty=qty)


def update_project_demand(demand, qty):
    """Правка кол-ва приборов в потребности (автосейв)."""
    _editable_project(demand.project)
    if qty is None or qty <= ZERO:
        raise ValidationError('Кол-во приборов должно быть больше нуля.')
    demand.qty = qty
    demand.save(update_fields=['qty'])
    return demand


def remove_project_demand(demand):
    """Убрать прибор из потребности проекта."""
    _editable_project(demand.project)
    demand.delete()


# --------------------------------------------------------------------------- #
#  Состав изделия / BOM (редактор на экране изделия)
# --------------------------------------------------------------------------- #
def _bom_would_cycle(parent, component):
    """True, если component (через свой BOM вглубь) содержит parent → цикл."""
    seen = set()
    stack = [component]
    while stack:
        cur = stack.pop()
        if cur.id == parent.id:
            return True
        if cur.id in seen:
            continue
        seen.add(cur.id)
        stack.extend(bl.component for bl in cur.bom_lines.select_related('component'))
    return False


def add_bom_line(parent, component, qty):
    """Добавить компонент в состав изделия. Без самоссылки, циклов и дублей.
    Гейт фиксации: у зафиксированного изделия состав не правят (волна 17).

    Аудит-1 (Б1а-3): `position` (позиционное обозначение — C1, R12) снят. Поле
    приговорил ещё FIELD_MATRIX: год оно жило в схеме, движке и API, но форма
    изделия его не показывала — а состав приезжает из библиотеки Altium, где
    позиции живут в схемотехнике, а не в PLM. Хранить то, чего никто не вводит и
    не видит, — не «задел», а расхождение схемы и вью.
    """
    _require_item_unlocked(parent)
    if qty is None or qty <= ZERO:
        raise ValidationError('Кол-во должно быть больше нуля.')
    if component.id == parent.id:
        raise ValidationError('Изделие не может входить само в себя.')
    if models.BomLine.objects.filter(parent=parent, component=component).exists():
        raise ValidationError(f'Компонент {component.code} уже в составе.')
    if _bom_would_cycle(parent, component):
        raise ValidationError(f'Цикл в составе: {component.code} уже содержит {parent.code}.')
    return models.BomLine.objects.create(
        parent=parent, component=component, qty=qty)


def update_bom_line(line, qty=None):
    """Правка строки состава (кол-во, автосейв). Гейт фиксации у изделия-
    владельца (волна 17)."""
    _require_item_unlocked(line.parent)
    if qty is not None:
        if qty <= ZERO:
            raise ValidationError('Кол-во должно быть больше нуля.')
        line.qty = qty
        line.save(update_fields=['qty'])
    return line


def remove_bom_line(line):
    """Убрать строку из состава изделия. Гейт фиксации у изделия-владельца (волна 17)."""
    _require_item_unlocked(line.parent)
    line.delete()


# --------------------------------------------------------------------------- #
#  Вложения (волна 11): PDF/сканы к документам и изделиям (exclusive-arc владелец)
# --------------------------------------------------------------------------- #
# Владелец вложения. API-контракт `owner_type` неизменён (стабильные строки:
# 'item' + виды ордера) — но после коллапса дуги (Ф2b) физических владельцев два:
# `Attachment.item` (изделие) и `Attachment.document` (ордер, любой вид). Разрешаем
# owner_type в КОНКРЕТНУЮ модель (строгая проверка «не найден»/несовпадение вида),
# а храним в `item` (для 'item') или `document` (для видов ордера).
ATTACHMENT_OWNER_MODELS = {
    'item': models.Item, 'receipt': models.Receipt, 'transfer': models.Transfer,
    'kitting': models.Kitting, 'inventory': models.Inventory,
    'writeoff': models.Writeoff, 'requisition': models.Requisition,
    'relocation': models.Relocation,
    # Волна 19, Ф12b: не-ордерные владельцы. Вид ордера остаётся именем в API
    # (`receipt`/`transfer`/…) и ложится в общее поле `document`; у этих —
    # своё поле, имя типа = имя поля.
    'project': models.Project, 'procurement': models.Procurement,
    'purchase': models.Purchase, 'counterparty': models.Counterparty,
}

# Имя типа в API → поле-владелец в `Attachment`. Совпадают у всех, кроме шести
# видов ордера: они схлопнуты в один FK `document` (волна 13, Ф2b).
_ATTACHMENT_OWNER_FIELD_BY_TYPE = {
    t: (t if t in models.ATTACHMENT_OWNER_FIELDS else 'document')
    for t in ATTACHMENT_OWNER_MODELS
}


def _attachment_owner_field(owner_type):
    """Поле-владелец под owner_type: 'project' → project; вид ордера → document."""
    return _ATTACHMENT_OWNER_FIELD_BY_TYPE[owner_type]


def resolve_attachment_owner(owner_type, owner_id):
    """Найти владельца по типу (имя из API) и id. Ошибка на неизвестный тип."""
    model = ATTACHMENT_OWNER_MODELS.get(owner_type)
    if model is None:
        raise ValidationError(f'Неизвестный тип владельца вложения: {owner_type}.')
    try:
        return model.objects.get(pk=owner_id)
    except model.DoesNotExist:
        raise ValidationError('Документ-владелец вложения не найден.')


# Файл на диске может разойтись с записью в БД — вручную удалили, перезалили мимо
# приложения, не доехал при переносе. Витрина показывает это ЦВЕТОМ глифа (волна 19,
# Ф12a): ok — совпадает, changed — на месте, но размер/время не те, missing — записи
# есть, файла нет. Допуск по времени: mtime записывается диском чуть позже, чем БД
# штампует `uploaded_at`.
_ATTACHMENT_MTIME_SLACK = timedelta(minutes=1)


def attachment_state(att):
    """Состояние файла на диске относительно записи в БД: ok | changed | missing."""
    try:
        path = att.file.path
        stat = os.stat(path)
    except (ValueError, OSError):
        return 'missing'
    if stat.st_size != att.size:
        return 'changed'
    if att.uploaded_at:
        touched = datetime.fromtimestamp(stat.st_mtime, tz=dt_timezone.utc)
        if touched > att.uploaded_at + _ATTACHMENT_MTIME_SLACK:
            return 'changed'
    return 'ok'


def attachment_row(att):
    """Проекция вложения для витрины (путь к файлу не отдаём — качаем эндпоинтом)."""
    return {
        'id': att.id, 'filename': att.filename or att.file.name,
        'size': att.size, 'content_type': att.content_type,
        'description': att.description, 'uploaded_at': att.uploaded_at,
        # Человеческое имя, а не логин: в списке «Загрузил» читают людей (Ф12a).
        # Тот же выбор, что у авторства документов (`_author`).
        'user': (att.user.get_full_name() or att.user.get_username()) if att.user_id else '',
        'state': attachment_state(att),
        'url': f'/api/attachments/{att.id}/download/',
    }


def attachments_for(owner_type, owner_id):
    """Список вложений владельца (свежие сверху)."""
    if owner_type not in ATTACHMENT_OWNER_MODELS:
        raise ValidationError(f'Неизвестный тип владельца вложения: {owner_type}.')
    field = _attachment_owner_field(owner_type)
    if field == 'document':
        # id ордеров глобально уникален (Ф2a) → document_id однозначен; фильтр по
        # kind сохраняет прежнюю строгость (несовпадение вида → пусто).
        flt = {'document_id': owner_id, 'document__kind': owner_type}
    else:
        flt = {f'{field}_id': owner_id}
    qs = (models.Attachment.objects.filter(**flt)
          .select_related('user').order_by('-id'))
    return [attachment_row(a) for a in qs]


def _fit(value, field_name):
    """Подрезать служебную строку вложения под длину её поля (лимит берём у модели)."""
    limit = models.Attachment._meta.get_field(field_name).max_length
    return (value or '')[:limit]


def add_attachment(owner_type, owner, upload, user, description=''):
    """Прикрепить файл к владельцу: файл на диск, метаданные из upload (не с клиента).

    filename/size/content_type заполняет сервер из загруженного файла. Владелец
    ровно один (exclusive arc item↔document) — поле задаётся по owner_type. Синхронно.

    Служебные строки (имя, MIME) приходят от браузера и подрезаются по длине поля:
    метаданные описывают файл, а не решают, приняли его или нет — длинный MIME не
    повод отказать в загрузке (грабля xlsx, 65 символов против прежних 64).
    """
    if owner_type not in ATTACHMENT_OWNER_MODELS:
        raise ValidationError(f'Неизвестный тип владельца вложения: {owner_type}.')
    if upload is None:
        raise ValidationError('Нужен файл вложения.')
    limit = settings.MAX_ATTACHMENT_SIZE
    if upload.size and upload.size > limit:
        raise ValidationError(f'Файл больше лимита ({limit // (1024 * 1024)} МБ).')
    att = models.Attachment(
        file=upload,
        filename=_fit(upload.name, 'filename'), size=upload.size or 0,
        content_type=_fit(getattr(upload, 'content_type', ''), 'content_type'),
        description=(description or '').strip(), user=user,
        **{_attachment_owner_field(owner_type): owner})
    att.full_clean(exclude=['file'])   # exclusive-arc + длины полей (file уже валиден)
    att.save()
    return att


def update_attachment(att, description=None):
    """Правка описания вложения. Метаданные файла неизменны."""
    if description is not None:
        att.description = (description or '').strip()
        att.save(update_fields=['description'])
    return att


def delete_attachment(att):
    """Удалить вложение: строку в БД и физический файл с диска."""
    att.file.delete(save=False)
    att.delete()


# --------------------------------------------------------------------------- #
#  Аккаунт: форма пользователя + тема интерфейса (волна 21)
# --------------------------------------------------------------------------- #
# Пользователь становится сущностью со своей каноничной формой. Три вещи, которые
# здесь важнее кода:
#
# 1. **Фиксации у пользователя нет.** `locked` в продукте значит «документ стал фактом,
#    движок дальше не даст его менять». Человек не документ — он не рождает факта,
#    который движок считает, и замораживать нечего. Остаётся личный замок ФОРМЫ
#    (интерфейсный, живёт во вью). Прецедент — контрагент (волна 20).
# 2. **Титул формы = `username`, и полем он не рисуется.** Пара `code`+`description`
#    не нарушена, а расщеплена по ДНК Django: идентичность — `username`, литературное
#    имя — `first_name`+`last_name` (`get_full_name` склеивает обратно). `username`
#    задаётся админкой, значит степенью свободы формы не является — и полем не бывает
#    (тот же приём, что снял `Item.native`).
# 3. **Про темы движок знает ровно слаг и список допустимых** (`models.THEMES`). Ярлык
#    («Тёмная»), палитра, набор CSS-файлов — знание ВЬЮ: `core/theme.ts` +
#    `themes/registry.ts`. Движок не знает даже, какая из тем светлая.


def profile_of(user):
    """Приставка настроек пользователя (`UserProfile`), рождаемая ЛЕНИВО.

    Сигнала на `post_save` пользователя намеренно нет: сигнал — магия на расстоянии
    ради экономии одной строки, и он ломает `loaddata` (снимки прода). Единственный
    вход к профилю — эта функция, поэтому «профиля ещё нет» состоянием продукта не
    является: первый же запрос его создаёт.
    """
    profile, _ = models.UserProfile.objects.get_or_create(user=user)
    return profile


def _document_rows(queryset):
    """Строки ленты ордеров — СМЕШАННЫЙ фид семи видов с колонкой типа.

    Один таб, а не семь (как список режима «Ордера»): вид ордера здесь колонка, а не
    раздел. Счётчиков в строке нет намеренно — общей меры у семи видов не существует
    (у поставки объём это партии, у передачи — строки), а счёт, который человеку нужен
    («сколько всего»), живёт в мете формы.
    """
    return [{
        'id': d.id, 'kind': d.kind, 'code': d.code, 'number': d.number,
        'description': d.description, 'date': d.date, 'locked': d.locked,
        'project_code': d.project.code,
    } for d in queryset.select_related('project').order_by('-id')]


def user_form(user):
    """Проекция формы аккаунта: ДНК Django + тема + три ленты «своих» документов.

    Ленты — теми же строками, что отдаёт форма контрагента (`_procurement_rows` /
    `_purchase_rows` берут queryset именно для этого). «Своих» — буквально: реверс
    `user.procurements` / `user.purchases` / `user.documents`, то есть авторство, а не
    права. Модели «кто чьи документы видит» в продукте нет и не нужно.

    Списка допустимых тем здесь нет: дропдаун наполняет `themes/registry.ts` (ярлыки —
    знание вью), а движок стережёт слаг в `set_theme`. Два источника одного словаря
    разошлись бы.
    """
    return {
        # ДНК Django. `username` идёт в титул формы и полем не рисуется (см. шапку).
        'username': user.username,
        'first_name': user.first_name,
        'last_name': user.last_name,
        'email': user.email,
        'full_name': user.get_full_name() or user.username,
        'theme': profile_of(user).theme,
        'procurements': _procurement_rows(user.procurements),
        'purchases': _purchase_rows(user.purchases),
        'documents': _document_rows(user.documents),
    }


def update_user(user, first_name=None, last_name=None, email=None):
    """Правка ДНК пользователя (имя / фамилия / почта) — под замком формы.

    Пустые значения законны (Django своё пустое имя и почту допускает), поэтому
    часового `_UNSET` здесь нет: `None` = «ключа в PATCH не было», пустая строка =
    «очистили», и запрет очистки означал бы «заполнить можно, передумать нельзя».
    """
    if first_name is not None:
        user.first_name = first_name.strip()
    if last_name is not None:
        user.last_name = last_name.strip()
    if email is not None:
        email = email.strip()
        if email:
            EmailValidator(message='Почта задана неверно.')(email)
        user.email = email
    user.save()
    return user


def set_theme(user, slug):
    """Единственный вход к теме интерфейса. Неизвестный слаг — отказ.

    Валидация живёт ЗДЕСЬ, а не `CheckConstraint`'ом в БД (единственное осознанное
    исключение из привычки продукта стеречь `choices` схемой): тема — это набор файлов
    вью, и требовать под новую тему миграцию значило бы вписать вью в схему.
    """
    slug = (slug or '').strip()
    if slug not in models.THEMES:
        raise ValidationError(f'Неизвестная тема интерфейса: «{slug}».')
    profile = profile_of(user)
    profile.theme = slug
    profile.save(update_fields=['theme'])
    return profile


def change_password(user, current, new, repeat):
    """Смена пароля: текущий → повтор → штатные валидаторы Django → запись.

    Своих правил стойкости не изобретаем — `AUTH_PASSWORD_VALIDATORS` уже настроены и
    те же, что у админки. Порядок проверок = порядок, в котором человек ошибается.
    Сессию после смены пароля поддерживает вьюха (`update_session_auth_hash`): это
    знание про HTTP-сессию, а не про пользователя.
    """
    if not user.check_password(current or ''):
        raise ValidationError('Текущий пароль неверен.')
    if not new:
        raise ValidationError('Новый пароль пуст.')
    if new != repeat:
        raise ValidationError('Новый пароль и повтор не совпадают.')
    validate_password(new, user)
    user.set_password(new)
    user.save(update_fields=['password'])
    return user
